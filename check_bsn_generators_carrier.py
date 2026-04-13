import openpyxl

wb = openpyxl.load_workbook('interface.xlsx')
ws = wb['지역_BSN']

print('='*70)
print('지역_BSN 시트 generators carrier 확인')
print('='*70)

# 발전기 섹션 찾기
for row in range(1, min(100, ws.max_row + 1)):
    cell = ws.cell(row=row, column=1)
    if cell.value and '발전기' in str(cell.value):
        print(f'\n발전기 섹션 찾음: 행 {row}')
        header_row = row + 2
        print(f'헤더 행: {header_row}')
        
        # 헤더 읽기
        headers = []
        for i in range(1, 30):
            h = ws.cell(row=header_row, column=i).value
            if h:
                headers.append(h)
            else:
                break
        
        print(f'\n헤더: {headers}')
        
        # carrier 컬럼 위치 찾기
        carrier_idx = None
        p_nom_idx = None
        name_idx = 1  # 이름은 첫 번째 컬럼
        
        for i, h in enumerate(headers):
            if '캐리어' in str(h) or 'carrier' in str(h).lower():
                carrier_idx = i + 1
            if '정격용량' in str(h) or 'p_nom' in str(h).lower():
                p_nom_idx = i + 1
        
        print(f'\n컬럼 위치: name={name_idx}, carrier={carrier_idx}, p_nom={p_nom_idx}')
        
        # 데이터 행 읽기
        print(f'\n발전기 데이터:')
        for data_row in range(header_row+1, min(header_row+20, ws.max_row+1)):
            name = ws.cell(row=data_row, column=name_idx).value
            if not name:
                break
            
            carrier = ws.cell(row=data_row, column=carrier_idx).value if carrier_idx else None
            p_nom = ws.cell(row=data_row, column=p_nom_idx).value if p_nom_idx else None
            
            print(f'  {name}: carrier={carrier}, p_nom={p_nom}')
        
        break

print('\n' + '='*70)
