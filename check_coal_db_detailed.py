import pandas as pd
import openpyxl

# openpyxl로 직접 읽기
wb = openpyxl.load_workbook('interface.xlsx', read_only=True)
ws = wb['Coal_DB']

# D2 셀의 선택연도
selected_year = ws['D2'].value
print(f"선택연도 (D2): {selected_year}")

# 상위 10행의 A, D, E, F~L, AM 열 확인
print("\n상위 10행 샘플:")
print(f"{'Row':<5} {'A열(지역)':<10} {'D열(발전소)':<20} {'E열(호기)':<10} {'F열':<10} {'G열':<10} {'AM열(효율)':<10}")
print("-" * 100)

for row in range(1, 11):
    a_val = ws.cell(row, 1).value  # A열
    d_val = ws.cell(row, 4).value  # D열
    e_val = ws.cell(row, 5).value  # E열
    f_val = ws.cell(row, 6).value  # F열 (2024년?)
    g_val = ws.cell(row, 7).value  # G열
    am_val = ws.cell(row, 39).value  # AM열
    
    print(f"{row:<5} {str(a_val):<10} {str(d_val):<20} {str(e_val):<10} {str(f_val):<10} {str(g_val):<10} {str(am_val):<10}")

# 헤더 행 찾기 - F열부터 AF열까지 연도 확인
print("\n\nF열~AF열의 1~5행 값 (연도 헤더 찾기):")
for row in range(1, 6):
    print(f"Row {row}: ", end="")
    for col in range(6, 32, 2):  # F, H, J, ... 2열씩 건너뛰며 샘플링
        val = ws.cell(row, col).value
        col_letter = chr(64 + col)
        print(f"{col_letter}={val}, ", end="")
    print()

wb.close()

# pandas로 skiprows를 조정하며 읽기
print("\n\n=== pandas로 읽기 (skiprows=3) ===")
df = pd.read_excel('interface.xlsx', sheet_name='Coal_DB', skiprows=3)
print(f"Shape: {df.shape}")
print(f"Columns: {df.columns.tolist()}")
print("\n상위 10행:")
print(df.head(10))
