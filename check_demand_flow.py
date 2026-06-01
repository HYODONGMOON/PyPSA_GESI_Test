import re

content = open('PyPSA_GUI.py', encoding='utf-8').read()
lines = content.split('\n')

# 1) _apply_scenario_to_loads_in_input 호출 위치
print('=== _apply_scenario_to_loads_in_input 호출 위치 ===')
for i, line in enumerate(lines, 1):
    if '_apply_scenario_to_loads_in_input' in line:
        print(f'  L{i}: {line.strip()[:120]}')

# 2) 시나리오_에너지수요 관련 코드
print('\n=== 시나리오_에너지수요 관련 코드 ===')
for i, line in enumerate(lines, 1):
    if '시나리오_에너지수요' in line or 'scenario_demand' in line.lower():
        print(f'  L{i}: {line.strip()[:120]}')

# 3) Demand 값 읽기 관련 (인터페이스 시트)
print('\n=== interface.xlsx Demand 읽기 관련 ===')
for i, line in enumerate(lines, 1):
    if ('인터페이스' in line and 'demand' in line.lower()) or ('SheetData' in line and 'load' in line.lower()):
        print(f'  L{i}: {line.strip()[:120]}')

# 4) p_set 설정 전체 흐름 (int_data['loads'] 관련)
print('\n=== input_data loads p_set 주입 흐름 ===')
for i, line in enumerate(lines, 1):
    if 'p_set' in line and ('load' in line.lower() or 'demand' in line.lower() or 'input_data' in line.lower()):
        print(f'  L{i}: {line.strip()[:120]}')

import openpyxl
# 5) interface.xlsx 시트 목록
wb = openpyxl.load_workbook('interface.xlsx', data_only=True)
print(f'\n=== interface.xlsx 시트 목록 ===')
print(' ', wb.sheetnames)

# 시나리오_에너지수요 시트가 있는지
if '시나리오_에너지수요' in wb.sheetnames:
    ws = wb['시나리오_에너지수요']
    print('\n  [시나리오_에너지수요 시트 헤더]')
    for cell in ws[1]:
        if cell.value:
            print(f'    Col {cell.column_letter}: {cell.value}')
    print('\n  [첫 5행]')
    for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        print(f'    {row}')
else:
    print('\n  시나리오_에너지수요 시트 없음!')
