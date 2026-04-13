import pandas as pd
import openpyxl

# interface.xlsx 파일 열기
wb = openpyxl.load_workbook('interface.xlsx', read_only=False)

# Coal_DB 시트 확인
if 'Coal_DB' in wb.sheetnames:
    ws = wb['Coal_DB']
    
    # D2 셀의 선택연도 확인
    selected_year = ws['D2'].value
    print(f"선택연도 (D2): {selected_year}")
    
    # 헤더 행 찾기 (F열부터 AF열까지 연도 정보)
    print("\n연도 헤더 (F1~AF1):")
    for col in range(6, 32):  # F=6, AF=32
        cell = ws.cell(1, col)
        if cell.value:
            print(f"  {chr(64+col)}열: {cell.value}")
    
    # 효율 열 확인 (AM열 = 39)
    print(f"\nAM1 (효율 헤더): {ws['AM1'].value}")
    
    wb.close()
    
    # pandas로 데이터 읽기
    df = pd.read_excel('interface.xlsx', sheet_name='Coal_DB', header=0)
    print(f"\nCoal_DB 데이터프레임:")
    print(f"  Shape: {df.shape}")
    print(f"  Columns: {df.columns.tolist()[:20]}")  # 처음 20개 컬럼
    print(f"\n상위 5개 행:")
    print(df.head())
    
    # 지역별 발전기 수 확인
    if 'A' in df.columns or df.columns[0]:
        region_col = df.columns[0]
        print(f"\n지역별 발전기 수:")
        print(df[region_col].value_counts())
else:
    print("Coal_DB 시트가 없습니다.")
    print(f"사용 가능한 시트: {wb.sheetnames}")
    wb.close()
