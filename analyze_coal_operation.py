import pandas as pd
import os
import glob

print("=" * 80)
print("1. 최근 결과 파일 확인")
print("=" * 80)

# 최근 결과 파일 찾기
result_files = glob.glob("*.xlsx")
result_files = [f for f in result_files if '결과' in f or 'result' in f.lower()]
if result_files:
    result_files.sort(key=os.path.getmtime, reverse=True)
    latest_result = result_files[0]
    print(f"\n최근 결과 파일: {latest_result}")
    print(f"수정 시간: {pd.Timestamp.fromtimestamp(os.path.getmtime(latest_result))}")
else:
    print("\n결과 파일을 찾을 수 없습니다.")
    latest_result = None

print("\n" + "=" * 80)
print("2. integrated_input_data.xlsx의 석탄발전기 설정 확인")
print("=" * 80)

df_gens = pd.read_excel('integrated_input_data.xlsx', sheet_name='generators')

# 석탄발전기 필터링 (발전소 이름 기반)
coal_gens = df_gens[df_gens['name'].str.contains('당진|영흥|하동|삼척|태안|보령|삼천포|호남|동해', na=False)]

print(f"\n세분화된 석탄발전기 수: {len(coal_gens)}")
print(f"총 용량: {coal_gens['p_nom'].sum():.1f} MW")

# Committable 설정 확인
print(f"\nCommittable 설정:")
print(coal_gens['committable'].value_counts())

# 샘플 확인
print(f"\n샘플 (상위 5개):")
print(coal_gens[['name', 'carrier', 'p_nom', 'committable', 'p_min_pu', 'marginal_cost']].head())

# LNG 발전기 확인
lng_gens = df_gens[df_gens['name'].str.contains('LNG', case=False)]
print(f"\n\nLNG 발전기 수: {len(lng_gens)}")
if not lng_gens.empty:
    print(f"LNG Committable 설정:")
    print(lng_gens['committable'].value_counts())
    print(f"\nLNG 샘플:")
    print(lng_gens[['name', 'carrier', 'p_nom', 'committable', 'marginal_cost']].head())

print("\n" + "=" * 80)
print("3. 결과 파일 분석 (있는 경우)")
print("=" * 80)

if latest_result:
    try:
        # FinalEnergy_Supply 시트 확인
        xls = pd.ExcelFile(latest_result)
        print(f"\n결과 파일의 시트: {xls.sheet_names}")
        
        if 'FinalEnergy_Supply' in xls.sheet_names:
            df_supply = pd.read_excel(latest_result, sheet_name='FinalEnergy_Supply')
            print(f"\n\nFinalEnergy_Supply 시트:")
            print(df_supply.head(20))
            
            # Coal 관련 행 찾기
            coal_rows = df_supply[df_supply.iloc[:, 0].astype(str).str.contains('Coal|coal|석탄', case=False, na=False)]
            if not coal_rows.empty:
                print(f"\n\nCoal 관련 데이터:")
                print(coal_rows)
            else:
                print(f"\n\n[경고] FinalEnergy_Supply에 Coal 데이터가 없습니다!")
        
        if 'generation_output' in xls.sheet_names:
            df_gen_output = pd.read_excel(latest_result, sheet_name='generation_output', nrows=10)
            print(f"\n\ngeneration_output 시트 (상위 10행):")
            print(f"컬럼: {df_gen_output.columns.tolist()}")
            
            # Coal 관련 컬럼 찾기
            coal_cols = [col for col in df_gen_output.columns if 'coal' in str(col).lower() or '석탄' in str(col) or any(plant in str(col) for plant in ['당진', '영흥', '하동', '삼척', '태안'])]
            print(f"\nCoal 관련 컬럼 수: {len(coal_cols)}")
            if coal_cols:
                print(f"샘플 컬럼: {coal_cols[:10]}")
            else:
                print(f"[경고] generation_output에 Coal 관련 컬럼이 없습니다!")
                
    except Exception as e:
        print(f"\n결과 파일 읽기 오류: {e}")

print("\n" + "=" * 80)
print("4. interface.xlsx의 지역별 Coal 설정 확인")
print("=" * 80)

import openpyxl
wb = openpyxl.load_workbook('interface.xlsx', data_only=True)

# CND 지역의 Coal 설정 확인
if '지역_CND' in wb.sheetnames:
    ws = wb['지역_CND']
    print(f"\n지역_CND의 Coal 발전기 설정:")
    
    for row in range(1, min(ws.max_row + 1, 100)):
        cell_val = ws.cell(row, 1).value
        if cell_val and 'Coal' in str(cell_val):
            # 해당 행의 모든 값 출력
            row_data = []
            for col in range(1, 15):
                row_data.append(ws.cell(row, col).value)
            print(f"\nRow {row}: {row_data}")
            
            # 헤더 찾기
            header_row = row - 1
            if header_row > 0:
                headers = []
                for col in range(1, 15):
                    headers.append(ws.cell(header_row, col).value)
                print(f"Headers: {headers}")
            break

wb.close()
