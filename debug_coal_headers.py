import openpyxl

wb = openpyxl.load_workbook('interface.xlsx', data_only=True)
ws = wb['Coal_DB']

print("Row 5의 모든 값 확인:")
for col_idx in range(1, 40):
    val = ws.cell(5, col_idx).value
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    print(f"  {col_letter}5 (col {col_idx}): {val} (type: {type(val).__name__})")

wb.close()
