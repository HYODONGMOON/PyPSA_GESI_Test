import openpyxl

wb = openpyxl.load_workbook('interface.xlsx', data_only=True)

print("=" * 80)
print("지역_CND 시트의 발전기 섹션 상세 확인")
print("=" * 80)

if '지역_CND' in wb.sheetnames:
    ws = wb['지역_CND']
    
    # Coal과 LNG 근처의 행들을 모두 출력
    print("\n20~30행 전체 확인:")
    for row in range(20, 31):
        row_data = []
        for col in range(1, 12):
            val = ws.cell(row, col).value
            row_data.append(str(val) if val is not None else '')
        print(f"Row {row:2d}: {' | '.join(row_data)}")
    
    # committable 관련 열 찾기
    print("\n\nCommittable 관련 정보 찾기:")
    for row in range(15, 35):
        for col in range(1, 15):
            val = ws.cell(row, col).value
            if val and ('committable' in str(val).lower() or 'commit' in str(val).lower()):
                print(f"  Row {row}, Col {col}: {val}")

wb.close()
