import openpyxl

wb = openpyxl.load_workbook('integrated_input_data.xlsx', data_only=True)

# 1) loads 시트 확인
ws_loads = wb['loads']
headers = [c.value for c in next(ws_loads.iter_rows(min_row=1, max_row=1))]
print('=== loads 시트 헤더 ===')
print(' ', headers)
print('\n  [전체 load 목록]')
for row in ws_loads.iter_rows(min_row=2, values_only=True):
    if row[0]:
        print(f'    {row}')

# 2) timeseries 시트 확인 (p_set 포함 여부)
ws_ts = wb['timeseries']
ts_headers = [c.value for c in next(ws_ts.iter_rows(min_row=1, max_row=1))]
print(f'\n=== timeseries 시트 헤더 (총 {len(ts_headers)}개 컬럼) ===')
print('  처음 30개:', ts_headers[:30])

# p_set 관련 컬럼 찾기
load_cols = [h for h in ts_headers if h and 'demand' in str(h).lower() or 
             (h and any(r in str(h) for r in ['BSN','CBD','CND','GBD','GGD','GND','GWD','ICN','JBD','JJD','JND']))]
print(f'\n  부하(Demand) 관련 컬럼 ({len(load_cols)}개):')
for c in load_cols[:20]:
    print(f'    {c}')

# 3) 실제 첫 5행 데이터 확인
print('\n  [timeseries 처음 3개 데이터 행]')
for i, row in enumerate(ws_ts.iter_rows(min_row=2, max_row=4, values_only=True)):
    print(f'  row{i+2}: {row[:10]}...')

# 4) interface.xlsx 인터페이스 시트 - Demand 값 확인
wb2 = openpyxl.load_workbook('interface.xlsx', data_only=True)
ws_iface = wb2['인터페이스']
print('\n=== interface.xlsx 인터페이스 시트 - Demand 행 ===')
headers2 = [c.value for c in next(ws_iface.iter_rows(min_row=1, max_row=1))]
print('  헤더:', headers2[:15])
cur_cat = None
for row_cells in ws_iface.iter_rows(min_row=2):
    col_b = row_cells[1].value
    col_c = row_cells[2].value
    if col_b in ['Generators','Storages','Links','Demand']:
        cur_cat = col_b
    if cur_cat == 'Demand' and col_c:
        vals = [c.value for c in row_cells[:12]]
        print(f'  {vals}')
