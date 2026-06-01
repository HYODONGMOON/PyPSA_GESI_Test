"""
visualize_results.py
====================
PyPSA 분석 결과 시각화 스크립트
- results/ 폴더에서 가장 최근 .nc 파일 자동 탐색·로드
- 5가지 유형의 시각화 PNG 파일 생성

사용법:
    python visualize_results.py
    또는 결과폴더를 직접 지정:
    python visualize_results.py results/20260512_160606
"""

import os
import sys
import re
import glob
import warnings
import numpy as np
import pandas as pd
import matplotlib
# Agg 백엔드 설정 — 이미 다른 백엔드가 초기화된 경우(PyPSA_GUI 등) 경고 없이 무시
try:
    matplotlib.use('Agg')
except Exception:
    pass
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.font_manager as fm
from matplotlib.lines import Line2D
from matplotlib.transforms import blended_transform_factory
import pypsa

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────
#  기본 설정
# ─────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
RESULTS_BASE = os.path.join(SCRIPT_DIR, 'results')
INPUT_XLSX   = os.path.join(SCRIPT_DIR, 'integrated_input_data.xlsx')
# 수정 작업 중 임시 저장 폴더 (확정 전까지 이 폴더에 덮어씌워 가며 확인)
VIZ_TEMP_DIR = os.path.join(SCRIPT_DIR, 'viz_temp')

RE_CARRIERS    = {'solar', 'wind'}
NONRE_CARRIERS = {'coal', 'gas', 'nuclear', 'electricity'}   # electricity = 수력
DR_CARRIERS    = {'DR'}

REGIONS = ['BSN', 'CBD', 'CND', 'GBD', 'GGD', 'GND', 'GWD', 'ICN', 'JBD', 'JJD', 'JND']

# 그래프1·3 복도 고정 순서 (지리적 위치 기반, 조건 비교 일관성 확보)
CORRIDOR_ORDER = [
    'GWD_GGD',   # 강원–경기
    'GGD_ICN',   # 경기–인천
    'GGD_CND',   # 경기–충남
    'GGD_CBD',   # 경기–충북
    'GWD_CBD',   # 강원–충북
    'CND_CBD',   # 충남–충북
    'JBD_CND',   # 전북–충남
    'GBD_CBD',   # 경북–충북
    'GWD_GBD',   # 강원–경북
    'GND_GBD',   # 경남–경북
    'GBD_BSN',   # 경북–부산
    'GND_BSN',   # 경남–부산
    'GND_JBD',   # 경남–전북
    'JND_JBD',   # 전남–전북
    'GND_JND',   # 경남–전남
    'JND_JJD',   # 전남–제주
]

SEASON_MONTHS = {
    '봄':  [3, 4, 5],
    '여름': [6, 7, 8],
    '가을': [9, 10, 11],
    '겨울': [12, 1, 2],
}
SEASON_ORDER  = ['봄', '여름', '가을', '겨울']
SEASON_COLORS = {'봄': '#4CAF50', '여름': '#FF5722', '가을': '#FF9800', '겨울': '#2196F3'}

_MONTH_SEASON = {}
for _s, _ms in SEASON_MONTHS.items():
    for _m in _ms:
        _MONTH_SEASON[_m] = _s

# ─────────────────────────────────────────────
#  한국어 폰트 설정
# ─────────────────────────────────────────────
def _setup_font():
    candidates = ['Malgun Gothic', 'NanumGothic', 'NanumBarunGothic',
                  'Apple Gothic', 'AppleGothic', 'Gulim', 'Batang']
    available = {f.name for f in fm.fontManager.ttflist}
    for c in candidates:
        if c in available:
            plt.rcParams['font.family'] = c
            print(f"[폰트] {c} 사용")
            return
    plt.rcParams['font.family'] = 'DejaVu Sans'
    print("[폰트] 한국어 폰트를 찾지 못했습니다. DejaVu Sans 사용")

_setup_font()
plt.rcParams.update({
    'axes.unicode_minus': False,
    'figure.dpi': 150,
    'axes.titlesize': 11,
    'axes.labelsize': 9,
    'xtick.labelsize': 8,
    'ytick.labelsize': 8,
    'legend.fontsize': 7,
})

# ─────────────────────────────────────────────
#  헬퍼 함수
# ─────────────────────────────────────────────

def find_latest_nc(base_dir=None):
    base = base_dir or RESULTS_BASE
    folders = sorted(glob.glob(os.path.join(base, '2*')))
    for folder in reversed(folders):
        ncs = glob.glob(os.path.join(folder, 'optimization_result_*.nc'))
        if ncs:
            return folder, ncs[0]
    raise FileNotFoundError(f".nc 파일을 찾을 수 없습니다: {base}")


def load_network(nc_path):
    n = pypsa.Network()
    n.import_from_netcdf(nc_path)
    print(f"  스냅샷: {len(n.snapshots)}개  |  선로: {len(n.lines)}개  |  발전기: {len(n.generators)}개")
    return n


def load_excel_line_caps(xlsx_path=None):
    """integrated_input_data.xlsx 'lines' 시트에서 원본 s_nom(MVA) 읽기"""
    path = xlsx_path or INPUT_XLSX
    if not os.path.exists(path):
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = None
        for sname in wb.sheetnames:
            if 'line' in sname.lower():
                ws = wb[sname]
                break
        if ws is None:
            return {}
        headers = [c.value for c in ws[1]]
        ni = next((i for i, h in enumerate(headers) if h == 'name'), None)
        si = next((i for i, h in enumerate(headers) if h == 's_nom'), None)
        if ni is None or si is None:
            return {}
        caps = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[ni]
            snom = row[si]
            if name and snom and pd.notna(snom):
                try:
                    caps[str(name)] = float(snom)
                except (ValueError, TypeError):
                    pass
        print(f"  [용량] Excel에서 {len(caps)}개 선로 용량 로드")
        return caps
    except Exception as e:
        print(f"  [용량] Excel 로드 실패: {e} — nc 파일 s_nom 사용")
        return {}


def get_corridor(line_name):
    """선로명에서 복도명 추출: GWD_GGD1 → GWD_GGD"""
    return re.sub(r'\d+$', '', str(line_name)).rstrip('_')


def get_season_series(timestamps):
    return pd.Series(timestamps.month.map(_MONTH_SEASON).values,
                     index=timestamps, name='season')


def _loads_p(n):
    if hasattr(n.loads_t, 'p_set') and not n.loads_t.p_set.empty:
        return n.loads_t.p_set
    if hasattr(n.loads_t, 'p') and not n.loads_t.p.empty:
        return n.loads_t.p
    return pd.DataFrame(index=n.snapshots)


def _gens_p(n):
    if hasattr(n.generators_t, 'p') and not n.generators_t.p.empty:
        return n.generators_t.p
    return pd.DataFrame(index=n.snapshots)


def _get_load_ts(n, load_name):
    lp = _loads_p(n)
    for name in [load_name, load_name.strip(), load_name + ' ']:
        if not lp.empty and name in lp.columns:
            return lp[name]
    for name in [load_name, load_name.strip(), load_name + ' ']:
        if name in n.loads.index:
            pset = n.loads.at[name, 'p_set']
            if pd.notna(pset):
                return pd.Series(float(pset), index=n.snapshots)
    return pd.Series(0.0, index=n.snapshots)


def _get_load_annual(n, load_name):
    return float(_get_load_ts(n, load_name).sum())


def util_level_color(mean_pct):
    if mean_pct >= 70:   return '#C62828'
    elif mean_pct >= 50: return '#EF6C00'
    elif mean_pct >= 30: return '#F9A825'
    elif mean_pct >= 15: return '#2E7D32'
    else:                return '#1565C0'


# ─────────────────────────────────────────────
#  interface.xlsx 에서 s_nom_basic / s_nom_flex 로드
# ─────────────────────────────────────────────

def load_interface_flex_info(interface_path=None):
    """
    interface.xlsx '인터페이스_1' 시트(헤더 5행)에서
    선로별 {line: {'basic': float, 'flex': float}} 반환.
    s_nom_basic = G열, s_nom_flex = F열 (F > G > 0 아니어도 모두 반환).
    """
    path = interface_path or os.path.join(SCRIPT_DIR, 'interface.xlsx')
    if not os.path.exists(path):
        return {}
    try:
        import openpyxl as _opxl
        wb = _opxl.load_workbook(path, data_only=True)
        if '인터페이스_1' not in wb.sheetnames:
            return {}
        ws = wb['인터페이스_1']
        # 헤더 행(5행)에서 열 위치 파악
        hdrs = {str(c.value).strip().lower(): c.column
                for c in ws[5] if c.value is not None}
        col_name  = 1                              # A 열
        col_basic = hdrs.get('s_nom_basic', 7)     # G 열 기본
        col_flex  = hdrs.get('s_nom',       6)     # F 열 기본
        result = {}
        for row in ws.iter_rows(min_row=6, max_row=100, values_only=True):
            nm = row[col_name - 1]
            if not nm:
                continue
            nm = str(nm).strip()
            try:
                basic = float(row[col_basic - 1]) if row[col_basic - 1] is not None else None
                flex  = float(row[col_flex  - 1]) if row[col_flex  - 1] is not None else None
                if basic is not None and basic > 0:
                    result[nm] = {
                        'basic': basic,
                        'flex':  flex if (flex is not None and flex >= basic) else basic,
                    }
            except (TypeError, ValueError):
                pass
        return result
    except Exception as _e:
        print(f"  [interface.xlsx] 로드 실패: {_e} — nc s_nom 사용")
        return {}


# ─────────────────────────────────────────────
#  복도별 s_nom_basic / s_nom_flex 집계
# ─────────────────────────────────────────────

def compute_corridor_caps(n, flex_info=None):
    """
    복도별 s_nom_basic 합계 및 s_nom_flex 합계 반환.
    {corridor: {'basic': float, 'flex': float, 'has_flex': bool}}

    우선순위:
      1) interface.xlsx s_nom_basic (nc s_nom 과 같은 스케일인지 확인 후 사용)
      2) nc s_nom (fallback)
    스케일 확인: interface 값 합이 최대 흐름의 0.5~5배 범위 내일 때 유효로 판정.
    """
    if flex_info is None:
        flex_info = load_interface_flex_info()

    _rv = set(REGIONS)
    def _valid(c):
        p = c.split('_')
        return len(p) >= 2 and p[0] in _rv and p[-1] in _rv and p[0] != p[-1]

    caps = {}   # {corridor: {'basic': float, 'flex': float,
                #              'iface_basic': float, 'iface_flex': float, 'nc_cap': float}}

    if not n.lines.empty and hasattr(n.lines_t, 'p0') and not n.lines_t.p0.empty:
        for line in n.lines_t.p0.columns:
            if line not in n.lines.index:
                continue
            nc_snom = float(n.lines.at[line, 's_nom'])
            if nc_snom <= 0:
                continue
            corr = get_corridor(line)
            if not _valid(corr):
                continue
            if corr not in caps:
                caps[corr] = {'iface_basic': 0.0, 'iface_flex': 0.0,
                              'nc_cap': 0.0, 'has_iface': False}
            caps[corr]['nc_cap'] += nc_snom
            fi = flex_info.get(line)
            if fi:
                caps[corr]['iface_basic'] += fi['basic']
                caps[corr]['iface_flex']  += fi['flex']
                caps[corr]['has_iface']    = True

    # HVDC 링크 (항상 nc p_nom 사용)
    if not n.links.empty and hasattr(n.links_t, 'p0') and not n.links_t.p0.empty:
        inter = n.links[
            n.links.apply(
                lambda r: (pd.notna(r['bus0']) and pd.notna(r['bus1']) and
                           r['bus0'][:3] != r['bus1'][:3]), axis=1)
        ]
        for lk in inter.index:
            if lk not in n.links_t.p0.columns:
                continue
            p_nom = n.links.at[lk, 'p_nom']
            if p_nom <= 0:
                continue
            corr = get_corridor(lk)
            if not _valid(corr):
                continue
            if corr not in caps:
                caps[corr] = {'iface_basic': 0.0, 'iface_flex': 0.0,
                              'nc_cap': 0.0, 'has_iface': False}
            caps[corr]['nc_cap']      += p_nom
            caps[corr]['iface_basic'] += p_nom   # HVDC는 nc 값 그대로
            caps[corr]['iface_flex']  += p_nom

    # 스케일 판정 후 최종 basic/flex 결정
    result = {}
    for corr, d in caps.items():
        if d['has_iface'] and d['iface_basic'] > 0:
            # interface 값이 nc 값의 0.01~2배 범위이면 같은 스케일 → interface 사용
            ratio = d['nc_cap'] / d['iface_basic']
            if 0.5 <= ratio <= 5.0:
                result[corr] = {
                    'basic':     d['iface_basic'],
                    'flex':      d['iface_flex'],
                    'has_flex':  d['iface_flex'] > d['iface_basic'] * 1.01,
                }
            else:
                # 스케일 불일치(구버전 nc 파일 등) → nc s_nom 사용
                result[corr] = {
                    'basic':    d['nc_cap'],
                    'flex':     d['nc_cap'],
                    'has_flex': False,
                }
        else:
            result[corr] = {
                'basic':    d['nc_cap'],
                'flex':     d['nc_cap'],
                'has_flex': False,
            }
    return result


# ─────────────────────────────────────────────
#  핵심: 복도별 시간별 이용률 계산 (s_nom_basic 기준)
#  100% 초과 = 유연 운영(과부하) 구간
# ─────────────────────────────────────────────

def compute_corridor_utilization(n, excel_caps=None):
    """
    복도별 시간별 이용률(%) DataFrame 반환

    이용률 = sum(|p0_i|) / sum(s_nom_basic_i) × 100%
    - s_nom_basic: interface.xlsx G열 (nc s_nom 과 스케일 일치 시 우선 사용)
    - 스케일 불일치(구버전 nc 파일) → nc s_nom 사용
    - 100% 초과 = 유연 운영(과부하) 구간
    """
    if n.lines.empty or not (hasattr(n.lines_t, 'p0') and not n.lines_t.p0.empty):
        return pd.DataFrame(index=n.snapshots)

    corridor_caps = compute_corridor_caps(n)

    _rv = set(REGIONS)
    def _valid(c):
        p = c.split('_')
        return len(p) >= 2 and p[0] in _rv and p[-1] in _rv and p[0] != p[-1]

    corridors = {}

    for line in n.lines_t.p0.columns:
        if line not in n.lines.index:
            continue
        corr = get_corridor(line)
        if not _valid(corr) or corr not in corridor_caps:
            continue
        if corr not in corridors:
            corridors[corr] = pd.Series(0.0, index=n.snapshots)
        corridors[corr] += n.lines_t.p0[line].abs()

    # HVDC 링크 흐름 추가
    if not n.links.empty and hasattr(n.links_t, 'p0') and not n.links_t.p0.empty:
        inter = n.links[
            n.links.apply(
                lambda r: (pd.notna(r['bus0']) and pd.notna(r['bus1']) and
                           r['bus0'][:3] != r['bus1'][:3]), axis=1)
        ]
        for lk in inter.index:
            if lk not in n.links_t.p0.columns:
                continue
            corr = get_corridor(lk)
            if not _valid(corr) or corr not in corridor_caps:
                continue
            if corr not in corridors:
                corridors[corr] = pd.Series(0.0, index=n.snapshots)
            corridors[corr] += n.links_t.p0[lk].abs()

    result = {
        corr: flow / corridor_caps[corr]['basic'] * 100.0
        for corr, flow in corridors.items()
        if corridor_caps[corr]['basic'] > 0
    }
    return pd.DataFrame(result, index=n.snapshots)


def compute_corridor_max_line_util(n):
    """
    복도별 시간별 '개별 선로 최대 이용률'(%) DataFrame 반환.

    각 시간대마다 복도 내 여러 병렬 선로 중 가장 높게 이용된 선로의
    이용률(= |p0_i| / s_nom_basic_i × 100)을 반환합니다.

    집계 이용률 대신 이 값을 포화 시간 계산에 사용하면,
    일부 선로만 포화 상태여도 정확하게 검출됩니다.
    """
    flex_info = load_interface_flex_info()
    _rv = set(REGIONS)
    def _valid(c):
        p = c.split('_')
        return len(p) >= 2 and p[0] in _rv and p[-1] in _rv and p[0] != p[-1]

    corr_line_series = {}   # {corridor: [Series, ...]}

    if not n.lines.empty and hasattr(n.lines_t, 'p0') and not n.lines_t.p0.empty:
        for line in n.lines_t.p0.columns:
            if line not in n.lines.index:
                continue
            corr = get_corridor(line)
            if not _valid(corr):
                continue
            nc_snom = float(n.lines.at[line, 's_nom'])
            fi = flex_info.get(line)
            if fi and fi['basic'] > 0:
                ratio = nc_snom / fi['basic']
                cap = fi['basic'] if 0.5 <= ratio <= 5.0 else nc_snom
            else:
                cap = nc_snom
            if cap <= 0:
                continue
            line_util = n.lines_t.p0[line].abs() / cap * 100.0
            corr_line_series.setdefault(corr, []).append(line_util)

    result = {}
    for corr, series_list in corr_line_series.items():
        if not series_list:
            continue
        result[corr] = pd.concat(series_list, axis=1).max(axis=1)

    return pd.DataFrame(result, index=n.snapshots)


def compute_re_output_ts(n):
    gp  = _gens_p(n)
    re  = n.generators[n.generators['carrier'].isin(RE_CARRIERS)].index
    re  = [g for g in re if g in gp.columns]
    if not re:
        return pd.Series(0.0, index=n.snapshots)
    return gp[re].sum(axis=1)


# ─────────────────────────────────────────────
#  복도별 연간 순 송전 방향 계산
#  반환: {corridor: signed_annual_flow_MWh}
#  양수 = bus0→bus1 방향 우세, 음수 = bus1→bus0 방향 우세
# ─────────────────────────────────────────────

def compute_corridor_net_flow(n):
    """복도별 연간 순 송전량(부호 있음, MWh) → 방향성 파악용"""
    _rv = set(REGIONS)
    def _valid(c):
        p = c.split('_')
        return len(p) >= 2 and p[0] in _rv and p[-1] in _rv and p[0] != p[-1]

    net = {}

    # AC 선로
    if not n.lines.empty and hasattr(n.lines_t, 'p0') and not n.lines_t.p0.empty:
        for line in n.lines_t.p0.columns:
            if line not in n.lines.index:
                continue
            corr = get_corridor(line)
            if not _valid(corr):
                continue
            net[corr] = net.get(corr, 0.0) + float(n.lines_t.p0[line].sum())

    # 지역간 HVDC 링크
    if not n.links.empty and hasattr(n.links_t, 'p0') and not n.links_t.p0.empty:
        inter = n.links[
            n.links.apply(
                lambda r: (pd.notna(r['bus0']) and pd.notna(r['bus1']) and
                           r['bus0'][:3] != r['bus1'][:3]),
                axis=1,
            )
        ]
        for lk in inter.index:
            if lk not in n.links_t.p0.columns:
                continue
            corr = get_corridor(lk)
            if not _valid(corr):
                continue
            net[corr] = net.get(corr, 0.0) + float(n.links_t.p0[lk].sum())

    return net


# ─────────────────────────────────────────────
#  차트 1: 복도별 이용률 수직 Boxplot
# ─────────────────────────────────────────────

def chart1_line_utilization(n, out_path):
    print("[차트1] 송전망 이용률 Boxplot 생성 중...")
    util = compute_corridor_utilization(n)
    if util.empty:
        print("  송전망 이용률 데이터 없음 — 스킵")
        return

    # 고정 순서로 정렬: CORRIDOR_ORDER 에 없는 복도는 뒤에 추가
    avail = util.columns.tolist()
    ordered = [c for c in CORRIDOR_ORDER if c in avail] + \
              [c for c in avail if c not in CORRIDOR_ORDER]
    util = util[ordered]

    stats = pd.DataFrame({
        'min':    util.min(),
        'p25':    util.quantile(0.25),
        'median': util.median(),
        'mean':   util.mean(),
        'p75':    util.quantile(0.75),
        'max':    util.max(),
    })

    # 복도별 최대 가능 이용률 = s_nom_flex / s_nom_basic × 100
    corr_caps   = compute_corridor_caps(n)
    max_cap_pct = {
        corr: (d['flex'] / d['basic'] * 100.0 if d['basic'] > 0 else 100.0)
        for corr, d in corr_caps.items()
    }

    # 포화 시간: 복도 내 개별 선로 이용률 기준 (집계가 아닌 개별 선로 최대값)
    # 이유: 병렬 선로 중 일부만 포화여도 집계 이용률은 99.5%에 미달할 수 있음
    max_line_util = compute_corridor_max_line_util(n)
    sat_hours = pd.Series({
        corr: int((max_line_util[corr] >= max_cap_pct.get(corr, 100.0) * 0.995).sum())
        if corr in max_line_util.columns else 0
        for corr in util.columns
    })
    zero_hours = (util <= 0.1).sum()

    # 복도별 연간 순 송전 방향 (부호 있음)
    net_flow = compute_corridor_net_flow(n)

    n_corr = len(stats)
    fig, ax = plt.subplots(figsize=(max(8, n_corr * 0.68), 8))

    x = np.arange(n_corr)
    colors = [util_level_color(v) for v in stats['mean']]

    # min-max 범위 선
    for i, (_, row) in enumerate(stats.iterrows()):
        ax.plot([i, i], [row['min'], row['max']],
                color='#999', lw=1.2, zorder=2, solid_capstyle='round')
        ax.plot(i, row['min'], '_', color='#777', ms=7, zorder=3)
        ax.plot(i, row['max'], '_', color='#777', ms=7, zorder=3)

    # IQR 박스
    ax.bar(x,
           (stats['p75'] - stats['p25']).values,
           bottom=stats['p25'].values,
           width=0.55, color=colors, alpha=0.65, zorder=4)

    # 중앙값 선
    ax.hlines(stats['median'].values, x - 0.28, x + 0.28,
              colors='white', lw=2.5, zorder=5)
    ax.hlines(stats['median'].values, x - 0.28, x + 0.28,
              colors='#333',  lw=1.5, zorder=6)

    # 평균 점
    ax.scatter(x, stats['mean'].values,
               s=50, c=colors, edgecolors='black', linewidths=0.6,
               zorder=7, marker='D')

    # ── s_nom_basic 100% 기준선 ─────────────────────────────────────────
    ax.axhline(100, color='#B71C1C', lw=1.0, ls='--', alpha=0.7, zorder=3)
    ax.text(-0.5, 100.5, '기본 용량 100%', fontsize=6.5,
            color='#B71C1C', va='bottom', ha='left', zorder=9)

    # ── X축 레이블: bus1 항상 위, bus0 항상 아래 (방향은 화살표로만 표시) ──
    trans = blended_transform_factory(ax.transData, ax.transAxes)
    ax.set_xticks(x)
    ax.set_xticklabels([''] * n_corr)
    ax.tick_params(axis='x', length=0)

    Y_TOP  = -0.050   # bus1 텍스트 (위 고정)
    Y_AHD  = -0.100   # 화살표 머리
    Y_ATL  = -0.161   # 화살표 꼬리
    Y_BOT  = -0.204   # bus0 텍스트 (아래 고정)

    for i, (corr, col) in enumerate(zip(stats.index, colors)):
        flow  = net_flow.get(corr, 0.0)
        parts = corr.split('_')
        bus0  = parts[0]
        bus1  = parts[-1] if len(parts) >= 2 else parts[0]

        # 라벨 위치 고정: bus1 위, bus0 아래
        # 화살표 방향: flow>=0 → bus0→bus1 (↑), flow<0 → bus1→bus0 (↓)
        if flow >= 0:
            head_y, tail_y = Y_AHD, Y_ATL   # ↑
        else:
            head_y, tail_y = Y_ATL, Y_AHD   # ↓

        ax.text(i, Y_TOP, bus1,
                ha='center', va='top', fontsize=8, color='#333333',
                transform=trans, clip_on=False)
        ax.annotate('',
            xy=(i, head_y), xytext=(i, tail_y),
            xycoords=trans, textcoords=trans,
            arrowprops=dict(arrowstyle='-|>', color=col, lw=2.0,
                            mutation_scale=16, shrinkA=0, shrinkB=0),
            annotation_clip=False)
        ax.text(i, Y_BOT, bus0,
                ha='center', va='top', fontsize=8, color='#333333',
                transform=trans, clip_on=False)

    ax.set_ylabel('이용률 (%, s_nom_basic 기준)')
    ax.set_title('권역간 송전망 이용률 분포 (s_nom_basic 기준)\n'
                 '(■ IQR 범위, ◆ 평균, ─ 중앙값, | 최소~최대 | 화살표: 연간 순 송전 방향)')
    ax.grid(axis='y', alpha=0.3, zorder=1)

    # Y축 범위 — 100% 초과 영역 + 레이블 여백
    raw_ymax    = stats['max'].max()
    has_over100 = raw_ymax > 101
    has_zero    = zero_hours.reindex(stats.index).fillna(0).gt(0).any()
    ymax = max(raw_ymax * 1.10, 105)
    ymin = -6 if has_zero else -1
    ax.set_ylim(ymin, ymax)

    # 포화 시간 레이블 (max 막대 바로 위)
    label_gap = (ymax - ymin) * 0.018
    for i, corr in enumerate(stats.index):
        sh = int(sat_hours.get(corr, 0))
        zh = int(zero_hours.get(corr, 0))
        mx = stats.at[corr, 'max']
        if sh > 0:
            ax.text(i, mx + label_gap,
                    f'{sh}h',
                    ha='center', va='bottom', fontsize=6,
                    color='#B71C1C', fontweight='bold', zorder=8,
                    bbox=dict(fc='#FFEBEE', ec='#EF9A9A', lw=0.5,
                              boxstyle='round,pad=0.15', alpha=0.85))
        if zh > 0:
            ax.text(i, 0 - label_gap,
                    f'{zh}h',
                    ha='center', va='top', fontsize=6,
                    color='#0D47A1', fontweight='bold', zorder=8,
                    bbox=dict(fc='#E3F2FD', ec='#90CAF9', lw=0.5,
                              boxstyle='round,pad=0.15', alpha=0.85))

    legend_patches = [
        mpatches.Patch(fc='#C62828', alpha=0.7, label='>=70%'),
        mpatches.Patch(fc='#EF6C00', alpha=0.7, label='50~70%'),
        mpatches.Patch(fc='#F9A825', alpha=0.7, label='30~50%'),
        mpatches.Patch(fc='#2E7D32', alpha=0.7, label='15~30%'),
        mpatches.Patch(fc='#1565C0', alpha=0.7, label='<15%'),
    ]
    if has_over100:
        legend_patches.append(
            mpatches.Patch(fc='none', ec='#B71C1C', lw=1, ls='--',
                           label='100% = s_nom_basic 한계'))
    ax.legend(handles=legend_patches, title='평균 이용률',
              loc='upper right', fontsize=7, title_fontsize=7)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  저장 -> {os.path.basename(out_path)}")


# ─────────────────────────────────────────────
#  차트 2: RE 발전량 vs 송전망 이용률 + 계절 평균
# ─────────────────────────────────────────────

def chart2_re_vs_tx(n, out_path):
    print("[차트2] RE 발전량 vs 송전망 이용률 산점도 생성 중...")
    re_ts = compute_re_output_ts(n)
    util  = compute_corridor_utilization(n)
    if util.empty or re_ts.sum() == 0:
        print("  데이터 없음 — 스킵")
        return

    # 전체 평균 이용률: 복도 수 기준 단순 평균 (각 복도를 동등하게 취급)
    mean_util = util.mean(axis=1)

    seasons = get_season_series(n.snapshots)
    df = pd.DataFrame({'re_mw': re_ts, 'tx_util': mean_util,
                       'season': seasons}).dropna()

    fig, ax = plt.subplots(figsize=(10, 6))

    # 시간별 산점도 (배경)
    for s in SEASON_ORDER:
        m = df['season'] == s
        ax.scatter(df.loc[m, 're_mw'], df.loc[m, 'tx_util'],
                   c=SEASON_COLORS[s], alpha=0.20, s=5, label=s, zorder=2)

    # 10% 구간 중앙값 곡선
    re_max = df['re_mw'].max()
    bins   = np.linspace(0, re_max, 11)
    mids, meds = [], []
    for i in range(len(bins) - 1):
        m = (df['re_mw'] >= bins[i]) & (df['re_mw'] < bins[i + 1])
        if m.sum() >= 5:
            mids.append((bins[i] + bins[i + 1]) / 2)
            meds.append(df.loc[m, 'tx_util'].median())
    if mids:
        ax.plot(mids, meds, 'k-o', ms=5, lw=2,
                label='구간 중앙값 (RE 10% 구간)', zorder=5)

    # 계절별 평균 마커 (테두리 없음, 텍스트 없음)
    for s in SEASON_ORDER:
        m = df['season'] == s
        if m.sum() == 0:
            continue
        sx = df.loc[m, 're_mw'].mean()
        sy = df.loc[m, 'tx_util'].mean()
        ax.scatter(sx, sy,
                   c=SEASON_COLORS[s], s=220, zorder=8,
                   edgecolors='none',
                   marker='*', label=f'{s} 평균')

    ax.set_xlabel('재생에너지 발전량 (MW)')
    ax.set_ylabel('송전망 평균 이용률 (%)')
    ax.set_title('재생에너지 발전량 vs 송전망 이용률\n(8,760시간, 계절별 색상 / ★ 계절 평균)')

    # 중복 레전드 제거 + 별 마커만 크기 30%로 축소
    # 산점도 별(s=220, markerscale=1.5 → 등가 ~22pt) → 30% = ~7pt 의 Line2D 로 교체
    handles, labels = ax.get_legend_handles_labels()
    seen = {}
    unique_h, unique_l = [], []
    for h, l in zip(handles, labels):
        if l not in seen:
            seen[l] = True
            if '평균' in l:   # 계절 평균 별 마커 → 작은 Line2D 로 교체
                season_key = l.replace(' 평균', '')
                c = SEASON_COLORS.get(season_key, 'gray')
                h = Line2D([0], [0], linestyle='none', marker='*',
                           color=c, markersize=7, label=l)
            unique_h.append(h)
            unique_l.append(l)
    ax.legend(unique_h, unique_l, loc='upper left')
    ax.grid(alpha=0.3)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  저장 -> {os.path.basename(out_path)}")


# ─────────────────────────────────────────────
#  차트 3: 계절 x 시간대별 이용률 히트맵
# ─────────────────────────────────────────────

def chart3_seasonal_heatmap(n, out_path):
    print("[차트3] 계절x시간대별 이용률 히트맵 생성 중...")
    util = compute_corridor_utilization(n)
    if util.empty:
        print("  데이터 없음 — 스킵")
        return

    seasons = get_season_series(n.snapshots)
    hours   = pd.Series(n.snapshots.hour, index=n.snapshots)

    col_tuples = [(s, h) for s in SEASON_ORDER for h in range(24)]
    hmap_dict  = {}
    for s in SEASON_ORDER:
        for h in range(24):
            mask = (seasons == s) & (hours == h)
            if mask.sum() > 0:
                hmap_dict[(s, h)] = util.loc[mask].mean()
            else:
                hmap_dict[(s, h)] = pd.Series(0.0, index=util.columns)

    heatmap = pd.DataFrame(hmap_dict, index=util.columns)[col_tuples]
    # 고정 순서로 정렬: CORRIDOR_ORDER 에 없는 복도는 뒤에 추가
    avail_h = heatmap.index.tolist()
    ordered_h = [c for c in CORRIDOR_ORDER if c in avail_h] + \
                [c for c in avail_h if c not in CORRIDOR_ORDER]
    heatmap = heatmap.loc[ordered_h]

    n_corr  = len(heatmap)
    fig_h   = max(5, n_corr * 0.38)
    fig, ax = plt.subplots(figsize=(20, fig_h))

    # 100% 초과 값은 100%와 동일 색상으로 표시 (s_nom_basic 기준)
    heatmap_display = heatmap.clip(upper=100.0)
    vmax = 100.0
    im = ax.imshow(heatmap_display.values, aspect='auto',
                   cmap='plasma_r', vmin=0, vmax=vmax, interpolation='nearest')
    plt.colorbar(im, ax=ax, label='이용률 (%, 100% = s_nom_basic 한계)', fraction=0.015, pad=0.01)

    ax.set_yticks(range(n_corr))
    ax.set_yticklabels(heatmap.index, fontsize=7)

    for s_idx in range(1, 4):
        ax.axvline(s_idx * 24 - 0.5, color='white', lw=2)
    ax.set_xticks([s_idx * 24 + 11.5 for s_idx in range(4)])
    ax.set_xticklabels(SEASON_ORDER, fontsize=9)

    hour_labels_pos  = list(range(96))
    hour_labels_text = [f'{h+1:02d}h' if h % 6 == 0 else '' for h in range(24)] * 4
    ax2 = ax.twiny()
    ax2.set_xlim(ax.get_xlim())
    ax2.set_xticks(hour_labels_pos)
    ax2.set_xticklabels(hour_labels_text, fontsize=6, rotation=90)
    ax2.tick_params(axis='x', length=2)

    ax.set_title('계절 x 시간대별 송전망 이용률 히트맵 (복도별 평균, 내림차순)')
    ax.set_xlabel('계절 / 시간대')

    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  저장 -> {os.path.basename(out_path)}")


# ─────────────────────────────────────────────
#  차트 4 헬퍼: 권역별 시간별 전력 순수지(MW)
# ─────────────────────────────────────────────

def _compute_hourly_regional_balance(n):
    """권역별 시간별 전력 순수지(MW) DataFrame 반환 — 상단 chart4 와 동일 항목"""
    gp  = _gens_p(n)
    lp  = _loads_p(n)
    all_carriers = RE_CARRIERS | NONRE_CARRIERS | DR_CARRIERS

    result = {}
    for region in REGIONS:
        net = pd.Series(0.0, index=n.snapshots)

        # 발전 (RE + 비RE + DR)
        region_gens = n.generators[
            n.generators.index.str.startswith(region + '_') &
            n.generators['carrier'].isin(all_carriers)
        ].index
        for g in region_gens:
            if not gp.empty and g in gp.columns:
                net = net + gp[g]

        # CHP 전기 출력 (link p1 < 0 → -p1)
        chp = f'{region}_CHP'
        if (chp in n.links.index
                and hasattr(n.links_t, 'p1')
                and not n.links_t.p1.empty
                and chp in n.links_t.p1.columns):
            net = net + (-n.links_t.p1[chp]).clip(lower=0)

        # 수요 (EL + EV + FAB + DC)
        for load_name in [f'{region}_Demand_EL', f'{region}_Demand_EV',
                          f'{region}_Demand_Fab']:
            net = net - _get_load_ts(n, load_name)
        # DC 수요 (공백 변형 처리)
        dc_ts = _get_load_ts(n, f'{region}_Demand_DC ')
        if dc_ts.sum() == 0:
            dc_ts = _get_load_ts(n, f'{region}_Demand_DC')
        net = net - dc_ts

        result[region] = net

    return pd.DataFrame(result, index=n.snapshots)


# ─────────────────────────────────────────────
#  차트 4: 권역별 에너지 수지 누적막대 + 분포 boxplot
# ─────────────────────────────────────────────

def chart4_regional_balance(n, out_path):
    print("[차트4] 권역별 에너지 수지 생성 중...")
    gp = _gens_p(n)

    rows = []
    for region in REGIONS:
        row = {'region': region}

        # ── 생산 측 ──────────────────────────────────────
        # RE 발전
        re_gens = n.generators[
            n.generators.index.str.startswith(region + '_') &
            n.generators['carrier'].isin(RE_CARRIERS)
        ].index
        row['RE발전'] = sum(
            float(gp[g].sum()) for g in re_gens if g in gp.columns
        ) if not gp.empty else 0.0

        # 출력제한 (RE 포텐셜 - 실발전)
        curtail = 0.0
        for g in re_gens:
            if gp.empty or g not in gp.columns:
                continue
            p_actual = gp[g]
            p_nom    = n.generators.at[g, 'p_nom']
            if (hasattr(n.generators_t, 'p_max_pu')
                    and not n.generators_t.p_max_pu.empty
                    and g in n.generators_t.p_max_pu.columns):
                p_max = n.generators_t.p_max_pu[g] * p_nom
            else:
                p_max = pd.Series(p_nom, index=n.snapshots)
            curtail += float((p_max - p_actual).clip(lower=0).sum())
        row['출력제한'] = curtail

        # 비RE 발전 (coal, gas, nuclear, electricity/수력) + CHP 통합
        nonre_gens = n.generators[
            n.generators.index.str.startswith(region + '_') &
            n.generators['carrier'].isin(NONRE_CARRIERS)
        ].index
        nonre_gen = sum(
            float(gp[g].sum()) for g in nonre_gens if g in gp.columns
        ) if not gp.empty else 0.0

        # CHP 발전 (link p1 <0 → 전력 공급 → -p1)
        chp = f'{region}_CHP'
        chp_gen = 0.0
        if (chp in n.links.index
                and hasattr(n.links_t, 'p1')
                and not n.links_t.p1.empty
                and chp in n.links_t.p1.columns):
            chp_gen = float((-n.links_t.p1[chp]).clip(lower=0).sum())
        row['비RE발전'] = nonre_gen + chp_gen   # CHP 포함

        # DR
        dr_gens = n.generators[
            n.generators.index.str.startswith(region + '_') &
            n.generators['carrier'].isin(DR_CARRIERS)
        ].index
        row['DR'] = sum(
            float(gp[g].sum()) for g in dr_gens if g in gp.columns
        ) if not gp.empty else 0.0

        # ── 소비 측 ──────────────────────────────────────
        row['전력수요'] = _get_load_annual(n, f'{region}_Demand_EL')
        row['EV수요']   = _get_load_annual(n, f'{region}_Demand_EV')
        row['FAB수요']  = _get_load_annual(n, f'{region}_Demand_Fab')
        dc = _get_load_annual(n, f'{region}_Demand_DC ')
        if dc == 0:
            dc = _get_load_annual(n, f'{region}_Demand_DC')
        row['DC수요'] = dc

        rows.append(row)

    df = pd.DataFrame(rows).set_index('region')
    df = df / 1e3   # MWh → GWh

    PROD_COLS   = ['RE발전', '비RE발전', 'DR', '출력제한']
    CONS_COLS   = ['전력수요', 'EV수요', 'FAB수요', 'DC수요']
    PROD_COLORS = {
        'RE발전':  '#4CAF50',
        '비RE발전': '#795548',
        'DR':      '#03A9F4',
        '출력제한': '#FFEB3B',
    }
    CONS_COLORS = {
        '전력수요': '#F44336',
        'EV수요':   '#FF9800',
        'FAB수요':  '#673AB7',
        'DC수요':   '#E91E63',
    }

    # ── 상단·하단 서브플롯 (2:1) ──────────────────────────────────────
    fig, (ax, ax_bot) = plt.subplots(
        2, 1, figsize=(14, 10.5),
        gridspec_kw={'height_ratios': [2, 1]}
    )
    x = np.arange(len(REGIONS))

    # ── 상단: 연간 에너지 수지 누적막대 ──────────────────────────────

    # 생산 측 (양수) 누적 막대
    bottom_pos = np.zeros(len(REGIONS))
    for col in PROD_COLS:
        vals = df[col].reindex(REGIONS).fillna(0).values
        hatch = '///' if col == '출력제한' else None
        ax.bar(x, vals, bottom=bottom_pos, label=col,
               color=PROD_COLORS[col], alpha=0.85,
               hatch=hatch, edgecolor='white', linewidth=0.4)
        bottom_pos += vals

    # 소비 측 (음수) 누적 막대
    bottom_neg = np.zeros(len(REGIONS))
    for col in CONS_COLS:
        vals = df[col].reindex(REGIONS).fillna(0).values
        ax.bar(x, -vals, bottom=bottom_neg, label=col,
               color=CONS_COLORS[col], alpha=0.85,
               edgecolor='white', linewidth=0.4)
        bottom_neg -= vals

    # 0 기준선
    ax.axhline(0, color='black', lw=0.8)

    # 순수지(net) 계산 및 라벨 (0 기준선 인근에 배치)
    prod_total = df[PROD_COLS].reindex(REGIONS).fillna(0).sum(axis=1)
    cons_total = df[CONS_COLS].reindex(REGIONS).fillna(0).sum(axis=1)
    net_annual = prod_total - cons_total   # + = 공급우위, - = 수요우위

    YLIM   = 300_000   # GWh 고정 축 범위
    offset = YLIM * 0.015

    for i, region in enumerate(REGIONS):
        v = net_annual.loc[region]
        if abs(v) < 1e-3:
            continue
        if v > 0:
            ax.text(i, offset, f'+{v:.0f}',
                    ha='center', va='bottom', fontsize=7,
                    color='#1B5E20', fontweight='bold')
        else:
            ax.text(i, -offset, f'{v:.0f}',
                    ha='center', va='top', fontsize=7,
                    color='#B71C1C', fontweight='bold')

    ax.set_xticks(x)
    ax.set_xticklabels(REGIONS, fontsize=9)
    ax.set_ylim(-YLIM, YLIM)
    ax.set_ylabel('에너지 (GWh/년)')
    ax.set_title('권역별 에너지 수지\n( + : 생산 관련 | - : 소비 관련 | 숫자: 순수지 )')
    ax.legend(loc='upper right', ncol=4, fontsize=7,
              framealpha=0.85, edgecolor='#cccccc')
    ax.grid(axis='y', alpha=0.3)

    # ── 하단: 시간별 전력 순수지 분포 boxplot ─────────────────────────
    hourly_bal = _compute_hourly_regional_balance(n)   # MW, 8760×len(REGIONS)

    # 권역별 통계
    stats_bot = pd.DataFrame({
        'min':    hourly_bal.reindex(columns=REGIONS).min(),
        'p25':    hourly_bal.reindex(columns=REGIONS).quantile(0.25),
        'median': hourly_bal.reindex(columns=REGIONS).median(),
        'p75':    hourly_bal.reindex(columns=REGIONS).quantile(0.75),
        'max':    hourly_bal.reindex(columns=REGIONS).max(),
        'mean':   hourly_bal.reindex(columns=REGIONS).mean(),
    })

    # 평균 순수지에 따른 색상: 공급 우위=녹색 계열, 수요 우위=적색 계열
    def _region_color_bot(mean_val):
        if mean_val >= 3000:
            return '#1B5E20'
        elif mean_val >= 1000:
            return '#4CAF50'
        elif mean_val >= 0:
            return '#A5D6A7'
        elif mean_val >= -1000:
            return '#FFCDD2'
        elif mean_val >= -3000:
            return '#EF9A9A'
        else:
            return '#B71C1C'

    colors_bot = [_region_color_bot(stats_bot.at[r, 'mean']) for r in REGIONS]

    xb = np.arange(len(REGIONS))

    # IQR 박스
    for i, region in enumerate(REGIONS):
        s = stats_bot.loc[region]
        ax_bot.bar(xb[i],
                   s['p75'] - s['p25'],
                   bottom=s['p25'],
                   color=colors_bot[i], alpha=0.75, width=0.6, zorder=3)

    # 최소~최대 선
    ax_bot.vlines(xb,
                  stats_bot['min'].values,
                  stats_bot['max'].values,
                  colors=[c for c in colors_bot],
                  lw=1.5, zorder=2)

    # 최소·최대 캡
    cap_w = 0.2
    ax_bot.hlines(stats_bot['min'].values, xb - cap_w, xb + cap_w,
                  colors='#444', lw=1.2, zorder=4)
    ax_bot.hlines(stats_bot['max'].values, xb - cap_w, xb + cap_w,
                  colors='#444', lw=1.2, zorder=4)

    # 중앙값 선
    ax_bot.hlines(stats_bot['median'].values, xb - 0.28, xb + 0.28,
                  colors='white', lw=2.5, zorder=5)
    ax_bot.hlines(stats_bot['median'].values, xb - 0.28, xb + 0.28,
                  colors='#333',  lw=1.5, zorder=6)

    # 평균 점
    ax_bot.scatter(xb, stats_bot['mean'].values,
                   s=50, c=colors_bot, edgecolors='black', linewidths=0.6,
                   zorder=7, marker='D')

    # 0 기준선
    ax_bot.axhline(0, color='black', lw=0.8, linestyle='--', alpha=0.6)

    ax_bot.set_xticks(xb)
    ax_bot.set_xticklabels(REGIONS, fontsize=9)
    ax_bot.set_ylabel('시간별 순수지 (MW)')
    ax_bot.set_title('권역별 에너지 수지 분포\n(■ IQR 범위, ◆ 평균, ─ 중앙값, | 최소~최대)')
    ax_bot.grid(axis='y', alpha=0.3, zorder=1)

    # 색상 범례 (하단)
    bot_patches = [
        mpatches.Patch(fc='#1B5E20', alpha=0.75, label='공급 우위 >3,000 MW'),
        mpatches.Patch(fc='#4CAF50', alpha=0.75, label='공급 우위 1,000~3,000 MW'),
        mpatches.Patch(fc='#A5D6A7', alpha=0.75, label='공급 우위 0~1,000 MW'),
        mpatches.Patch(fc='#FFCDD2', alpha=0.75, label='수요 우위 0~1,000 MW'),
        mpatches.Patch(fc='#EF9A9A', alpha=0.75, label='수요 우위 1,000~3,000 MW'),
        mpatches.Patch(fc='#B71C1C', alpha=0.75, label='수요 우위 >3,000 MW'),
    ]
    ax_bot.legend(handles=bot_patches, title='평균 순수지',
                  loc='upper right', fontsize=6, title_fontsize=6, ncol=2)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  저장 -> {os.path.basename(out_path)}")


# ─────────────────────────────────────────────
#  차트 5: 섹터커플링 이용률 분석
# ─────────────────────────────────────────────

SC_DEF = {
    'Electrolyser': {'label': '수전해(P2H2)',  'color': '#1565C0'},
    'HP':           {'label': '히트펌프(P2H)', 'color': '#FF5722'},
    'ESS_Ch':       {'label': 'ESS 충전',      'color': '#388E3C'},
    'ESS_Disch':    {'label': 'ESS 방전',      'color': '#81C784'},
    'EV_Charger':   {'label': 'EV 충전',       'color': '#F57C00'},
    'EV_V2G':       {'label': 'EV V2G',        'color': '#FFB74D'},
}


def _build_sc_ts(n):
    if not (hasattr(n.links_t, 'p0') and not n.links_t.p0.empty):
        return pd.DataFrame(index=n.snapshots)
    sc = {}
    for comp in SC_DEF:
        cols = [c for c in n.links_t.p0.columns if c.endswith(f'_{comp}')]
        if cols:
            sc[comp] = n.links_t.p0[cols].sum(axis=1).abs()
    return pd.DataFrame(sc, index=n.snapshots) if sc else pd.DataFrame(index=n.snapshots)


def chart5_sector_coupling(n, out_path):
    print("[차트5] 섹터커플링 이용률 분석 생성 중...")
    sc_df = _build_sc_ts(n)
    if sc_df.empty:
        print("  섹터커플링 데이터 없음 — 스킵")
        return

    active = [c for c in sc_df.columns if sc_df[c].sum() > 0]
    if not active:
        print("  모든 섹터커플링 출력 = 0 — 스킵")
        return

    seasons = get_season_series(n.snapshots)
    fig, axes = plt.subplots(2, 2, figsize=(15, 10))
    fig.suptitle('섹터커플링 이용률 분석', fontsize=13, fontweight='bold')

    # (0,0) 지속곡선
    ax = axes[0, 0]
    for comp in active:
        sv = np.sort(sc_df[comp].values)[::-1]
        ax.plot(np.arange(1, len(sv) + 1), sv,
                label=SC_DEF[comp]['label'],
                color=SC_DEF[comp]['color'], lw=1.4, alpha=0.9)
    ax.set_xlabel('시간 (내림차순 정렬)')
    ax.set_ylabel('출력 (MW)')
    ax.set_title('섹터커플링 지속곡선 (Duration Curve)')
    ax.legend()
    ax.grid(alpha=0.3)
    ax.set_xlim(0, len(n.snapshots))

    # (0,1) 월별 평균 출력
    ax = axes[0, 1]
    sc_monthly = sc_df.copy()
    sc_monthly['month'] = n.snapshots.month
    monthly_avg = sc_monthly.groupby('month')[active].mean()
    bottom = np.zeros(12)
    months = list(range(1, 13))
    for comp in active:
        vals = monthly_avg[comp].reindex(months).fillna(0).values
        ax.bar(months, vals, bottom=bottom,
               label=SC_DEF[comp]['label'],
               color=SC_DEF[comp]['color'], alpha=0.85,
               edgecolor='white', lw=0.3)
        bottom += vals
    ax.set_xlabel('월')
    ax.set_ylabel('평균 출력 (MW)')
    ax.set_title('월별 섹터커플링 활용 분해')
    ax.set_xticks(months)
    ax.set_xticklabels([f'{m}월' for m in months], fontsize=7)
    ax.legend()
    ax.grid(axis='y', alpha=0.3)

    # (1,0) RE 잉여 vs SC 소비 산점도
    ax = axes[1, 0]
    re_ts = compute_re_output_ts(n)
    el_loads = [c for c in n.loads.index if '_Demand_EL' in c]
    lp = _loads_p(n)
    el_demand = pd.Series(0.0, index=n.snapshots)
    for ld in el_loads:
        if not lp.empty and ld in lp.columns:
            el_demand += lp[ld]
        else:
            pset = n.loads.at[ld, 'p_set'] if ld in n.loads.index else 0.0
            el_demand += float(pset if pd.notna(pset) else 0.0)

    re_surplus = re_ts - el_demand
    sc_total   = sc_df[active].sum(axis=1)

    for s in SEASON_ORDER:
        m = seasons == s
        ax.scatter(re_surplus[m], sc_total[m],
                   c=SEASON_COLORS[s], alpha=0.18, s=5, label=s)
    ax.axvline(0, color='#888', lw=0.8, linestyle='--')
    ax.set_xlabel('RE 잉여량 (MW)\n(RE 발전량 - 전력 수요)')
    ax.set_ylabel('섹터커플링 총 소비 (MW)')
    ax.set_title('RE 잉여량 vs 섹터커플링 소비\n(계절별 색상)')
    ax.legend(markerscale=3)
    ax.grid(alpha=0.3)

    # (1,1) 계절별 SC 소비 분해
    ax = axes[1, 1]
    sc_seas = sc_df[active].copy()
    sc_seas['season'] = seasons.values
    seas_sum = sc_seas.groupby('season')[active].sum() / 1e3
    seas_sum = seas_sum.reindex(SEASON_ORDER)
    x = np.arange(4)
    bottom = np.zeros(4)
    for comp in active:
        vals = seas_sum[comp].values
        ax.bar(x, vals, bottom=bottom,
               label=SC_DEF[comp]['label'],
               color=SC_DEF[comp]['color'], alpha=0.85,
               edgecolor='white', lw=0.3)
        bottom += vals
    ax.set_xticks(x)
    ax.set_xticklabels(SEASON_ORDER)
    ax.set_ylabel('에너지 소비 (GWh)')
    ax.set_title('계절별 섹터커플링 소비 분해')
    ax.legend()
    ax.grid(axis='y', alpha=0.3)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  저장 -> {os.path.basename(out_path)}")


# ─────────────────────────────────────────────
#  공개 API: PyPSA_GUI.py 에서 직접 호출 가능
# ─────────────────────────────────────────────

_CHART_DEFS = [
    ('viz_01_line_utilization.png', chart1_line_utilization),
    ('viz_02_re_vs_tx.png',         chart2_re_vs_tx),
    ('viz_03_seasonal_heatmap.png', chart3_seasonal_heatmap),
    ('viz_04_regional_balance.png', chart4_regional_balance),
    ('viz_05_sector_coupling.png',  chart5_sector_coupling),
]


def run_all_charts(network, out_dir):
    """
    5개 차트를 out_dir 에 일괄 생성.

    Parameters
    ----------
    network : pypsa.Network
        최적화 완료된 네트워크 객체 (nc 파일 로딩 불필요).
    out_dir : str
        저장 폴더 경로 (없으면 자동 생성).
    """
    os.makedirs(out_dir, exist_ok=True)
    _setup_font()
    print(f"\n{'='*60}")
    print(f"[커스텀 시각화] 5개 차트 생성 → {out_dir}")
    print(f"{'='*60}")

    ok = fail = 0
    for fname, fn in _CHART_DEFS:
        try:
            fn(network, os.path.join(out_dir, fname))
            ok += 1
        except Exception as _e:
            print(f"  [{fname}] 오류 (스킵): {_e}")
            fail += 1

    print(f"{'='*60}")
    print(f"[커스텀 시각화] 완료 {ok}개 생성" +
          (f", {fail}개 오류" if fail else ""))
    print(f"{'='*60}\n")


# ─────────────────────────────────────────────
#  메인
# ─────────────────────────────────────────────

def main():
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        results_dir = arg if os.path.isabs(arg) else os.path.join(SCRIPT_DIR, arg)
        ncs = glob.glob(os.path.join(results_dir, 'optimization_result_*.nc'))
        if not ncs:
            raise FileNotFoundError(f".nc 파일 없음: {results_dir}")
        nc_path = ncs[0]
    else:
        results_dir, nc_path = find_latest_nc()

    print(f"\n{'='*60}")
    print(f"결과 폴더 : {results_dir}")
    print(f"NC 파일   : {os.path.basename(nc_path)}")
    print(f"{'='*60}")

    n = load_network(nc_path)

    # 임시 저장 폴더 생성 (수정 확인용 — 덮어씌우며 반복 검토)
    os.makedirs(VIZ_TEMP_DIR, exist_ok=True)
    print(f"[저장 위치] {VIZ_TEMP_DIR}")

    def out(fname):
        return os.path.join(VIZ_TEMP_DIR, fname)

    # nc 파일의 s_nom(num_parallel 반영 실제 제약 용량) 기준으로 이용률 계산
    chart1_line_utilization(n,  out('viz_01_line_utilization.png'))
    chart2_re_vs_tx(n,          out('viz_02_re_vs_tx.png'))
    chart3_seasonal_heatmap(n,  out('viz_03_seasonal_heatmap.png'))
    chart4_regional_balance(n,  out('viz_04_regional_balance.png'))
    chart5_sector_coupling(n,   out('viz_05_sector_coupling.png'))

    print(f"\n{'='*60}")
    print(f"[완료] 시각화 5개 파일 -> {VIZ_TEMP_DIR}")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
