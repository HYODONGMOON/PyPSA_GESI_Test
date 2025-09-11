import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def copy_region_data_improved():
    """
    지역_BSN 시트를 정확히 복제하여 다른 지역 시트를 만들고,
    버스 이름 및 참조를 각 지역 코드에 맞게 변경합니다.
    """
    try:
        # 파일 경로 설정
        template_file = 'regional_input_template.xlsx'
        
        # 지역 이름 매핑 (지역 코드 -> 지역 이름)
        region_names = {
            'SEL': '서울특별시',
            'BSN': '부산광역시',
            'DGU': '대구광역시',
            'ICN': '인천광역시',
            'GWJ': '광주광역시',
            'DJN': '대전광역시',
            'USN': '울산광역시',
            'SJN': '세종특별자치시',
            'GGD': '경기도',
            'GWD': '강원도',
            'CBD': '충청북도',
            'CND': '충청남도',
            'JBD': '전라북도',
            'JND': '전라남도',
            'GBD': '경상북도',
            'GND': '경상남도',
            'JJD': '제주특별자치도'
        }
        
        # 파일 존재 확인
        if not os.path.exists(template_file):
            print(f"오류: '{template_file}' 파일을 찾을 수 없습니다.")
            return False
            
        # 워크북 열기
        print(f"'{template_file}' 파일을 열고 있습니다...")
        wb = openpyxl.load_workbook(template_file)
        
        # 지역_BSN 시트 확인
        if '지역_BSN' not in wb.sheetnames:
            print(f"오류: '지역_BSN' 시트를 찾을 수 없습니다.")
            return False
        
        # 모든 지역 시트 목록 가져오기
        region_sheets = []
        for sheet in wb.sheetnames:
            if sheet.startswith('지역_') and sheet != '지역_BSN':
                region_sheets.append(sheet)
                
        print(f"찾은 지역 시트: {', '.join(region_sheets)}")
        
        # 지역 코드 추출
        region_codes = [sheet.split('_')[1] for sheet in region_sheets]
        print(f"지역 코드: {', '.join(region_codes)}")
        
        # 지역_BSN 시트 가져오기
        bsn_sheet = wb['지역_BSN']
        
        # A22:A25 셀의 값을 저장 (발전기 이름)
        bsn_generator_names = {}
        for row in range(22, 26):
            cell_value = bsn_sheet.cell(row=row, column=1).value
            if cell_value:
                bsn_generator_names[row] = cell_value
                
        print(f"BSN 발전기 이름 (A22:A25): {bsn_generator_names}")
        
        # 각 섹션의 시작 행과 끝 행 찾기
        sections = {}
        current_section = None
        
        for row_idx, row in enumerate(bsn_sheet.iter_rows(min_row=1, max_row=bsn_sheet.max_row, values_only=True), 1):
            if row[0] == "버스":
                sections["buses"] = {"start": row_idx}
                if current_section:
                    sections[current_section]["end"] = row_idx - 1
                current_section = "buses"
            elif row[0] == "발전기":
                sections["generators"] = {"start": row_idx}
                if current_section:
                    sections[current_section]["end"] = row_idx - 1
                current_section = "generators"
            elif row[0] == "부하":
                sections["loads"] = {"start": row_idx}
                if current_section:
                    sections[current_section]["end"] = row_idx - 1
                current_section = "loads"
            elif row[0] == "저장장치":
                sections["stores"] = {"start": row_idx}
                if current_section:
                    sections[current_section]["end"] = row_idx - 1
                current_section = "stores"
            elif row[0] == "링크":
                sections["links"] = {"start": row_idx}
                if current_section:
                    sections[current_section]["end"] = row_idx - 1
                current_section = "links"
        
        # 마지막 섹션의 끝 설정
        if current_section:
            sections[current_section]["end"] = bsn_sheet.max_row
            
        print(f"섹션 정보: {sections}")
        
        # 각 지역 시트 처리
        for sheet_name in region_sheets:
            region_code = sheet_name.split('_')[1]
            print(f"'{sheet_name}' 시트 처리 중... (지역 코드: {region_code})")
            
            # 기존 시트가 있으면 삭제
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            
            # 새 시트 생성
            new_sheet = wb.create_sheet(sheet_name)
            
            # 지역_BSN 시트의 모든 내용과 서식 복사
            for row_idx in range(1, bsn_sheet.max_row + 1):
                for col_idx in range(1, bsn_sheet.max_column + 1):
                    # 셀 복사
                    source_cell = bsn_sheet.cell(row=row_idx, column=col_idx)
                    target_cell = new_sheet.cell(row=row_idx, column=col_idx)
                    
                    # 값 복사
                    target_cell.value = source_cell.value
                    
                    # 스타일 복사
                    if source_cell.has_style:
                        target_cell.font = Font(
                            name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            color=source_cell.font.color
                        )
                        target_cell.alignment = Alignment(
                            horizontal=source_cell.alignment.horizontal,
                            vertical=source_cell.alignment.vertical,
                            wrap_text=source_cell.alignment.wrap_text
                        )
                        if source_cell.border:
                            target_cell.border = Border(
                                left=Side(style=source_cell.border.left.style, color=source_cell.border.left.color) if source_cell.border.left else None,
                                right=Side(style=source_cell.border.right.style, color=source_cell.border.right.color) if source_cell.border.right else None,
                                top=Side(style=source_cell.border.top.style, color=source_cell.border.top.color) if source_cell.border.top else None,
                                bottom=Side(style=source_cell.border.bottom.style, color=source_cell.border.bottom.color) if source_cell.border.bottom else None
                            )
                        if source_cell.fill and source_cell.fill.fill_type:
                            if source_cell.fill.fill_type == 'solid':
                                target_cell.fill = PatternFill(
                                    fill_type=source_cell.fill.fill_type,
                                    start_color=source_cell.fill.start_color,
                                    end_color=source_cell.fill.end_color
                                )
            
            # 병합된 셀 복사
            for merged_cell_range in bsn_sheet.merged_cells.ranges:
                new_sheet.merge_cells(str(merged_cell_range))
            
            # 열 너비 복사
            for col_idx in range(1, bsn_sheet.max_column + 1):
                col_letter = get_column_letter(col_idx)
                new_sheet.column_dimensions[col_letter].width = bsn_sheet.column_dimensions[col_letter].width
            
            # 행 높이 복사
            for row_idx in range(1, bsn_sheet.max_row + 1):
                if row_idx in bsn_sheet.row_dimensions:
                    new_sheet.row_dimensions[row_idx].height = bsn_sheet.row_dimensions[row_idx].height
                    
            # 1행 제목 변경 (지역 이름으로)
            region_name = region_names.get(region_code, f"{region_code} 지역")
            title_cell = new_sheet.cell(row=1, column=1)
            if title_cell.value and "에너지시스템 입력" in str(title_cell.value):
                title_cell.value = f"{region_name} 에너지시스템 입력"
            
            # A3:B5 셀의 값 비우기 (정보 제공용 셀)
            print(f"  A3:B5 셀의 값 비우기")
            for row in range(3, 6):
                for col in range(1, 3):  # A, B 열
                    cell = new_sheet.cell(row=row, column=col)
                    # 셀 값은 비우지만 서식은 유지
                    cell.value = None
            
            # 특정 셀 위치의 버스 이름 변경
            # A10-A12 (버스 섹션의 버스 이름)
            bus_names = {
                "A10": f"{region_code}_EL",  # 전력 버스
                "A11": f"{region_code}_H2",  # 수소 버스
                "A12": f"{region_code}_H"    # 열 버스
            }
            for cell_addr, new_value in bus_names.items():
                cell = new_sheet[cell_addr]
                cell.value = new_value
            
            # A22:A25 셀의 값을 BSN에서 복사 (발전기 이름 유지)
            for row, value in bsn_generator_names.items():
                cell = new_sheet.cell(row=row, column=1)
                cell.value = value
                print(f"  A{row} 셀 값 설정: {value}")
            
            # B19-B25 (발전기 섹션의 버스 참조)
            for row in range(19, 26):
                cell = new_sheet.cell(row=row, column=2)
                if cell.value:
                    cell.value = f"{region_code}_EL"
            
            # B32-B34 (부하 섹션의 버스 참조)
            load_buses = {
                32: f"{region_code}_EL",  # 전력 부하
                33: f"{region_code}_H",   # 열 부하
                34: f"{region_code}_H2"   # 수소 부하
            }
            for row, new_value in load_buses.items():
                cell = new_sheet.cell(row=row, column=2)
                cell.value = new_value
            
            # B41-B43 (저장장치 섹션의 버스 참조)
            store_buses = {
                41: f"{region_code}_EL",  # 전력 저장장치
                42: f"{region_code}_H2",  # 수소 저장장치
                43: f"{region_code}_H"    # 열 저장장치
            }
            for row, new_value in store_buses.items():
                cell = new_sheet.cell(row=row, column=2)
                cell.value = new_value
            
            # B52-B53 (링크 섹션의 버스0 참조)
            for row in range(52, 54):
                cell = new_sheet.cell(row=row, column=2)
                cell.value = f"{region_code}_EL"
            
            # C52-C53 (링크 섹션의 버스1 참조)
            link_buses = {
                52: f"{region_code}_H2",  # 전력->수소 변환
                53: f"{region_code}_H"    # 전력->열 변환
            }
            for row, new_value in link_buses.items():
                cell = new_sheet.cell(row=row, column=3)
                cell.value = new_value
            
            # 이제 각 섹션의 데이터에서 버스 이름과 참조를 지역 코드로 업데이트
            for section_name, section_info in sections.items():
                # 헤더 행 찾기 (섹션 시작 후 5번째 행)
                header_row = section_info["start"] + 5
                data_start = header_row + 1
                data_end = section_info["end"]
                
                # 데이터 행이 있는지 확인
                if data_start <= data_end:
                    print(f"  {section_name} 섹션 업데이트 중... (행 {data_start}-{data_end})")
                    
                    # 각 데이터 행 처리
                    for row_idx in range(data_start, data_end + 1):
                        # A22:A25 범위는 건너뛰기 (이미 위에서 처리함)
                        if 22 <= row_idx <= 25:
                            continue
                            
                        # 첫 번째 열이 비어있지 않은 경우만 처리
                        name_cell = new_sheet.cell(row=row_idx, column=1)
                        if name_cell.value:
                            old_name = str(name_cell.value)
                            
                            # 이름에 지역 코드 추가
                            if "_" in old_name:  # 이미 지역 코드가 있는 경우
                                parts = old_name.split('_')
                                new_name = f"{region_code}_{parts[1]}"
                            else:
                                new_name = f"{region_code}_{old_name}"
                            
                            # 이름 업데이트
                            name_cell.value = new_name
                            
                            # 버스 참조 업데이트 (발전기, 부하, 저장장치, 링크)
                            if section_name in ["generators", "loads", "stores"]:
                                bus_cell = new_sheet.cell(row=row_idx, column=2)
                                if bus_cell.value:
                                    old_bus = str(bus_cell.value)
                                    if "_" in old_bus:
                                        parts = old_bus.split('_')
                                        new_bus = f"{region_code}_{parts[1]}"
                                    else:
                                        new_bus = f"{region_code}_{old_bus}"
                                    bus_cell.value = new_bus
                            
                            # 링크의 경우 여러 버스 참조 업데이트
                            elif section_name == "links":
                                # 버스0, 버스1, 버스2, 버스3 업데이트 (열 2,3,4,5)
                                for bus_col in range(2, 6):
                                    bus_cell = new_sheet.cell(row=row_idx, column=bus_col)
                                    if bus_cell.value:
                                        old_bus = str(bus_cell.value)
                                        if "_" in old_bus:
                                            parts = old_bus.split('_')
                                            new_bus = f"{region_code}_{parts[1]}"
                                        else:
                                            new_bus = f"{region_code}_{old_bus}"
                                        bus_cell.value = new_bus
        
        # 변경사항 저장
        wb.save(template_file)
        print(f"모든 지역 시트가 성공적으로 업데이트되었습니다.")
        return True
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    copy_region_data_improved() 