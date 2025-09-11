#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
송전선 데이터를 regional_input_template.xlsx 파일에 추가하는 스크립트

create_lines_data.py에서 생성한 송전선 데이터를 regional_input_template.xlsx 파일의 '지역간 연결' 시트에 추가합니다.
"""

import pandas as pd
import numpy as np
import math
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 한국 행정구역 정보
KOREA_REGIONS = {
    'SEL': {
        'name': '서울특별시',
        'name_eng': 'Seoul',
        'center': (126.986, 37.566),
        'area': 605.21,  # km²
        'population': 9720846,  # 2021년 기준
        'color': 'lightcoral'
    },
    'BSN': {
        'name': '부산광역시',
        'name_eng': 'Busan',
        'center': (129.075, 35.180),
        'area': 770.04,
        'population': 3385601,
        'color': 'lightskyblue'
    },
    'DGU': {
        'name': '대구광역시',
        'name_eng': 'Daegu',
        'center': (128.601, 35.871),
        'area': 883.56,
        'population': 2418346,
        'color': 'lightgreen'
    },
    'ICN': {
        'name': '인천광역시',
        'name_eng': 'Incheon',
        'center': (126.705, 37.456),
        'area': 1063.27,
        'population': 2947217,
        'color': 'yellow'
    },
    'GWJ': {
        'name': '광주광역시',
        'name_eng': 'Gwangju',
        'center': (126.851, 35.160),
        'area': 501.24,
        'population': 1441611,
        'color': 'pink'
    },
    'DJN': {
        'name': '대전광역시',
        'name_eng': 'Daejeon',
        'center': (127.385, 36.351),
        'area': 539.84,
        'population': 1463882,
        'color': 'lightseagreen'
    },
    'USN': {
        'name': '울산광역시',
        'name_eng': 'Ulsan',
        'center': (129.311, 35.539),
        'area': 1062.04,
        'population': 1136017,
        'color': 'plum'
    },
    'SJN': {
        'name': '세종특별자치시',
        'name_eng': 'Sejong',
        'center': (127.289, 36.480),
        'area': 465.23,
        'population': 365309,
        'color': 'orange'
    },
    'GGD': {
        'name': '경기도',
        'name_eng': 'Gyeonggi',
        'center': (127.013, 37.275),
        'area': 10191.79,
        'population': 13530519,
        'color': 'lightsalmon'
    },
    'GWD': {
        'name': '강원도',
        'name_eng': 'Gangwon',
        'center': (128.318, 37.883),
        'area': 16826.37,
        'population': 1538492,
        'color': 'lightcyan'
    },
    'CBD': {
        'name': '충청북도',
        'name_eng': 'Chungbuk',
        'center': (127.705, 36.801),
        'area': 7407.30,
        'population': 1600957,
        'color': 'wheat'
    },
    'CND': {
        'name': '충청남도',
        'name_eng': 'Chungnam',
        'center': (126.799, 36.658),
        'area': 8245.41,
        'population': 2118670,
        'color': 'lightsteelblue'
    },
    'JBD': {
        'name': '전라북도',
        'name_eng': 'Jeonbuk',
        'center': (127.144, 35.820),
        'area': 8069.07,
        'population': 1792712,
        'color': 'mediumaquamarine'
    },
    'JND': {
        'name': '전라남도',
        'name_eng': 'Jeonnam',
        'center': (126.991, 34.868),
        'area': 12344.06,
        'population': 1851549,
        'color': 'thistle'
    },
    'GBD': {
        'name': '경상북도',
        'name_eng': 'Gyeongbuk',
        'center': (128.744, 36.566),
        'area': 19030.80,
        'population': 2641823,
        'color': 'paleturquoise'
    },
    'GND': {
        'name': '경상남도',
        'name_eng': 'Gyeongnam',
        'center': (128.241, 35.462),
        'area': 10540.29,
        'population': 3340216,
        'color': 'peachpuff'
    },
    'JJD': {
        'name': '제주특별자치도',
        'name_eng': 'Jeju',
        'center': (126.542, 33.387),
        'area': 1849.16,
        'population': 674635,
        'color': 'mediumpurple'
    },
}

# 인접 지역 정의 (지리적으로 인접한 지역들)
ADJACENT_REGIONS = {
    'SEL': ['GGD', 'ICN'],
    'BSN': ['GND', 'USN'],
    'DGU': ['GBD', 'GND'],
    'ICN': ['SEL', 'GGD'],
    'GWJ': ['JND', 'JBD'],
    'DJN': ['CND', 'CBD', 'SJN'],
    'USN': ['BSN', 'GBD', 'GND'],
    'SJN': ['CND', 'CBD', 'DJN'],
    'GGD': ['SEL', 'ICN', 'GWD', 'CBD', 'CND'],
    'GWD': ['GGD', 'CBD', 'GBD'],
    'CBD': ['GGD', 'GWD', 'GBD', 'CND', 'DJN', 'SJN'],
    'CND': ['GGD', 'CBD', 'JBD', 'SJN', 'DJN'],
    'JBD': ['CND', 'JND', 'GND', 'GWJ'],
    'JND': ['JBD', 'GND', 'GWJ'],
    'GBD': ['GWD', 'CBD', 'GND', 'USN', 'DGU'],
    'GND': ['GBD', 'JBD', 'JND', 'BSN', 'USN', 'DGU'],
    'JJD': []  # 제주도는 육지와 인접하지 않음
}

def calculate_distance(region_code1, region_code2):
    """두 지역 간 거리 계산 (하버사인 공식)"""
    if region_code1 not in KOREA_REGIONS or region_code2 not in KOREA_REGIONS:
        return None
    
    # 두 지역의 중심 좌표
    lon1, lat1 = KOREA_REGIONS[region_code1]['center']
    lon2, lat2 = KOREA_REGIONS[region_code2]['center']
    
    # 하버사인 공식
    R = 6371.0  # 지구 반지름 (km)
    
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)
    
    dlon = lon2_rad - lon1_rad
    dlat = lat2_rad - lat1_rad
    
    a = math.sin(dlat / 2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    
    distance = R * c
    return round(distance, 1)  # 소수점 첫째자리까지 반올림

def create_lines_data():
    """지역 간 송전선 데이터 생성"""
    lines_data = []
    
    # 모든 인접 지역 쌍에 대해 송전선 생성
    for region1, adjacent_regions in ADJACENT_REGIONS.items():
        for region2 in adjacent_regions:
            # 중복 방지 (A-B와 B-A는 동일한 연결)
            if region1 < region2:
                distance = calculate_distance(region1, region2)
                
                # 송전선 데이터 생성
                line_data = {
                    'name': f"{region1}_{region2}",
                    'region1': region1,
                    'bus1': f"{region1}_EL",
                    'region2': region2,
                    'bus2': f"{region2}_EL",
                    'capacity': 1000,  # 기본 용량 1000MW
                    'voltage': 345,    # 기본 전압 345kV
                    'distance': distance,
                    'reactance': round(0.0004 * distance, 5),
                    'resistance': round(0.0001 * distance, 5)
                }
                lines_data.append(line_data)
    
    # 제주도는 해저 케이블로 육지와 연결 (전남과 연결)
    jeju_haenam_distance = calculate_distance('JJD', 'JND')
    jeju_line = {
        'name': "JJD_JND",
        'region1': 'JJD',
        'bus1': "JJD_EL",
        'region2': 'JND',
        'bus2': "JND_EL",
        'capacity': 700,     # HVDC 용량
        'voltage': 500,      # HVDC 전압
        'distance': jeju_haenam_distance,
        'reactance': 0,      # DC는 리액턴스 없음
        'resistance': round(0.0002 * jeju_haenam_distance, 5)
    }
    lines_data.append(jeju_line)
    
    return lines_data

def add_lines_to_template():
    """송전선 데이터를 regional_input_template.xlsx 파일에 추가"""
    template_file = 'regional_input_template.xlsx'
    
    # 템플릿 파일이 없으면 오류
    if not os.path.exists(template_file):
        print(f"오류: '{template_file}' 파일이 존재하지 않습니다.")
        return False
    
    try:
        # 워크북 로드
        wb = openpyxl.load_workbook(template_file)
        
        # '지역간 연결' 시트 확인
        if '지역간 연결' not in wb.sheetnames:
            print(f"오류: '{template_file}' 파일에 '지역간 연결' 시트가 없습니다.")
            return False
        
        # 시트 선택
        ws = wb['지역간 연결']
        
        # 기존 데이터 지우기 (헤더 아래부터)
        for row in range(6, ws.max_row + 1):
            for col in range(1, 11):
                ws.cell(row=row, column=col).value = None
        
        # 송전선 데이터 생성
        lines_data = create_lines_data()
        
        # 데이터 추가
        for i, line in enumerate(lines_data):
            row = i + 6  # 데이터는 6행부터 시작
            
            # 데이터 입력
            ws.cell(row=row, column=1).value = line['name']
            ws.cell(row=row, column=2).value = line['region1']
            ws.cell(row=row, column=3).value = line['bus1']
            ws.cell(row=row, column=4).value = line['region2']
            ws.cell(row=row, column=5).value = line['bus2']
            ws.cell(row=row, column=6).value = line['capacity']
            ws.cell(row=row, column=7).value = line['voltage']
            ws.cell(row=row, column=8).value = line['distance']
            ws.cell(row=row, column=9).value = line['reactance']
            ws.cell(row=row, column=10).value = line['resistance']
        
        # 파일 저장
        wb.save(template_file)
        
        print(f"송전선 데이터가 '{template_file}'의 '지역간 연결' 시트에 추가되었습니다.")
        print(f"총 {len(lines_data)}개의 송전선이 추가되었습니다.")
        return True
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    add_lines_to_template() 