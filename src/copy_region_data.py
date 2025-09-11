import pandas as pd
import os
import openpyxl

def copy_region_data():
    """
    지역_BSN 시트의 데이터를 모든 지역 시트에 복사하고, 
    버스 이름 및 참조를 각 지역 코드에 맞게 변경합니다.
    """
    try:
        # 파일 경로 설정
        template_file = 'regional_input_template.xlsx'
        
        # 파일 존재 확인
        if not os.path.exists(template_file):
            print(f"오류: '{template_file}' 파일을 찾을 수 없습니다.")
            return False
            
        # pandas로 먼저 데이터 읽기 (병합된 셀 문제 회피)
        print("BSN 시트에서 데이터 읽는 중...")
        
        # 각 섹션의 시작 행 찾기
        xl = pd.ExcelFile(template_file)
        sheet_data = pd.read_excel(xl, sheet_name='지역_BSN', header=None)
        
        section_starts = {}
        for idx, row in sheet_data.iterrows():
            if row.iloc[0] == "버스":
                section_starts["buses"] = idx
            elif row.iloc[0] == "발전기":
                section_starts["generators"] = idx
            elif row.iloc[0] == "부하":
                section_starts["loads"] = idx
            elif row.iloc[0] == "저장장치":
                section_starts["stores"] = idx
            elif row.iloc[0] == "링크":
                section_starts["links"] = idx
        
        print(f"섹션 시작 행: {section_starts}")
        
        # 각 섹션 데이터 읽기
        header_row = 5  # 헤더는 각 섹션 제목 이후 5번째 행
        
        buses_df = None
        generators_df = None
        loads_df = None
        stores_df = None
        links_df = None
        
        if "buses" in section_starts and "generators" in section_starts:
            start_row = section_starts["buses"] + header_row
            end_row = section_starts["generators"]
            buses_df = pd.read_excel(template_file, sheet_name='지역_BSN', header=None, skiprows=start_row, nrows=end_row-start_row)
            # 비어 있는 행 제거
            if not buses_df.empty and buses_df.shape[1] > 0:
                buses_df = buses_df.dropna(subset=[0])
        
        if "generators" in section_starts and "loads" in section_starts:
            start_row = section_starts["generators"] + header_row
            end_row = section_starts["loads"]
            generators_df = pd.read_excel(template_file, sheet_name='지역_BSN', header=None, skiprows=start_row, nrows=end_row-start_row)
            if not generators_df.empty and generators_df.shape[1] > 0:
                generators_df = generators_df.dropna(subset=[0])
        
        if "loads" in section_starts and "stores" in section_starts:
            start_row = section_starts["loads"] + header_row
            end_row = section_starts["stores"]
            loads_df = pd.read_excel(template_file, sheet_name='지역_BSN', header=None, skiprows=start_row, nrows=end_row-start_row)
            if not loads_df.empty and loads_df.shape[1] > 0:
                loads_df = loads_df.dropna(subset=[0])
        
        if "stores" in section_starts and "links" in section_starts:
            start_row = section_starts["stores"] + header_row
            end_row = section_starts["links"]
            stores_df = pd.read_excel(template_file, sheet_name='지역_BSN', header=None, skiprows=start_row, nrows=end_row-start_row)
            if not stores_df.empty and stores_df.shape[1] > 0:
                stores_df = stores_df.dropna(subset=[0])
        
        if "links" in section_starts:
            start_row = section_starts["links"] + header_row
            links_df = pd.read_excel(template_file, sheet_name='지역_BSN', header=None, skiprows=start_row, nrows=100)  # 최대 100행 읽기
            if not links_df.empty and links_df.shape[1] > 0:
                links_df = links_df.dropna(subset=[0])
                
        print("모든 지역 시트 목록 가져오는 중...")
        region_sheets = []
        for sheet in xl.sheet_names:
            if sheet.startswith('지역_') and sheet != '지역_BSN':
                region_sheets.append(sheet)
        
        print(f"찾은 지역 시트: {', '.join(region_sheets)}")
        print(f"'지역_BSN' 시트의 데이터를 다른 지역 시트에 복사합니다.")
        
        # 각 지역 시트에 데이터 쓰기
        for sheet_name in region_sheets:
            region_code = sheet_name.split('_')[1]  # 지역 코드 추출
            print(f"'{sheet_name}' 시트에 데이터 복사 중... (지역 코드: {region_code})")
            
            # 새 워크북에 데이터 쓰기 (병합된 셀 문제 회피)
            with pd.ExcelWriter(f'temp_{region_code}.xlsx', engine='openpyxl') as writer:
                # 원본 시트 복사 (원본 구조 유지)
                original_df = pd.read_excel(template_file, sheet_name=sheet_name)
                original_df.to_excel(writer, sheet_name='Sheet1', index=False)
                
                # 워크북 가져오기
                wb = writer.book
                ws = wb['Sheet1']
                
                # 버스 데이터 업데이트
                if isinstance(buses_df, pd.DataFrame) and not buses_df.empty:
                    row_offset = section_starts.get("buses", 0) + header_row
                    
                    for idx, row in buses_df.iterrows():
                        row_num = idx + row_offset
                        
                        # 이름에서 지역 코드 변경
                        if pd.notna(row.iloc[0]):
                            old_name = str(row.iloc[0])
                            if "_" in old_name:  # 지역 코드 포함된 이름
                                parts = old_name.split('_')
                                new_name = f"{region_code}_{parts[1]}"
                            else:
                                new_name = f"{region_code}_{old_name}"
                            
                            # 새로운 워크시트에 값 쓰기
                            ws.cell(row=row_num, column=1).value = new_name
                            for col_idx in range(1, min(len(row), 20)):  # 최대 20열까지
                                if pd.notna(row.iloc[col_idx]):  # 값이 있는 경우만 복사
                                    ws.cell(row=row_num, column=col_idx+1).value = row.iloc[col_idx]
                
                # 발전기 데이터 업데이트
                if isinstance(generators_df, pd.DataFrame) and not generators_df.empty:
                    row_offset = section_starts.get("generators", 0) + header_row
                    
                    for idx, row in generators_df.iterrows():
                        row_num = idx + row_offset
                        
                        # 이름에서 지역 코드 변경
                        if pd.notna(row.iloc[0]):
                            old_name = str(row.iloc[0])
                            if "_" in old_name:  # 지역 코드 포함된 이름
                                parts = old_name.split('_')
                                new_name = f"{region_code}_{parts[1]}"
                            else:
                                new_name = f"{region_code}_{old_name}"
                            
                            # 버스 참조 변경
                            if pd.notna(row.iloc[1]):
                                old_bus = str(row.iloc[1])
                                if "_" in old_bus:  # 지역 코드 포함된 이름
                                    parts = old_bus.split('_')
                                    new_bus = f"{region_code}_{parts[1]}"
                                else:
                                    new_bus = f"{region_code}_{old_bus}"
                                
                                # 새로운 워크시트에 값 쓰기
                                ws.cell(row=row_num, column=1).value = new_name
                                ws.cell(row=row_num, column=2).value = new_bus
                                for col_idx in range(2, min(len(row), 20)):  # 최대 20열까지
                                    if pd.notna(row.iloc[col_idx]):  # 값이 있는 경우만 복사
                                        ws.cell(row=row_num, column=col_idx+1).value = row.iloc[col_idx]
                
                # 부하 데이터 업데이트
                if isinstance(loads_df, pd.DataFrame) and not loads_df.empty:
                    row_offset = section_starts.get("loads", 0) + header_row
                    
                    for idx, row in loads_df.iterrows():
                        row_num = idx + row_offset
                        
                        # 이름에서 지역 코드 변경
                        if pd.notna(row.iloc[0]):
                            old_name = str(row.iloc[0])
                            if "_" in old_name:  # 지역 코드 포함된 이름
                                parts = old_name.split('_')
                                new_name = f"{region_code}_{parts[1]}"
                            else:
                                new_name = f"{region_code}_{old_name}"
                            
                            # 버스 참조 변경
                            if pd.notna(row.iloc[1]):
                                old_bus = str(row.iloc[1])
                                if "_" in old_bus:  # 지역 코드 포함된 이름
                                    parts = old_bus.split('_')
                                    new_bus = f"{region_code}_{parts[1]}"
                                else:
                                    new_bus = f"{region_code}_{old_bus}"
                                
                                # 새로운 워크시트에 값 쓰기
                                ws.cell(row=row_num, column=1).value = new_name
                                ws.cell(row=row_num, column=2).value = new_bus
                                for col_idx in range(2, min(len(row), 20)):  # 최대 20열까지
                                    if pd.notna(row.iloc[col_idx]):  # 값이 있는 경우만 복사
                                        ws.cell(row=row_num, column=col_idx+1).value = row.iloc[col_idx]
                
                # 저장장치 데이터 업데이트
                if isinstance(stores_df, pd.DataFrame) and not stores_df.empty:
                    row_offset = section_starts.get("stores", 0) + header_row
                    
                    for idx, row in stores_df.iterrows():
                        row_num = idx + row_offset
                        
                        # 이름에서 지역 코드 변경
                        if pd.notna(row.iloc[0]):
                            old_name = str(row.iloc[0])
                            if "_" in old_name:  # 지역 코드 포함된 이름
                                parts = old_name.split('_')
                                new_name = f"{region_code}_{parts[1]}"
                            else:
                                new_name = f"{region_code}_{old_name}"
                            
                            # 버스 참조 변경
                            if pd.notna(row.iloc[1]):
                                old_bus = str(row.iloc[1])
                                if "_" in old_bus:  # 지역 코드 포함된 이름
                                    parts = old_bus.split('_')
                                    new_bus = f"{region_code}_{parts[1]}"
                                else:
                                    new_bus = f"{region_code}_{old_bus}"
                                
                                # 새로운 워크시트에 값 쓰기
                                ws.cell(row=row_num, column=1).value = new_name
                                ws.cell(row=row_num, column=2).value = new_bus
                                for col_idx in range(2, min(len(row), 20)):  # 최대 20열까지
                                    if pd.notna(row.iloc[col_idx]):  # 값이 있는 경우만 복사
                                        ws.cell(row=row_num, column=col_idx+1).value = row.iloc[col_idx]
                
                # 링크 데이터 업데이트
                if isinstance(links_df, pd.DataFrame) and not links_df.empty:
                    row_offset = section_starts.get("links", 0) + header_row
                    
                    for idx, row in links_df.iterrows():
                        row_num = idx + row_offset
                        
                        # 이름에서 지역 코드 변경
                        if pd.notna(row.iloc[0]):
                            old_name = str(row.iloc[0])
                            if "_" in old_name:  # 지역 코드 포함된 이름
                                parts = old_name.split('_')
                                new_name = f"{region_code}_{parts[1]}"
                            else:
                                new_name = f"{region_code}_{old_name}"
                            
                            # 버스0 참조 변경
                            new_bus0 = None
                            if len(row) > 1 and pd.notna(row.iloc[1]):
                                old_bus0 = str(row.iloc[1])
                                if "_" in old_bus0:  # 지역 코드 포함된 이름
                                    parts = old_bus0.split('_')
                                    new_bus0 = f"{region_code}_{parts[1]}"
                                else:
                                    new_bus0 = f"{region_code}_{old_bus0}"
                            
                            # 버스1 참조 변경
                            new_bus1 = None
                            if len(row) > 2 and pd.notna(row.iloc[2]):
                                old_bus1 = str(row.iloc[2])
                                if "_" in old_bus1:  # 지역 코드 포함된 이름
                                    parts = old_bus1.split('_')
                                    new_bus1 = f"{region_code}_{parts[1]}"
                                else:
                                    new_bus1 = f"{region_code}_{old_bus1}"
                            
                            # 버스2 참조 변경
                            new_bus2 = None
                            if len(row) > 3 and pd.notna(row.iloc[3]):
                                old_bus2 = str(row.iloc[3])
                                if "_" in old_bus2:  # 지역 코드 포함된 이름
                                    parts = old_bus2.split('_')
                                    new_bus2 = f"{region_code}_{parts[1]}"
                                else:
                                    new_bus2 = f"{region_code}_{old_bus2}"
                            
                            # 버스3 참조 변경
                            new_bus3 = None
                            if len(row) > 4 and pd.notna(row.iloc[4]):
                                old_bus3 = str(row.iloc[4])
                                if "_" in old_bus3:  # 지역 코드 포함된 이름
                                    parts = old_bus3.split('_')
                                    new_bus3 = f"{region_code}_{parts[1]}"
                                else:
                                    new_bus3 = f"{region_code}_{old_bus3}"
                            
                            # 새로운 워크시트에 값 쓰기
                            ws.cell(row=row_num, column=1).value = new_name
                            ws.cell(row=row_num, column=2).value = new_bus0
                            ws.cell(row=row_num, column=3).value = new_bus1
                            ws.cell(row=row_num, column=4).value = new_bus2
                            ws.cell(row=row_num, column=5).value = new_bus3
                            for col_idx in range(5, min(len(row), 20)):  # 최대 20열까지
                                if col_idx < len(row) and pd.notna(row.iloc[col_idx]):  # 값이 있는 경우만 복사
                                    ws.cell(row=row_num, column=col_idx+1).value = row.iloc[col_idx]
            
            # 임시 파일을 원본 파일의 시트로 복사
            temp_wb = openpyxl.load_workbook(f'temp_{region_code}.xlsx')
            main_wb = openpyxl.load_workbook(template_file)
            
            # 워크시트 복사
            source_sheet = temp_wb.active
            if sheet_name in main_wb.sheetnames:
                # 기존 시트 제거
                std = main_wb.get_sheet_by_name(sheet_name)
                main_wb.remove(std)
            
            # 새 시트 생성
            target_sheet = main_wb.create_sheet(sheet_name)
            
            # 셀 복사
            for row in source_sheet.rows:
                for cell in row:
                    if cell.value is not None:  # 빈 셀은 무시
                        target_sheet.cell(row=cell.row, column=cell.column).value = cell.value
            
            # 저장
            main_wb.save(template_file)
            
            # 임시 파일 삭제
            os.remove(f'temp_{region_code}.xlsx')
        
        print(f"모든 지역 시트에 데이터가 성공적으로 복사되었습니다.")
        return True
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    copy_region_data() 