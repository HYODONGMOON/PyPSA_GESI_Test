import openpyxl

def clear_a3b5_cells():
    """
    모든 지역 시트의 A3:B5 셀 값을 비웁니다.
    """
    try:
        # 파일 경로 설정
        template_file = 'regional_input_template.xlsx'
        
        # 워크북 열기
        print(f"'{template_file}' 파일을 열고 있습니다...")
        wb = openpyxl.load_workbook(template_file)
        
        # 모든 지역 시트 목록 가져오기
        region_sheets = []
        for sheet in wb.sheetnames:
            if sheet.startswith('지역_') and sheet != '지역_BSN':
                region_sheets.append(sheet)
                
        print(f"찾은 지역 시트: {', '.join(region_sheets)}")
        
        # 각 지역 시트 처리
        for sheet_name in region_sheets:
            print(f"'{sheet_name}' 시트 처리 중...")
            sheet = wb[sheet_name]
            
            # A3:B5 셀의 값 비우기 (정보 제공용 셀)
            print(f"  A3:B5 셀의 값 비우기")
            for row in range(3, 6):
                for col in range(1, 3):  # A, B 열
                    cell = sheet.cell(row=row, column=col)
                    # 셀 값은 비우지만 서식은 유지
                    cell.value = None
        
        # 변경사항 저장
        wb.save(template_file)
        print(f"모든 지역 시트의 A3:B5 셀이 성공적으로 비워졌습니다.")
        return True
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    clear_a3b5_cells() 