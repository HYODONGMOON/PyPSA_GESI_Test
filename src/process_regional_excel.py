#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
지역별 데이터 Excel 처리기

지역별 데이터 입력 Excel 파일을 읽어 PyPSA-HD 모델에 필요한 통합 데이터를 생성합니다.
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import traceback

from regional_selector import RegionalSelector, KOREA_REGIONS
from regional_data_manager import RegionalDataManager, COMPONENT_TEMPLATES

# PyPSA-HD 모듈 임포트
try:
    import PyPSA_GUI  # 기존 GUI 모듈
except Exception:
    print("경고: PyPSA_GUI 모듈을 임포트할 수 없습니다. 일부 기능이 제한될 수 있습니다.")

# 기본 파일 경로
DEFAULT_TEMPLATE_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'interface.xlsx'))
DEFAULT_OUTPUT_PATH = "integrated_input_data.xlsx"

class RegionalExcelProcessor:
    """지역별 Excel 데이터 처리 클래스"""
    
    def __init__(self, excel_path=None):
        """초기화 함수
        
        Args:
            excel_path (str, optional): Excel 파일 경로
        """
        self.excel_path = excel_path or DEFAULT_TEMPLATE_PATH
        self.output_path = DEFAULT_OUTPUT_PATH
        
        # 지역 선택기 및 데이터 관리자 초기화
        self.region_selector = RegionalSelector()
        self.data_manager = RegionalDataManager(self.region_selector)
        
        # 선택된 지역 목록
        self.selected_regions = []
        
        # 연결 정보
        self.connections = []
        
        # 시간 설정
        self.timeseries = {}
        
        # 발전 패턴
        self.patterns = {}
    
    def process_excel(self):
        """Excel 파일 처리 및 통합 데이터 생성"""
        try:
            # 엑셀 파일이 존재하는지 확인
            if not os.path.exists(self.excel_path):
                print(f"오류: 파일 '{self.excel_path}'가 존재하지 않습니다.")
                return False
            
            # 엑셀 파일 열기
            print(f"Excel 파일 '{self.excel_path}' 읽는 중...")
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            # 1. 선택된 지역 읽기
            if not self.read_selected_regions(wb):
                return False
            
            # 2. 지역별 데이터 읽기
            if not self.read_regional_data(wb):
                return False
            
            # 3. 지역간 연결 읽기
            if not self.read_connections(wb):
                return False
            
            # 4. 시간 설정 읽기
            if not self.read_timeseries(wb):
                return False
            
            # 5. 패턴 데이터 읽기
            if not self.read_patterns(wb):
                return False
            
            # 6. 통합 데이터 생성
            if not self.create_integrated_data():
                return False
            
            # 7. PyPSA 실행
            if not self.run_pypsa_model():
                return False
            
            print("처리가 완료되었습니다.")
            return True
            
        except Exception as e:
            print(f"오류 발생: {str(e)}")
            traceback.print_exc()
            return False
    
    def read_selected_regions(self, wb):
        """선택된 지역 읽기
        
        Args:
            wb (openpyxl.Workbook): 워크북 객체
            
        Returns:
            bool: 성공 여부
        """
        try:
            if "지역 선택" not in wb.sheetnames:
                print("오류: '지역 선택' 시트가 없습니다.")
                return False
            
            ws = wb["지역 선택"]
            self.selected_regions = []
            
            # 데이터 범위 찾기
            max_row = ws.max_row
            
            # 코드와 선택 여부 읽기
            for row in range(6, max_row + 1):  # 데이터는 6행부터 시작
                code = ws.cell(row=row, column=1).value
                selected = ws.cell(row=row, column=4).value
                
                if code and selected == "O":
                    self.selected_regions.append(code)
                    print(f"선택된 지역: {KOREA_REGIONS.get(code, {}).get('name', code)}")
            
            if not self.selected_regions:
                print("경고: 선택된 지역이 없습니다.")
                return False
            
            # 지역 선택기에 선택된 지역 설정
            for region_code in self.selected_regions:
                self.region_selector.select_region(region_code)
            
            return True
            
        except Exception as e:
            print(f"선택된 지역 읽기 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return False
    
    def read_regional_data(self, wb):
        """지역별 데이터 읽기
        
        Args:
            wb (openpyxl.Workbook): 워크북 객체
            
        Returns:
            bool: 성공 여부
        """
        try:
            # 각 선택된 지역에 대해
            for region_code in self.selected_regions:
                print(f"지역 '{region_code}' 데이터 읽는 중...")
                
                # 지역 초기화
                self.data_manager.initialize_region(region_code)
                
                # 새로운 통합 지역 시트 확인
                integrated_sheet_name = f"지역_{region_code}"
                
                if integrated_sheet_name in wb.sheetnames:
                    # 통합 지역 시트에서 데이터 읽기
                    self._read_integrated_region_sheet(wb, region_code, integrated_sheet_name)
                else:
                    # 구 버전 호환성을 위해 개별 구성요소 시트 읽기 시도
                    self._read_separate_component_sheets(wb, region_code)
            
            return True
            
        except Exception as e:
            print(f"지역별 데이터 읽기 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return False
    
    def _read_integrated_region_sheet(self, wb, region_code, sheet_name):
        """통합 지역 시트에서 데이터 읽기
        
        Args:
            wb (openpyxl.Workbook): 워크북 객체
            region_code (str): 지역 코드
            sheet_name (str): 시트 이름
        """
        ws = wb[sheet_name]
        
        # 전체 시트 내용 읽기
        max_row = ws.max_row
        current_component = None
        header_row = None
        headers = []
        
        # 시트의 각 행 처리
        for row in range(1, max_row + 1):
            # 구성요소 제목 행 찾기 (굵은 글꼴, 배경색 있음)
            cell = ws.cell(row=row, column=1)
            if cell.font.bold and cell.fill.start_color.index != '00000000':
                component_title = cell.value
                if component_title:
                    # 한글 구성요소 이름에서 영문 컴포넌트 이름 찾기
                    for comp_name, display_name in {
                        'buses': '버스',
                        'generators': '발전기',
                        'lines': '송전선',
                        'loads': '부하',
                        'stores': '저장장치',
                        'links': '링크'
                    }.items():
                        if display_name in component_title:
                            current_component = comp_name
                            header_row = row + 2  # 제목 다음 두 번째 행이 헤더
                            break
            
            # 헤더 행 처리
            if current_component and row == header_row:
                header_cells = [ws.cell(row=row, column=i).value for i in range(1, ws.max_column + 1)]
                header_cells = [h for h in header_cells if h]  # None 제거
                
                # 한글 필드명을 영문으로 변환
                headers = []
                for header in header_cells:
                    english_field = get_english_field_name(header)
                    if english_field:
                        headers.append(english_field)
                    else:
                        headers.append(header)  # 변환할 수 없는 경우 원래 이름 사용
            
            # 데이터 행 처리
            if current_component and headers and row > header_row:
                # 빈 행 또는 구분선 건너뛰기
                if not ws.cell(row=row, column=1).value:
                    # 구분선 확인 (테두리가 있으면 구분선으로 간주)
                    if ws.cell(row=row, column=1).border.bottom.style:
                        current_component = None  # 다음 구성요소 준비
                        headers = []
                    continue
                
                # 데이터 읽기
                item_data = {}
                for i, field in enumerate(headers):
                    if i < len(headers):  # 인덱스 범위 확인
                        value = ws.cell(row=row, column=i + 1).value
                        if value is not None:  # 빈 셀은 건너뛰기
                            item_data[field] = value
                
                # 지역 코드 추가
                item_data['region'] = region_code
                
                # 이름 필드가 있고 지역 접두사가 없는 경우 추가
                if 'name' in item_data and item_data['name'] and not str(item_data['name']).startswith(f"{region_code}_"):
                    item_data['name'] = f"{region_code}_{item_data['name']}"
                
                # 버스 참조 필드에 지역 접두사 추가
                for bus_field in ['bus', 'bus0', 'bus1', 'bus2', 'bus3']:
                    if bus_field in item_data and item_data[bus_field] and not str(item_data[bus_field]).startswith(f"{region_code}_"):
                        item_data[bus_field] = f"{region_code}_{item_data[bus_field]}"
                
                # 데이터 관리자에 추가
                if item_data and current_component:
                    self.data_manager.add_component(region_code, current_component, item_data)
    
    def _read_separate_component_sheets(self, wb, region_code):
        """개별 구성요소 시트에서 데이터 읽기 (이전 버전 호환성)
        
        Args:
            wb (openpyxl.Workbook): 워크북 객체
            region_code (str): 지역 코드
        """
        # 각 구성요소 시트 읽기
        for component in COMPONENT_TEMPLATES.keys():
            sheet_name = f"{region_code}_{component}"
            
            if sheet_name not in wb.sheetnames:
                print(f"경고: '{sheet_name}' 시트가 없습니다. 건너뜁니다.")
                continue
            
            ws = wb[sheet_name]
            
            # 헤더 행 찾기 (일반적으로 7행)
            header_row = 7
            header_cells = [ws.cell(row=header_row, column=i).value for i in range(1, ws.max_column + 1)]
            header_cells = [h for h in header_cells if h]  # None 제거
            
            if not header_cells:
                print(f"경고: '{sheet_name}' 시트에 헤더가 없습니다. 건너뜁니다.")
                continue
            
            # 한글 필드명을 영문으로 변환
            headers = []
            for header in header_cells:
                english_field = get_english_field_name(header)
                if english_field:
                    headers.append(english_field)
                else:
                    headers.append(header)  # 변환할 수 없는 경우 원래 이름 사용
            
            # 데이터 행 처리
            for row in range(header_row + 1, ws.max_row + 1):
                # 빈 행 건너뛰기
                if not ws.cell(row=row, column=1).value:
                    continue
                
                # 데이터 읽기
                item_data = {}
                for i, field in enumerate(headers):
                    value = ws.cell(row=row, column=i + 1).value
                    if value is not None:  # 빈 셀은 건너뛰기
                        item_data[field] = value
                
                # 지역 코드 추가
                item_data['region'] = region_code
                
                # 이름 필드가 있고 지역 접두사가 없는 경우 추가
                if 'name' in item_data and item_data['name'] and not str(item_data['name']).startswith(f"{region_code}_"):
                    item_data['name'] = f"{region_code}_{item_data['name']}"
                
                # 버스 참조 필드에 지역 접두사 추가
                for bus_field in ['bus', 'bus0', 'bus1', 'bus2', 'bus3']:
                    if bus_field in item_data and item_data[bus_field] and not str(item_data[bus_field]).startswith(f"{region_code}_"):
                        item_data[bus_field] = f"{region_code}_{item_data[bus_field]}"
                
                # 데이터 관리자에 추가
                if item_data:
                    self.data_manager.add_component(region_code, component, item_data)
    
    def read_connections(self, wb):
        """지역간 연결 읽기
        
        Args:
            wb (openpyxl.Workbook): 워크북 객체
            
        Returns:
            bool: 성공 여부
        """
        try:
            if "지역간 연결" not in wb.sheetnames:
                print("경고: '지역간 연결' 시트가 없습니다. 건너뜁니다.")
                return True
            
            ws = wb["지역간 연결"]
            
            # 헤더 행 찾기 (일반적으로 5행)
            header_row = 5
            
            # 데이터 행 처리
            for row in range(header_row + 1, ws.max_row + 1):
                # 빈 행 건너뛰기
                if not ws.cell(row=row, column=1).value:
                    continue
                
                # 연결 데이터 읽기
                name = ws.cell(row=row, column=1).value
                region1 = ws.cell(row=row, column=2).value
                bus1 = ws.cell(row=row, column=3).value
                region2 = ws.cell(row=row, column=4).value
                bus2 = ws.cell(row=row, column=5).value
                capacity = ws.cell(row=row, column=6).value
                voltage = ws.cell(row=row, column=7).value
                distance = ws.cell(row=row, column=8).value
                x = ws.cell(row=row, column=9).value
                r = ws.cell(row=row, column=10).value
                
                # 필수 필드 확인
                if not all([region1, region2, bus1, bus2]):
                    print(f"경고: 행 {row}에 필수 필드가 누락되었습니다. 건너뜁니다.")
                    continue
                
                # 이미 지역 접두사가 포함된 경우를 처리
                if not bus1.startswith(f"{region1}_"):
                    bus1 = f"{region1}_{bus1}"
                
                if not bus2.startswith(f"{region2}_"):
                    bus2 = f"{region2}_{bus2}"
                
                # 연결 데이터 구성
                connection_data = {
                    'name': name,
                    'bus0': bus1,
                    'bus1': bus2,
                    'carrier': 'AC',
                    's_nom': float(capacity) if capacity else 1000.0,
                    'v_nom': float(voltage) if voltage else 345.0,
                    'length': float(distance) if distance else None,
                    'x': float(x) if x else None,
                    'r': float(r) if r else None
                }
                
                # 거리, x, r 자동 계산 (값이 없는 경우)
                if not distance or not x or not r:
                    calc_distance = self.region_selector.calculate_distance(region1, region2)
                    if calc_distance:
                        if not distance:
                            connection_data['length'] = calc_distance
                        if not x:
                            connection_data['x'] = calc_distance * 0.0004
                        if not r:
                            connection_data['r'] = calc_distance * 0.0001
                
                # 데이터 관리자에 연결 추가
                self.data_manager.add_connection(region1, region2, connection_data)
            
            return True
            
        except Exception as e:
            print(f"지역간 연결 읽기 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return False
    
    def read_timeseries(self, wb):
        """시간 설정 정보 읽기"""
        try:
            # 먼저 "시간 설정" 시트 확인
            if "시간 설정" in wb.sheetnames:
                print("'시간 설정' 시트에서 시간 정보를 읽습니다.")
                ws = wb["시간 설정"]
                
                # 첫 번째 행의 헤더 찾기
                headers = {}
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers[cell_value.lower()] = col
                
                # 필요한 열 찾기
                start_col = None
                end_col = None
                freq_col = None
                
                for header, col in headers.items():
                    if "시작" in header or "start" in header:
                        start_col = col
                    elif "종료" in header or "end" in header:
                        end_col = col
                    elif "간격" in header or "주기" in header or "freq" in header:
                        freq_col = col
                
                # 값 읽기
                if start_col and end_col:
                    start_time = ws.cell(row=2, column=start_col).value
                    end_time = ws.cell(row=2, column=end_col).value
                    frequency = "1h"  # 기본값
                    
                    if freq_col:
                        freq_value = ws.cell(row=2, column=freq_col).value
                        if freq_value:
                            frequency = freq_value
                    
                    # datetime 객체로 변환
                    if isinstance(start_time, str):
                        start_time = pd.to_datetime(start_time)
                    if isinstance(end_time, str):
                        end_time = pd.to_datetime(end_time)
                    
                    # 시간 설정 데이터프레임 생성
                    timeseries_data = pd.DataFrame({
                        'start_time': [start_time],
                        'end_time': [end_time],
                        'frequency': [frequency]
                    })
                    
                    self.data_manager.timeseries = timeseries_data
                    print(f"시간 설정: {start_time} ~ {end_time}, 간격: {frequency}")
                    return True
            
            # "timeseries" 시트 확인 (이전 방식)
            if 'timeseries' in wb.sheetnames:
                print("'timeseries' 시트에서 시간 정보를 읽습니다.")
                # 기존 코드 유지
                # ...
                return True
            
            # 시트가 모두 없으면 기본값 설정
            print("시간 설정 시트를 찾을 수 없어 기본값 사용: 2023-01-01 00:00:00 ~ 2023-12-30 00:00:00, 간격: 1h")
            self.data_manager.timeseries = pd.DataFrame({
                'start_time': [pd.Timestamp('2023-01-01 00:00:00')],
                'end_time': [pd.Timestamp('2023-12-30 00:00:00')],
                'frequency': ['1h']
            })
            return True
            
        except Exception as e:
            print(f"시간 설정 읽기 실패: {str(e)}")
            # 오류 발생 시 기본값 설정
            self.data_manager.timeseries = pd.DataFrame({
                'start_time': [pd.Timestamp('2023-01-01 00:00:00')],
                'end_time': [pd.Timestamp('2023-12-30 00:00:00')],
                'frequency': ['1h']
            })
            return False
    
    def read_patterns(self, wb):
        """발전 패턴 데이터 읽기
        
        Args:
            wb (openpyxl.Workbook): 워크북 객체
            
        Returns:
            bool: 성공 여부
        """
        try:
            # 재생에너지 패턴 읽기
            self.read_renewable_patterns(wb)
            
            # 부하 패턴 읽기
            self.read_load_patterns(wb)
            
            # 지역별 부하 패턴 읽기
            self.read_regional_load_patterns(wb)
            
            # 제약사항 읽기
            self.read_constraints(wb)
            
            return True
            
        except Exception as e:
            print(f"패턴 및 제약사항 데이터 읽기 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return False
    
    def read_renewable_patterns(self, wb):
        """재생에너지 패턴 데이터 읽기"""
        try:
            if "renewable_patterns" not in wb.sheetnames:
                print("경고: 'renewable_patterns' 시트가 없습니다. 기본 패턴을 사용합니다.")
                self._create_default_patterns()
                return True
            
            ws = wb["renewable_patterns"]
            
            # 헤더 행 찾기 (일반적으로 5행)
            header_row = 5
            header_cells = [ws.cell(row=header_row, column=i).value for i in range(1, ws.max_column + 1)]
            
            # 데이터 초기화
            hour_data = []
            pv_data = []
            wt_data = []
            
            # 데이터 행 처리
            for row in range(header_row + 1, ws.max_row + 1):
                # 빈 행이면 종료
                if not ws.cell(row=row, column=1).value:
                    break
                
                # 데이터 읽기
                hour = ws.cell(row=row, column=1).value
                pv = ws.cell(row=row, column=2).value
                wt = ws.cell(row=row, column=3).value
                
                if hour is not None:
                    hour_data.append(int(hour))
                if pv is not None:
                    pv_data.append(float(pv))
                if wt is not None:
                    wt_data.append(float(wt))
            
            # 데이터가 충분한지 확인
            if len(hour_data) < 24:
                print("경고: 재생에너지 패턴 데이터가 불충분합니다. 기본 패턴을 사용합니다.")
                self._create_default_patterns()
                return True
            
            # 패턴 데이터 생성
            self.patterns = {
                'hour': hour_data,
                'PV': pv_data,
                'WT': wt_data
            }
            
            # 8760시간 데이터로 확장
            if len(hour_data) < 8760:
                print(f"참고: 재생에너지 패턴 데이터({len(hour_data)}시간)를 8760시간으로 확장합니다.")
                self._extend_patterns()
            
            return True
            
        except Exception as e:
            print(f"재생에너지 패턴 데이터 읽기 중 오류 발생: {str(e)}")
            traceback.print_exc()
            self._create_default_patterns()
            return True  # 오류가 있어도 계속 진행 (기본값 사용)
    
    def read_load_patterns(self, wb):
        """부하 패턴 데이터 읽기"""
        try:
            # 부하 패턴 초기화
            self.load_patterns = {
                'hour': list(range(24)),
                'electricity': [0.7] * 24,
                'heat': [0.7] * 24,
                'hydrogen': [0.7] * 24
            }
            
            if "load_patterns" not in wb.sheetnames:
                print("경고: 'load_patterns' 시트가 없습니다. 기본 부하 패턴을 사용합니다.")
                return True
            
            ws = wb["load_patterns"]
            
            # 헤더 행 찾기 (일반적으로 5행)
            header_row = 5
            header_cells = [ws.cell(row=header_row, column=i).value for i in range(1, ws.max_column + 1)]
            
            # 헤더 인덱스 찾기
            hour_idx = None
            elec_idx = None
            heat_idx = None
            hydrogen_idx = None
            
            for i, header in enumerate(header_cells):
                if header == '시간' or header == 'hour':
                    hour_idx = i
                elif header == '전력(electricity)' or header == 'electricity':
                    elec_idx = i
                elif header == '열(heat)' or header == 'heat':
                    heat_idx = i
                elif header == '수소(hydrogen)' or header == 'hydrogen':
                    hydrogen_idx = i
            
            # 필수 인덱스 확인
            if hour_idx is None:
                print("경고: 'load_patterns' 시트에 시간 열이 없습니다. 기본 부하 패턴을 사용합니다.")
                return True
            
            # 데이터 초기화
            hour_data = []
            elec_data = []
            heat_data = []
            hydrogen_data = []
            
            # 데이터 행 처리
            for row in range(header_row + 1, ws.max_row + 1):
                # 빈 행이면 종료
                if not ws.cell(row=row, column=hour_idx + 1).value:
                    break
                
                # 데이터 읽기
                hour = ws.cell(row=row, column=hour_idx + 1).value
                hour_data.append(int(hour))
                
                if elec_idx is not None:
                    elec = ws.cell(row=row, column=elec_idx + 1).value
                    elec_data.append(float(elec) if elec is not None else 0.7)
                
                if heat_idx is not None:
                    heat = ws.cell(row=row, column=heat_idx + 1).value
                    heat_data.append(float(heat) if heat is not None else 0.7)
                
                if hydrogen_idx is not None:
                    hydrogen = ws.cell(row=row, column=hydrogen_idx + 1).value
                    hydrogen_data.append(float(hydrogen) if hydrogen is not None else 0.7)
            
            # 데이터가 충분한지 확인
            if len(hour_data) < 24:
                print("경고: 부하 패턴 데이터가 불충분합니다. 기본 부하 패턴을 사용합니다.")
                return True
            
            # 패턴 데이터 생성
            self.load_patterns = {
                'hour': hour_data
            }
            
            if elec_data:
                self.load_patterns['electricity'] = elec_data
            if heat_data:
                self.load_patterns['heat'] = heat_data
            if hydrogen_data:
                self.load_patterns['hydrogen'] = hydrogen_data
            
            # 8760시간 데이터로 확장
            if len(hour_data) < 8760:
                print(f"참고: 부하 패턴 데이터({len(hour_data)}시간)를 8760시간으로 확장합니다.")
                self._extend_load_patterns()
            
            return True
            
        except Exception as e:
            print(f"부하 패턴 데이터 읽기 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return True  # 오류가 있어도 계속 진행 (기본값 사용)
    
    def read_regional_load_patterns(self, wb):
        """지역별 부하 패턴 데이터 읽기"""
        try:
            # 지역별 부하 패턴 초기화
            self.regional_load_patterns = {}
            
            if "regional_load_patterns" not in wb.sheetnames:
                print("경고: 'regional_load_patterns' 시트가 없습니다. 기본 부하 패턴만 사용합니다.")
                return True
            
            ws = wb["regional_load_patterns"]
            
            # 지역 및 에너지 유형 정보 읽기 (일반적으로 7행부터)
            region_rows = {}  # {(region_code, energy_type): row_index}
            
            for row in range(7, ws.max_row + 1):
                region_code = ws.cell(row=row, column=1).value
                energy_type = ws.cell(row=row, column=2).value
                apply = ws.cell(row=row, column=3).value
                
                if not region_code or not energy_type or apply != "O":
                    continue
                
                region_rows[(region_code, energy_type)] = row
            
            # 패턴 데이터 읽기
            for (region_code, energy_type), row_idx in region_rows.items():
                # 해당 지역 코드가 선택된 지역인지 확인
                if region_code not in self.selected_regions:
                    continue
                
                # 패턴 데이터 초기화
                if region_code not in self.regional_load_patterns:
                    self.regional_load_patterns[region_code] = {}
                
                # 시간 및 패턴값 읽기
                hours = []
                values = []
                
                for i in range(24):  # 24시간 데이터 읽기
                    data_row = i + 7  # 시작 행 (7행부터)
                    hour = ws.cell(row=data_row, column=5).value  # E열: 시간
                    value = ws.cell(row=data_row, column=6).value  # F열: 패턴값
                    
                    if hour is not None and value is not None:
                        hours.append(int(hour))
                        values.append(float(value))
                
                # 데이터가 있으면 저장
                if hours and values:
                    self.regional_load_patterns[region_code][energy_type] = {
                        'hour': hours,
                        'value': values
                    }
                    print(f"지역 '{region_code}'의 '{energy_type}' 부하 패턴을 읽었습니다.")
            
            return True
            
        except Exception as e:
            print(f"지역별 부하 패턴 데이터 읽기 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return True  # 오류가 있어도 계속 진행
    
    def read_constraints(self, wb):
        """제약사항 데이터 읽기"""
        try:
            # 제약사항 초기화
            self.constraints = []
            
            if "constraints" not in wb.sheetnames:
                print("경고: 'constraints' 시트가 없습니다. 제약사항 없이 진행합니다.")
                return True
            
            ws = wb["constraints"]
            
            # 헤더 행 찾기 (일반적으로 6행)
            header_row = 6
            
            # 데이터 행 처리
            for row in range(header_row + 1, ws.max_row + 1):
                # 빈 행이면 종료
                if not ws.cell(row=row, column=1).value:
                    break
                
                # 데이터 읽기
                name = ws.cell(row=row, column=1).value
                constraint_type = ws.cell(row=row, column=2).value
                target_attr = ws.cell(row=row, column=3).value
                condition = ws.cell(row=row, column=4).value
                constant = ws.cell(row=row, column=5).value
                region = ws.cell(row=row, column=6).value
                description = ws.cell(row=row, column=7).value
                
                # 필수 필드 확인
                if not all([name, constraint_type, target_attr, condition, constant]):
                    print(f"경고: 행 {row}에 필수 제약사항 필드가 누락되었습니다. 건너뜁니다.")
                    continue
                
                # 제약사항 추가
                constraint = {
                    'name': name,
                    'type': constraint_type,
                    'carrier_attribte': target_attr,  # 오타 주의: attribte
                    'sense': condition,
                    'constant': float(constant)
                }
                
                if region and region != "ALL":
                    constraint['region'] = region
                
                if description:
                    constraint['description'] = description
                
                self.constraints.append(constraint)
                print(f"제약사항 '{name}'을 읽었습니다.")
            
            return True
            
        except Exception as e:
            print(f"제약사항 데이터 읽기 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return True  # 오류가 있어도 계속 진행
    
    def _extend_load_patterns(self):
        """부하 패턴 데이터를 8760시간으로 확장"""
        original_hours = len(self.load_patterns['hour'])
        
        # 시간 데이터 확장
        self.load_patterns['hour'] = list(range(8760))
        
        # 각 에너지 유형별 패턴 데이터 확장
        for energy_type in ['electricity', 'heat', 'hydrogen']:
            if energy_type in self.load_patterns:
                extended_data = []
                
                for h in range(8760):
                    # 원본 데이터에서 해당하는 인덱스 계산
                    idx = h % original_hours
                    
                    # 확장된 데이터에 추가
                    extended_data.append(self.load_patterns[energy_type][idx])
                
                self.load_patterns[energy_type] = extended_data
    
    def create_integrated_data(self):
        """통합 데이터 생성 및 저장"""
        try:
            print("통합 데이터 생성 중...")
            
            # 데이터 관리자에서 통합 데이터 가져오기
            merged_data = self.data_manager.merge_data()
            if not merged_data:
                print("오류: 통합 데이터를 생성할 수 없습니다.")
                return False
            
            # 통합 데이터에 시간 설정 추가
            merged_data['timeseries'] = pd.DataFrame([self.timeseries])
            
            # 통합 데이터에 재생에너지 패턴 추가
            merged_data['renewable_patterns'] = pd.DataFrame(self.patterns)
            
            # 통합 데이터에 부하 패턴 추가
            if hasattr(self, 'load_patterns') and self.load_patterns:
                merged_data['load_patterns'] = pd.DataFrame(self.load_patterns)
            
            # 통합 데이터에 제약사항 추가
            if hasattr(self, 'constraints') and self.constraints:
                merged_data['constraints'] = pd.DataFrame(self.constraints)
            
            # (선택) 추가 시트 복사: load_patterns / renewable_patterns / 시나리오_에너지수요
            try:
                if os.path.exists(self.excel_path):
                    # load_patterns
                    try:
                        lp_df = pd.read_excel(self.excel_path, sheet_name='load_patterns')
                        if lp_df is not None and not lp_df.empty:
                            merged_data['load_patterns'] = lp_df
                            print("'load_patterns' 시트를 통합 파일에 포함했습니다.")
                    except Exception:
                        pass
                    # renewable_patterns
                    try:
                        rp_df = pd.read_excel(self.excel_path, sheet_name='renewable_patterns')
                        if rp_df is not None and not rp_df.empty:
                            merged_data['renewable_patterns'] = rp_df
                            print("'renewable_patterns' 시트를 통합 파일에 포함했습니다.")
                    except Exception:
                        pass
                    # 시나리오_에너지수요
                    try:
                        scenario_df = pd.read_excel(self.excel_path, sheet_name='시나리오_에너지수요')
                        if scenario_df is not None and not scenario_df.empty:
                            merged_data['시나리오_에너지수요'] = scenario_df
                            print("'시나리오_에너지수요' 시트를 통합 파일에 포함했습니다.")
                    except Exception:
                        pass
                    # 시나리오_링크
                    try:
                        link_scn_df = pd.read_excel(self.excel_path, sheet_name='시나리오_링크')
                        if link_scn_df is not None and not link_scn_df.empty:
                            merged_data['시나리오_링크'] = link_scn_df
                            print("'시나리오_링크' 시트를 통합 파일에 포함했습니다.")
                    except Exception:
                        pass
            except Exception:
                pass
            
            # 파일로 저장
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                # 각 구성요소 저장
                for component, df in merged_data.items():
                    if not df.empty:
                        df.to_excel(writer, sheet_name=component, index=False)
            
            print(f"통합 데이터가 '{self.output_path}'에 저장되었습니다.")
            return True
            
        except Exception as e:
            print(f"통합 데이터 생성 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return False
    
    def run_pypsa_model(self):
        """PyPSA 모델 실행"""
        try:
            print("\nPyPSA 모델 실행 중...")
            
            if 'PyPSA_GUI' not in sys.modules:
                print("경고: PyPSA_GUI 모듈을 찾을 수 없습니다. 모델 실행을 건너뜁니다.")
                return True
            
            # 입력 파일 경로 설정
            PyPSA_GUI.INPUT_FILE = self.output_path
            
            # 데이터 로드
            input_data = PyPSA_GUI.read_input_data(self.output_path)
            if input_data is None:
                print("오류: 입력 데이터를 로드할 수 없습니다.")
                return False
            
            # 데이터 유효성 검사
            try:
                input_data = PyPSA_GUI.validate_input_data(input_data)
            except Exception as e:
                print(f"데이터 유효성 검사 오류: {str(e)}")
                return False
            
            # 네트워크 생성
            network = PyPSA_GUI.create_network(input_data)
            if network is None:
                print("오류: 네트워크를 생성할 수 없습니다.")
                return False
            
            # 네트워크 최적화
            if not PyPSA_GUI.optimize_network(network):
                print("오류: 네트워크 최적화에 실패했습니다.")
                return False
            
            # 결과 저장
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            result_file = f'results_({current_time}).xlsx'
            if not PyPSA_GUI.save_results(network, result_file):
                print("오류: 결과를 저장할 수 없습니다.")
                return False
            
            print(f"모델 실행이 완료되었습니다. 결과가 '{result_file}'에 저장되었습니다.")
            return True
            
        except Exception as e:
            print(f"모델 실행 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return False
    
    def _create_default_patterns(self):
        """기본 패턴 데이터 생성"""
        # 시간 데이터 (8760시간)
        hour_data = list(range(1, 8761))
        
        # 태양광 패턴 (낮에 높고 밤에 0)
        pv_data = []
        for h in range(1, 8761):
            hour_of_day = (h - 1) % 24
            if 6 <= hour_of_day < 18:  # 낮 시간
                value = 0.5 + 0.5 * np.sin(np.pi * (hour_of_day - 6) / 12)
            else:  # 밤 시간
                value = 0.0
            pv_data.append(value)
        
        # 풍력 패턴 (더 랜덤하게)
        np.random.seed(42)  # 재현성을 위한 시드 설정
        wt_data = 0.2 + 0.6 * np.random.rand(8760)
        
        self.patterns = {
            'hour': hour_data,
            'PV': pv_data,
            'WT': wt_data
        }
    
    def _extend_patterns(self):
        """패턴 데이터를 8760시간으로 확장"""
        original_hours = len(self.patterns['hour'])
        
        # 시간 데이터 확장
        self.patterns['hour'] = list(range(1, 8761))
        
        # 패턴 데이터 확장
        extended_pv = []
        extended_wt = []
        
        for h in range(1, 8761):
            # 원본 데이터에서 해당하는 인덱스 계산
            idx = (h - 1) % original_hours
            
            # 확장된 데이터에 추가
            extended_pv.append(self.patterns['PV'][idx])
            extended_wt.append(self.patterns['WT'][idx])
        
        self.patterns['PV'] = extended_pv
        self.patterns['WT'] = extended_wt

def get_english_field_name(korean_name):
    """한글 필드명을 영문으로 변환"""
    mapping = {
        '이름': 'name',
        '지역': 'region',
        '전압(kV)': 'v_nom',
        '캐리어': 'carrier',
        'X좌표': 'x',
        'Y좌표': 'y',
        '버스': 'bus',
        '정격용량(MW)': 'p_nom',
        '용량확장가능': 'p_nom_extendable',
        '최소용량(MW)': 'p_nom_min',
        '최대용량(MW)': 'p_nom_max',
        '한계비용': 'marginal_cost',
        '설비비용': 'capital_cost',
        '효율': 'efficiency',
        '최대출력비율': 'p_max_pu',
        '최소출력비율': 'p_min_pu',
        '기동정지가능': 'committable',
        '증발속도제한': 'ramp_limit_up',
        '최소가동시간': 'min_up_time',
        '기동비용': 'start_up_cost',
        '수명(년)': 'lifetime',
        '시작버스': 'bus0',
        '종료버스': 'bus1',
        '버스2': 'bus2',
        '버스3': 'bus3',
        '저항(p.u.)': 'r',
        '정격용량(MVA)': 's_nom',
        '정격용량확장가능': 's_nom_extendable',
        '최소용량(MVA)': 's_nom_min',
        '최대용량(MVA)': 's_nom_max',
        '길이(km)': 'length',
        '부하량(MW)': 'p_set',
        '저장용량(MWh)': 'e_nom',
        '주기적운전': 'e_cyclic',
        '자체손실': 'standing_loss',
        '충전효율': 'efficiency_store',
        '방전효율': 'efficiency_dispatch',
        '초기저장량': 'e_initial',
        '최대저장용량': 'e_nom_max',
        '최소저장용량': 'e_nom_min',
        '효율0': 'efficiency0',
        '효율1': 'efficiency1',
        '효율2': 'efficiency2',
        '효율3': 'efficiency3',
        '리액턴스(p.u.)': 'x'
    }
    return mapping.get(korean_name)

def main():
    """메인 함수"""
    import argparse
    
    # 명령행 인수 파싱
    parser = argparse.ArgumentParser(description='지역별 데이터 Excel 처리기')
    parser.add_argument('--input', type=str, help='입력 Excel 파일 경로')
    parser.add_argument('--output', type=str, help='출력 Excel 파일 경로')
    args = parser.parse_args()
    
    # 프로세서 생성
    processor = RegionalExcelProcessor(args.input)
    
    # 출력 파일 경로 설정
    if args.output:
        processor.output_path = args.output
    
    # 처리 실행
    success = processor.process_excel()
    
    # 종료 코드 설정
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main() 