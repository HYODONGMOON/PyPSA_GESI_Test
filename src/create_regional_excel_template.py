#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
지역별 데이터 입력 Excel 템플릿 생성기

지역별 에너지시스템 데이터를 입력하기 위한 Excel 템플릿을 생성합니다.
템플릿에는 입력 시트와 분석 시트가 포함되며, 매크로를 사용하여 자동 반영이 가능합니다.
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 지역 선택기 임포트 (지역 정보를 가져오기 위함)
from regional_selector import RegionalSelector, KOREA_REGIONS
from regional_data_manager import COMPONENT_TEMPLATES, REGION_TEMPLATES

# Excel 파일 경로
TEMPLATE_PATH = "regional_input_template.xlsx"

# 색상 정의
COLORS = {
    'header': 'DDEBF7',          # 밝은 파란색 (헤더)
    'input_cell': 'FFFFFF',      # 흰색 (입력 셀)
    'formula_cell': 'F2F2F2',    # 연한 회색 (수식 셀)
    'highlight': 'FFD966',       # 노란색 (강조)
    'required': 'FCE4D6',        # 연한 주황색 (필수 필드)
    'optional': 'E2EFDA',        # 연한 녹색 (선택 필드)
    'sel_region': 'D9E1F2',      # 연한 푸른색 (서울)
    'jbd_region': 'E2EFDA',      # 연한 녹색 (전북)
    'error': 'FF9999',           # 연한 빨간색 (오류)
}

def create_excel_template(output_path='regional_input_template.xlsx'):
    """지역별 입력 템플릿 생성"""
    wb = openpyxl.Workbook()
    
    # 기본 시트 삭제
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 제어 시트 생성
    create_control_sheet(wb)
    
    # 지역 선택 시트 생성
    create_region_selection_sheet(wb)
    
    # 지역별 데이터 시트 생성
    create_region_data_sheets(wb)
    
    # 지역간 연결 시트 생성
    create_connection_sheet(wb)
    
    # 시간 설정 시트 생성
    create_timeseries_sheet(wb)
    
    # 패턴 및 제약사항 시트 생성
    create_pattern_sheets(wb)
    
    # 통합 데이터 시트 생성
    create_integrated_sheets(wb)
    
    # 기본 데이터 추가
    add_default_data(wb)
    
    # 템플릿 저장
    wb.save(output_path)
    print(f"Excel 템플릿이 '{output_path}'에 생성되었습니다.")
    return True

def create_control_sheet(wb):
    """제어 시트 생성"""
    ws = wb.create_sheet("제어판", 0)
    
    # 제목 및 설명
    ws['A1'] = "PyPSA-HD 지역 기반 에너지시스템 입력 도구"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    
    ws['A3'] = "이 Excel 파일은 지역 기반 에너지시스템 데이터를 입력하기 위한 템플릿입니다."
    ws['A4'] = "각 지역의 에너지시스템 구성요소를 입력하고, 지역간 연결을 정의한 후, 통합 데이터를 생성할 수 있습니다."
    ws.merge_cells('A3:H3')
    ws.merge_cells('A4:H4')
    
    # 작업 순서 안내
    ws['A6'] = "작업 순서:"
    ws['A6'].font = Font(bold=True)
    ws['A7'] = "1. '지역 선택' 시트에서 분석할 지역을 선택하고 클릭하면 해당 지역 입력 시트로 이동합니다."
    ws['A8'] = "2. 각 지역별 시트에서 에너지시스템 구성요소 데이터를 입력합니다."
    ws['A9'] = "3. 입력 완료 후 '저장' 버튼을 클릭하여 데이터를 반영하거나, '추가 지역 입력'을 클릭하여 다른 지역을 입력합니다."
    ws['A10'] = "4. '지역간 연결' 시트에서 지역 간 연결을 정의합니다."
    ws['A11'] = "5. '시간 설정' 시트에서 분석 기간을 설정합니다."
    ws['A12'] = "6. 아래 '데이터 통합' 버튼을 클릭하여 입력한 데이터를 통합합니다."
    
    # 버튼 추가 (실제로는 VBA 매크로나 Python 스크립트와 연결해야 함)
    ws['B15'] = "데이터 통합"
    ws['B15'].font = Font(bold=True, color="FFFFFF")
    ws['B15'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws['B15'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('B15:D15')
    
    # 지역 선택 바로가기 버튼 추가
    ws['F15'] = "지역 선택으로 이동"
    ws['F15'].font = Font(bold=True)
    ws['F15'].fill = PatternFill(start_color=COLORS['highlight'], end_color=COLORS['highlight'], fill_type="solid")
    ws['F15'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F15'].hyperlink = "#'지역 선택'!A1"  # 지역 선택 시트로 연결
    ws.merge_cells('F15:H15')
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 12
    
    return ws

def create_region_selection_sheet(wb):
    """지역 선택 시트 생성"""
    ws = wb.create_sheet("지역 선택", 1)
    
    # 제목
    ws['A1'] = "분석할 지역 선택"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # 설명
    ws['A3'] = "아래 목록에서 분석할 지역을 선택하세요. '선택' 열에 'O'를 입력하면 해당 지역이 분석에 포함됩니다."
    ws['A4'] = "각 지역 코드를 클릭하면 해당 지역 입력 시트로 이동합니다."
    ws.merge_cells('A3:F3')
    ws.merge_cells('A4:F4')
    
    # 헤더
    headers = ['코드', '지역명', '영문명', '선택', '인구(명)', '면적(km²)']
    for i, header in enumerate(headers):
        cell = ws.cell(row=5, column=i+1)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 지역 목록
    region_selector = RegionalSelector()
    row = 6
    for code, region in sorted(KOREA_REGIONS.items()):
        # 지역 코드 셀에 하이퍼링크 설정
        code_cell = ws.cell(row=row, column=1)
        code_cell.value = code
        code_cell.font = Font(color="0000FF", underline="single")  # 파란색 밑줄
        code_cell.hyperlink = f"#'지역_{code}'!A1"
        
        ws.cell(row=row, column=2).value = region['name']
        ws.cell(row=row, column=3).value = region['name_eng']
        
        # 선택 셀 설정
        select_cell = ws.cell(row=row, column=4)
        select_cell.value = ""  # 여기에 'O'를 입력하면 선택됨
        
        ws.cell(row=row, column=5).value = region['population']
        ws.cell(row=row, column=6).value = region['area']
        
        # 서울과 전북은 기본적으로 선택되도록 설정
        if code in ['SEL', 'JBD']:
            select_cell.value = "O"
        
        row += 1
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # 데이터 유효성 검사 (선택 열에는 'O' 또는 빈 셀만 허용)
    dv = DataValidation(type="list", formula1='"O,"')
    ws.add_data_validation(dv)
    dv.add(f'D6:D{row-1}')
    
    return ws

def create_region_data_sheets(wb):
    """지역별 데이터 시트 생성 - 모든 지역에 대한 시트 생성"""
    # 모든 지역에 대한 시트 생성
    for region_code, region in sorted(KOREA_REGIONS.items()):
        create_region_sheet(wb, region_code, region['name'])
    
    return

def create_region_sheet(wb, region_code, region_name):
    """특정 지역의 데이터 시트 생성 - 모든 구성요소를 한 시트에 통합"""
    # 지역 정보 가져오기
    region_info = KOREA_REGIONS[region_code]
    
    # 시트 이름 생성 (예: 지역_SEL)
    sheet_name = f"지역_{region_code}"
    ws = wb.create_sheet(sheet_name)
    
    # 지역별 배경색 설정
    region_color = COLORS['sel_region'] if region_code == 'SEL' else COLORS['jbd_region']
    
    # 제목
    ws['A1'] = f"{region_name} 에너지시스템 입력"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].fill = PatternFill(start_color=region_color, end_color=region_color, fill_type="solid")
    ws.merge_cells('A1:H1')
    
    # 지역 정보
    ws['A3'] = "지역:"
    ws['B3'] = region_name
    ws['A4'] = "코드:"
    ws['B4'] = region_code
    ws['A5'] = "중심좌표:"
    ws['B5'] = f"({region_info['center'][0]}, {region_info['center'][1]})"
    
    # 선택 상자 추가
    ws['F3'] = "입력 완료 후 선택:"
    ws['F3'].font = Font(bold=True)
    ws['G3'] = "추가 지역 입력"
    ws['G3'].fill = PatternFill(start_color=COLORS['highlight'], end_color=COLORS['highlight'], fill_type="solid")
    ws['G3'].alignment = Alignment(horizontal='center')
    # 추가지역입력 선택 시 지역 선택 시트로 이동하는 링크 (VBA 매크로에서 사용)
    ws['G3'].hyperlink = f"#'지역 선택'!A1"
    
    ws['H3'] = "저장"
    ws['H3'].fill = PatternFill(start_color=COLORS['highlight'], end_color=COLORS['highlight'], fill_type="solid")
    ws['H3'].alignment = Alignment(horizontal='center')
    # 저장 선택 시 통합 데이터에 반영하는 링크 (VBA 매크로에서 사용)
    ws['H3'].hyperlink = f"#'제어판'!A1"
    
    # 현재 행 위치
    current_row = 7
    
    # 각 구성요소 섹션 추가 (송전선 제외)
    for component, template in COMPONENT_TEMPLATES.items():
        # 송전선 건너뛰기
        if component == 'lines':
            continue
            
        # 구성요소 제목
        ws.cell(row=current_row, column=1).value = f"{get_component_display_name(component)}"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=region_color, end_color=region_color, fill_type="solid")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        
        current_row += 2  # 헤더 행을 위한 간격
        
        # 헤더 행
        headers = template['columns']
        # 'region' 필드는 표시하지 않음 (자동으로 처리됨)
        if 'region' in headers:
            headers = [h for h in headers if h != 'region']
            
        for i, header in enumerate(headers):
            cell = ws.cell(row=current_row, column=i+1)
            cell.value = get_field_display_name(header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            
            # 필수 필드 표시
            if header in template['required']:
                cell.fill = PatternFill(start_color=COLORS['required'], end_color=COLORS['required'], fill_type="solid")
        
        current_row += 1  # 데이터 입력 행 시작
        
        # 샘플 데이터 추가 (템플릿에 있는 경우)
        if region_code in REGION_TEMPLATES and component in REGION_TEMPLATES[region_code]:
            items = REGION_TEMPLATES[region_code][component]
            for i, item in enumerate(items):
                for j, header in enumerate(headers):
                    cell = ws.cell(row=current_row+i, column=j+1)
                    
                    # 템플릿 아이템에서 값 가져오기
                    if header in item:
                        value = item[header]
                        # 버스 참조에는 지역 접두사 제거 (UI에서만)
                        if header in ['bus', 'bus0', 'bus1'] and value:
                            prefix = f"{region_code}_"
                            if str(value).startswith(prefix):
                                value = str(value)[len(prefix):]
                        cell.value = value
            
            # 아이템 수에 따라 행 증가
            current_row += len(items)
        
        # 빈 행 추가 (최소 5개)
        for i in range(5):
            current_row += 1
        
        # 구성요소 구분선 추가
        for col in range(1, 9):
            ws.cell(row=current_row, column=col).border = Border(bottom=Side(style='thin'))
        
        current_row += 3  # 다음 구성요소를 위한 간격
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    return ws

def create_connection_sheet(wb):
    """지역간 연결 시트 생성"""
    ws = wb.create_sheet("지역간 연결", 50)
    
    # 제목
    ws['A1'] = "지역간 연결 정의"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # 설명
    ws['A3'] = "이 시트에서는 선택한 지역 간의 연결(송전선 등)을 정의합니다."
    ws.merge_cells('A3:G3')
    
    # 헤더
    headers = ['이름', '시작 지역', '시작 버스', '도착 지역', '도착 버스', '용량(MW)', '전압(kV)', '거리(km)', '리액턴스(p.u.)', '저항(p.u.)']
    for i, header in enumerate(headers):
        cell = ws.cell(row=5, column=i+1)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 기본 연결 예시 (서울-전북)
    if 'SEL' in KOREA_REGIONS and 'JBD' in KOREA_REGIONS:
        distance = calculate_distance(
            KOREA_REGIONS['SEL']['center'],
            KOREA_REGIONS['JBD']['center']
        )
        
        row = 6
        ws.cell(row=row, column=1).value = "SEL_to_JBD_Line"
        ws.cell(row=row, column=2).value = "SEL"
        ws.cell(row=row, column=3).value = "Main_EL"
        ws.cell(row=row, column=4).value = "JBD"
        ws.cell(row=row, column=5).value = "Main_EL"
        ws.cell(row=row, column=6).value = 1000
        ws.cell(row=row, column=7).value = 345
        ws.cell(row=row, column=8).value = round(distance, 1)
        ws.cell(row=row, column=9).value = round(distance * 0.0004, 5)
        ws.cell(row=row, column=10).value = round(distance * 0.0001, 5)
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    
    return ws

def create_timeseries_sheet(wb):
    """시간 설정 시트 생성"""
    ws = wb.create_sheet("시간 설정", 51)
    
    # 제목
    ws['A1'] = "시간 설정"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # 설명
    ws['A3'] = "시간 설정을 통해 분석 기간과 시간 간격을 정의합니다."
    ws.merge_cells('A3:F3')
    
    # 헤더
    headers = ['시작 시간', '종료 시간', '시간 간격']
    for i, header in enumerate(headers):
        cell = ws.cell(row=5, column=i+1)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 기본값 설정
    ws.cell(row=6, column=1).value = datetime(2023, 1, 1, 0, 0, 0)
    ws.cell(row=6, column=2).value = datetime(2024, 1, 1, 0, 0, 0)
    ws.cell(row=6, column=3).value = "1h"
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    
    return ws

def create_pattern_sheets(wb):
    """패턴 및 제약사항 시트 생성"""
    # 1. 재생에너지 패턴 시트
    create_renewable_patterns_sheet(wb)
    
    # 2. 부하 패턴 시트 (전력, 열, 수소)
    create_load_patterns_sheet(wb)
    
    # 3. 지역별 부하 패턴 시트
    create_regional_load_patterns_sheet(wb)
    
    # 4. 제약사항 시트
    create_constraints_sheet(wb)
    
    return

def create_renewable_patterns_sheet(wb):
    """재생에너지 패턴 시트 생성"""
    ws = wb.create_sheet("renewable_patterns")
    
    # 제목
    ws['A1'] = "재생에너지 발전 패턴 설정"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # 설명
    ws['A3'] = "재생에너지 발전 패턴을 정의합니다. 시간별 정규화된 값(0~1)을 입력하세요."
    ws.merge_cells('A3:F3')
    
    # 헤더
    headers = ['시간', '태양광(PV)', '풍력(WT)']
    for i, header in enumerate(headers):
        cell = ws.cell(row=5, column=i+1)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 기본 패턴 예시 (24시간)
    for hour in range(24):
        row = hour + 6
        ws.cell(row=row, column=1).value = hour
        
        # 태양광 패턴 (낮에 높고 밤에 0)
        if 6 <= hour < 18:  # 낮 시간
            pv_value = 0.5 + 0.5 * np.sin(np.pi * (hour - 6) / 12)
        else:  # 밤 시간
            pv_value = 0.0
        ws.cell(row=row, column=2).value = round(pv_value, 2)
        
        # 풍력 패턴 (임의의 값)
        np.random.seed(hour)  # 재현성을 위한 시드 설정
        wt_value = 0.2 + 0.6 * np.random.rand()
        ws.cell(row=row, column=3).value = round(wt_value, 2)
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # 참고 사항
    ws['A31'] = "참고: 실제 분석에서는 8760시간(1년) 데이터가 필요합니다."
    ws['A32'] = "여기서는 예시로 24시간만 표시합니다."
    ws.merge_cells('A31:E31')
    ws.merge_cells('A32:E32')
    
    return ws

def create_load_patterns_sheet(wb):
    """부하 패턴 시트 생성 (전력, 열, 수소)"""
    ws = wb.create_sheet("load_patterns")
    
    # 제목
    ws['A1'] = "부하 패턴 설정 (국가 기본값)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # 설명
    ws['A3'] = "전력, 열, 수소 부하의 시간별 패턴을 정의합니다. 지역별 패턴이 없는 경우 이 기본값이 사용됩니다."
    ws.merge_cells('A3:G3')
    
    # 헤더
    headers = ['시간', '전력(electricity)', '열(heat)', '수소(hydrogen)']
    for i, header in enumerate(headers):
        cell = ws.cell(row=5, column=i+1)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 기본 패턴 예시 (24시간)
    for hour in range(24):
        row = hour + 6
        ws.cell(row=row, column=1).value = hour
        
        # 전력 부하 패턴 (낮에 높고 밤에 낮음)
        if 8 <= hour < 20:  # 주간 시간
            elec_value = 0.7 + 0.3 * np.sin(np.pi * (hour - 8) / 12)
        else:  # 야간 시간
            elec_value = 0.5 + 0.1 * np.sin(np.pi * hour / 12)
        ws.cell(row=row, column=2).value = round(elec_value, 2)
        
        # 열 부하 패턴 (아침, 저녁에 높음)
        if 6 <= hour < 9 or 17 <= hour < 22:  # 아침, 저녁
            heat_value = 0.8 + 0.2 * np.random.rand()
        else:  # 그 외 시간
            heat_value = 0.3 + 0.2 * np.random.rand()
        ws.cell(row=row, column=3).value = round(heat_value, 2)
        
        # 수소 부하 패턴 (비교적 일정)
        hydrogen_value = 0.7 + 0.1 * np.sin(2 * np.pi * hour / 24)
        ws.cell(row=row, column=4).value = round(hydrogen_value, 2)
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    return ws

def create_regional_load_patterns_sheet(wb):
    """지역별 부하 패턴 시트 생성"""
    ws = wb.create_sheet("regional_load_patterns")
    
    # 제목
    ws['A1'] = "지역별 부하 패턴 설정"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # 설명
    ws['A3'] = "각 지역별 부하 패턴을 정의합니다. 입력하지 않은 지역은 기본 국가 패턴이 적용됩니다."
    ws['A4'] = "지역 코드와 에너지 유형을 선택한 후, 해당 지역의 패턴을 입력하세요."
    ws.merge_cells('A3:G3')
    ws.merge_cells('A4:G4')
    
    # 헤더 - 지역 선택 부분
    ws['A6'] = "지역 코드"
    ws['B6'] = "에너지 유형"
    ws['C6'] = "패턴 적용"
    
    ws['A6'].font = Font(bold=True)
    ws['B6'].font = Font(bold=True)
    ws['C6'].font = Font(bold=True)
    
    ws['A6'].fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
    ws['B6'].fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
    ws['C6'].fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
    
    # 예시 행 추가
    ws['A7'] = "SEL"  # 서울
    ws['B7'] = "electricity"
    ws['C7'] = "O"  # 적용함
    
    ws['A8'] = "JBD"  # 전북
    ws['B8'] = "heat"
    ws['C8'] = "O"  # 적용함
    
    # 데이터 유효성 검사 설정
    # 지역 코드 드롭다운
    region_codes = [code for code in KOREA_REGIONS.keys()]
    region_dv = DataValidation(type="list", formula1=f'"{",".join(region_codes)}"')
    ws.add_data_validation(region_dv)
    region_dv.add('A7:A100')  # A7부터 A100까지 적용
    
    # 에너지 유형 드롭다운
    energy_types = ["electricity", "heat", "hydrogen"]
    energy_dv = DataValidation(type="list", formula1=f'"{",".join(energy_types)}"')
    ws.add_data_validation(energy_dv)
    energy_dv.add('B7:B100')  # B7부터 B100까지 적용
    
    # 패턴 적용 여부 드롭다운
    apply_dv = DataValidation(type="list", formula1='"O,X"')
    ws.add_data_validation(apply_dv)
    apply_dv.add('C7:C100')  # C7부터 C100까지 적용
    
    # 패턴 입력 부분 - 헤더
    ws['E6'] = "시간"
    ws['F6'] = "패턴값"
    
    ws['E6'].font = Font(bold=True)
    ws['F6'].font = Font(bold=True)
    
    ws['E6'].fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
    ws['F6'].fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
    
    # 시간별 패턴 입력란
    for hour in range(24):
        ws.cell(row=hour+7, column=5).value = hour
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    
    return ws

def create_constraints_sheet(wb):
    """제약사항 시트 생성"""
    ws = wb.create_sheet("constraints")
    
    # 제목
    ws['A1'] = "에너지시스템 제약사항 설정"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # 설명
    ws['A3'] = "에너지시스템 모델링에 적용할 제약사항을 정의합니다."
    ws['A4'] = "예: 특정 지역의 재생에너지 비율, CO2 배출 제한, 특정 발전원의 최소/최대 용량 등"
    ws.merge_cells('A3:G3')
    ws.merge_cells('A4:G4')
    
    # 헤더
    headers = ['이름', '유형', '대상 속성', '조건', '상수값', '적용 지역', '설명']
    for i, header in enumerate(headers):
        cell = ws.cell(row=6, column=i+1)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 예시 데이터
    example_data = [
        ['CO2_limit', 'global', 'co2_emissions', 'less_equal', 1000000, 'ALL', 'CO2 배출량 제한'],
        ['RES_min_SEL', 'regional', 'p_nom_opt', 'greater_equal', 0.3, 'SEL', '서울 재생에너지 비율 30% 이상'],
        ['Coal_max', 'carrier', 'p_nom_opt', 'less_equal', 5000, 'ALL', '석탄 발전 최대 5GW']
    ]
    
    for i, row_data in enumerate(example_data):
        for j, value in enumerate(row_data):
            ws.cell(row=i+7, column=j+1).value = value
    
    # 데이터 유효성 검사 설정
    # 유형 드롭다운
    constraint_types = ["global", "regional", "carrier", "component"]
    type_dv = DataValidation(type="list", formula1=f'"{",".join(constraint_types)}"')
    ws.add_data_validation(type_dv)
    type_dv.add('B7:B100')  # B7부터 B100까지 적용
    
    # 조건 드롭다운
    conditions = ["equal", "less_equal", "greater_equal", "less", "greater"]
    condition_dv = DataValidation(type="list", formula1=f'"{",".join(conditions)}"')
    ws.add_data_validation(condition_dv)
    condition_dv.add('D7:D100')  # D7부터 D100까지 적용
    
    # 열 너비 조정
    for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], [20, 15, 15, 15, 15, 15, 30]):
        ws.column_dimensions[col].width = width
    
    return ws

def get_component_display_name(component):
    """구성요소 이름의 표시용 이름 반환"""
    display_names = {
        'buses': '버스',
        'generators': '발전기',
        'lines': '송전선',
        'loads': '부하',
        'stores': '저장장치',
        'links': '링크'
    }
    return display_names.get(component, component)

def get_field_display_name(field):
    """필드 이름의 표시용 이름 반환"""
    display_names = {
        'name': '이름',
        'region': '지역',
        'v_nom': '전압(kV)',
        'carrier': '캐리어',
        'x': 'X좌표',
        'y': 'Y좌표',
        'bus': '버스',
        'p_nom': '정격용량(MW)',
        'p_nom_extendable': '용량확장가능',
        'p_nom_min': '최소용량(MW)',
        'p_nom_max': '최대용량(MW)',
        'marginal_cost': '한계비용',
        'capital_cost': '설비비용',
        'efficiency': '효율',
        'p_max_pu': '최대출력비율',
        'p_min_pu': '최소출력비율',
        'committable': '기동정지가능',
        'ramp_limit_up': '증발속도제한',
        'min_up_time': '최소가동시간',
        'start_up_cost': '기동비용',
        'lifetime': '수명(년)',
        'bus0': '시작버스',
        'bus1': '종료버스',
        'bus2': '버스2',
        'bus3': '버스3',
        'r': '저항(p.u.)',
        's_nom': '정격용량(MVA)',
        's_nom_extendable': '용량확장가능',
        's_nom_min': '최소용량(MVA)',
        's_nom_max': '최대용량(MVA)',
        'length': '길이(km)',
        'p_set': '부하량(MW)',
        'e_nom': '저장용량(MWh)',
        'e_cyclic': '주기적운전',
        'standing_loss': '자체손실',
        'efficiency_store': '충전효율',
        'efficiency_dispatch': '방전효율',
        'e_initial': '초기저장량',
        'e_nom_max': '최대저장용량',
        'e_nom_min': '최소저장용량',
        'efficiency0': '효율0',
        'efficiency1': '효율1',
        'efficiency2': '효율2',
        'efficiency3': '효율3'
    }
    return display_names.get(field, field)

def calculate_distance(coord1, coord2):
    """두 좌표 간의 거리 계산 (하버사인 공식)"""
    lon1, lat1 = coord1
    lon2, lat2 = coord2
    
    # 하버사인 공식
    R = 6371.0  # 지구 반지름 (km)
    
    lat1_rad = np.radians(lat1)
    lon1_rad = np.radians(lon1)
    lat2_rad = np.radians(lat2)
    lon2_rad = np.radians(lon2)
    
    dlon = lon2_rad - lon1_rad
    dlat = lat2_rad - lat1_rad
    
    a = np.sin(dlat / 2)**2 + np.cos(lat1_rad) * np.cos(lat2_rad) * np.sin(dlon / 2)**2
    c = 2 * np.arctan2(np.sqrt(a), np.sqrt(1 - a))
    
    distance = R * c
    return distance

def create_integrated_sheets(wb):
    """통합 데이터 시트 생성"""
    # 통합 데이터의 각 구성요소별로 시트 생성
    for component, template in COMPONENT_TEMPLATES.items():
        sheet_name = f"통합_{component}"
        ws = wb.create_sheet(sheet_name, 100 + list(COMPONENT_TEMPLATES.keys()).index(component))
        
        # 제목
        ws['A1'] = f"통합 데이터 - {get_component_display_name(component)}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # 설명
        ws['A3'] = "이 시트는 선택한 모든 지역의 데이터가 통합되어 표시됩니다. 직접 수정하지 마세요."
        ws.merge_cells('A3:G3')
        
        # 헤더
        headers = template['columns']
        # 'region' 필드는 표시하지 않음 (자동으로 처리됨)
        if 'region' in headers:
            headers = [h for h in headers if h != 'region']
            
        for i, header in enumerate(headers):
            cell = ws.cell(row=5, column=i+1)
            cell.value = get_field_display_name(header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # 열 너비 조정
        for i, header in enumerate(headers):
            col_letter = get_column_letter(i+1)
            width = 20 if header in ['name', 'bus', 'bus0', 'bus1'] else 15
            ws.column_dimensions[col_letter].width = width
    
    return

def add_default_data(wb):
    """각 지역 시트에 기본 데이터를 추가합니다."""
    print("기본 데이터는 사용자가 직접 입력합니다.")
    return wb

if __name__ == "__main__":
    try:
        wb = create_excel_template()
        print("Excel 템플릿 생성 완료!")
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        import traceback
        traceback.print_exc() 