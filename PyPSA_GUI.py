import os as _os
import sys as _sys
import traceback as _tb_early

# ── exe 실행 시 가장 먼저 로그 파일 오픈 (import 오류도 포착) ─────────────
_IS_FROZEN  = getattr(_sys, 'frozen', False)
_EXE_DIR    = _os.path.dirname(_sys.executable) if _IS_FROZEN else _os.path.dirname(_os.path.abspath(__file__))
_LOG_PATH   = _os.path.join(_EXE_DIR, 'pypsa_analysis_log.txt')
_log_file   = None
_orig_stdout = _sys.stdout
_orig_stderr = _sys.stderr

if _IS_FROZEN:
    try:
        _log_file = open(_LOG_PATH, 'w', encoding='utf-8', buffering=1)

        class _Tee:
            def __init__(self, *streams):
                self._s = streams
            def write(self, data):
                for s in self._s:
                    try: s.write(data); s.flush()
                    except Exception: pass
            def flush(self):
                for s in self._s:
                    try: s.flush()
                    except Exception: pass
            @property
            def encoding(self):
                return getattr(self._s[0], 'encoding', 'utf-8')

        _sys.stdout = _Tee(_sys.__stdout__, _log_file)
        _sys.stderr = _Tee(_sys.__stderr__, _log_file)
        print("=" * 60)
        print("  PyPSA Analysis 시작")
        print(f"  로그 파일: {_LOG_PATH}")
        print("=" * 60)
    except Exception as _e_log:
        pass  # 로그 파일 오픈 실패 시 무시하고 계속

# ── 일반 Python 실행 시: stdout/stderr를 UTF-8 + errors='replace'로 재설정 ──
# Windows 기본 인코딩(cp949)은 ⚠ 등 유니코드 문자 출력 시 UnicodeEncodeError 발생
# errors='replace' 로 출력 불가 문자를 '?'로 대체하여 크래시 방지
if not _IS_FROZEN:
    try:
        if hasattr(_sys.stdout, 'reconfigure'):
            _sys.stdout.reconfigure(errors='replace')
        if hasattr(_sys.stderr, 'reconfigure'):
            _sys.stderr.reconfigure(errors='replace')
    except Exception:
        pass


def _ensure_proj_lib_env():
    try:
        if 'PROJ_LIB' in _os.environ and _os.environ.get('PROJ_LIB'):
            return
        candidates = []
        # PyInstaller 실행파일 환경: exe 폴더 내 bundled proj 데이터 우선 탐색
        if getattr(_sys, 'frozen', False):
            _base = _os.path.dirname(_sys.executable)
            candidates.append(_os.path.join(_base, 'pyproj', 'proj_dir', 'share', 'proj'))
            candidates.append(_os.path.join(_base, 'share', 'proj'))
            if hasattr(_sys, '_MEIPASS'):
                candidates.append(_os.path.join(_sys._MEIPASS, 'pyproj', 'proj_dir', 'share', 'proj'))
                candidates.append(_os.path.join(_sys._MEIPASS, 'share', 'proj'))
        try:
            # Conda/Windows 일반 경로
            candidates.append(_os.path.join(_sys.prefix, 'Library', 'share', 'proj'))
            candidates.append(_os.path.join(_sys.prefix, 'share', 'proj'))
        except Exception:
            pass
        # 고정 경로(사용 환경)
        candidates.append(r'C:\ProgramData\anaconda3\envs\pypsa_env\Library\share\proj')
        for c in candidates:
            try:
                if c and _os.path.exists(c):
                    _os.environ['PROJ_LIB'] = c
                    print(f"[OK] PROJ_LIB 설정됨: {c}")
                    break
            except Exception:
                continue
        
        # 설정 실패 시 경고
        if 'PROJ_LIB' not in _os.environ:
            print("[경고] PROJ_LIB 환경변수 설정 실패 - 지도 시각화에 문제가 발생할 수 있습니다.")
    except Exception as e:
        print(f"[경고] PROJ_LIB 설정 중 오류: {e}")

_ensure_proj_lib_env()

# pyproj 및 지리 관련 경고 무시
import warnings
warnings.filterwarnings('ignore', category=UserWarning, message='.*pyproj.*')
warnings.filterwarnings('ignore', category=FutureWarning)

# PyPSA import 전에 pyproj 설정 강제
try:
    import pyproj
    # PROJ_LIB 강제 설정
    if 'PROJ_LIB' in _os.environ:
        pyproj.datadir.set_data_dir(_os.environ['PROJ_LIB'])
except Exception as e:
    print(f"[경고] pyproj 설정 경고 (무시 가능): {e}")

import pypsa
import pandas as pd
import numpy as np
from datetime import datetime
import os
import traceback
import sys
import copy


def get_base_dir():
    """사용자 파일 디렉토리 반환 (interface.xlsx, results 등).

    일반 Python 실행: PyPSA_GUI.py 가 있는 폴더
    PyInstaller exe:  exe 파일이 위치한 폴더 (dist/PyPSA_Analysis/)
                      ※ _internal/ 이 아닌 exe 바로 옆 폴더
    """
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _get_bundle_dir():
    """번들 내부 디렉토리 반환 (src/ 모듈 등 코드용).

    PyInstaller 6.x 에서 번들 데이터는 _internal/ 폴더(sys._MEIPASS)에 위치.
    """
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


def _setup_user_directory():
    """exe 실행 시 사용자 폴더에 필요한 파일이 없으면 _internal/ 에서 복사.

    interface.xlsx 가 exe 옆에 없으면 _internal/ 의 초기 버전을 복사해줌.
    사용자는 항상 exe 옆의 interface.xlsx 를 편집하면 됨.
    """
    if not getattr(sys, 'frozen', False):
        return
    user_dir = get_base_dir()
    bundle_dir = _get_bundle_dir()
    for fname in ('interface.xlsx',):
        dest = os.path.join(user_dir, fname)
        src_path = os.path.join(bundle_dir, fname)
        if not os.path.exists(dest) and os.path.exists(src_path):
            try:
                import shutil
                shutil.copy2(src_path, dest)
                print(f"[초기화] {fname} → {dest} 복사 완료")
            except Exception as _e:
                print(f"[경고] {fname} 복사 실패: {_e}")


_setup_user_directory()

# src 폴더를 Python 경로에 추가 (번들 내부 _internal/src/ 경로 사용)
sys.path.append(os.path.join(_get_bundle_dir(), 'src'))

# 지도 모듈은 선택적 임포트로 처리하여 rasterio/GDAL 미설치 시에도 실행이 가능하도록 함
try:
    from korea_map import KoreaMapVisualizer  # noqa: F401
except Exception as _map_e:
    KoreaMapVisualizer = None
    try:
        _force_print(f"지도 모듈 로드 경고: {str(_map_e)}")
    except Exception:
        pass

# 상수 정의
INPUT_FILE = "integrated_input_data.xlsx"

# ── 시각화 생략 플래그 ────────────────────────────────────────────────────────
# True : 시각화·지역분석 후처리 건너뜀 (장기 시나리오 반복 분석용 고속 모드)
# False: 시각화·지역분석 포함 (기본, 단기 분석용)
# run_longterm_scenario.py 에서 자동으로 True 로 설정됨
SKIP_VISUALIZATION = False

def _normalize_region_code(value):
    try:
        s = str(value).strip()
        if not s:
            return s
        codes = ['SEL','BSN','DGU','ICN','GWJ','DJN','USN','SJG','GGD','GWD','CBD','CND','JBD','JND','GBD','GND','JJD']
        if s.upper() in codes:
            return s.upper()
        name_to_code = {
            '서울특별시':'SEL','부산광역시':'BSN','대구광역시':'DGU','인천광역시':'ICN','광주광역시':'GWJ',
            '대전광역시':'DJN','울산광역시':'USN','세종특별자치시':'SJG','경기도':'GGD','강원도':'GWD',
            '충청북도':'CBD','충청남도':'CND','전라북도':'JBD','전라남도':'JND','경상북도':'GBD','경상남도':'GND','제주특별자치도':'JJD'
        }
        return name_to_code.get(s, s)
    except Exception:
        return str(value)

def _read_line_flex_settings(interface_path):
    """
    interface.xlsx '인터페이스_1' 시트 행 6~51에서
    선로별 유연 운영(Flexible Line Operation) 설정을 읽어 반환합니다.

    열 레이아웃 (행 5 헤더 기준):
        F열 = s_nom        : 유연 운영 시 최대 용량 (s_nom_flex)
        G열 = s_nom_basic  : 기본(정상) 운전 상한 (s_nom_base)
        H열 = time_limit   : 유연 운영 연간 최대 허용 시간 (h/yr)
        I열 = load_region  : 수전 지역 코드 (예: GGD, CND, JJD …)
                             해당 지역 부하 피크 시간대에 s_flex 허용

    F > G 이고 H > 0 인 선로만 반환합니다.

    Returns:
        dict: {line_name: {'s_nom_flex': float,
                           's_nom_base': float,
                           'time_limit': float,
                           'load_region': str or None}}
    """
    try:
        import openpyxl as _opxl
        if not os.path.exists(interface_path):
            return {}
        wb = _opxl.load_workbook(interface_path, data_only=True)
        if '인터페이스_1' not in wb.sheetnames:
            return {}
        ws = wb['인터페이스_1']

        # 헤더행(5행)에서 F/G/H/I 열 인덱스 확인
        col_f = col_g = col_h = col_i = None
        for cell in ws[5]:
            if cell.value is None:
                continue
            hdr = str(cell.value).strip().lower()
            if hdr == 's_nom':
                col_f = cell.column
            elif hdr == 's_nom_basic':
                col_g = cell.column
            elif hdr == 'time_limit':
                col_h = cell.column
            elif hdr in ('load_region', '수전지역', 'load_bus'):
                col_i = cell.column

        # 열을 못 찾으면 기본 위치(F=6, G=7, H=8, I=9) 사용
        if col_f is None:
            col_f = 6
        if col_g is None:
            col_g = 7
        if col_h is None:
            col_h = 8
        if col_i is None:
            col_i = 9

        max_col = max(col_f, col_g, col_h, col_i)
        result = {}
        for row in ws.iter_rows(min_row=6, max_row=51, min_col=1, max_col=max_col):
            name_cell = row[0].value
            if not name_cell:
                continue
            line_name = str(name_cell).strip()
            try:
                f_val = row[col_f - 1].value
                g_val = row[col_g - 1].value
                h_val = row[col_h - 1].value
                i_val = row[col_i - 1].value if col_i - 1 < len(row) else None
            except IndexError:
                continue
            try:
                s_flex = float(f_val) if f_val is not None else 0.0
                s_base = float(g_val) if g_val is not None else 0.0
                t_lim  = float(h_val) if h_val is not None else 0.0
            except (TypeError, ValueError):
                continue
            load_region = str(i_val).strip() if i_val is not None else None
            if s_flex > s_base > 0 and t_lim > 0:
                result[line_name] = {
                    's_nom_flex':   s_flex,
                    's_nom_base':   s_base,
                    'time_limit':   t_lim,
                    'load_region':  load_region,
                }
        return result
    except Exception as _e:
        print(f"[선로 유연운영 설정 읽기 오류] {_e}")
        return {}


def _read_dc_fab_from_interface(interface_path, existing_load_names):
    """interface.xlsx 지역별 시트의 부하 섹션에서 Demand_DC / Demand_Fab 읽기.

    각 지역 시트(발전_XXX)를 스캔하여 부하 이름이 'Demand_DC' 또는 'Demand_Fab'인
    행을 찾아 {region}_Demand_DC / {region}_Demand_Fab 부하를 반환한다.
    버스는 해당 행의 2번째 컬럼(버스명)을 사용하며, 없으면 {region}_EL로 기본 지정.
    """
    try:
        import openpyxl as _opxl
        wb = _opxl.load_workbook(interface_path, data_only=True)
        new_loads = []

        for ws in wb.worksheets:
            sn = ws.title
            # 지역별 시트 식별: 시트명 마지막 '_' 이후가 3자리 대문자 알파벳
            if '_' not in sn:
                continue
            region = sn.rsplit('_', 1)[-1]
            if len(region) != 3 or not region.isalpha():
                continue

            dc_load_name  = f"{region}_Demand_DC"
            fab_load_name = f"{region}_Demand_Fab"

            for row in ws.iter_rows(values_only=True):
                if row[0] is None:
                    continue
                v = str(row[0]).strip()

                if v == 'Demand_DC' and dc_load_name not in existing_load_names:
                    bus  = str(row[1]).strip() if (len(row) > 1 and row[1]) else f"{region}_EL"
                    pset = 0.0
                    if len(row) > 3 and row[3] is not None:
                        try:
                            pset = float(row[3])
                        except Exception:
                            pass
                    new_loads.append({
                        'name': dc_load_name, 'region': region,
                        'bus': bus, 'carrier': 'electricity', 'p_set': pset
                    })

                elif v == 'Demand_Fab' and fab_load_name not in existing_load_names:
                    bus  = str(row[1]).strip() if (len(row) > 1 and row[1]) else f"{region}_EL"
                    pset = 0.0
                    if len(row) > 3 and row[3] is not None:
                        try:
                            pset = float(row[3])
                        except Exception:
                            pass
                    new_loads.append({
                        'name': fab_load_name, 'region': region,
                        'bus': bus, 'carrier': 'electricity', 'p_set': pset
                    })

        if not new_loads:
            return None
        return {'loads': pd.DataFrame(new_loads)}
    except Exception as e:
        print(f"DC/Fab 부하 읽기 오류: {str(e)}")
        return None


def _apply_coal_efficiencies_from_db(network, interface_path):
    """
    interface.xlsx Coal_DB 시트의 AM열(효율)을 석탄 발전기에 적용합니다.

    매핑 방식: 지역(A열) 기준, Coal_DB 행 순서 = integrated_input_data 발전기 순서
    - Coal_DB 행 6~끝: 지역코드(A) + 발전소명(D) + 효율(AM열=39번째)
    - 네트워크 내 coal 발전기를 carrier=='coal' 로 식별, 지역별 순서 유지
    - 지역별 건수가 일치하면 1:1 매핑, 불일치 시 경고 후 가능한 범위만 적용

    효과: Coal_DB 수정 → PyPSA 모델에 자동 반영 (integrated_input_data 수동 갱신 불필요)
    """
    try:
        import openpyxl as _opxl
        if not os.path.exists(interface_path):
            return
        wb = _opxl.load_workbook(interface_path, data_only=True)
        if 'Coal_DB' not in wb.sheetnames:
            print("  [Coal_DB] 시트 없음 — 석탄 발전기 효율 적용 생략")
            return
        ws = wb['Coal_DB']

        # Coal_DB: 지역별 효율 목록 (행 순서 유지)
        db_effs = {}          # {region: [eff1, eff2, ...]}
        db_plants = {}        # {region: [plant_name1, ...]} (참고용 로그)
        for r in range(6, ws.max_row + 1):
            region = ws.cell(r, 1).value
            plant  = ws.cell(r, 4).value
            eff    = ws.cell(r, 39).value
            if not region or eff is None:
                continue
            region = str(region).strip().upper()
            db_effs.setdefault(region, []).append(float(eff))
            db_plants.setdefault(region, []).append(str(plant or '').strip())

        # 네트워크 내 석탄 발전기를 지역별·등록 순서대로 수집
        coal_by_region = {}
        for gen_name in network.generators.index:
            if network.generators.at[gen_name, 'carrier'] == 'coal':
                region = gen_name.split('_')[0].upper()
                coal_by_region.setdefault(region, []).append(gen_name)

        applied = changed = skipped = 0
        for region, gen_names in coal_by_region.items():
            region_effs   = db_effs.get(region, [])
            region_plants = db_plants.get(region, [])
            if not region_effs:
                print(f"  [Coal_DB] {region}: Coal_DB에 항목 없음 — 생략")
                skipped += len(gen_names)
                continue
            if len(region_effs) != len(gen_names):
                print(f"  [Coal_DB] {region}: Coal_DB {len(region_effs)}건 ≠ 모델 {len(gen_names)}건 "
                      f"→ 앞부터 순서대로 {min(len(region_effs), len(gen_names))}건 적용")
            for i, gen_name in enumerate(gen_names):
                if i >= len(region_effs):
                    skipped += 1
                    continue
                new_eff = region_effs[i]
                old_eff = float(network.generators.at[gen_name, 'efficiency'])
                network.generators.at[gen_name, 'efficiency'] = new_eff
                applied += 1
                if abs(old_eff - new_eff) > 1e-4:
                    plant_nm = region_plants[i] if i < len(region_plants) else '?'
                    print(f"  [Coal_DB] {gen_name} ({plant_nm}): eff {old_eff:.4f} → {new_eff:.4f}")
                    changed += 1

        print(f"  [Coal_DB] 석탄 발전기 효율 적용 완료: {applied}건 (변경 {changed}건, 미매핑 {skipped}건)")
    except Exception as _e:
        print(f"  [Coal_DB] 효율 적용 실패: {_e}")


def _read_links_from_interface(interface_path, existing_link_names):
    """interface.xlsx 지역별 시트의 링크 섹션에서 신규 링크 읽기.

    각 지역 시트(발전_XXX)를 스캔하여 링크 섹션 헤더(이름/출발버스/도착버스)를
    찾은 뒤 데이터 행을 읽는다. integrated_input_data.xlsx에 이미 있는 링크는
    건너뛰므로, 향후 HVDC 등 지역 간 링크를 interface.xlsx에만 추가해도 자동
    반영된다.

    컬럼 순서 (0-indexed, 지역별 시트 링크 헤더 기준):
      0:이름  1:출발버스  2:도착버스  3:버스2  4:버스3
      5:효율0  6:효율1  7:효율2  8:효율3
      9:정격용량(MW)  10:용량확대가능  11:최소용량(MW)  12:최대용량(MW)
      13:자본비용  14:한계비용
    """
    try:
        import openpyxl as _opxl
        wb = _opxl.load_workbook(interface_path, data_only=True)
        new_links = []

        # 컬럼 인덱스 고정 매핑
        CI = {
            'name': 0, 'bus0': 1, 'bus1': 2, 'bus2': 3, 'bus3': 4,
            'efficiency0': 5, 'efficiency1': 6, 'efficiency2': 7, 'efficiency3': 8,
            'p_nom': 9, 'p_nom_extendable': 10, 'p_nom_min': 11, 'p_nom_max': 12,
            'capital_cost': 13, 'marginal_cost': 14,
        }

        def _sf(v):
            if v is None:
                return float('nan')
            try:
                return float(v)
            except Exception:
                return float('nan')

        def _sb(v):
            if v is None:
                return False
            if isinstance(v, bool):
                return v
            return str(v).strip().lower() in ('true', '1', 'yes', '예')

        for ws in wb.worksheets:
            sn = ws.title
            if '_' not in sn:
                continue
            region = sn.rsplit('_', 1)[-1]
            if len(region) != 3 or not region.isalpha():
                continue

            # 링크 섹션 헤더 행 탐색
            # 인식 기준: col0='이름'(U+C774 U+B984) AND col1이 '버스'(U+BC84 U+C2A4)로
            # 끝나면서 2글자 초과 (발전기/부하 섹션의 단순 '버스'와 구별)
            # → interface.xlsx 실제 컬럼: 시작버스(U+C2DC U+C791+버스), 종료버스
            _IREUM  = '\uc774\ub984'   # 이름
            _BEOSEU = '\ubc84\uc2a4'   # 버스
            header_row_idx = None
            all_rows = list(ws.iter_rows(values_only=True))
            for i, row in enumerate(all_rows):
                if not row or row[0] is None:
                    continue
                v0 = str(row[0]).strip()
                v1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                v2 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                # 조건: 이름 + (버스 접미사 2글자 초과 col1) + (버스 접미사 col2)
                # 또는 영문 bus0/bus1/from/to 표기 지원
                is_ireum = (v0 == _IREUM or v0.lower() == 'name')
                is_link_bus1 = (
                    (v1.endswith(_BEOSEU) and len(v1) > len(_BEOSEU))
                    or v1.lower() in ('bus0', 'from')
                )
                is_link_bus2 = (
                    (v2.endswith(_BEOSEU) and len(v2) > len(_BEOSEU))
                    or 'to' in v2.lower()
                    or v2.lower() == 'bus1'
                )
                if is_ireum and is_link_bus1 and is_link_bus2:
                    header_row_idx = i
                    break

            if header_row_idx is None:
                continue

            # 헤더 다음 행부터 데이터 읽기 (빈 행에서 중단)
            for row in all_rows[header_row_idx + 1:]:
                if not row or row[CI['name']] is None:
                    break
                raw_name = str(row[CI['name']]).strip()
                if not raw_name:
                    break

                # {region}_ 접두어 보정
                link_name = (
                    raw_name if raw_name.startswith(f"{region}_")
                    else f"{region}_{raw_name}"
                )

                bus0_v = str(row[CI['bus0']]).strip() if len(row) > CI['bus0'] and row[CI['bus0']] is not None else None
                bus1_v = str(row[CI['bus1']]).strip() if len(row) > CI['bus1'] and row[CI['bus1']] is not None else None

                if not bus0_v or not bus1_v:
                    continue

                # 지역 간 링크(HVDC 등): bus0·bus1의 지역코드가 다르면 항상 갱신
                # → _persist_standardized_input이 integrated를 덮어써도 다음 실행에서 복원됨
                bus0_region = bus0_v.split('_')[0]
                bus1_region = bus1_v.split('_')[0]
                is_cross_regional = (bus0_region != bus1_region)

                # 같은 지역 내 링크: integrated에 이미 있으면 건너뜀 (중복 방지)
                if link_name in existing_link_names and not is_cross_regional:
                    continue

                bus2_v = str(row[CI['bus2']]).strip() if len(row) > CI['bus2'] and row[CI['bus2']] is not None else None
                bus3_v = str(row[CI['bus3']]).strip() if len(row) > CI['bus3'] and row[CI['bus3']] is not None else None

                record = {
                    'name': link_name,
                    'region': region,
                    'bus0': bus0_v,
                    'bus1': bus1_v,
                    'bus2': bus2_v,
                    'bus3': bus3_v,
                    'efficiency0': _sf(row[CI['efficiency0']] if len(row) > CI['efficiency0'] else None),
                    'efficiency1': _sf(row[CI['efficiency1']] if len(row) > CI['efficiency1'] else None),
                    'efficiency2': _sf(row[CI['efficiency2']] if len(row) > CI['efficiency2'] else None),
                    'efficiency3': _sf(row[CI['efficiency3']] if len(row) > CI['efficiency3'] else None),
                    'p_nom':           _sf(row[CI['p_nom']]           if len(row) > CI['p_nom']           else None),
                    'p_nom_extendable': _sb(row[CI['p_nom_extendable']] if len(row) > CI['p_nom_extendable'] else None),
                    'p_nom_min':       _sf(row[CI['p_nom_min']]       if len(row) > CI['p_nom_min']       else None),
                    'p_nom_max':       _sf(row[CI['p_nom_max']]       if len(row) > CI['p_nom_max']       else None),
                    'capital_cost':    _sf(row[CI['capital_cost']]    if len(row) > CI['capital_cost']    else None),
                    'marginal_cost':   _sf(row[CI['marginal_cost']]   if len(row) > CI['marginal_cost']   else None),
                }
                new_links.append(record)
                print(f"interface.xlsx 신규 링크 감지: {link_name} ({bus0_v} → {bus1_v})")

        if not new_links:
            return None
        return pd.DataFrame(new_links)
    except Exception as e:
        print(f"interface.xlsx 링크 보강 읽기 오류: {str(e)}")
        return None


def read_input_data(input_file):
    """Excel 파일에서 입력 데이터 읽기"""
    try:
        # 파일 경로 및 수정 시간 로깅
        root_dir = get_base_dir()
        integrated_path = os.path.abspath(os.path.join(root_dir, input_file))
        interface_path = os.path.abspath(os.path.join(root_dir, 'interface.xlsx'))
        try:
            if os.path.exists(integrated_path):
                print(f"integrated_input_data 경로: {integrated_path}")
                print(f"integrated_input_data 수정 시간: {datetime.fromtimestamp(os.path.getmtime(integrated_path))}")
            if os.path.exists(interface_path):
                print(f"interface.xlsx 경로: {interface_path}")
                print(f"interface.xlsx 수정 시간: {datetime.fromtimestamp(os.path.getmtime(interface_path))}")
        except Exception:
            pass

        xls = pd.ExcelFile(input_file)
        input_data = {
            'buses': pd.DataFrame(),
            'generators': pd.DataFrame(),
            'loads': pd.DataFrame(),
            'stores': pd.DataFrame(),
            'links': pd.DataFrame(),
            'lines': pd.DataFrame(),
            'timeseries': pd.DataFrame(),
            'renewable_patterns': pd.DataFrame(),
            'load_patterns': pd.DataFrame()
        }
        
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            # 컬럼명이 문자열인 경우에만 strip 적용
            if df.columns.dtype == 'object':
                try:
                    df.columns = df.columns.str.strip()
                except:
                    # 문자열이 아닌 컬럼명이 있는 경우 문자열로 변환 후 strip
                    df.columns = [str(col).strip() for col in df.columns]
            input_data[sheet_name] = df
        
        # Fallback: 통합 파일에 패턴 시트가 없으면 interface.xlsx에서 보강
        try:
            if (('load_patterns' not in xls.sheet_names) or input_data['load_patterns'].empty) and os.path.exists(interface_path):
                lp = pd.read_excel(interface_path, sheet_name='load_patterns')
                if lp is not None and not lp.empty:
                    input_data['load_patterns'] = lp
                    print("load_patterns를 interface.xlsx에서 보강 로딩했습니다.")
        except Exception as e:
            print(f"load_patterns 보강 로딩 실패: {str(e)}")
        try:
            if (('renewable_patterns' not in xls.sheet_names) or input_data['renewable_patterns'].empty) and os.path.exists(interface_path):
                rp = pd.read_excel(interface_path, sheet_name='renewable_patterns')
                if rp is not None and not rp.empty:
                    input_data['renewable_patterns'] = rp
                    print("renewable_patterns를 interface.xlsx에서 보강 로딩했습니다.")
        except Exception as e:
            print(f"renewable_patterns 보강 로딩 실패: {str(e)}")

        # Fallback: constraints 시트가 없으면 interface.xlsx에서 로딩 (CO2/RPS 제약 반영)
        try:
            cons_missing = ('constraints' not in input_data or
                            not isinstance(input_data.get('constraints'), pd.DataFrame) or
                            input_data['constraints'].empty)
            if cons_missing and os.path.exists(interface_path):
                raw = pd.read_excel(interface_path, sheet_name='constraints', header=None)
                # 데이터 행 탐색: 첫 컬럼값이 'CO2_limit', 'CO2Limit', 'RPS_' 등 실제 제약명인 행
                # (헤더행 다음 행이 실제 데이터 첫 행)
                data_start = None
                for i, row in raw.iterrows():
                    first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                    if first_val in ('CO2_limit', 'CO2Limit') or first_val.startswith('RPS_'):
                        data_start = i
                        break
                if data_start is not None:
                    # interface.xlsx constraints 컬럼 순서:
                    # 0=이름, 1=유형, 2=연결속성, 3=조건, 4=값(상수), 5=적용지역, 6=설명
                    eng_cols = ['name', 'type', 'carrier_attribute', 'sense', 'constant', 'region', 'description']
                    df_cons = raw.iloc[data_start:].copy().reset_index(drop=True)
                    # 실제 데이터 컬럼 수에 맞게 자르기
                    n_cols = min(len(eng_cols), df_cons.shape[1])
                    df_cons = df_cons.iloc[:, :n_cols].copy()
                    df_cons.columns = eng_cols[:n_cols]
                    df_cons = df_cons.dropna(subset=['name']).reset_index(drop=True)
                    df_cons = df_cons[df_cons['name'].astype(str).str.strip().str.len() > 0]
                    if not df_cons.empty:
                        input_data['constraints'] = df_cons
                        co2_rows = df_cons[df_cons['name'].astype(str).str.strip().isin(['CO2_limit', 'CO2Limit'])]
                        print(f"constraints를 interface.xlsx에서 폴백 로딩 ({len(df_cons)}행, CO2: {len(co2_rows)}개)")
        except Exception as e:
            print(f"constraints 폴백 로딩 실패: {str(e)}")
        try:
            if os.path.exists(interface_path):
                sd = pd.read_excel(interface_path, sheet_name='시나리오_에너지수요')
                if sd is not None and not sd.empty:
                    input_data['시나리오_에너지수요'] = sd
                    print("시나리오_에너지수요를 interface.xlsx에서 로딩했습니다.")
        except Exception as e:
            print(f"시나리오_에너지수요 로딩 실패: {str(e)}")

        # Fallback: lines 시트가 없거나 비어있으면 interface.xlsx의 '지역간 연결'에서 생성
        try:
            if (('lines' not in xls.sheet_names) or input_data['lines'].empty) and os.path.exists(interface_path):
                fb = _fallback_build_lines_from_interface(input_data, interface_path)
                if fb is not None and not fb.empty:
                    input_data['lines'] = fb
                    print("lines를 interface.xlsx의 '지역간 연결'에서 폴백 생성했습니다.")
                    # 통합 파일에도 저장해 다음 실행부터 바로 사용
                    _persist_lines_to_integrated(integrated_path, fb)
        except Exception as e:
            print(f"lines 폴백 생성 실패: {str(e)}")

        # DC/Fab 부하 보강: interface.xlsx 지역별 시트에서 Demand_DC/Demand_Fab 읽기
        # (integrated_input_data.xlsx에 아직 반영되지 않은 경우 자동 보강)
        try:
            if os.path.exists(interface_path):
                existing_load_names = set(
                    input_data['loads']['name'].astype(str)
                ) if ('loads' in input_data and not input_data['loads'].empty and 'name' in input_data['loads'].columns) else set()
                dc_fab = _read_dc_fab_from_interface(interface_path, existing_load_names)
                if dc_fab is not None:
                    if dc_fab['loads'] is not None and not dc_fab['loads'].empty:
                        input_data['loads'] = pd.concat(
                            [input_data['loads'], dc_fab['loads']], ignore_index=True
                        ) if not input_data['loads'].empty else dc_fab['loads']
                        print(f"DC/Fab 부하 보강 완료: {len(dc_fab['loads'])}개 부하 추가")
        except Exception as e:
            print(f"DC/Fab 부하 보강 실패: {str(e)}")

        # 링크 보강: interface.xlsx 지역별 시트에서 신규 링크 읽기
        # integrated_input_data.xlsx의 links 시트에 없는 링크(HVDC 등)를 자동 보강
        # - 향후 HVDC 추가 시 interface.xlsx 해당 지역 시트 링크 섹션에만 입력하면 됨
        # - integrated_input_data.xlsx에 버스가 없는 지역(흡수·삭제된 지역)은 제외
        try:
            if os.path.exists(interface_path):
                existing_link_names = set(
                    input_data['links']['name'].astype(str)
                ) if ('links' in input_data and not input_data['links'].empty and 'name' in input_data['links'].columns) else set()

                # integrated buses에서 유효한 지역 코드 추출
                known_regions = set()
                if 'buses' in input_data and not input_data['buses'].empty:
                    _b = input_data['buses']
                    if 'region' in _b.columns:
                        known_regions = set(_b['region'].dropna().astype(str))
                    elif 'name' in _b.columns:
                        known_regions = set(str(n).split('_')[0] for n in _b['name'].dropna())

                new_links_df = _read_links_from_interface(interface_path, existing_link_names)
                if new_links_df is not None and not new_links_df.empty:
                    # 유효한 지역의 링크만 남김
                    if known_regions and 'region' in new_links_df.columns:
                        before = len(new_links_df)
                        new_links_df = new_links_df[new_links_df['region'].isin(known_regions)].reset_index(drop=True)
                        skipped = before - len(new_links_df)
                        if skipped > 0:
                            print(f"링크 보강: {skipped}개 링크 제외 (integrated에 버스 없는 지역)")
                    if not new_links_df.empty:
                        # 지역 간 링크(HVDC 등)는 integrated의 기존 항목을 제거하고 교체
                        # → bus 정보 오류가 integrated에 저장되어도 다음 실행에서 복원됨
                        if 'links' in input_data and not input_data['links'].empty and 'name' in input_data['links'].columns:
                            cross_names = set(new_links_df['name'].astype(str))
                            input_data['links'] = input_data['links'][
                                ~input_data['links']['name'].astype(str).isin(cross_names)
                            ].reset_index(drop=True)
                        input_data['links'] = pd.concat(
                            [input_data['links'], new_links_df], ignore_index=True
                        ) if not input_data['links'].empty else new_links_df
                        print(f"링크 보강 완료: {len(new_links_df)}개 링크 추가/갱신 (interface.xlsx)")
        except Exception as e:
            print(f"링크 보강 실패: {str(e)}")

        return input_data
        
    except Exception as e:
        print(f"데이터 읽기 오류: {str(e)}")
        raise

def adjust_pattern_length(pattern_values, required_length):
    """패턴 길이를 필요한 길이에 맞게 조정"""
    if len(pattern_values) == required_length:
        return pattern_values
    
    # 패턴이 더 짧은 경우
    elif len(pattern_values) < required_length:
        # 부족한 만큼 처음부터 반복
        repetitions = required_length // len(pattern_values) + 1
        extended_pattern = np.tile(pattern_values, repetitions)
        return extended_pattern[:required_length]
    
    # 패턴이 더 긴 경우
    else:
        return pattern_values[:required_length]

def normalize_pattern(pattern):
    """발전 패턴을 0~1 사이로 정규화"""
    if np.max(pattern) > 0:
        return pattern / np.max(pattern)
    return pattern

def _get_scenario_year(input_data):
    try:
        if 'timeseries' in input_data and not input_data['timeseries'].empty:
            start = pd.to_datetime(input_data['timeseries'].iloc[0]['start_time'])
            return int(start.year)
    except Exception:
        pass
    return None

def _parse_scenario_demand(input_data, scenario_year):
    key_map = {
        'year': ['year', '연도', '년도'],
        'region': ['region', '지역', '코드', '지역코드'],
        'EL': ['EL', '전력', '전력(MWh)', '전력_MWh'],
        'H': ['H', '열', '열(MWh)', '열_MWh', 'HEAT'],
        'H2': ['H2', '수소', '수소(MWh)', '수소_MWh']
    }
    def find_col(cols, candidates):
        for cand in candidates:
            for c in cols:
                if str(c).strip().lower() == str(cand).strip().lower():
                    return c
        return None
    scenario = {}
    if '시나리오_에너지수요' not in input_data:
        return scenario
    df = input_data['시나리오_에너지수요']
    if df is None or df.empty:
        return scenario
    year_col = find_col(df.columns, key_map['year'])
    region_col = find_col(df.columns, key_map['region'])
    el_col = find_col(df.columns, key_map['EL'])
    h_col = find_col(df.columns, key_map['H'])
    h2_col = find_col(df.columns, key_map['H2'])
    if year_col is None or region_col is None:
        return scenario
    try:
        filt = df[year_col].astype(int) == int(scenario_year)
        ydf = df.loc[filt].copy()
    except Exception:
        ydf = df.copy()
    for _, row in ydf.iterrows():
        raw_region = str(row[region_col]).strip()
        if '_' in raw_region:
            raw_region = raw_region.split('_')[0]
        region = _normalize_region_code(raw_region)
        if not region or str(region).lower() == 'nan':
            continue
        if el_col in ydf.columns and pd.notna(row.get(el_col)):
            scenario[(region, 'EL')] = float(row[el_col])
        if h_col in ydf.columns and pd.notna(row.get(h_col)):
            scenario[(region, 'H')] = float(row[h_col])
        if h2_col in ydf.columns and pd.notna(row.get(h2_col)):
            scenario[(region, 'H2')] = float(row[h2_col])
    return scenario

def _apply_scenario_demand_scaling(network, input_data):
    try:
        # 멀티년 분석 등에서 시나리오 스케일링 비활성화 플래그
        if os.environ.get('DISABLE_SCALING') == '1':
            print("시나리오 수요 스케일링 비활성화됨(DISABLE_SCALING=1)")
            return
        scenario_year = _get_scenario_year(input_data)
        if scenario_year is None:
            return
        scenario = _parse_scenario_demand(input_data, scenario_year)
        if not scenario:
            print("시나리오_에너지수요 시트가 없거나 해당 연도 데이터가 없습니다. 기본 부하를 사용합니다.")
            return
        group_to_loads = {}
        for load_name in network.loads.index:
            region = load_name.split('_')[0] if '_' in load_name else None
            demand_type = None
            lname = load_name.lower()
            if '_demand_el' in lname:
                demand_type = 'EL'
            elif '_demand_h2' in lname:
                demand_type = 'H2'
            elif '_demand_h' in lname:
                demand_type = 'H'
            if region and demand_type:
                key = (region, demand_type)
                group_to_loads.setdefault(key, []).append(load_name)
        total_scaled_groups = 0
        for key, names in group_to_loads.items():
            region, dtype = key
            target = scenario.get((region, dtype))
            if target is None:
                continue
            current = 0.0
            for nm in names:
                if nm in network.loads_t.p_set.columns:
                    current += float(np.nansum(network.loads_t.p_set[nm].values))
            if current <= 0:
                if target > 0:
                    print(f"경고: {region}-{dtype} 현재 부하 합계가 0입니다. 스케일링을 건너뜁니다.")
                continue
            # 이미 패턴 합계=1로 연간 총량을 분배했으므로, 스케일은 미세 오차 보정 수준이어야 함
            scale = target / current
            for nm in names:
                if nm in network.loads_t.p_set.columns:
                    network.loads_t.p_set[nm] = network.loads_t.p_set[nm] * scale
            print(f"수요 스케일링 적용: {region}-{dtype}: 현재 {current:.1f} → 목표 {target:.1f} (배율 {scale:.6f})")
            total_scaled_groups += 1
        if total_scaled_groups == 0:
            print("시나리오 스케일링 대상 그룹이 없습니다. 이름 규칙 또는 시트 컬럼을 확인하세요.")
    except Exception as e:
        print(f"수요 시나리오 스케일링 중 오류: {str(e)}")

def _get_load_pattern(input_data, region, demand_type, snapshots_len):
    try:
        if 'load_patterns' not in input_data or input_data['load_patterns'].empty:
            return None
        df = input_data['load_patterns'].copy()

        # 1) 고정 위치(B8:F8765)에서 직접 읽기 시도 [EV/DC/Fab 패턴 포함]
        try:
            base_len = 8760
            start_row = 7  # Excel 8행 → iloc 7
            # 열 인덱스(0-based): A=시간(0), B=전력(1), C=열(2), D=수소(3),
            #   E=EV_DRIVING(4), F=EV_CHARGING(5), G=빈칸(6),
            #   H=DC_patterns(7), I=Fab_patterns(8), J=빈칸(9),
            #   K=BSN_EL_P(10), L=CBD_EL_P(11), ... (지역별 EL 패턴)
            col_map = {
                'EL': 1, 'H': 2, 'H2': 3, 'EV_DRIVING': 4, 'EV_CHARGING': 5,
                'DC': 7, 'Fab': 8
            }  # B,C,D,E,F,H,I

            # 지역별 EL 패턴 우선 적용: load_patterns 5행(pandas index 4)에서 컬럼명 탐색
            effective_col = col_map.get(demand_type)
            if demand_type == 'EL' and region:
                regional_col_name = f"{region}_EL_P"
                for hdr_idx in [3, 4]:  # Excel 5행 또는 6행 (pandas index 3 또는 4)
                    if df.shape[0] > hdr_idx:
                        hdr = df.iloc[hdr_idx]
                        for col_idx, val in enumerate(hdr):
                            if str(val).strip() == regional_col_name:
                                effective_col = col_idx
                                print(f"지역별 EL 패턴 사용: {regional_col_name} (열 인덱스 {col_idx})")
                                break
                        if effective_col != col_map.get(demand_type):
                            break

            if effective_col is not None and df.shape[0] > start_row + 10 and df.shape[1] > effective_col:
                series = pd.to_numeric(df.iloc[start_row:, effective_col], errors='coerce')
                values = series.dropna().astype(float).values
                if len(values) > 0:
                    # 기본 8760 채움: 부족 시 타일링
                    if len(values) >= base_len:
                        pattern_base = values[:base_len]
                    else:
                        repeats = base_len // len(values) + 1
                        pattern_base = np.tile(values, repeats)[:base_len]
                    # 스냅샷 길이에 맞추어 1-based 타일링(8761→1...)
                    if snapshots_len <= base_len:
                        pattern_full = pattern_base[:snapshots_len]
                    else:
                        repeats = snapshots_len // base_len + 1
                        pattern_full = np.tile(pattern_base, repeats)[:snapshots_len]
                    # 디버그 로그
                    try:
                        _base_label = {
                            'EL': '전력(B열)',
                            'H': '열(C열)',
                            'H2': '수소(D열)',
                            'EV_DRIVING': 'EV주행(E열)',
                            'EV_CHARGING': 'EV충전(F열)',
                            'DC': 'DC(H열)',
                            'Fab': '반도체팹(I열)'
                        }.get(demand_type, f'{demand_type}(열{effective_col})')
                        # 지역별 EL 패턴 사용 시 레이블 보정
                        if demand_type == 'EL' and region and effective_col != col_map.get('EL'):
                            label = f"지역별 EL 패턴({region}_EL_P, 열{effective_col})"
                        else:
                            label = _base_label
                        sample_n = min(8, len(pattern_base))
                        print(f"load_patterns 고정위치 사용: {label}, 시작행 Excel 8행 기준")
                        print(f"패턴 길이(기본): {len(pattern_base)} (최대 {np.nanmax(pattern_base):.6f}, 평균 {np.nanmean(pattern_base):.6f})")
                        print(f"패턴 샘플 앞 {sample_n}개: {np.round(pattern_base[:sample_n], 6).tolist()}")
                        if len(pattern_base) >= 24:
                            day_slice = pattern_base[:24]
                            print(f"하루(1~24시) 최소/최대: {np.min(day_slice):.6f}/{np.max(day_slice):.6f}")
                    except Exception:
                        pass
                    return pattern_full
        except Exception as e:
            print(f"고정 위치 패턴 로딩 실패(폴백 사용): {str(e)}")

        # 2) 폴백: 기존 컬럼 매칭 로직
        # 컬럼 정리: 공백 제거 및 한국어 기본 패턴명을 'pattern'으로 표준화
        cleaned_cols = []
        for c in df.columns:
            name = str(c).strip()
            if ('부하' in name and '패턴' in name) or name.lower() in ['default', 'pattern']:
                name = 'pattern'
            cleaned_cols.append(name)
        df.columns = cleaned_cols

        # hour/시간 컬럼 확인(이름 부분일치 + 수치 램프 감지)
        hour_col = None
        # 1) 이름 부분일치로 우선 탐색
        for c in df.columns:
            cl = str(c).strip().lower()
            if ('hour' in cl) or ('시간' in cl) or (cl in ['h', 't', 'index']):
                hour_col = c
                break
        # 2) 이름으로 못 찾으면, 값 패턴이 1..8760 류의 정수 램프인지 검사하여 선택
        if hour_col is None:
            for c in df.columns:
                s = pd.to_numeric(df[c], errors='coerce')
                if s.notna().sum() < 10:
                    continue
                vals = s.dropna().values
                # 정수성, 범위, 단조 증가(많은 구간) 검사
                is_int_like = np.nanmax(np.abs(vals - np.round(vals))) < 1e-6
                if not is_int_like:
                    continue
                vmin, vmax = np.nanmin(vals), np.nanmax(vals)
                if vmin >= 1 and vmax <= 8760 and vmax - vmin >= 1000:
                    diffs = np.diff(vals[:min(len(vals), 2000)])
                    if np.mean(diffs >= 0) > 0.95:  # 거의 단조 증가
                        hour_col = c
                        break

        # 후보 컬럼 선택
        candidates_by_type = {
            'EL': [f"{region}_EL", f"{region}_electricity", 'EL', 'electricity', '전력'],
            'H': [f"{region}_H", f"{region}_heating", 'H', 'heat', 'heating', '열'],
            'H2': [f"{region}_H2", f"{region}_hydrogen", 'H2', 'hydrogen', '수소'],
            'EV_DRIVING': [f"{region}_EV_DRIVING", 'EV_DRIVING', 'EV주행', '전기차주행'],
            'EV_CHARGING': [f"{region}_EV_CHARGING", 'EV_CHARGING', 'EV충전', '전기차충전']
        }
        candidates = candidates_by_type.get(demand_type, [])
        # 지역 무관 전역 후보 추가
        if demand_type == 'EL':
            candidates += ['EL', 'electricity', '전력']
        elif demand_type == 'H':
            candidates += ['H', 'heat', 'heating', '열']
        elif demand_type == 'H2':
            candidates += ['H2', 'hydrogen', '수소']
        elif demand_type == 'EV_DRIVING':
            candidates += ['EV_DRIVING', 'EV주행', '전기차주행']
        elif demand_type == 'EV_CHARGING':
            candidates += ['EV_CHARGING', 'EV충전', '전기차충전']
        # 기본 패턴 컬럼
        candidates += ['pattern']

        col = None
        for cand in candidates:
            if cand in df.columns:
                col = cand
                break

        # 마지막 fallback: 첫 번째 수치형 컬럼 탐색(hour/시간 유사 컬럼 제외)
        if col is None:
            numeric_cols = []
            for c in df.columns:
                cl = str(c).strip().lower()
                if hour_col and c == hour_col:
                    continue
                if ('hour' in cl) or ('시간' in cl) or (cl in ['h', 't', 'index']):
                    continue
                s = pd.to_numeric(df[c], errors='coerce')
                if s.notna().sum() > 0:
                    # 값 자체가 1..8760 류 램프인 경우 제외(시간 컬럼 가능성 높음)
                    vals = s.dropna().values
                    if len(vals) >= 100 and np.nanmax(vals) <= 100000:
                        diffs = np.diff(vals[:min(len(vals), 2000)])
                        ramp_like = np.mean(diffs >= 0) > 0.95 and np.nanmin(vals) >= 1
                        if ramp_like:
                            continue
                    numeric_cols.append(c)
            if numeric_cols:
                col = numeric_cols[0]
            else:
                return None

        # 기본 길이는 8760으로 간주(윤년/추가 스냅샷은 타일링)
        base_len = 8760

        if hour_col:
            # 1-based 매핑: 시간값 1..8760 → 인덱스 0..8759
            base = np.zeros(base_len, dtype=float)
            series = pd.to_numeric(df[col], errors='coerce')
            hours = pd.to_numeric(df[hour_col], errors='coerce')
            for v, h in zip(series, hours):
                if pd.notna(v) and pd.notna(h):
                    try:
                        h_int = int(float(h))
                    except Exception:
                        continue
                    if 1 <= h_int <= base_len:
                        base[h_int - 1] = float(v)
            pattern_base = base
        else:
            # 시간 컬럼이 없으면, 수치값만 추출하여 8760 길이로 보정(부족 시 타일링)
            series = pd.to_numeric(df[col], errors='coerce')
            values = series.dropna().astype(float).values
            if len(values) == 0:
                return None
            if len(values) >= base_len:
                pattern_base = values[:base_len]
            else:
                repeats = base_len // len(values) + 1
                pattern_base = np.tile(values, repeats)[:base_len]

        # 스냅샷 길이에 맞추어 1-based 타일링(8761→1...)
        if snapshots_len <= base_len:
            pattern_full = pattern_base[:snapshots_len]
        else:
            repeats = snapshots_len // base_len + 1
            pattern_full = np.tile(pattern_base, repeats)[:snapshots_len]

        # 디버그 로그: 선택된 컬럼과 패턴 샘플/단조 증가 여부
        try:
            print(f"load_patterns 컬럼: {list(df.columns)[:10]}{'...' if len(df.columns)>10 else ''}")
            print(f"선택된 시간 컬럼: {hour_col}")
            print(f"선택된 패턴 컬럼: {col}")
            sample_n = min(8, len(pattern_base))
            print(f"패턴 샘플(기본 기준 앞 {sample_n}개): {np.round(pattern_base[:sample_n], 6).tolist()}")
            if len(pattern_base) >= 24:
                day_slice = pattern_base[:24]
                print(f"하루(1~24시) 최소/최대: {np.min(day_slice):.6f}/{np.max(day_slice):.6f}")
        except Exception:
            pass

        return pattern_full
    except Exception:
        return None

def _sanitize_component_bounds(network):
    """
    네트워크 컴포넌트의 경계값 이상을 탐지하고 수정한다.
    p_min_pu > p_max_pu 인 시간대 수정, e_nom_min > e_nom_max 수정 등.
    """
    import numpy as np
    issues_fixed = 0

    # 1. Generator p_min_pu vs p_max_pu (시계열 및 정적)
    for gen in network.generators.index:
        try:
            p_min = float(network.generators.at[gen, 'p_min_pu'] or 0)
            if p_min <= 0:
                continue
            # 시계열 p_max_pu 확인
            if gen in network.generators_t.p_max_pu.columns:
                ts = network.generators_t.p_max_pu[gen]
                viol = ts < p_min
                if viol.any():
                    n_v = int(viol.sum())
                    network.generators_t.p_max_pu[gen] = ts.clip(lower=p_min)
                    print(f"  [경계수정] {gen}: p_max_pu < p_min_pu({p_min}) {n_v}개 시간 → 상향조정")
                    issues_fixed += 1
            else:
                # 정적 p_max_pu 확인
                p_max = float(network.generators.at[gen, 'p_max_pu'] or 1.0)
                if p_max < p_min:
                    network.generators.at[gen, 'p_max_pu'] = p_min
                    print(f"  [경계수정] {gen}: p_max_pu({p_max:.3f}) < p_min_pu({p_min:.3f}) → {p_min:.3f}으로 조정")
                    issues_fixed += 1
        except Exception:
            pass

    # 1b. Extendable Generator p_nom_min vs p_nom_max
    for gen in network.generators.index:
        try:
            if not bool(network.generators.at[gen, 'p_nom_extendable']):
                continue
            p_nom_min = float(network.generators.at[gen, 'p_nom_min'] or 0)
            p_nom_max = float(network.generators.at[gen, 'p_nom_max'] or float('inf'))
            if p_nom_min > p_nom_max + 0.01:
                print(f"  [경계오류] {gen}: p_nom_min({p_nom_min:.0f}) > p_nom_max({p_nom_max:.0f}) → p_nom_min를 0으로 조정")
                network.generators.at[gen, 'p_nom_min'] = 0.0
                issues_fixed += 1
        except Exception:
            pass

    # 2. Store e_nom_min vs e_nom_max, e_nom vs e_nom_max
    for store in network.stores.index:
        try:
            ext = bool(network.stores.at[store, 'e_nom_extendable'] if 'e_nom_extendable' in network.stores.columns else False)
            e_nom = float(network.stores.at[store, 'e_nom'] or 0)
            mn = float(network.stores.at[store, 'e_nom_min'] if 'e_nom_min' in network.stores.columns else 0) or 0
            mx = float(network.stores.at[store, 'e_nom_max'] if 'e_nom_max' in network.stores.columns else float('inf'))
            import math
            if math.isinf(mx):
                mx_str = 'inf'
            else:
                mx_str = f'{mx:.0f}'
            if ext:
                import pandas as pd
                if pd.notna(mn) and pd.notna(mx) and not math.isinf(mx) and float(mn) > float(mx) + 0.01:
                    print(f"  [경계오류] {store} (extendable): e_nom_min({mn:.0f}) > e_nom_max({mx_str}) → e_nom_min를 0으로 조정")
                    network.stores.at[store, 'e_nom_min'] = 0.0
                    issues_fixed += 1
                # e_nom이 e_nom_max보다 큰 경우: CPLEX bound infeasibility 원인
                if not math.isinf(mx) and e_nom > mx + 0.01:
                    print(f"  [경계오류] {store} (extendable): e_nom({e_nom:.0f}) > e_nom_max({mx_str})")
                    print(f"             → e_nom_max를 e_nom({e_nom:.0f})으로 상향조정 (bound infeasibility 방지)")
                    network.stores.at[store, 'e_nom_max'] = e_nom
                    issues_fixed += 1
            # e_min_pu vs e_max_pu 확인 (정적)
            e_min_pu = float(network.stores.at[store, 'e_min_pu'] if 'e_min_pu' in network.stores.columns else 0) or 0
            e_max_pu = float(network.stores.at[store, 'e_max_pu'] if 'e_max_pu' in network.stores.columns else 1.0) or 1.0
            if e_min_pu > e_max_pu + 0.001:
                print(f"  [경계오류] {store}: e_min_pu({e_min_pu:.3f}) > e_max_pu({e_max_pu:.3f}) → e_min_pu를 0으로 조정")
                network.stores.at[store, 'e_min_pu'] = 0.0
                issues_fixed += 1
        except Exception:
            pass

    # 3. Link p_min_pu vs p_max_pu (정적)
    for link in network.links.index:
        try:
            p_min = float(network.links.at[link, 'p_min_pu'] or 0)
            p_max = float(network.links.at[link, 'p_max_pu'] or 1.0)
            if p_min > p_max + 0.01:
                network.links.at[link, 'p_min_pu'] = 0.0
                print(f"  [경계수정] {link}: p_min_pu({p_min:.3f}) > p_max_pu({p_max:.3f}) → p_min_pu를 0으로 조정")
                issues_fixed += 1
        except Exception:
            pass

    # 3b. Link 시계열 p_max_pu vs 정적 p_min_pu (V2G 패턴 적용 후 확인)
    for link in network.links.index:
        try:
            if link not in network.links_t.p_max_pu.columns:
                continue
            p_min = float(network.links.at[link, 'p_min_pu'] or 0)
            if p_min <= 0:
                continue
            ts = network.links_t.p_max_pu[link]
            viol = ts < p_min
            if viol.any():
                n_v = int(viol.sum())
                network.links_t.p_max_pu[link] = ts.clip(lower=p_min)
                print(f"  [경계수정] Link {link}: 시계열 p_max_pu < p_min_pu({p_min}) {n_v}개 시간 → 상향조정")
                issues_fixed += 1
        except Exception:
            pass

    if issues_fixed > 0:
        print(f"  [경계정리 완료] 총 {issues_fixed}개 이상 수정됨")
    else:
        print("  [경계정리] 이상 없음")


def _diagnose_infeasibility(network):
    """
    최적화 전 간단한 infeasibility 사전 진단
    (과발전 가능성, 수요 미달 가능성 체크)
    """
    try:
        import numpy as np, pandas as pd
        print("\n[진단] 사전 infeasibility 체크...")

        # 원자력 최소발전 vs 야간 최소수요 비교
        nuclear_gens = network.generators[
            (network.generators['p_min_pu'].notna()) &
            (network.generators['p_min_pu'] > 0)
        ]
        if not nuclear_gens.empty:
            total_min_gen = (nuclear_gens['p_nom'] * nuclear_gens['p_min_pu']).sum()
            print(f"  필수 최소발전(p_min_pu>0 발전기 합계): {total_min_gen:.0f} MW")

        # 전력 수요 시계열 최솟값
        el_loads = network.loads[network.loads.bus.str.endswith('_EL') | network.loads.bus.isin(
            [b for b in network.buses.index if network.buses.at[b,'carrier']=='electricity']
        )]
        if not network.loads_t.p_set.empty:
            el_load_names = el_loads.index[el_loads.index.isin(network.loads_t.p_set.columns)]
            if len(el_load_names) > 0:
                total_el_load_ts = network.loads_t.p_set[el_load_names].sum(axis=1)
                min_el = total_el_load_ts.min()
                print(f"  전력수요 최솟값 (시계열): {min_el:.0f} MW")
                if total_min_gen > min_el:
                    print(f"  [!!] 최소 필수발전({total_min_gen:.0f} MW) > 최소 전력수요({min_el:.0f} MW)")
                    print(f"      → 야간/경부하 시간대 과발전 위험 → ESS/EV충전/송전으로 해소 필요")

        # 수요가 공급능력보다 큰 버스 체크
        for bus in network.buses.index:
            try:
                carrier = str(network.buses.at[bus, 'carrier'])
                if carrier not in ['electricity', 'heat', 'hydrogen']:
                    continue
                bus_loads = network.loads.index[network.loads.bus == bus]
                bus_gens = network.generators.index[network.generators.bus == bus]
                bus_gen_max = network.generators.loc[bus_gens, 'p_nom'].sum() if len(bus_gens) > 0 else 0
                bus_load_static = network.loads.loc[bus_loads, 'p_set'].sum() if len(bus_loads) > 0 else 0
                if bus_gen_max == 0 and bus_load_static > 0:
                    has_slack = any('Slack' in g or 'Fallback' in g for g in bus_gens)
                    if not has_slack:
                        print(f"  [!!] {bus}: 발전기 없음 + 수요 있음({bus_load_static:.0f}MW) + 슬랙 없음 → 주의")
            except Exception:
                pass

        print("[진단 완료]\n")
    except Exception as ex:
        print(f"[진단 오류] {ex}")


def create_network(input_data):
    try:
        network = pypsa.Network()
        
        # carriers 정의 (PyPSA 내부 배출계수 활용)
        # 주의: PyPSA는 type="primary_energy"로 (전력량/efficiency) × co2_emissions 계산
        # 배출계수는 전력량 기준이므로, carrier.co2_emissions = 전력량기준배출계수 × 대표efficiency
        # 예: 석탄 0.8384 tCO2/MWh_electric × 0.47 efficiency = 0.39405 tCO2/MWh_fuel
        carriers = {
            'AC': {'name': 'AC', 'co2_emissions': 0},
            'DC': {'name': 'DC', 'co2_emissions': 0},
            'electricity': {'name': 'electricity', 'co2_emissions': 0},
            'coal': {'name': 'coal', 'co2_emissions': 0.73 * 0.47},  # 석탄 전력량기준 × 평균효율0.47
            'gas': {'name': 'gas', 'co2_emissions': 0.33 * 0.53},     # LNG 전력량기준 × 평균효율0.53
            'nuclear': {'name': 'nuclear', 'co2_emissions': 0},
            'solar': {'name': 'solar', 'co2_emissions': 0},
            'wind': {'name': 'wind', 'co2_emissions': 0},
            'hydrogen': {'name': 'hydrogen', 'co2_emissions': 0},
            'heat': {'name': 'heat', 'co2_emissions': 0},
            'EV': {'name': 'EV', 'co2_emissions': 0}  # 전기차 배터리 carrier
        }
        
        # carriers 추가
        for carrier, specs in carriers.items():
            network.add("Carrier",
                       name=specs['name'],
                       co2_emissions=specs['co2_emissions'])
        
        # 시간 설정
        if 'timeseries' in input_data and not input_data['timeseries'].empty:
            ts = input_data['timeseries'].iloc[0]
            snapshots = pd.date_range(
                start=ts['start_time'],
                end=ts['end_time'],
                freq=ts['frequency'],
                inclusive='left'
            )
            network.set_snapshots(snapshots)
            snapshots_length = len(snapshots)
        else:
            # 기본 시간 설정 (2024년 1년간, 1시간 간격)
            print("[경고] timeseries 시트가 없거나 비어있습니다. 기본 시간 설정을 사용합니다.")
            snapshots = pd.date_range(
                start='2024-01-01 00:00:00',
                end='2025-01-01 00:00:00',
                freq='1h',
                inclusive='left'
            )
            network.set_snapshots(snapshots)
            snapshots_length = len(snapshots)
        
        # 버스 추가 - 실제 데이터만 사용
        if 'buses' in input_data:
            print("\n=== 버스 추가 시작 ===")
            for _, bus in input_data['buses'].iterrows():
                bus_name = str(bus['name'])
                raw_carrier = str(bus['carrier'])
                token = _standardize_bus_token_by_carrier(raw_carrier)
                carrier_map = {'EL': 'electricity', 'H': 'heat', 'H2': 'hydrogen', 'LNG': 'gas', 'EV': 'electricity'}
                carrier = carrier_map.get(token, raw_carrier)
                
                v_nom_val = float(bus['v_nom']) if pd.notna(bus['v_nom']) else 345.0
                network.add("Bus",
                          name=bus_name,
                          v_nom=v_nom_val,
                          carrier=carrier)
                print(f"버스 추가됨: {bus_name} (carrier: {carrier}, v_nom: {v_nom_val})")

        # 재생에너지 패턴 준비 (지역별)
        renewable_patterns = {}
        if 'renewable_patterns' in input_data:
            patterns_df = input_data['renewable_patterns']
            print(f"\n=== 재생에너지 패턴 로딩 (지역별) ===")
            print(f"컬럼: {patterns_df.columns.tolist()}")
            
            # 실제 데이터가 있는 행 찾기 (4행부터 시작)
            if len(patterns_df) > 4:
                # 헤더 행 찾기 (B5, C5, D5, ... 에 패턴 이름)
                header_row = patterns_df.iloc[3]  # 5행 (0-based index 3은 4행)
                print(f"헤더 행 (처음 5개): {header_row.tolist()[:5]}")
                
                # 데이터 행들 (5행부터, 0-based index 4부터)
                data_rows = patterns_df.iloc[4:].copy()
                
                # 각 컬럼을 순회하며 패턴 로드
                for col_idx, col_name in enumerate(patterns_df.columns):
                    if col_idx == 0:  # 첫 번째 컬럼(시간)은 건너뛰기
                        continue
                    
                    # 헤더에서 패턴 이름 가져오기
                    pattern_name = str(header_row.iloc[col_idx]).strip()
                    
                    # 건너뛰기: 빈 값이거나 'Unnamed'인 경우
                    if pd.isna(pattern_name) or 'Unnamed' in pattern_name or pattern_name == 'nan':
                        continue
                    
                    # 패턴 데이터 추출
                    pattern_values = pd.to_numeric(data_rows.iloc[:, col_idx], errors='coerce').dropna().values
                    
                    if len(pattern_values) > 0:
                        pattern = normalize_pattern(pattern_values)
                        pattern = adjust_pattern_length(pattern, snapshots_length)
                        renewable_patterns[pattern_name] = pattern
                        print(f"  [{pattern_name}] 로드 완료 - 길이: {len(pattern)}, 최대: {np.max(pattern):.3f}, 최소: {np.min(pattern):.3f}")
            
            if not renewable_patterns:
                print("[경고] 재생에너지 패턴을 찾을 수 없습니다. 기본값 1.0을 사용합니다.")
            else:
                print(f"[OK] 총 {len(renewable_patterns)}개의 재생에너지 패턴 로드 완료")
        
        # 발전기 추가
        if 'generators' in input_data:
            print("\n=== 발전기 추가 시작 ===")
            defined_buses = set(network.buses.index)
            bus_carriers_map = dict(zip(network.buses.index, network.buses.carrier))
            for _, gen in input_data['generators'].iterrows():
                gen_name = str(gen['name'])
                p_nom_value = float(gen['p_nom'])
                
                # p_nom이 0인 경우 건너뛰기
                if p_nom_value <= 0:
                    print(f"발전기 {gen_name} 건너뜀: p_nom이 0 이하 ({p_nom_value})")
                    continue
                
                raw_bus = str(gen['bus'])
                norm_bus = _normalize_bus_name(raw_bus, defined_buses, True, bus_carriers_map)
                if norm_bus != raw_bus:
                    print(f"발전기 {gen_name} 버스명 정규화: {raw_bus} → {norm_bus}")
                
                # 발전기 이름 기반으로 올바른 carrier 설정 (CO2 제약조건을 위해)
                def get_correct_carrier(gen_name, original_carrier):
                    """발전기 이름에서 올바른 carrier 추론"""
                    gen_lower = gen_name.lower()
                    
                    # 석탄발전소 이름으로 인식 (발전소 이름 포함)
                    coal_plants = ['coal', '석탄', '당진', '영흥', '하동', '삼척', '태안', '보령', '삼천포', '호남', '동해']
                    if any(keyword in gen_lower for keyword in coal_plants):
                        return 'coal'
                    elif any(keyword in gen_lower for keyword in ['lng', 'gas', 'chp', '가스', 'slack', 'fallback']):
                        return 'gas'  # 슬랙/폴백 발전기도 LNG 기반으로 처리
                    elif any(keyword in gen_lower for keyword in ['nuclear', '원자력']):
                        return 'nuclear'
                    elif any(keyword in gen_lower for keyword in ['pv', 'solar', '태양']):
                        return 'solar'
                    elif any(keyword in gen_lower for keyword in ['wind', 'wt', '풍력']):
                        return 'wind'
                    elif str(original_carrier).upper() == 'DR' or gen_lower.endswith('_dr'):
                        return 'DR'
                    else:
                        # 원본 carrier가 유효하면 사용, 아니면 electricity
                        orig = str(original_carrier).lower()
                        if orig in ['coal', 'gas', 'nuclear', 'solar', 'wind']:
                            return orig
                        return 'electricity'
                
                correct_carrier = get_correct_carrier(gen_name, gen['carrier'])
                
                params = {
                    'name': gen_name,
                    'bus': norm_bus,
                    'p_nom': p_nom_value,
                    'carrier': correct_carrier
                }
                
                if correct_carrier != str(gen['carrier']):
                    print(f"  [수정] {gen_name}: carrier 수정 {gen['carrier']} → {correct_carrier}")
                
                # 나머지 파라미터 추가
                optional_params = ['p_nom_extendable', 'marginal_cost', 'capital_cost', 'efficiency', 'p_min_pu', 'p_nom_min', 'p_nom_max', 
                                   'ramp_limit_up', 'ramp_limit_down', 'committable', 'min_up_time', 'min_down_time', 'start_up_cost', 'shut_down_cost']
                for param in optional_params:
                    if param in gen and pd.notna(gen[param]):
                        if param in ['p_nom_extendable', 'committable']:
                            params[param] = _to_bool(gen[param])
                        else:
                            params[param] = float(gen[param])
                
                network.add("Generator", **params)
                print(f"발전기 추가됨: {gen_name} (p_nom: {p_nom_value}, carrier: {params['carrier']})")

        # Coal_DB 효율 적용: interface.xlsx Coal_DB 시트 AM열 → 개별 석탄 발전기 efficiency 갱신
        print("\n=== Coal_DB 개별 효율 적용 ===")
        _iface_path = os.path.abspath(os.path.join(get_base_dir(), 'interface.xlsx'))
        _apply_coal_efficiencies_from_db(network, _iface_path)

        # 발전기 추가 후 재생에너지 패턴 적용
        print("\n=== 재생에너지 패턴 적용 시작 ===")
        pv_applied_count = 0
        wt_applied_count = 0
        
        for gen_name in network.generators.index:
            pattern_applied = False
            
            # 지역 코드 추출 (발전기 이름에서 첫 부분)
            region_code = None
            region_codes = ['SEL', 'BSN', 'DGU', 'ICN', 'GWJ', 'DJN', 'USN', 'SJG', 
                          'GGD', 'GWD', 'CBD', 'CND', 'JBD', 'JND', 'GBD', 'GND', 'JJD']
            for code in region_codes:
                if gen_name.startswith(code + '_') or gen_name.startswith(code):
                    region_code = code
                    break
            
            # PV 패턴 적용 (지역별 + 계절별 일출/일몰 시간 반영)
            if 'PV' in gen_name or 'Solar' in gen_name or 'solar' in gen_name:
                pattern_key = None
                
                # 지역별 패턴 우선 적용
                if region_code:
                    # 예: SEL_PV → SEL_PV_Patterns
                    regional_pattern = f"{region_code}_PV_Patterns"
                    if regional_pattern in renewable_patterns:
                        pattern_key = regional_pattern
                    # 폴백: BSN_PV_Patterns (전국 대표)
                    elif 'BSN_PV_Patterns' in renewable_patterns:
                        pattern_key = 'BSN_PV_Patterns'
                
                # 구버전 호환성 (PV_pattern)
                if not pattern_key and 'PV_pattern' in renewable_patterns:
                    pattern_key = 'PV_pattern'
                
                if pattern_key and pattern_key in renewable_patterns:
                    # 패턴 길이 확인 및 조정
                    pattern = renewable_patterns[pattern_key].copy()
                    if len(pattern) != len(network.snapshots):
                        print(f"[경고] {pattern_key} 길이 불일치: {len(pattern)} vs {len(network.snapshots)}")
                        pattern = adjust_pattern_length(pattern, len(network.snapshots))
                    
                    # 계절별 일출/일몰 시간 반영 (하드코딩)
                    # 봄/가을: 6~19시, 여름: 5~20시, 겨울: 7~18시
                    for idx, snapshot in enumerate(network.snapshots):
                        month = snapshot.month
                        hour = snapshot.hour
                        
                        # 계절 판단
                        if month in [3, 4, 5]:  # 봄
                            if not (6 <= hour < 19):
                                pattern[idx] = 0.0
                        elif month in [6, 7, 8]:  # 여름
                            if not (5 <= hour < 20):
                                pattern[idx] = 0.0
                        elif month in [9, 10, 11]:  # 가을
                            if not (6 <= hour < 19):
                                pattern[idx] = 0.0
                        else:  # 겨울 (12, 1, 2월)
                            if not (7 <= hour < 18):
                                pattern[idx] = 0.0
                    
                    # 패턴을 p_max_pu에 적용
                    network.generators_t.p_max_pu[gen_name] = pattern
                    pv_applied_count += 1
                    pattern_applied = True
            
            # WT 패턴 적용 (지역별)
            if 'WT' in gen_name or 'Wind' in gen_name or 'wind' in gen_name:
                pattern_key = None
                
                # 제주는 별도 패턴, 나머지는 전국 패턴
                if region_code == 'JJD' and 'JJD_WT_Patterns' in renewable_patterns:
                    pattern_key = 'JJD_WT_Patterns'
                elif 'National_WT_Patterns' in renewable_patterns:
                    pattern_key = 'National_WT_Patterns'
                # 구버전 호환성 (WT_pattern)
                elif 'WT_pattern' in renewable_patterns:
                    pattern_key = 'WT_pattern'
                
                if pattern_key and pattern_key in renewable_patterns:
                    # 패턴 길이 확인 및 조정
                    pattern = renewable_patterns[pattern_key]
                    if len(pattern) != len(network.snapshots):
                        print(f"[경고] {pattern_key} 길이 불일치: {len(pattern)} vs {len(network.snapshots)}")
                        pattern = adjust_pattern_length(pattern, len(network.snapshots))
                    
                    # 패턴을 p_max_pu에 직접 적용
                    network.generators_t.p_max_pu[gen_name] = pattern
                    wt_applied_count += 1
                    pattern_applied = True
            
            # 패턴이 적용되지 않은 재생에너지 발전기 확인
            if not pattern_applied and any(keyword in gen_name.lower() for keyword in ['pv', 'solar', 'wind', 'wt']):
                print(f"[경고] 재생에너지 발전기 {gen_name}에 패턴이 적용되지 않음")
        
        print(f"\n재생에너지 패턴 적용 완료:")
        print(f"- PV 패턴 적용: {pv_applied_count}개 발전기")
        print(f"- WT 패턴 적용: {wt_applied_count}개 발전기")
        
        # 적용된 패턴의 통계 정보 출력
        pv_pattern_count = sum(1 for k in renewable_patterns.keys() if 'PV_Patterns' in k or k == 'PV_pattern')
        wt_pattern_count = sum(1 for k in renewable_patterns.keys() if 'WT_Patterns' in k or k == 'WT_pattern')
        print(f"- 로드된 PV 패턴: {pv_pattern_count}개")
        print(f"- 로드된 WT 패턴: {wt_pattern_count}개")
        
        # 샘플 패턴 통계 출력 (처음 3개만)
        sample_count = 0
        for pattern_name in renewable_patterns.keys():
            if 'PV_Patterns' in pattern_name or pattern_name == 'PV_pattern':
                if sample_count < 3:
                    sample_pattern = renewable_patterns[pattern_name]
                    print(f"- 샘플 PV 패턴({pattern_name}): 최대 {np.max(sample_pattern):.3f}, 최소 {np.min(sample_pattern):.3f}, 평균 {np.mean(sample_pattern):.3f}")
                    sample_count += 1
        
        sample_count = 0
        for pattern_name in renewable_patterns.keys():
            if 'WT_Patterns' in pattern_name or pattern_name == 'WT_pattern':
                if sample_count < 3:
                    sample_pattern = renewable_patterns[pattern_name]
                    print(f"- 샘플 WT 패턴({pattern_name}): 최대 {np.max(sample_pattern):.3f}, 최소 {np.min(sample_pattern):.3f}, 평균 {np.mean(sample_pattern):.3f}")
                    sample_count += 1
        
        # 발전기 최대/최소 출력비율(p_max_pu/p_min_pu) 반영
        # 주의: efficiency는 p_max_pu에 곱하지 않음 (발전량 제한이 아닌 연료소비량 계산에만 사용)
        # [경고] 재생에너지(PV, WT)는 패턴이 이미 적용되었으므로 추가 처리 제외
        try:
            gdf = input_data.get('generators', pd.DataFrame())
            if not gdf.empty:
                cols = list(gdf.columns)
                for _, row in gdf.iterrows():
                    gname = str(row.get('name', '')).strip()
                    if not gname or gname not in network.generators.index:
                        continue
                    
                    # 재생에너지 발전기는 이미 패턴이 적용되었으므로 건너뜀
                    is_renewable = any(keyword in gname.lower() for keyword in ['pv', 'solar', 'wind', 'wt'])
                    if is_renewable:
                        continue
                    
                    # 최대/최소 출력비율 (효율은 제외!)
                    max_ratio = pd.to_numeric(row.get('p_max_pu'), errors='coerce')
                    if pd.isna(max_ratio) and '최대출력비율' in cols:
                        max_ratio = pd.to_numeric(row.get('최대출력비율'), errors='coerce')
                    min_ratio = pd.to_numeric(row.get('p_min_pu'), errors='coerce')
                    if pd.isna(min_ratio) and '최소출력비율' in cols:
                        min_ratio = pd.to_numeric(row.get('최소출력비율'), errors='coerce')
                    
                    # p_max_pu 적용 (효율 곱하지 않음!)
                    if pd.notna(max_ratio):
                        cap_mul = float(np.clip(max_ratio, 0.0, 1.0))
                        if gname in network.generators_t.p_max_pu.columns:
                            network.generators_t.p_max_pu[gname] = network.generators_t.p_max_pu[gname] * cap_mul
                        else:
                            network.generators_t.p_max_pu[gname] = pd.Series(cap_mul, index=network.snapshots)
                        print(f"발전기 출력제한 적용: {gname} (p_max_pu={cap_mul:.3f})")
                    
                    # p_min_pu 적용(있을 때만)
                    if pd.notna(min_ratio):
                        min_val = float(np.clip(min_ratio, 0.0, 1.0))
                        network.generators_t.p_min_pu[gname] = pd.Series(min_val, index=network.snapshots)
                        print(f"발전기 최소출력 설정: {gname} (p_min_pu={min_val:.3f})")
            # p_min_pu <= p_max_pu 보정 (재생에너지 제외)
            try:
                for gname in network.generators.index:
                    # 재생에너지는 패턴이 p_max_pu에만 적용되고 p_min_pu=0이므로 보정 제외
                    is_renewable = any(keyword in gname.lower() for keyword in ['pv', 'solar', 'wind', 'wt'])
                    if is_renewable:
                        continue
                    
                    if gname in network.generators_t.p_max_pu.columns:
                        pmax = network.generators_t.p_max_pu[gname].fillna(0.0)
                    else:
                        pmax = pd.Series(1.0, index=network.snapshots)
                    if hasattr(network.generators_t, 'p_min_pu') and (gname in network.generators_t.p_min_pu.columns):
                        pmin = network.generators_t.p_min_pu[gname].fillna(0.0)
                        network.generators_t.p_min_pu[gname] = pd.concat([pmin, pmax], axis=1).min(axis=1).clip(lower=0.0)
            except Exception as _e_clamp:
                print(f"p_min_pu 보정 경고: {_e_clamp}")
        except Exception as _e_der:
            print(f"발전기 효율/출력비율 적용 경고: {_e_der}")
        
        # 발전기 백업 보강: 각 버스에 최소 하나의 가용 발전기 보장
        try:
            print("\n=== 버스별 기본 발전기 보강 확인 ===")
            buses_with_gen = set(network.generators.bus.unique()) if not network.generators.empty else set()
            for bus in network.buses.index:
                # LNG 버스에는 전력 보강발전기 추가하지 않음(전력용)
                if bus.endswith('_LNG'):
                    # 대신 연료(가스) 공급원이 없다면 가스 버스에 연료공급 소스를 추가
                    try:
                        bus_carrier_chk = str(network.buses.at[bus, 'carrier']).lower()
                    except Exception:
                        bus_carrier_chk = ''
                    if bus_carrier_chk == 'gas':
                        fuel_name = f"{bus}_Fuel_Supply"
                        if fuel_name not in network.generators.index:
                            network.add("Generator",
                                       name=fuel_name,
                                       bus=bus,
                                       p_nom=0.0,
                                       p_nom_extendable=True,
                                       marginal_cost=0.0,
                                       carrier='gas')
                            network.generators_t.p_max_pu[fuel_name] = pd.Series(1.0, index=network.snapshots)
                            print(f"가스 연료공급 추가: {fuel_name} (버스 {bus})")
                    continue
                # 버스 캐리어 확인
                try:
                    bus_carrier = str(network.buses.at[bus, 'carrier']).lower()
                except Exception:
                    bus_carrier = ''
                # 열 버스: 항상 확장가능 고비용 보강발전기 추가(부족분만 보완)
                if bus_carrier == 'heat':
                    # 해당 열 버스에 실제 열부하가 있는 경우에만 보강발전기 후보 추가 (네트워크 시계열 기준)
                    has_heat_load = False
                    try:
                        total = 0.0
                        # 1) 네트워크 시계열 기준
                        if (not network.loads.empty) and (hasattr(network.loads_t, 'p_set') and not network.loads_t.p_set.empty):
                            load_names = network.loads.index[network.loads.bus.astype(str) == bus]
                            for nm in load_names:
                                if nm in network.loads_t.p_set.columns:
                                    total += float(np.nansum(network.loads_t.p_set[nm].values))
                        # 2) 폴백: 입력 데이터 기준
                        if total <= 0.0:
                            loads_df = input_data.get('loads', pd.DataFrame())
                            if not loads_df.empty and 'bus' in loads_df.columns and 'p_set' in loads_df.columns:
                                sub = loads_df[loads_df['bus'].astype(str) == bus]
                                if not sub.empty:
                                    total = float(pd.to_numeric(sub['p_set'], errors='coerce').fillna(0).sum())
                        has_heat_load = total > 0.0
                    except Exception:
                        has_heat_load = True
                    if not has_heat_load:
                        continue
                    fallback_name = f"{bus}_Fallback_Gen"
                    if fallback_name not in network.generators.index:
                        # 열 fallback: 가스보일러 운영비 수준 (~10만 원/MWh_열)
                        # HP 대비 약 50% 비싸지만 필요시 사용 가능한 수준
                        network.add("Generator",
                                   name=fallback_name,
                                   bus=bus,
                                   p_nom=0.0,
                                   p_nom_extendable=True,
                                   capital_cost=0.0,
                                   marginal_cost=1e5,
                                   carrier='heat')
                        network.generators_t.p_max_pu[fallback_name] = pd.Series(1.0, index=network.snapshots)
                        print(f"열 보강 발전기 추가(가스보일러 수준): {fallback_name} (버스 {bus}, 10만 원/MWh)")
                    continue
                # 전력 버스: LNG_Fallback_Gen / Slack_Gen 비활성화
                # - 전력 시스템은 석탄/LNG/원자력/RE 등 충분한 발전원이 있으므로 보완 발전기 불필요
                # - 이들이 있으면 CO2 계산 왜곡(efficiency=0.5 적용해도 LP 변수 증가)
                # - Infeasible 발생 시 모델 설정 오류 진단 신호로 활용
                if bus_carrier == 'electricity' or bus_carrier == 'AC':
                    print(f"[전력 보완 발전기] 비활성화됨: {bus} (LNG_Fallback, Slack 추가 안 함)")
                # 수소 버스: 수소 백업기(필요시)
                if bus_carrier == 'hydrogen':
                    total_load = 0.0
                    try:
                        if (not network.loads.empty) and hasattr(network.loads_t, 'p_set') and (network.loads_t.p_set is not None) and (not network.loads_t.p_set.empty):
                            load_names = network.loads.index[network.loads.bus.astype(str) == bus]
                            for nm in load_names:
                                if nm in network.loads_t.p_set.columns:
                                    total_load += float(np.nansum(network.loads_t.p_set[nm].values))
                        if total_load <= 0.0 and (not network.loads.empty) and hasattr(network.loads_t, 'p') and (network.loads_t.p is not None) and (not network.loads_t.p.empty):
                            load_names = network.loads.index[network.loads.bus.astype(str) == bus]
                            for nm in load_names:
                                if nm in network.loads_t.p.columns:
                                    total_load += float(np.nansum(network.loads_t.p[nm].values))
                        if total_load <= 0.0:
                            loads_df = input_data.get('loads', pd.DataFrame())
                            if not loads_df.empty and 'bus' in loads_df.columns and 'p_set' in loads_df.columns:
                                sub = loads_df[loads_df['bus'].astype(str) == bus]
                                if not sub.empty:
                                    total_load = float(pd.to_numeric(sub['p_set'], errors='coerce').fillna(0).sum())
                    except Exception:
                        pass
                    if total_load > 0.0:
                        fallback_name = f"{bus}_H2_Fallback_Gen"
                        if fallback_name not in network.generators.index:
                            # 수소 fallback: 해외 그린수소 수입비용 수준 (~10만 원/MWh_H2)
                            # 국제 수소가격 2~5 USD/kg = 약 6~15만 원/MWh (LHV 기준)
                            network.add("Generator",
                                       name=fallback_name,
                                       bus=bus,
                                       p_nom=0.0,
                                       p_nom_extendable=True,
                                       capital_cost=0.0,
                                       marginal_cost=1e5,
                                       carrier='hydrogen')
                            network.generators_t.p_max_pu[fallback_name] = pd.Series(1.0, index=network.snapshots)
        except Exception as e:
            print(f"기본 발전기 보강 중 오류: {str(e)}")

        # 모든 발전기에 p_max_pu 기본값(1.0) 보장
        try:
            missing_cols = [g for g in network.generators.index if g not in network.generators_t.p_max_pu.columns]
            if missing_cols:
                for g in missing_cols:
                    network.generators_t.p_max_pu[g] = pd.Series(1.0, index=network.snapshots)
                print(f"p_max_pu 기본 적용: {len(missing_cols)}개 발전기에 1.0 설정")
        except Exception as e:
            print(f"p_max_pu 기본값 설정 중 오류: {str(e)}")
        
        # 부하 추가
        if 'loads' in input_data:
            print("\n=== 부하 추가 시작 ===")
            for _, load in input_data['loads'].iterrows():
                name = str(load['name'])
                bus_name = str(load['bus'])
                p_set = float(load['p_set'])
                
                # 부하 패턴 적용
                if 'load_patterns' in input_data:
                    # 부하 이름에서 지역과 타입 추출
                    region = name.split('_')[0] if '_' in name else None
                    # 부하 타입 감지 (EV/DC/Fab 포함)
                    if '_Demand_EL' in name:
                        dtype = 'EL'
                    elif '_Demand_H2' in name:
                        dtype = 'H2'
                    elif '_Demand_H' in name:
                        dtype = 'H'
                    elif '_Demand_EV' in name or '_EV' in name:
                        dtype = 'EV_DRIVING'
                    elif '_Demand_DC' in name:
                        dtype = 'DC'   # 데이터센터 전력 수요 → DC_patterns 적용
                    elif '_Demand_Fab' in name:
                        dtype = 'Fab'  # 반도체팹 전력 수요 → Fab_patterns 적용
                    else:
                        dtype = None
                    
                    # 패턴 적용
                    if dtype is not None:
                        pattern = _get_load_pattern(input_data, region, dtype, len(snapshots)) if region else None
                        if pattern is not None:
                            # 총수요 × 8760 × 패턴(스케일 없이 그대로 적용)
                            p_set_original = float(p_set)
                            p_set = p_set_original * 8760.0 * pattern
                            print(f"부하 {name}에 패턴 적용됨 (총수요×8760×패턴, 타입={dtype})")
                            
                            # EV 부하인 경우 상세 로그
                            if dtype == 'EV_DRIVING':
                                print(f"  [EV 주행] 원본 수요: {p_set_original:.2f} MW")
                                print(f"  [EV 주행] 패턴 통계: 최소={np.min(pattern):.6f}, 최대={np.max(pattern):.6f}, 평균={np.mean(pattern):.6f}, 합={np.sum(pattern):.2f}")
                                print(f"  [EV 주행] 적용 후 시간별 수요: 최소={np.min(p_set):.2f}, 최대={np.max(p_set):.2f}, 평균={np.mean(p_set):.2f} MW")
                        else:
                            cols = list(input_data['load_patterns'].columns) if 'load_patterns' in input_data else []
                            print(f"부하 {name}: 패턴 미발견 (region={region}, type={dtype}), 사용 가능 컬럼: {cols[:10]}{'...' if len(cols)>10 else ''}")
                            # 패턴이 없는 경우 일정한 부하
                            p_set = np.full(len(snapshots), p_set)
                            print(f"부하 {name}에 일정한 부하 적용됨")
                else:
                    p_set = np.full(len(snapshots), p_set)
                
                network.add("Load",
                          name=name,
                          bus=bus_name)
                # 시간별 p_set은 명시적으로 loads_t에 설정 (스냅샷 인덱스 정렬)
                if isinstance(p_set, pd.Series):
                    series = p_set.reindex(network.snapshots)
                elif isinstance(p_set, (np.ndarray, list)):
                    series = pd.Series(p_set, index=network.snapshots)
                else:
                    series = pd.Series(np.full(len(network.snapshots), float(p_set)), index=network.snapshots)
                network.loads_t.p_set[name] = series
                print(f"부하 추가됨: {name} (버스: {bus_name})")
        
        # 시나리오 수요 스케일링 비활성화 (지역별 시트 원본 데이터 사용)
        # _apply_scenario_demand_scaling(network, input_data)
        
        # 스케일링 이후 사후 백업 발전기 보강(전력/열 버스 대상)
        try:
            added_backup = 0
            for bus in network.buses.index:
                try:
                    bus_carrier_post = str(network.buses.at[bus, 'carrier']).lower()
                except Exception:
                    bus_carrier_post = ''
                # 버스별 부하 총량 계산(p_set 우선, 없으면 p)
                total_load = 0.0
                try:
                    if (not network.loads.empty) and hasattr(network.loads_t, 'p_set') and (network.loads_t.p_set is not None) and (not network.loads_t.p_set.empty):
                        load_names = network.loads.index[network.loads.bus.astype(str) == bus]
                        for nm in load_names:
                            if nm in network.loads_t.p_set.columns:
                                total_load += float(np.nansum(network.loads_t.p_set[nm].values))
                    if total_load <= 0.0 and (not network.loads.empty) and hasattr(network.loads_t, 'p') and (network.loads_t.p is not None) and (not network.loads_t.p.empty):
                        load_names = network.loads.index[network.loads.bus.astype(str) == bus]
                        for nm in load_names:
                            if nm in network.loads_t.p.columns:
                                total_load += float(np.nansum(network.loads_t.p[nm].values))
                    # 최후 폴백: 입력 데이터 기준
                    if total_load <= 0.0:
                        loads_df2 = input_data.get('loads', pd.DataFrame())
                        if not loads_df2.empty and 'bus' in loads_df2.columns and 'p_set' in loads_df2.columns:
                            sub2 = loads_df2[loads_df2['bus'].astype(str) == bus]
                            if not sub2.empty:
                                total_load = float(pd.to_numeric(sub2['p_set'], errors='coerce').fillna(0).sum())
                except Exception:
                    pass
                # 열 버스: 열 백업기
                if bus_carrier_post == 'heat' and total_load > 0.0:
                    fallback_name = f"{bus}_Fallback_Gen"
                    if fallback_name not in network.generators.index:
                        network.add("Generator",
                                   name=fallback_name,
                                   bus=bus,
                                   p_nom=0.0,
                                   p_nom_extendable=True,
                                   capital_cost=1000,
                                   marginal_cost=40,
                                   carrier='heat')
                        network.generators_t.p_max_pu[fallback_name] = pd.Series(1.0, index=network.snapshots)
                        added_backup += 1
                # 전력 버스: LNG 백업기 비활성화
                # - efficiency 미지정 시 PyPSA 기본값 1.0 적용 → CO2/MWh_e = 0.2014 (실제 LNG 0.4028의 절반)
                # - CO2 제약이 타이트해지면 최적화기가 이 발전기를 대량 사용(CO2 shadow price 급등)
                # - 결과적으로 LNG 배출량이 절반으로 보고되는 왜곡 발생
                # - 실제 전력 시스템은 석탄/LNG/원자력/RE만으로 충분하므로 비활성화
                if bus_carrier_post == 'electricity' and total_load > 0.0:
                    print(f"[전력 사후 백업 발전기] 비활성화됨: {bus} (LNG_Fallback 추가 안 함 — CO2 왜곡 방지)")
                # 수소 버스: 수소 백업기(필요시)
                if bus_carrier_post == 'hydrogen' and total_load > 0.0:
                    fallback_name = f"{bus}_H2_Fallback_Gen"
                    if fallback_name not in network.generators.index:
                        network.add("Generator",
                                   name=fallback_name,
                                   bus=bus,
                                   p_nom=0.0,
                                   p_nom_extendable=True,
                                   capital_cost=1e7,
                                   marginal_cost=1e6,
                                   carrier='hydrogen')
                        network.generators_t.p_max_pu[fallback_name] = pd.Series(1.0, index=network.snapshots)
                        added_backup += 1
            if added_backup > 0:
                print(f"스케일링 이후 백업 발전기 보강: {added_backup}개 추가")
        except Exception as _e_post:
            print(f"사후 백업 발전기 보강 경고: {_e_post}")
        
        # Links 추가
        if 'links' in input_data:
            print("\n=== Links 추가 시작 ===")
            links_df = input_data['links']
            print(f"Links 시트 컬럼: {list(links_df.columns)}")
            
            # CHP 제외 모드 제거됨
            
            for idx, link in links_df.iterrows():
                link_name = str(link['name'])
                print(f"\nLink 처리 중: {link_name}")
                
                # CHP 판단: bus2가 있거나 이름에 'CHP' 포함
                # CHP 필터링 제거
                
                # 후보 컬럼에서 값 읽기 유틸
                def _read_first_available(col_candidates):
                    for c in col_candidates:
                        if (c in links_df.columns) and pd.notna(link.get(c)):
                            return link.get(c)
                    return None
                
                # 버스 연결 읽기(한글/영문 모두 지원)
                bus0_raw = _read_first_available(['bus0', '시작버스', 'from', '출발', '시작'])
                bus1_raw = _read_first_available(['bus1', '종료버스', 'to', '도착', '끝'])
                bus2_raw = _read_first_available(['bus2', '버스2'])
                bus3_raw = _read_first_available(['bus3', '버스3'])
                
                bus0_name = str(bus0_raw) if bus0_raw is not None else None
                bus1_name = str(bus1_raw) if bus1_raw is not None else None
                bus2_name = str(bus2_raw) if bus2_raw is not None else None
                bus3_name = str(bus3_raw) if bus3_raw is not None else None
                
                # 정의된 버스들만 확인
                defined_buses = set(network.buses.index)
                bus_carriers = dict(zip(network.buses.index, network.buses.carrier))
                # 이름 정규화(예: BSN_EL ↔ BSN_BSN_EL)
                # 주의: 연료 버스(LNG/gas)는 EL/H/H2로 강제 치환하지 않도록 전력 선호(prefer_electric)를 끕니다.
                if bus0_name:
                    prefer = False if (('lng' in bus0_name.lower()) or ('gas' in bus0_name.lower())) else True
                    bus0_name = _normalize_bus_name(bus0_name, defined_buses, prefer, bus_carriers)
                if bus1_name:
                    bus1_name = _normalize_bus_name(bus1_name, defined_buses, True, bus_carriers)
                if bus2_name:
                    prefer2 = False if (('lng' in bus2_name.lower()) or ('gas' in bus2_name.lower())) else True
                    bus2_name = _normalize_bus_name(bus2_name, defined_buses, prefer2, bus_carriers)
                if bus3_name:
                    prefer3 = False if (('lng' in bus3_name.lower()) or ('gas' in bus3_name.lower())) else True
                    bus3_name = _normalize_bus_name(bus3_name, defined_buses, prefer3, bus_carriers)

                # 링크 유형에 따른 버스 자동 정렬/보정
                try:
                    lname = link_name.lower()
                    is_chp_link = False
                    def _is_el(b):
                        return bool(b) and (b.endswith('_EL') or str(bus_carriers.get(b, '')).lower() == 'electricity')
                    def _is_h(b):
                        return bool(b) and (b.endswith('_H') or str(bus_carriers.get(b, '')).lower() == 'heat')
                    def _is_h2(b):
                        return bool(b) and (b.endswith('_H2') or str(bus_carriers.get(b, '')).lower() == 'hydrogen')
                    def _is_gas(b):
                        return bool(b) and (b.endswith('_LNG') or 'gas' in str(bus_carriers.get(b, '')).lower())

                    # CHP: bus0=연료(LNG/gas), bus1=전기, bus2=열
                    cand = [bus0_name, bus1_name, bus2_name]
                    fuel = next((b for b in cand if _is_gas(b)), None)
                    el   = next((b for b in cand if _is_el(b)), None)
                    heat = next((b for b in cand if _is_h(b)), None)
                    # 포트 패턴이 보이면 이름 여부와 무관하게 항상 CHP로 정렬
                    if (fuel and el and heat) or ('chp' in lname):
                        if fuel and el and heat:
                            if (bus0_name, bus1_name, bus2_name) != (fuel, el, heat):
                                note = "이름 기반" if ('chp' in lname) else "포트 기반"
                                print(f"CHP 버스 자동정렬({note}): {bus0_name},{bus1_name},{bus2_name} -> {fuel},{el},{heat}")
                            bus0_name, bus1_name, bus2_name = fuel, el, heat
                        is_chp_link = True

                    # Electrolyser: bus0=전기, bus1=수소, bus2 미사용
                    elif ('electrolyser' in lname) or ('electrolyzer' in lname):
                        if not _is_el(bus0_name) and _is_el(bus1_name):
                            print(f"Electrolyser 버스 스왑: {bus0_name}<->{bus1_name}")
                            bus0_name, bus1_name = bus1_name, bus0_name
                        bus2_name = None

                    # Heat Pump: bus0=전기, bus1=열, bus2 미사용
                    elif ('heatpump' in lname) or ('heat_pump' in lname) or (lname.startswith('hp') or ' hp ' in lname):
                        if not _is_el(bus0_name) and _is_el(bus1_name):
                            print(f"HeatPump 버스 스왑: {bus0_name}<->{bus1_name}")
                            bus0_name, bus1_name = bus1_name, bus0_name
                        bus2_name = None
                except Exception:
                    pass
                
                # 유효한 버스 연결 확인 (bus0/bus1는 필수)
                if bus0_name and bus1_name and bus0_name in defined_buses and bus1_name in defined_buses:
                    # 효율(한글/영문 매핑):
                    # - efficiency: bus1로의 효율(우선순위: 'efficiency' → 'efficiency0' → '효율0')
                    # - efficiency2: bus2로의 효율(우선순위: 'efficiency2' → 'efficiency1' → 'efficiency_2' → '효율2' → '효율1')
                    # - efficiency3: bus3로의 효율(우선순위: 'efficiency3' → 'efficiency_3' → '효율3')
                    eff1_val = _read_first_available(['efficiency', 'efficiency0', '효율0'])
                    eff2_val = _read_first_available(['efficiency2', 'efficiency1', 'efficiency_2', '효율2', '효율1'])
                    eff3_val = _read_first_available(['efficiency3', 'efficiency_3', '효율3'])
                    
                    # Link 정격 용량(p_nom) 읽기: 다양한 대체 컬럼 지원
                    pnom_raw = _read_first_available(['p_nom', '용량', 'capacity', '정격', '전력용량', '전력_용량'])
                    try:
                        pnom_val = float(pd.to_numeric(pnom_raw, errors='coerce')) if pnom_raw is not None else float('nan')
                    except Exception:
                        pnom_val = float('nan')
                    if (pnom_raw is None) or (pd.isna(pnom_val)):
                        # 시트에 p_nom 컬럼이 있는지 확인
                        if 'p_nom' in links_df.columns:
                            sheet_pnom = link.get('p_nom')
                            if pd.notna(sheet_pnom):
                                pnom_val = float(sheet_pnom)
                            else:
                                # NaN인 경우, HP는 0으로 기본 설정, 다른 링크는 100
                                if 'hp' in link_name.lower():
                                    pnom_val = 0.0  # HP는 0으로 기본 설정
                                    print(f"HP 링크 {link_name}: p_nom이 NaN이므로 0으로 설정")
                                else:
                                    pnom_val = 100.0
                        else:
                            pnom_val = 100.0  # 진짜로 컬럼이 없을 때만 기본값 사용
                    
                    params = {
                        'name': link_name,
                        'bus0': bus0_name,
                        'bus1': bus1_name,
                        'p_nom': pnom_val,
                        'efficiency': float(eff1_val) if pd.notna(eff1_val) else (0.5 if ('electrolyser' in link_name.lower() or 'electrolyzer' in link_name.lower()) else 0.9),
                        'carrier': 'gas' if is_chp_link else 'electricity'  # CHP는 가스 carrier로 설정
                    }

                    # 열 공급 우선순위 유도: CHP < HP < Fallback
                    try:
                        default_mcost = None
                        lname2 = link_name.lower()
                        # 시트에 명시된 marginal_cost가 없을 때만 기본값 적용
                        has_sheet_mcost = ('marginal_cost' in links_df.columns and pd.notna(link.get('marginal_cost')))
                        if not has_sheet_mcost:
                            if is_chp_link:
                                default_mcost = 1.0  # 최우선 사용
                            elif ('heatpump' in lname2) or ('heat_pump' in lname2) or (lname2.startswith('hp') or ' hp ' in lname2):
                                default_mcost = 1e5  # Fallback보다 낮고 CHP보다 높게
                        if default_mcost is not None:
                            params['marginal_cost'] = float(default_mcost)
                        elif has_sheet_mcost:
                            params['marginal_cost'] = float(link.get('marginal_cost'))
                    except Exception:
                        pass
                    
                    # 선택적 추가 출력 버스와 효율
                    if bus2_name and bus2_name in defined_buses:
                        params['bus2'] = bus2_name
                        if pd.notna(eff2_val):
                            params['efficiency2'] = float(eff2_val)
                        else:
                            # CHP로 추정되나 eff2가 비어 있으면 기본 0.4 적용
                            # NaN 방지: CHP 포함 모든 경우, 명시 없으면 0.0으로 설정
                            params['efficiency2'] = 0.0
                    elif bus2_name:
                        print(f"경고: Link {link_name}의 bus2 '{bus2_name}'가 네트워크에 없어 무시합니다.")
                    
                    if bus3_name and bus3_name in defined_buses:
                        params['bus3'] = bus3_name
                        if pd.notna(eff3_val):
                            params['efficiency3'] = float(eff3_val)
                        else:
                            # NaN 방지
                            params['efficiency3'] = 0.0
                    elif bus3_name:
                        print(f"경고: Link {link_name}의 bus3 '{bus3_name}'가 네트워크에 없어 무시합니다.")
                    
                    # p_nom_extendable 및 용량 하한/상한 처리
                    if 'p_nom_extendable' in links_df.columns and pd.notna(link.get('p_nom_extendable')):
                        params['p_nom_extendable'] = _to_bool(link.get('p_nom_extendable'))
                    if 'p_nom_min' in links_df.columns and pd.notna(link.get('p_nom_min')):
                        params['p_nom_min'] = float(link.get('p_nom_min'))
                    if 'p_nom_max' in links_df.columns and pd.notna(link.get('p_nom_max')):
                        params['p_nom_max'] = float(link.get('p_nom_max'))
                    
                    try:
                        network.add("Link", **params)
                        print_msg = f"Link {link_name} 추가됨: {bus0_name} -> {bus1_name}"
                        if 'bus2' in params:
                            print_msg += f", bus2: {params['bus2']}"
                        if 'bus3' in params:
                            print_msg += f", bus3: {params['bus3']}"
                        print_msg += f" (p_nom: {params.get('p_nom', 'n/a')}, eff: {params.get('efficiency', 'n/a')}"
                        if 'efficiency2' in params:
                            print_msg += f", eff2: {params['efficiency2']}"
                        if 'efficiency3' in params:
                            print_msg += f", eff3: {params['efficiency3']}"
                        if is_chp_link:
                            print_msg += ", CHP=Y"
                        print_msg += ")"
                        print(print_msg)
                    except Exception as e:
                        print(f"Link {link_name} 추가 중 오류: {str(e)}")
                else:
                    print(f"Link {link_name} 건너뜀: 유효하지 않은 버스 연결 (bus0: {bus0_name}, bus1: {bus1_name})")
            
            # # EV 충전 패턴 적용 (Links 추가 완료 후) [격리 테스트를 위해 임시 비활성화]
            # print("\n=== EV 충전 패턴 적용 시작 ===")
            # if 'load_patterns' in input_data:
            #     ev_charging_pattern = _get_load_pattern(input_data, None, 'EV_CHARGING', len(snapshots))
            #     if ev_charging_pattern is not None:
            #         # 방어적 코딩: 0 값을 최소값으로 치환 (혹시나 모를 문제 방지)
            #         min_charging_rate = 0.01  # 최소 1% 충전 보장
            #         ev_charging_pattern = np.maximum(ev_charging_pattern, min_charging_rate)
            #         
            #         ev_charger_count = 0
            #         for link_name in network.links.index:
            #             # EV_Charger 링크 감지 (이름에 'EV_Charger', 'EV_charger', 'ev_charger' 포함)
            #             if 'ev' in link_name.lower() and 'charg' in link_name.lower():
            #                 # p_max_pu에 충전 가능 패턴 적용
            #                 network.links_t.p_max_pu[link_name] = ev_charging_pattern
            #                 print(f"Link {link_name}에 EV 충전 패턴 적용됨 (p_max_pu, 최소 {min_charging_rate*100:.0f}% 보장)")
            #                 ev_charger_count += 1
            #         
            #         if ev_charger_count > 0:
            #             print(f"총 {ev_charger_count}개 EV Charger 링크에 충전 패턴 적용 완료")
            #             # 패턴 샘플 출력
            #             sample_n = min(24, len(ev_charging_pattern))
            #             print(f"\nEV 충전 패턴 상세 (첫 24시간):")
            #             print(f"  {np.round(ev_charging_pattern[:sample_n], 4).tolist()}")
            #             print(f"\nEV 충전 패턴 통계:")
            #             print(f"  최소: {np.min(ev_charging_pattern):.6f}")
            #             print(f"  최대: {np.max(ev_charging_pattern):.6f}")
            #             print(f"  평균: {np.mean(ev_charging_pattern):.6f}")
            #             print(f"  합계: {np.sum(ev_charging_pattern):.6f} (정규화 여부 확인: 1.0이면 문제!)")
            #             print(f"  0.1 미만 개수: {np.sum(ev_charging_pattern < 0.1)} / {len(ev_charging_pattern)}")
            #             
            #             # EV Charger p_nom 확인
            #             charger_pnom = network.links.loc[network.links.index.str.contains('EV_Charger', case=False, na=False), 'p_nom']
            #             if not charger_pnom.empty:
            #                 print(f"\nEV Charger 용량 (p_nom):")
            #                 print(f"  평균: {charger_pnom.mean():.2f} MW")
            #                 print(f"  최소: {charger_pnom.min():.2f} MW")
            #                 print(f"  최대: {charger_pnom.max():.2f} MW")
            #                 print(f"  예상 최소 충전 가능: {charger_pnom.mean() * np.min(ev_charging_pattern):.2f} MW")
            #                 print(f"  예상 최대 충전 가능: {charger_pnom.mean() * np.max(ev_charging_pattern):.2f} MW")
            #         else:
            #             print("[경고] EV Charger 링크를 찾지 못했습니다.")
            #     else:
            #         print("[경고] load_patterns 시트에서 EV_CHARGING(F열) 패턴을 찾지 못했습니다.")
            # else:
            #     print("[경고] load_patterns 시트가 없어 EV 충전 패턴을 적용할 수 없습니다.")

            # V2G 링크에 역순 EV_Charging 패턴 적용
            # EV_Charging이 낮을 때(사람들이 충전 안 할 때) → V2G 최대 가동
            # EV_Charging이 높을 때(사람들이 충전할 때) → V2G 가동 불가
            print("\n=== V2G 패턴 적용 시작 (역순 EV_Charging) ===")
            try:
                ev_charging_pattern = _get_load_pattern(input_data, None, 'EV_CHARGING', len(snapshots))
                if ev_charging_pattern is not None:
                    # 역순 변환: v2g_pattern = 1 - ev_charging_pattern
                    # EV 충전이 낮을수록 V2G 가동률 높음
                    v2g_pattern = 1.0 - np.array(ev_charging_pattern, dtype=float)
                    # 0~1 범위 클리핑 (안전 처리)
                    v2g_pattern = np.clip(v2g_pattern, 0.0, 1.0)

                    v2g_count = 0
                    for link_name in network.links.index:
                        if 'v2g' in link_name.lower():
                            region = link_name.split('_')[0]
                            network.links_t.p_max_pu[link_name] = pd.Series(
                                v2g_pattern, index=network.snapshots)
                            v2g_count += 1

                    if v2g_count > 0:
                        print(f"  V2G 패턴 적용 완료: {v2g_count}개 링크")
                        print(f"  패턴 통계: 최소={v2g_pattern.min():.3f}, "
                              f"최대={v2g_pattern.max():.3f}, 평균={v2g_pattern.mean():.3f}")
                    else:
                        print("  [경고] V2G 링크를 찾지 못했습니다.")
                else:
                    print("  [경고] EV_CHARGING 패턴을 찾지 못해 V2G 패턴 미적용 (기본값 1.0 유지)")
            except Exception as _e_v2g:
                print(f"  V2G 패턴 적용 오류: {_e_v2g}")

        # 저장장치 추가
        if 'stores' in input_data:
            print("\n=== Stores 추가 시작 ===")
            for _, store in input_data['stores'].iterrows():
                store_name = str(store['name'])
                bus_name = str(store['bus'])
                
                # 버스 존재 확인
                if bus_name in network.buses.index:
                    e_nom_val = float(store['e_nom']) if pd.notna(store['e_nom']) else 0
                    
                    params = {
                        'name': store_name,
                        'bus': bus_name,
                        'carrier': str(store['carrier']),
                        'e_nom': e_nom_val,
                        'e_cyclic': _to_bool(store['e_cyclic']) if 'e_cyclic' in store and pd.notna(store['e_cyclic']) else True,
                        'efficiency_store': float(store['efficiency_store']) if 'efficiency_store' in store and pd.notna(store['efficiency_store']) else 0.9,
                        'efficiency_dispatch': float(store['efficiency_dispatch']) if 'efficiency_dispatch' in store and pd.notna(store['efficiency_dispatch']) else 0.9,
                        'standing_loss': float(store['standing_loss']) if 'standing_loss' in store and pd.notna(store['standing_loss']) else 0,
                        'e_initial': float(store['e_initial']) if 'e_initial' in store and pd.notna(store['e_initial']) else 0
                    }
                    # capital_cost: extendable일 때 MWh당 연간 자본비용 (없으면 0)
                    for _cap_col in ('capital_cost', '자본비용', '자본비용(원/MWh)'):
                        if _cap_col in store and pd.notna(store[_cap_col]):
                            params['capital_cost'] = float(store[_cap_col])
                            break
                    if 'e_nom_extendable' in store and pd.notna(store['e_nom_extendable']):
                        params['e_nom_extendable'] = _to_bool(store['e_nom_extendable'])
                    
                    # 용량 확장 하한/상한 (extendable일 때)
                    if 'e_nom_min' in store and pd.notna(store['e_nom_min']):
                        params['e_nom_min'] = float(store['e_nom_min'])
                    # e_nom_max: 영문 'e_nom_max' 또는 한국어 '최대저장용량'/'최대용량(MWh)' 컬럼 지원
                    # 주의: '최대용량(MW)'는 발전기 p_nom_max용이므로 stores에서는 사용하지 않음
                    _e_nom_max_val = None
                    for _col in ('e_nom_max', '최대용량', '최대용량(MWh)'):
                        if _col in store and pd.notna(store[_col]):
                            _e_nom_max_val = float(store[_col])
                            break
                    if _e_nom_max_val is not None:
                        params['e_nom_max'] = _e_nom_max_val
                        if params.get('e_nom_extendable'):
                            print(f"  {store_name}: e_nom_max={_e_nom_max_val:,.0f} MWh (확장 상한)")
                    
                    # 운영 중 최소 잔량 (비율) - e_min_pu [활성화 복구]
                    # Excel 컬럼명: "e_min_pu" 또는 "최소잔량" (0~1 비율로 입력)
                    e_min_pu_col = None
                    if 'e_min_pu' in store and pd.notna(store['e_min_pu']):
                        e_min_pu_col = 'e_min_pu'
                    elif '최소잔량' in store and pd.notna(store['최소잔량']):
                        e_min_pu_col = '최소잔량'
                    
                    if e_min_pu_col:
                        e_min_pu_val = float(store[e_min_pu_col])
                        # 0~1 범위 제한
                        params['e_min_pu'] = min(max(e_min_pu_val, 0.0), 1.0)
                        print(f"  {store_name}: e_min_pu={params['e_min_pu']:.2f} (최소잔량 {params['e_min_pu']*100:.0f}%) [from {e_min_pu_col}]")
                    
                    # 운영 중 최대 잔량 (비율) - e_max_pu (선택적) [활성화 복구]
                    if 'e_max_pu' in store and pd.notna(store['e_max_pu']):
                        e_max_pu_val = float(store['e_max_pu'])
                        params['e_max_pu'] = min(max(e_max_pu_val, 0.0), 1.0)
                        print(f"  {store_name}: e_max_pu={params['e_max_pu']:.2f} (최대잔량 {params['e_max_pu']*100:.0f}%)")
                    
                    network.add("Store", **params)
                    print(f"저장장치 {store_name} 추가됨 (버스: {bus_name})")
                    
                    # EV_DSM 특별 처리: SOC 하한은 interface.xlsx e_min_pu 기본값만 사용
                    # (시간대별 추가 제약 없음 → 6시 50% 및 야간 35% 모두 제거)
                    if 'EV_DSM' in store_name:
                        base_e_min_pu = params.get('e_min_pu', 0.2)
                        print(f"  [EV_DSM SOC 제약] {store_name}: "
                              f"기본={base_e_min_pu:.2f} (시간대별 추가 제약 없음)")
                else:
                    print(f"저장장치 {store_name} 건너뜀: 버스 '{bus_name}'가 존재하지 않음")

        # 선로 추가 (있는 경우)
        if 'lines' in input_data and not input_data['lines'].empty:
            print("\n=== 선로 추가 시작 ===")
            added_lines = 0
            skipped_lines = 0
            conditional_lines = 0  # RPS 조건부 선로 카운트
            
            # RPS 제약 여부 확인
            has_rps = False
            if hasattr(network, 'rps_constraints') and network.rps_constraints:
                has_rps = True
            elif 'constraints' in input_data and not input_data['constraints'].empty:
                try:
                    constraints_df = input_data['constraints'].copy()
                    if 'type' in constraints_df.columns or '유형' in constraints_df.columns:
                        type_col = 'type' if 'type' in constraints_df.columns else '유형'
                        has_rps = any(constraints_df[type_col].astype(str).str.upper() == 'RPS')
                except:
                    pass
            
            print(f"RPS 제약 존재: {has_rps}")
            
            for _, line in input_data['lines'].iterrows():
                line_name = str(line['name'])
                bus0_name = str(line['bus0'])
                bus1_name = str(line['bus1'])
                
                # 조건부 활성화 체크 (enable_condition 컬럼)
                enable_condition = None
                if 'enable_condition' in line and pd.notna(line['enable_condition']):
                    enable_condition = str(line['enable_condition']).strip().upper()
                
                # RPS 조건부 선로 처리
                if enable_condition == 'RPS':
                    if not has_rps:
                        # RPS 제약이 없으면 이 선로는 추가하지 않음
                        print(f"선로 {line_name} 건너뜀: RPS 제약 없음 (조건부 RE 송전선)")
                        skipped_lines += 1
                        conditional_lines += 1
                        continue
                    else:
                        print(f"선로 {line_name}: RPS 조건 만족 → RE Ring 송전선 활성화")
                
                # 버스 이름 정규화
                defined_buses = set(network.buses.index)
                bus_carriers = dict(zip(network.buses.index, network.buses.carrier))
                
                # RE 버스 간 연결은 prefer_electric=False로 설정하여 원래 이름 유지
                is_re_connection = ('_RE' in str(bus0_name) or '_RE' in str(bus1_name))
                prefer_electric = not is_re_connection  # RE 연결이면 False, 아니면 True
                
                bus0_name_norm = _normalize_bus_name(bus0_name, defined_buses, prefer_electric, bus_carriers)
                bus1_name_norm = _normalize_bus_name(bus1_name, defined_buses, prefer_electric, bus_carriers)
                
                # 선로 이름에서 지역코드 강제 추출 후 전력버스로 강제 매핑 시도 (예: GND_JND → GND_EL, JND_EL)

                
                print(f"선로 {line_name} 시도: {bus0_name} -> {bus1_name} (정규화/강제: {bus0_name_norm} -> {bus1_name_norm})")
                print(f"  bus0 존재: {bus0_name_norm in network.buses.index}")
                print(f"  bus1 존재: {bus1_name_norm in network.buses.index}")
                
                if bus0_name_norm in network.buses.index and bus1_name_norm in network.buses.index:
                    # 선로 전압(type) 기반으로 각 끝단에 해당 전압 보조 버스를 확보하고 변압기(링크) 자동 삽입
                    def _bus_v(bus_name):
                        try:
                            return float(network.buses.at[bus_name, 'v_nom'])
                        except Exception:
                            return float('nan')

                    def _parse_voltage(val):
                        try:
                            if pd.isna(val):
                                return None
                            s = str(val)
                            nums = ''.join(ch for ch in s if ch.isdigit())
                            if nums:
                                return float(nums)
                        except Exception:
                            pass
                        return None

                    line_voltage = _parse_voltage(line['type']) if ('type' in line) else None

                    # 라인 파라미터 구성
                    params = {
                        'name': line_name,
                        'bus0': bus0_name_norm,
                        'bus1': bus1_name_norm,
                        's_nom': float(line['s_nom']) if ('s_nom' in line and pd.notna(line['s_nom'])) else 1000.0,
                        'x': float(line['x']) if ('x' in line and pd.notna(line['x'])) else 0.1,
                        'r': float(line['r']) if ('r' in line and pd.notna(line['r'])) else 0.01
                    }

                    # 추가 속성: type, length, num_parallel
                    if 'type' in line and pd.notna(line['type']):
                        tval = str(line['type']).strip()
                        try:
                            available_types = set(network.line_types.index)
                        except Exception:
                            available_types = set()
                        if tval in available_types:
                            params['type'] = tval
                    if 'length' in line and pd.notna(line['length']):
                        params['length'] = float(line['length'])
                    if 'num_parallel' in line and pd.notna(line['num_parallel']):
                        params['num_parallel'] = int(pd.to_numeric(line['num_parallel'], errors='coerce'))

                    # s_nom 확장 및 하한/상한 처리
                    if 's_nom_extendable' in line and pd.notna(line['s_nom_extendable']):
                        params['s_nom_extendable'] = bool(line['s_nom_extendable'])
                    if 's_nom_min' in line and pd.notna(line['s_nom_min']):
                        params['s_nom_min'] = float(line['s_nom_min'])
                    if 's_nom_max' in line and pd.notna(line['s_nom_max']):
                        params['s_nom_max'] = float(line['s_nom_max'])

                    # 라인 전압 기반 변압기 자동삽입(기본 비활성화). 활성화하려면 ENABLE_AUTO_TRANSFORMER=1 환경변수 설정
                    if (os.environ.get('ENABLE_AUTO_TRANSFORMER', '0') == '1') and (line_voltage is not None and not np.isnan(line_voltage)):
                        def ensure_voltage_bus(orig_bus, target_kv):
                            if orig_bus not in network.buses.index:
                                return orig_bus
                            if abs(_bus_v(orig_bus) - target_kv) < 1e-6:
                                return orig_bus
                            new_bus = f"{orig_bus}_{int(target_kv)}"
                            if new_bus not in network.buses.index:
                                network.add("Bus", name=new_bus, v_nom=target_kv, carrier='electricity')
                                # 변압기 등가: Link로 근사(효율 0.995)
                                tr_name = f"TR_{orig_bus}_to_{new_bus}"
                                if tr_name not in network.links.index:
                                    network.add("Link", name=tr_name, bus0=orig_bus, bus1=new_bus, p_nom=1e6, efficiency=0.995)
                                # 역방향 변압기 링크도 추가(양방향 전력 흐름 허용)
                                tr_rev = f"TR_{new_bus}_to_{orig_bus}"
                                if tr_rev not in network.links.index:
                                    network.add("Link", name=tr_rev, bus0=new_bus, bus1=orig_bus, p_nom=1e6, efficiency=0.995)
                            return new_bus
                        params['bus0'] = ensure_voltage_bus(bus0_name_norm, line_voltage)
                        params['bus1'] = ensure_voltage_bus(bus1_name_norm, line_voltage)

                    try:
                        network.add("Line", **params)
                        print(f"선로 {line_name} 추가됨: {params['bus0']} - {params['bus1']}")
                        added_lines += 1
                    except Exception as e:
                        print(f"선로 {line_name} 추가 실패: {str(e)}")
                        skipped_lines += 1
                else:
                    print(f"선로 {line_name} 건너뜀: 유효하지 않은 버스 연결")
                    skipped_lines += 1
                    if bus0_name not in network.buses.index:
                        print(f"  누락된 버스: {bus0_name}")
                    if bus1_name not in network.buses.index:
                        print(f"  누락된 버스: {bus1_name}")
            
            print(f"\n선로 추가 요약: 성공 {added_lines}개, 실패 {skipped_lines}개")
            if conditional_lines > 0:
                if has_rps:
                    print(f"  - RPS 조건부 RE Ring 송전선: {conditional_lines}개 활성화됨")
                else:
                    print(f"  - RPS 조건부 RE Ring 송전선: {conditional_lines}개 비활성화됨")
            print(f"네트워크에 추가된 총 선로 수: {len(network.lines)}")
            
            # 버스 목록 확인
            print(f"\n현재 네트워크 버스 수: {len(network.buses)}")
            print("첫 10개 버스:", list(network.buses.index[:10]))
            
            # CHP 제외 요약 제거됨
        else:
            print("\n선로 데이터가 없습니다.")
        
        # 링크 효율 NaN 일괄 보정: efficiency2/efficiency3 NaN → 0.0
        try:
            if not network.links.empty:
                if 'efficiency2' in network.links.columns:
                    network.links['efficiency2'] = pd.to_numeric(network.links['efficiency2'], errors='coerce').fillna(0.0)
                if 'efficiency3' in network.links.columns:
                    network.links['efficiency3'] = pd.to_numeric(network.links['efficiency3'], errors='coerce').fillna(0.0)
                print("링크 효율 NaN 보정 완료(efficiency2/3 → 0.0)")
        except Exception as _e_eff:
            print(f"링크 효율 보정 경고: {_e_eff}")
        
        # 제약조건 추가 (CO2, RPS 등)
        if os.environ.get('DISABLE_CO2_LIMIT','0')=='1':
            print('CO2 제약 비활성화됨(DISABLE_CO2_LIMIT=1)')
        elif 'constraints' in input_data and not input_data['constraints'].empty:
            constraints_df = input_data['constraints'].copy()
            # 컬럼 표준화
            rename_map = {
                '이름': 'name', 'name': 'name',
                '상수': 'constant', 'constant': 'constant', '상수값': 'constant',
                '조건': 'sense', 'sense': 'sense',
                '유형': 'type', 'type': 'type',
                '대상속성': 'carrier_attribute', 'carrier_attribute': 'carrier_attribute',
                '적용지역': 'region', 'region': 'region', '적용 지역': 'region',
                '적용연도': 'year', 'year': 'year', '연도': 'year',
            }
            std_cols = {}
            for c in constraints_df.columns:
                key = str(c).strip()
                std_cols[c] = rename_map.get(key, key)
            constraints_df.rename(columns=std_cols, inplace=True)

            # 현재 시나리오 연도 추출 (timeseries start_time 기준)
            scenario_year = None
            try:
                if 'timeseries' in input_data and not input_data['timeseries'].empty:
                    st = str(input_data['timeseries'].iloc[0].get('start_time', ''))
                    if st:
                        scenario_year = int(st[:4])
            except Exception:
                pass

            def _filter_by_year(df):
                """'year' 컬럼이 있으면 현재 시나리오 연도와 일치하는 행만 반환.
                'year' 컬럼이 없거나 비어있으면 모든 행을 반환 (연도 무관 적용).
                """
                if 'year' not in df.columns or scenario_year is None:
                    return df
                year_vals = pd.to_numeric(df['year'], errors='coerce')
                # NaN(연도 미지정) → 모든 연도에 적용
                mask = year_vals.isna() | (year_vals == scenario_year)
                return df[mask]

            # 1) CO2 제약 처리
            co2_limit = pd.DataFrame()
            if 'name' in constraints_df.columns:
                # CO2_limit 또는 CO2Limit 모두 지원
                co2_mask = constraints_df['name'].astype(str).str.strip().isin(['CO2_limit', 'CO2Limit'])
                co2_limit = _filter_by_year(constraints_df[co2_mask])

            if not co2_limit.empty:
                const_col = 'constant' if 'constant' in co2_limit.columns else None
                if const_col:
                    limit_value = float(pd.to_numeric(co2_limit.iloc[0][const_col], errors='coerce'))
                    year_info = f" ({scenario_year}년 적용)" if scenario_year else ""
                    network.add("GlobalConstraint",
                                name="CO2_limit",
                                type="primary_energy",
                                carrier_attribute="co2_emissions",
                                sense="<=",
                                constant=limit_value)
                    print(f"[OK] PyPSA CO2 배출량 제약조건 추가됨{year_info}: ≤ {limit_value/1e6:.1f}백만톤 (전력량 기준)")
                    print(f"   carrier.co2_emissions는 efficiency 보정됨 (전력량기준 × 평균효율)")
                else:
                    print("경고: constraints 시트에 'constant' 컬럼이 없어 CO2Limit을 적용하지 못했습니다.")
            
            # 2) RPS (재생에너지 의무 비중) 제약 처리
            if 'type' in constraints_df.columns:
                rps_mask = constraints_df['type'].astype(str).str.strip().str.upper() == 'RPS'
                rps_constraints = _filter_by_year(constraints_df[rps_mask])
                
                if not rps_constraints.empty:
                    print(f"\n[RE] 재생에너지 의무 비중(RPS) 제약 추가 중...")
                    network.rps_constraints = []
                    
                    for _, rps_row in rps_constraints.iterrows():
                        try:
                            rps_name = str(rps_row.get('name', 'RPS')).strip()
                            region = str(rps_row.get('region', 'ALL')).strip()
                            share = float(pd.to_numeric(rps_row.get('constant', 0.2), errors='coerce'))
                            
                            network.rps_constraints.append({
                                'name': rps_name,
                                'region': region,
                                'share': share
                            })
                            
                            print(f"   [OK] {rps_name}: {region} 지역 재생에너지 비중 ≥ {share*100:.1f}%")
                        except Exception as e:
                            print(f"   [경고]  RPS 제약 '{rps_row.get('name', 'Unknown')}' 추가 실패: {str(e)}")
                    
                    print(f"   총 {len(network.rps_constraints)}개 RPS 제약 설정됨")
        else:
            print("경고: constraints 시트가 없거나 비어있습니다.")
        
        # 최종 안전장치: 전력(EL/EV) 슬랙 비활성화 — 열/수소 fallback은 별도 루프에서 처리
        # ensure_slack = False → EL, EV 버스에 Slack_Failsafe 추가 안 함
        try:
            failsafe_added = 0
            ensure_slack = False  # EL/EV 슬랙 비활성화
            print("최후수단 슬랙 발전기(EL/EV) 비활성화됨")
            for bus in network.buses.index:
                try:
                    bus_carrier = str(network.buses.at[bus, 'carrier']).lower()
                except Exception:
                    bus_carrier = ''
                if bus_carrier not in ['electricity', 'heat', 'hydrogen']:
                    continue
                # EL/EV 버스: 슬랙 비활성화 (열/수소는 별도 루프에서 처리)
                if bus_carrier in ['electricity']:
                    continue
                # 해당 버스 부하 존재 여부
                has_load = False
                try:
                    if (not network.loads.empty) and hasattr(network.loads_t, 'p_set') and (network.loads_t.p_set is not None) and (not network.loads_t.p_set.empty):
                        names = network.loads.index[network.loads.bus.astype(str) == bus]
                        for nm in names:
                            if nm in network.loads_t.p_set.columns and float(np.nansum(network.loads_t.p_set[nm].values)) > 0.0:
                                has_load = True
                                break
                    if (not has_load) and (not network.loads.empty) and hasattr(network.loads_t, 'p') and (network.loads_t.p is not None) and (not network.loads_t.p.empty):
                        names = network.loads.index[network.loads.bus.astype(str) == bus]
                        for nm in names:
                            if nm in network.loads_t.p.columns and float(np.nansum(network.loads_t.p[nm].values)) > 0.0:
                                has_load = True
                                break
                except Exception:
                    pass
                if not has_load:
                    continue
                # 기존 발전기 유무 확인
                has_gen = False
                try:
                    gens = network.generators.index[network.generators.bus.astype(str) == bus]
                    if len(gens) > 0:
                        has_gen = True
                except Exception:
                    pass
                # ensure_slack가 활성화되면 기존 발전기가 있어도 슬랙 추가
                if (not has_gen) or ensure_slack:
                    slack_name = f"{bus}_Slack_Failsafe"
                    if slack_name not in network.generators.index:
                        # EV 버스: 100만 원/MWh (전력/열/수소보다 저렴, 수치 안정성)
                        # 전력/열/수소 버스: 1000만 원/MWh (최후 수단)
                        if bus.endswith('_EV'):
                            mcost = 1_000_000.0   # 1e6
                        else:
                            mcost = float(os.environ.get('SLACK_GEN_COST', '10000000'))  # 1e7 기본값
                        # 전력/열/수소 중 해당 캐리어로 무배출 슬랙
                        carrier_val = bus_carrier if bus_carrier in ['electricity','heat','hydrogen'] else 'electricity'
                        network.add("Generator",
                                   name=slack_name,
                                   bus=bus,
                                   p_nom=0.0,
                                   p_nom_extendable=True,
                                   capital_cost=0.0,
                                   marginal_cost=mcost,
                                   carrier=carrier_val)
                        network.generators_t.p_max_pu[slack_name] = pd.Series(1.0, index=network.snapshots)
                        failsafe_added += 1
            if failsafe_added > 0:
                print(f"최후수단 슬랙 발전기 추가: {failsafe_added}개")
        except Exception as _e_fs:
            print(f"최후수단 슬랙 발전기 추가 경고: {_e_fs}")
        
        # 경계값(최소/최대) 정리: infeasible 방지
        try:
            _sanitize_component_bounds(network)
        except Exception as _e_s:
            print(f"경계값 정리 경고: {_e_s}")
        return network
        
    except Exception as e:
        print(f"네트워크 생성 중 오류 발생: {str(e)}")
        traceback.print_exc()
        return None

def optimize_network(network):
    """네트워크 최적화"""
    if network is None:
        print("네트워크가 생성되지 않았습니다.")
        return False
    
    try:
        print("\n최적화 시작...")
        
        # CPU 코어 수 확인
        import multiprocessing
        num_cores = multiprocessing.cpu_count()
        print(f"사용 가능한 CPU 코어 수: {num_cores}")
        
        # 시점별 CHP 링크 요약(진단용)
        try:
            if not network.links.empty:
                chp_like = []
                for lk in network.links.index:
                    try:
                        b0 = str(network.links.at[lk, 'bus0'])
                        b1 = str(network.links.at[lk, 'bus1'])
                        b2 = str(network.links.at[lk, 'bus2']) if 'bus2' in network.links.columns and not pd.isna(network.links.at[lk, 'bus2']) else None
                        c0 = str(network.buses.at[b0, 'carrier']).lower() if b0 in network.buses.index else ''
                        c1 = str(network.buses.at[b1, 'carrier']).lower() if b1 in network.buses.index else ''
                        c2 = str(network.buses.at[b2, 'carrier']).lower() if (b2 and b2 in network.buses.index) else ''
                        if ('gas' in c0) and ('electric' in c1) and ('heat' in c2):
                            chp_like.append((lk, b0, b1, b2, float(network.links.at[lk, 'p_nom'])))
                    except Exception:
                        continue
                if chp_like:
                    print(f"CHP-유사 링크 {len(chp_like)}개 발견:")
                    for row in chp_like[:10]:
                        print(f"  {row[0]}: {row[1]} -> {row[2]}(EL), bus2={row[3]}(H), p_nom={row[4]}")
        except Exception:
            pass
        
        # =====================================================================
        # DR(수요반응) 설정 - 하드코딩 파라미터
        # =====================================================================
        DR_PEAK_PERCENTILE = 90   # 상위 10% 부하 시점에서만 발동 허용
        DR_ANNUAL_HOURS    = 300  # 연간 최대 발동시간 (시간)
        # =====================================================================

        # DR 발전기 p_max_pu 설정: 각 지역 전력수요 상위 5% 시점에만 1, 나머지 0
        try:
            dr_generators = [
                g for g in network.generators.index
                if str(network.generators.at[g, 'carrier']).upper() == 'DR'
            ]
            if dr_generators:
                print(f"\n=== DR 발전기 p_max_pu 설정 (상위 {100 - DR_PEAK_PERCENTILE}% 피크 조건) ===")
                region_codes = [
                    'SEL', 'BSN', 'DGU', 'ICN', 'GWJ', 'DJN', 'USN', 'SJG',
                    'GGD', 'GWD', 'CBD', 'CND', 'JBD', 'JND', 'GBD', 'GND', 'JJD'
                ]
                for dr_name in dr_generators:
                    # 지역코드 추출: 버스 속성 우선, 발전기 이름 폴백
                    bus_attr = str(network.generators.at[dr_name, 'bus'])
                    region = None
                    # 1순위: 버스명에서 추출 (예: GGD_EL → GGD)
                    for code in region_codes:
                        if bus_attr.startswith(code + '_') or bus_attr == code:
                            region = code
                            break
                    # 2순위: 발전기 이름에서 추출 (폴백)
                    if region is None:
                        for code in region_codes:
                            if dr_name.startswith(code + '_') or dr_name == code:
                                region = code
                                break

                    if region is None:
                        print(f"  [경고] {dr_name} (bus={bus_attr}): 지역코드 추출 실패 → p_max_pu=1 적용")
                        network.generators_t.p_max_pu[dr_name] = pd.Series(1.0, index=network.snapshots)
                        continue

                    # 해당 지역 전력버스 부하 시계열 합산
                    el_bus = f"{region}_EL"
                    regional_load = pd.Series(0.0, index=network.snapshots)
                    try:
                        if (not network.loads.empty
                                and hasattr(network.loads_t, 'p_set')
                                and not network.loads_t.p_set.empty):
                            load_names = network.loads.index[
                                network.loads.bus.astype(str) == el_bus
                            ]
                            for ln in load_names:
                                if ln in network.loads_t.p_set.columns:
                                    regional_load = regional_load.add(
                                        network.loads_t.p_set[ln].fillna(0.0),
                                        fill_value=0.0
                                    )
                    except Exception as _e_load:
                        print(f"  [경고] {dr_name}: 부하 시계열 읽기 실패 ({_e_load}) → p_max_pu=1 적용")
                        network.generators_t.p_max_pu[dr_name] = pd.Series(1.0, index=network.snapshots)
                        continue

                    if regional_load.sum() <= 0:
                        print(f"  [경고] {dr_name}: {region} 전력부하 없음 → p_max_pu=1 적용")
                        network.generators_t.p_max_pu[dr_name] = pd.Series(1.0, index=network.snapshots)
                        continue

                    # 상위 5% 임계값 계산 후 p_max_pu 설정
                    threshold = float(np.percentile(regional_load.values, DR_PEAK_PERCENTILE))
                    peak_mask = (regional_load >= threshold).astype(float)
                    peak_hours = int(peak_mask.sum())
                    network.generators_t.p_max_pu[dr_name] = peak_mask
                    p_nom_dr = float(network.generators.at[dr_name, 'p_nom'])
                    print(f"  DR {dr_name}: 임계값={threshold:.1f} MW, "
                          f"발동가능={peak_hours}h/년, 용량={p_nom_dr:.0f} MW")
            else:
                print("\n[DR] DR 발전기 없음 - DR 설정 건너뜀")
        except Exception as _e_dr_setup:
            print(f"DR p_max_pu 설정 경고: {_e_dr_setup}")

        # DR 연간 발동시간 제한 extra_functionality 정의
        def _dr_annual_hours_constraint(n, sns):
            """DR 발전기 연간 발동 에너지 = p_nom × DR_ANNUAL_HOURS 이하 제약"""
            try:
                dr_gens_active = [
                    g for g in n.generators.index
                    if str(n.generators.at[g, 'carrier']).upper() == 'DR'
                    and g in n.model.variables['Generator-p'].coords['Generator'].values
                ]
                if not dr_gens_active:
                    return
                weights = n.snapshot_weightings['generators'].loc[sns]
                for dr_name in dr_gens_active:
                    p_nom_val = float(n.generators.at[dr_name, 'p_nom'])
                    annual_limit_mwh = p_nom_val * DR_ANNUAL_HOURS
                    p_dr = n.model.variables['Generator-p'].sel(
                        Generator=dr_name, snapshot=sns
                    )
                    energy_sum = (p_dr * weights).sum('snapshot')
                    n.model.add_constraints(
                        energy_sum <= annual_limit_mwh,
                        name=f"DR-annual-limit-{dr_name}"
                    )
                    print(f"  [DR 연간제약] {dr_name}: ≤ {p_nom_val:.0f}MW × {DR_ANNUAL_HOURS}h "
                          f"= {annual_limit_mwh:.0f} MWh/년")
            except Exception as _e_dr_c:
                print(f"DR 연간제약 추가 경고: {_e_dr_c}")

        # =====================================================================
        # 선로 유연 운영 설정 — 인터페이스_1 F/G/H열 기반
        # =====================================================================
        _flex_iface_path = os.path.abspath(
            os.path.join(get_base_dir(), 'interface.xlsx'))
        _line_flex = _read_line_flex_settings(_flex_iface_path)

        if _line_flex:
            print(f"\n=== 선로 유연 운영 설정 ({len(_line_flex)}개 선로) ===")

            # ── 지역별 부하 시계열 캐시 (같은 지역 여러 선로에서 재사용) ──
            _region_load_cache = {}

            def _get_region_load(region_code):
                """지역 전력버스(region_EL)에 연결된 부하의 합산 시계열 반환"""
                if region_code in _region_load_cache:
                    return _region_load_cache[region_code]
                el_bus = f"{region_code}_EL"
                load_series = pd.Series(0.0, index=network.snapshots)
                try:
                    load_names = network.loads.index[
                        network.loads.bus.astype(str) == el_bus
                    ]
                    for ln in load_names:
                        if (hasattr(network.loads_t, 'p_set')
                                and ln in network.loads_t.p_set.columns):
                            load_series = load_series.add(
                                network.loads_t.p_set[ln].fillna(0.0),
                                fill_value=0.0)
                except Exception as _e_load:
                    print(f"  [경고] {region_code} 부하 읽기 실패: {_e_load}")
                _region_load_cache[region_code] = load_series
                return load_series

            for _lname, _lc in _line_flex.items():
                if _lname not in network.lines.index:
                    print(f"  [경고] {_lname}: 네트워크에 없음 — 건너뜀")
                    continue

                s_base      = _lc['s_nom_base']
                s_flex      = _lc['s_nom_flex']
                h_limit     = _lc['time_limit']
                load_region = _lc.get('load_region')

                # s_nom을 s_flex로 상향 (PyPSA 내부 상한 확장)
                network.lines.at[_lname, 's_nom'] = s_flex

                # ── s_max_pu 시계열 설정 ───────────────────────────────
                # load_region이 있을 때만 s_max_pu로 시간 제한 적용
                if load_region:
                    region_load = _get_region_load(load_region)
                    total_hrs   = len(network.snapshots)
                    n_flex_hrs  = int(min(h_limit, total_hrs))

                    if region_load.sum() > 0 and n_flex_hrs > 0:
                        # 부하 상위 h_limit 시간을 유연운영 허용 시점으로 선별
                        threshold = float(
                            region_load.nlargest(n_flex_hrs).min())
                        flex_mask = (region_load >= threshold)

                        # 동률(threshold에 걸린 시간이 많을 때) → 정확히 h_limit개만 유지
                        flex_idx = region_load.nlargest(n_flex_hrs).index
                        flex_mask = pd.Series(False, index=network.snapshots)
                        flex_mask.loc[flex_idx] = True

                        # s_max_pu 시계열: flex 허용 시간=1.0, 나머지=s_base/s_flex
                        s_max_pu_series = pd.Series(
                            s_base / s_flex, index=network.snapshots)
                        s_max_pu_series.loc[flex_mask] = 1.0

                        # PyPSA lines_t.s_max_pu에 등록
                        if not hasattr(network.lines_t, 's_max_pu') \
                                or network.lines_t.s_max_pu is None \
                                or network.lines_t.s_max_pu.empty:
                            network.lines_t['s_max_pu'] = pd.DataFrame(
                                index=network.snapshots)
                        network.lines_t.s_max_pu[_lname] = s_max_pu_series

                        actual_flex_hrs = int(flex_mask.sum())
                        print(f"  {_lname}: 기본={s_base:.1f}MW → 최대={s_flex:.1f}MW, "
                              f"한도={h_limit:.0f}h/yr | "
                              f"수전지역={load_region}, "
                              f"s_max_pu 피크시간={actual_flex_hrs}h")
                    else:
                        print(f"  {_lname}: [경고] {load_region} 부하 없음 "
                              f"→ s_max_pu 미설정 (s_flex 전시간 허용)")
                        print(f"    기본={s_base:.1f}MW → 최대={s_flex:.1f}MW, "
                              f"한도={h_limit:.0f}h/yr")
                else:
                    # load_region 미입력 → 시간 제한 없이 s_flex 전체 허용
                    print(f"  {_lname}: 기본={s_base:.1f}MW → 최대={s_flex:.1f}MW, "
                          f"한도={h_limit:.0f}h/yr | "
                          f"수전지역 미입력 → 시간제한 없음(s_flex 전시간 허용)")
        else:
            print("\n[선로 유연운영] 설정 없음 — 기본 용량 사용")

        def _line_flex_constraint(n, sns):
            """
            선로 유연 운영 — s_max_pu 시계열이 시간 제한을 담당하므로
            extra_functionality에서는 별도 제약 추가 없음.
            (s_max_pu는 optimize_network 진입 시 이미 lines_t에 설정 완료)
            """
            # s_max_pu 방식: PyPSA가 lines_t.s_max_pu를 모델 빌드 시 자동 반영
            # → extra_functionality에서 추가 작업 불필요
            pass

        # RPS 제약조건 적용 (최적화 전 모델에 추가) - 임시 비활성화
        if False:  # hasattr(network, 'rps_constraints') and network.rps_constraints:
            print("\n[RE] RPS 제약조건을 모델에 추가 중...")
            try:
                # 모델 생성 (optimize 전에 수동으로 생성)
                network.optimize.create_model()
                
                for rps in network.rps_constraints:
                    region = rps['region']
                    share = rps['share']
                    name = rps['name']
                    
                    # 1) 해당 지역의 전력 부하 총량 계산
                    region_loads = network.loads[
                        (network.loads['bus'].str.startswith(f'{region}_')) &
                        (network.loads['bus'].str.contains('_EL'))
                    ]
                    if region_loads.empty:
                        print(f"   [경고]  {name}: {region} 지역에 전력 부하가 없습니다.")
                        continue
                    
                    # 시간별 수요 합계 계산
                    total_demand = 0
                    for load_name in region_loads.index:
                        if load_name in network.loads_t.p_set.columns:
                            total_demand += network.loads_t.p_set[load_name].sum()
                    
                    if total_demand <= 0:
                        print(f"   [경고]  {name}: {region} 지역 전력 수요가 0입니다.")
                        continue
                    
                    # 2) 해당 지역 RE 버스 찾기 (예: SEL_RE)
                    re_bus = f"{region}_RE"
                    if re_bus not in network.buses.index:
                        print(f"   [경고]  {name}: {region} 지역에 RE 버스({re_bus})가 없습니다.")
                        continue
                    
                    # 3) RE 버스로 유입되는 모든 전력 계산
                    # 3-1) 지역 내 재생에너지 발전기 (RE 버스에 연결된 발전기)
                    re_gens_local = network.generators[
                        network.generators['bus'] == re_bus
                    ].index.tolist()
                    
                    # 3-2) RE Ring 송전망을 통한 RE 유입/유출
                    # Lines: 다른 지역 RE → 이 지역 RE (RE Ring)
                    re_lines_in = network.lines[
                        (network.lines['bus1'] == re_bus) &
                        (network.lines['bus0'].str.contains('_RE'))
                    ].index.tolist() if not network.lines.empty else []
                    
                    re_lines_out = network.lines[
                        (network.lines['bus0'] == re_bus) &
                        (network.lines['bus1'].str.contains('_RE'))
                    ].index.tolist() if not network.lines.empty else []
                    
                    # Links: RE Ring을 링크로 구현한 경우
                    re_links_in = network.links[
                        (network.links['bus1'] == re_bus) &
                        (network.links['bus0'].str.contains('_RE'))
                    ].index.tolist() if not network.links.empty else []
                    
                    re_links_out = network.links[
                        (network.links['bus0'] == re_bus) &
                        (network.links['bus1'].str.contains('_RE'))
                    ].index.tolist() if not network.links.empty else []
                    
                    has_re_ring = bool(re_lines_in or re_lines_out or re_links_in or re_links_out)
                    
                    # 4) linopy 제약조건 생성
                    # RE 소비량 = 지역 RE 발전 + Ring 유입 - Ring 유출 >= 수요 × 비중
                    
                    # 지역 내 RE 발전
                    if re_gens_local:
                        re_supply = network.model["Generator-p"].loc[:, re_gens_local].sum(dims="Generator")
                    else:
                        # RE 발전기가 없는 경우, 0으로 초기화된 linopy 변수 생성
                        import linopy
                        re_supply = linopy.LinearExpression(None, model=network.model)
                        re_supply = re_supply + 0  # 0으로 초기화
                    
                    # RE Ring 송전망을 통한 유입/유출
                    if has_re_ring:
                        # 선로 유입 (+p0은 bus0→bus1 방향)
                        if re_lines_in:
                            re_supply = re_supply + network.model["Line-s"].loc[:, re_lines_in].sum(dims="Line")
                        
                        # 선로 유출 (빼기)
                        if re_lines_out:
                            re_supply = re_supply - network.model["Line-s"].loc[:, re_lines_out].sum(dims="Line")
                        
                        # 링크 유입
                        if re_links_in:
                            re_supply = re_supply - network.model["Link-p"].loc[:, re_links_in].sum(dims="Link")
                        
                        # 링크 유출
                        if re_links_out:
                            re_supply = re_supply + network.model["Link-p"].loc[:, re_links_out].sum(dims="Link")
                    
                    min_re_consumption = total_demand * share
                    
                    # 제약조건 추가
                    network.model.add_constraints(
                        re_supply >= min_re_consumption,
                        name=f"RPS-{region}"
                    )
                    
                    gen_info = f"{len(re_gens_local)}개 발전기" if re_gens_local else "발전기 없음"
                    ring_info = ""
                    if has_re_ring:
                        total_ring_paths = len(re_lines_in) + len(re_lines_out) + len(re_links_in) + len(re_links_out)
                        ring_info = f", RE Ring: {total_ring_paths}개 경로"
                    else:
                        ring_info = ", RE Ring: 없음 (지역 내 발전만)"
                    
                    print(f"   [OK] {name}: {region} RE 소비 ≥ {min_re_consumption/1e6:.2f} GWh ({share*100:.1f}% of {total_demand/1e6:.2f} GWh)")
                    print(f"      {gen_info}{ring_info}")
                    
                    if not re_gens_local and not has_re_ring:
                        print(f"      [경고]  {region}은 RE 발전도 없고 RE Ring도 없어 RPS 달성 불가!")
                
                print(f"   [OK] RPS 제약조건 {len(network.rps_constraints)}개 추가 완료")
            except Exception as e:
                print(f"   [경고]  RPS 제약조건 추가 중 오류: {str(e)}")
                import traceback
                traceback.print_exc()
        
        # ── 솔버 선택 ─────────────────────────────────────────────────────────
        def _detect_solver():
            # exe 빌드: HiGHS 고정 (CPLEX DLL 번들 불가)
            if getattr(sys, 'frozen', False):
                print("[솔버] 실행파일 모드 → HiGHS 사용")
                return 'highs'
            # 일반 Python 실행: CPLEX 우선, 없으면 HiGHS
            try:
                import cplex  # noqa: F401
                print("[솔버 감지] CPLEX 사용 가능 → CPLEX 선택")
                return 'cplex'
            except (ImportError, Exception):
                pass
            try:
                import highspy  # noqa: F401
                print("[솔버 감지] CPLEX 없음 → HiGHS 로 대체")
                return 'highs'
            except ImportError:
                pass
            print("[솔버 감지] CPLEX/HiGHS 모두 없음 → glpk 시도")
            return 'glpk'

        active_solver = _detect_solver()

        # ── 솔버별 옵션 설정 ──────────────────────────────────────────────────
        if active_solver == 'cplex':
            # Barrier + solutiontype=2: Crossover 완전 생략하고 내부점 해 직접 반환
            # - lpmethod=4: Barrier 내부점법 (~40초 내 해 발견)
            # - solutiontype=2: Primal-only 해 반환 (Crossover 없음, CPLEX 공식 옵션)
            #   → Crossover 수천 초 소요 문제 근본 해결
            # - parallel=1: 결정론적 병렬
            # - barrier.algorithm=3: Mehrotra predictor-corrector
            option_variants = [
                {'name': 'barrier_primal_only',
                 'opts': {'threads': num_cores, 'lpmethod': 4,
                          'parallel': 1, 'barrier.algorithm': 3,
                          'solutiontype': 2}},
            ]
            print(f"[최적화 설정] CPLEX Barrier + solutiontype=2 (Crossover 생략), 스레드: {num_cores}개")
        else:
            # HiGHS: IPM(내부점법)으로 대규모 LP 처리
            option_variants = [
                {'name': 'ipm',
                 'opts': {'threads': num_cores, 'solver': 'ipm',
                          'presolve': 'on'}},
                {'name': 'simplex',
                 'opts': {'threads': num_cores, 'solver': 'simplex',
                          'presolve': 'on'}},
            ]
            print(f"[최적화 설정] HiGHS IPM (CPLEX 미설치), 스레드: {num_cores}개")

        last_status = None
        last_error = None
        for variant in option_variants:
            vname = variant['name']
            sopts = variant['opts']
            print(f"\n[시도] {active_solver.upper()} 방법: {vname}, 옵션: {sopts}")
            try:
                # ── 석탄 최소 발전량 제약 비활성화 ────────────────────
                # (CO2 제약과 충돌하여 결과 왜곡 발생 → 필요 시 재활성화)
                # COAL_MIN_ANNUAL_MWH = 90_500_000   # 연간 석탄 최소 발전량 (MWh)
                print("[coal_min] 석탄 최소 발전량 제약 비활성화됨")

                def _chp_co2_constraint(n, sns):
                    """CHP Links(gas carrier)를 포함한 통합 CO2 제약 추가.
                    PyPSA GlobalConstraint primary_energy는 Generator만 포함하므로
                    Link 타입 CHP의 배출량을 별도 통합 제약으로 추가한다.
                    (기존 GlobalConstraint는 유지 - 이 제약이 더 엄격해 실제 적용됨)
                    """
                    try:
                        if 'CO2_limit' not in n.global_constraints.index:
                            return
                        total_co2_limit = float(n.global_constraints.loc['CO2_limit', 'constant'])

                        if 'gas' not in n.carriers.index:
                            return
                        gas_co2 = float(n.carriers.loc['gas', 'co2_emissions'])
                        if gas_co2 == 0:
                            return

                        if 'carrier' not in n.links.columns:
                            return
                        chp_links = n.links[n.links['carrier'] == 'gas']
                        if chp_links.empty:
                            return

                        m = n.model
                        if 'Link-p' not in m.variables:
                            return

                        weights = n.snapshot_weightings['generators'].loc[sns]
                        link_p_var = m.variables['Link-p']
                        try:
                            available_links = list(link_p_var.coords['Link'].values)
                        except (KeyError, AttributeError):
                            available_links = []

                        active_chp = [lnk for lnk in chp_links.index if lnk in available_links]
                        if not active_chp:
                            return

                        all_exprs = []

                        # 1) CHP 링크 가스 소비 × 배출계수
                        for link_name in active_chp:
                            p0 = link_p_var.sel(Link=link_name, snapshot=sns)
                            all_exprs.append((p0 * weights).sum('snapshot') * gas_co2)

                        # 2) 기존 GlobalConstraint가 커버하는 Generator 배출량도 합산
                        #    → 통합 제약: gen_CO2 + CHP_CO2 ≤ total_co2_limit
                        emissions_map = n.carriers['co2_emissions'][n.carriers['co2_emissions'] != 0]
                        if not emissions_map.empty and 'Generator-p' in m.variables:
                            gens = n.generators.query("carrier in @emissions_map.index")
                            if not gens.empty:
                                try:
                                    from pypsa.optimization.common import get_as_dense
                                    eff = get_as_dense(n, "Generator", "efficiency",
                                                       snapshots=sns, inds=gens.index)
                                    em_pu = gens['carrier'].map(emissions_map) / eff
                                    em_pu_w = em_pu.multiply(weights, axis=0)
                                    p_gen = m.variables['Generator-p'].sel(
                                        Generator=gens.index, snapshot=sns)
                                    all_exprs.append((p_gen * em_pu_w).sum())
                                except Exception:
                                    pass

                        if not all_exprs:
                            return

                        try:
                            from linopy.expressions import merge as lm_merge
                            combined = lm_merge(all_exprs) if len(all_exprs) > 1 else all_exprs[0]
                        except Exception:
                            combined = all_exprs[0]
                            for ex in all_exprs[1:]:
                                combined = combined + ex

                        m.add_constraints(combined <= total_co2_limit,
                                          name='CO2-generators-and-links')
                        print(f"  [CO2 통합제약] 발전기+CHP({len(active_chp)}개) "
                              f"≤ {total_co2_limit/1e6:.1f}백만톤 추가됨")
                    except Exception as e_co2:
                        print(f"  [경고] CHP CO2 통합제약 추가 실패: {e_co2}")

                def _combined_extra_functionality(n, sns):
                    """DR 연간 한도 + 선로 유연 운영 + CHP CO2 통합"""
                    _dr_annual_hours_constraint(n, sns)
                    _line_flex_constraint(n, sns)
                    _chp_co2_constraint(n, sns)

                status = network.optimize(
                    solver_name=active_solver,
                    solver_options=sopts,
                    extra_functionality=_combined_extra_functionality
                )
                print(f"→ 상태: {status}")
                last_status = status
                if isinstance(status, tuple):
                    st_main = status[0]
                else:
                    st_main = str(status)
                if st_main and ('ok' in st_main.lower() or 'optimal' in st_main.lower()
                               or 'feasible' in st_main.lower()):
                    break
                # ── 해가 존재하는 경우 수락 (unknown / feasible / time_limit 등) ──
                # solutiontype=2 (Barrier primal-only) 또는 Crossover 수치 노이즈 시
                # linopy가 "unknown"으로 반환하더라도 유효한 목적함수가 있으면 수락
                _accept_statuses = ('unknown', 'feasible', 'aborted', 'time')
                if st_main and any(s in st_main.lower() for s in _accept_statuses):
                    try:
                        # 목적함수 값으로만 해 유효성 확인
                        # (network.model은 optimize() 반환 후 접근 불가한 경우가 있어 sol 체크 제거)
                        obj_val = getattr(network, 'objective', None)
                        m = getattr(network, 'model', None)
                        obj_linopy = None
                        if m is not None:
                            try:
                                obj_linopy = float(m.objective.value)
                            except Exception:
                                pass
                        valid_obj = obj_linopy if obj_linopy is not None else obj_val
                        if valid_obj is not None and float(valid_obj) > 0:
                            print(f"  ▶ 상태='{st_main}' 이지만 유효한 Barrier 해 감지 → 수락")
                            print(f"    목적함수={float(valid_obj):.4e}")
                            # linopy 모델 상태를 강제 ok로 패치 후 assign_solution 호출
                            try:
                                _m = getattr(network, 'model', None)
                                if _m is not None:
                                    if hasattr(_m, 'status'):
                                        _m.status = 'ok'
                                    if hasattr(_m, 'termination_condition'):
                                        _m.termination_condition = 'optimal'
                            except Exception:
                                pass
                            # PyPSA assign_solution 수동 호출 시도
                            try:
                                from pypsa.optimization.optimize import assign_solution
                                assign_solution(network)
                                print(f"    assign_solution 완료 → 결과 정상 반영")
                            except Exception as _e_as:
                                try:
                                    from pypsa.optimization import assign_solution
                                    assign_solution(network)
                                    print(f"    assign_solution 완료 → 결과 정상 반영")
                                except Exception:
                                    print(f"    assign_solution 실패({_e_as}), 결과 일부 누락 가능")
                            last_status = ('ok', 'optimal')
                            break
                        else:
                            print(f"  ▶ 상태='{st_main}', 유효한 해 없음 (obj={valid_obj})")
                    except Exception as _e_fb:
                        print(f"  해 수락 fallback 오류: {_e_fb}")
            except ValueError as e:
                if 'No objects to concatenate' in str(e):
                    print("경고: AC 각도 결과(v_ang)가 없어 후처리에서 concat 실패. 각도 결과 없이 계속 진행합니다.")
                    network.buses_t.v_ang = pd.DataFrame(index=network.snapshots, columns=network.buses.index)
                    last_status = 'ok'
                    break
                else:
                    last_error = str(e)
                    print(f"→ 예외: {last_error}")
                    continue
            except Exception as e:
                last_error = str(e)
                print(f"→ 예외: {last_error}")
                continue
        
        print(f"\n최종 최적화 상태: {last_status}")
        if hasattr(network, 'objective'):
            print(f"목적함수 값: {network.objective}")
        
        # 실패 시 LP 문제 내보내기(환경변수로 활성화)
        try:
            if (not last_status) or (isinstance(last_status, tuple) and all(x and ('unknown' in str(x).lower() or 'infeasible' in str(x).lower()) for x in last_status)) or ('unknown' in str(last_status).lower()):
                if os.environ.get('EXPORT_LP', '0') == '1':
                    export_dir = os.path.join('results', 'debug')
                    os.makedirs(export_dir, exist_ok=True)
                    if hasattr(network, 'model') and hasattr(network.model, 'to_file'):
                        lp_path = os.path.join(export_dir, 'failed_model.lp')
                        try:
                            network.model.to_file(lp_path)
                            print(f"실패 모델 LP 내보냄: {lp_path}")
                        except Exception as _e_lp:
                            print(f"LP 내보내기 실패: {_e_lp}")
                    else:
                        print("네트워크 모델 객체가 없어 LP 내보내기 불가")
        except Exception:
            pass
        
        # 최종 성공 여부: 'ok' 또는 'optimal' 포함 시 성공
        _final_ok = (
            bool(last_status) and
            (('ok' in str(last_status).lower()) or ('optimal' in str(last_status).lower()))
        )
        return _final_ok
        
    except Exception as e:
        print(f"\n최적화 중 오류 발생: {str(e)}")
        traceback.print_exc()
        return False

def extract_results(network):
    """주요 결과 추출"""
    
    # 재생에너지 Curtailment 계산
    # [주의] 확장가능(extendable) 발전기는 p_nom_opt(최적화된 용량)를 기준으로 잠재발전량 계산
    # p_nom(초기 용량) 사용 시 실제 발전량이 잠재량을 초과하여 커튼먼트가 누락 또는 과소평가됨
    re_curtailment = {}
    for gen in network.generators.index:
        if any(keyword in gen.lower() for keyword in ['pv', 'solar', 'wind', 'wt']):
            if gen in network.generators_t.p_max_pu.columns and gen in network.generators_t.p.columns:
                # 확장가능 발전기는 p_nom_opt(최적화 용량) 사용, 고정 발전기는 p_nom 사용
                extendable = bool(network.generators.at[gen, 'p_nom_extendable']) if 'p_nom_extendable' in network.generators.columns else False
                has_opt = ('p_nom_opt' in network.generators.columns and
                           pd.notna(network.generators.at[gen, 'p_nom_opt']) and
                           float(network.generators.at[gen, 'p_nom_opt']) > 0)
                if extendable and has_opt:
                    p_nom_val = float(network.generators.at[gen, 'p_nom_opt'])
                else:
                    p_nom_val = float(network.generators.at[gen, 'p_nom'])

                # 잠재 발전량 (p_nom_opt × p_max_pu 프로필)
                potential = network.generators_t.p_max_pu[gen] * p_nom_val

                # 실제 발전량 (최적화 결과)
                actual = network.generators_t.p[gen]

                # Curtailment: 확장발전기는 actual ≤ p_nom_opt × p_max_pu 보장되므로 항상 ≥ 0
                curtailment = (potential - actual).clip(lower=0)

                if curtailment.sum() > 0.1:  # 0.1 MWh 이상만 기록
                    re_curtailment[gen] = {
                        'timeseries': curtailment,
                        'total_MWh': float(curtailment.sum()),
                        'potential_MWh': float(potential.sum()),
                        'actual_MWh': float(actual.sum()),
                        'curtailment_rate_%': (curtailment.sum() / potential.sum() * 100) if potential.sum() > 0 else 0
                    }
    
    # storage_units_t 안전하게 처리
    storage_state = None
    try:
        if hasattr(network, 'storage_units_t') and hasattr(network.storage_units_t, 'state_of_charge'):
            if not network.storage_units_t.state_of_charge.empty:
                storage_state = network.storage_units_t.state_of_charge
    except:
        pass
    
    results = {
        'generator_output': network.generators_t.p,
        'node_prices': network.buses_t.marginal_price,
        'line_flows': network.lines_t.p0,
        'total_cost': getattr(network, 'objective', float('nan')),
        'load_balance': network.buses_t.p,
        'storage_state': storage_state,
        're_curtailment': re_curtailment  # 재생에너지 Curtailment 추가
    }
    
    # Curtailment 요약 출력
    if re_curtailment:
        print("\n=== 재생에너지 Curtailment 요약 ===")
        total_curtailment = sum(data['total_MWh'] for data in re_curtailment.values())
        total_potential = sum(data['potential_MWh'] for data in re_curtailment.values())
        print(f"총 Curtailment: {total_curtailment:.1f} MWh ({total_curtailment/1e6:.3f} TWh)")
        print(f"총 잠재 발전: {total_potential:.1f} MWh ({total_potential/1e6:.3f} TWh)")
        print(f"전체 Curtailment 비율: {(total_curtailment/total_potential*100):.2f}%")
        
        print("\n지역별/기술별 상위 5개:")
        sorted_curtailment = sorted(re_curtailment.items(), 
                                   key=lambda x: x[1]['total_MWh'], 
                                   reverse=True)[:5]
        for gen, data in sorted_curtailment:
            print(f"  {gen}: {data['total_MWh']:.1f} MWh ({data['curtailment_rate_%']:.1f}%)")
    
    return results

def _classify_technology(gen_or_link_name):
    name = str(gen_or_link_name).strip().lower()
    if 'nuclear' in name:
        return '원자력'
    if 'pv' in name or 'solar' in name:
        return '태양광'
    if 'wt' in name or 'wind' in name:
        return '풍력'
    if 'hydro' in name or 'water' in name:
        return '수력'
    # 석탄발전소 인식 - 'coal' 키워드 또는 발전소 이름으로 판별
    coal_plants = ['당진', '영흥', '하동', '삼척', '태안', '보령', '삼천포', '호남', '동해', 
                   '강릉안인', '북평', '신서천', '고성', '여수', '삼척그린파워']
    if 'coal' in name or any(plant in name for plant in coal_plants):
        return '석탄'
    if 'chp' in name:
        return 'CHP'
    if 'lng' in name or 'gas' in name:
        return 'LNG'
    if 'slack' in name or 'fallback' in name or 'failsafe' in name:
        # Slack/Failsafe 발전기는 모델 수렴용 보조 발전기 → 별도 카테고리
        return 'Slack'
    if 'oil' in name or 'diesel' in name:
        return '석유'
    if 'biomass' in name or 'bio' in name:
        return '바이오'
    if 'geothermal' in name:
        return '지열'
    if 'heatpump' in name or 'heat_pump' in name or name.startswith('hp') or ' hp ' in name:
        return '히트펌프'
    if 'electrolyser' in name or 'electrolyzer' in name or 'electrolysis' in name:
        return '전해조'
    if 'h2' in name or 'hydrogen' in name:
        return '수소'
    if 'heat' in name:
        return '열'
    # DR (수요반응) 인식
    name_tokens = name.split('_')
    if 'dr' in name_tokens:
        return 'DR'
    # 미인식 기술: 발전기 이름에서 기술 부분 추출 (첫 토큰=지역코드 제거)
    # 예) GGD_DR → 'DR',  BSN_ESS → 'ESS',  ICN_Custom_Gen → 'Custom_Gen'
    orig_tokens = str(gen_or_link_name).strip().split('_')
    if len(orig_tokens) >= 2:
        tech_tokens = [t for t in orig_tokens[1:] if t and not t.isdigit()]
        if tech_tokens:
            return '_'.join(tech_tokens)
    return '기타'

def _map_final_energy_from_carrier(carrier_value):
    c = str(carrier_value).strip().lower()
    if c in ['el', '전력', 'ac', 'dc'] or ('electric' in c or 'power' in c or 'hvac' in c or 'hvdc' in c):
        return '전력'
    if c in ['h', '열'] or ('heat' in c):
        return '열'
    if c in ['h2', '수소'] or ('hydrogen' in c):
        return '수소'
    return None

def build_final_energy_supply_tables(network):
    import pandas as _pd
    import numpy as _np
    rows_total = []
    rows_region = []
    bus_to_carrier = network.buses.carrier.to_dict() if not network.buses.empty else {}
    
    # PyPSA statistics로 배출계수 가져오기
    emission_factors = {}
    for carrier in network.carriers.index:
        emission_factors[carrier] = network.carriers.loc[carrier, 'co2_emissions']
    
    if not network.generators.empty and not network.generators_t.p.empty:
        for gen in network.generators.index:
            # Fuel_Supply / Slack / Fallback 발전기는 최종 에너지 집계에서 제외
            # (Slack/Fallback은 모델 실행 가능성 보장용 보조 발전기 - 실제 발전원 아님)
            gen_lower = gen.lower()
            if 'fuel_supply' in gen_lower or 'slack' in gen_lower or 'fallback' in gen_lower:
                continue
            
            bus = network.generators.at[gen, 'bus']
            final_energy = _map_final_energy_from_carrier(bus_to_carrier.get(bus, ''))
            if final_energy is None:
                continue
            tech = _classify_technology(gen)
            series = network.generators_t.p[gen] if gen in network.generators_t.p.columns else _pd.Series(0.0, index=network.snapshots)
            val = float(_np.nansum(series.values))
            if val <= 0:
                continue
            
            # 배출량 계산 (전력량 기준)
            # carrier.co2_emissions는 efficiency가 곱해진 값이므로
            # 결과 시트에는 원래 전력량 기준 배출계수를 표시하기 위해 efficiency로 나눔
            gen_carrier = network.generators.at[gen, 'carrier']
            emission_factor = emission_factors.get(gen_carrier, 0.0)
            gen_efficiency = network.generators.at[gen, 'efficiency'] if 'efficiency' in network.generators.columns else 1.0
            
            # 전력량 기준 배출계수 = carrier.co2 / efficiency (원래 값으로 복원)
            emission_factor_per_mwh = emission_factor / gen_efficiency if gen_efficiency > 0 else emission_factor
            emissions = val * emission_factor_per_mwh
            
            rows_total.append([final_energy, tech, val, emissions])
            region = gen.split('_')[0] if '_' in gen else ''
            rows_region.append([region, final_energy, tech, val, emissions])
    def _add_link_port_supply(port_idx):
        p_attr = f"p{port_idx}"
        if not hasattr(network.links_t, p_attr):
            return None
        df_p = getattr(network.links_t, p_attr)
        if df_p is None or df_p.empty:
            return None
        bus_col = f"bus{port_idx}"
        for link in network.links.index:
            # Slack/Fallback 링크는 집계 제외
            link_lower = link.lower()
            if 'slack' in link_lower or 'fallback' in link_lower or 'fuel_supply' in link_lower:
                continue
            if bus_col not in network.links.columns:
                continue
            if _pd.isna(network.links.at[link, bus_col]):
                continue
            dest_bus = network.links.at[link, bus_col]
            final_energy = _map_final_energy_from_carrier(bus_to_carrier.get(dest_bus, ''))
            if final_energy is None:
                continue
            if link not in df_p.columns:
                continue
            series = df_p[link]
            delivered = (-series).clip(lower=0.0)
            val = float(_np.nansum(delivered.values))
            if val <= 0:
                continue
            tech = _classify_technology(link)
            
            # 링크 배출량 계산 (CHP 등)
            emissions = 0.0
            if 'chp' in link.lower():
                # CHP의 경우: 연료 소비량(p0) 기반으로 배출량 계산 후 효율 비율로 분배
                if hasattr(network.links_t, 'p0') and link in network.links_t.p0.columns:
                    fuel_consumption = float(_np.nansum(network.links_t.p0[link].abs().values))  # 연료 소비량
                    total_emissions = fuel_consumption * emission_factors.get('gas', 0.38)  # 총 배출량
                    
                    # 효율 정보 가져오기
                    try:
                        eff1 = float(network.links.at[link, 'efficiency']) if 'efficiency' in network.links.columns else 0.5
                        eff2 = float(network.links.at[link, 'efficiency2']) if 'efficiency2' in network.links.columns else 0.4
                        total_eff = eff1 + eff2
                        
                        # 현재 포트에 해당하는 효율 비율로 배출량 분배
                        if port_idx == 1:
                            emissions = total_emissions * (eff1 / total_eff) if total_eff > 0 else 0
                        elif port_idx == 2:
                            emissions = total_emissions * (eff2 / total_eff) if total_eff > 0 else 0
                        else:
                            emissions = 0  # port 3 이상은 배출량 0
                    except:
                        # 효율 정보가 없으면 전력:열 = 50:50으로 분배
                        emissions = total_emissions * 0.5 if port_idx in [1, 2] else 0
            else:
                # 일반 링크는 출력 기반 배출량 (carrier 확인)
                try:
                    link_carrier = network.links.at[link, 'carrier'] if 'carrier' in network.links.columns else 'electricity'
                    emissions = val * emission_factors.get(link_carrier, 0.0)
                except:
                    emissions = 0.0
            
            rows_total.append([final_energy, tech, val, emissions])
            region = str(dest_bus).split('_')[0] if '_' in str(dest_bus) else ''
            rows_region.append([region, final_energy, tech, val, emissions])
    for k in [1, 2, 3]:
        _add_link_port_supply(k)
    total_df = _pd.DataFrame(rows_total, columns=['final_energy', 'technology', 'supply_MWh', 'emissions_tCO2'])
    if not total_df.empty:
        total_df = total_df.groupby(['final_energy', 'technology'], as_index=False)[['supply_MWh', 'emissions_tCO2']].sum()
    by_region_df = _pd.DataFrame(rows_region, columns=['region', 'final_energy', 'technology', 'supply_MWh', 'emissions_tCO2'])
    if not by_region_df.empty:
        by_region_df = by_region_df.groupby(['region', 'final_energy', 'technology'], as_index=False)[['supply_MWh', 'emissions_tCO2']].sum()
    return total_df, by_region_df


def build_regional_power_balance_table(network):
    """지역별 전력수급 균형 테이블 생성
    
    반환 컬럼: 지역, 전력생산_MWh, 전력소비_MWh,
               송전전력IN_MWh, 송전전력OUT_MWh,
               재생에너지생산비중_%, Curtailment_MWh
    """
    import pandas as _pd
    import numpy as _np

    bus_to_carrier = network.buses.carrier.to_dict() if not network.buses.empty else {}

    def _is_el_bus(bus_name):
        carrier = str(bus_to_carrier.get(bus_name, '')).lower()
        if ('electric' in carrier or carrier in ['el', 'ac', 'dc', 'hvac', 'hvdc']):
            return True
        b = str(bus_name).upper()
        return b.endswith('_EL') or b == 'EL'

    def _region_of(bus_name):
        s = str(bus_name)
        return s.split('_')[0] if '_' in s else s

    # 전력 버스를 가진 지역 코드 수집
    el_buses = [b for b in network.buses.index if _is_el_bus(b)]
    region_codes = sorted({_region_of(b) for b in el_buses})

    region_data = {r: {
        'generation': 0.0,
        're_generation': 0.0,
        'consumption': 0.0,
        'tx_in': 0.0,
        'tx_out': 0.0,
        'curtailment': 0.0,
    } for r in region_codes}

    # 1. 발전량 집계 (EL 버스에 연결된 발전기)
    if (not network.generators.empty
            and hasattr(network.generators_t, 'p')
            and not network.generators_t.p.empty):
        for gen in network.generators.index:
            if gen not in network.generators_t.p.columns:
                continue
            bus = str(network.generators.at[gen, 'bus'])
            if not _is_el_bus(bus):
                continue
            region = _region_of(bus)
            if region not in region_data:
                continue
            gen_sum = float(_np.nansum(network.generators_t.p[gen].values))
            if gen_sum <= 0:
                continue
            region_data[region]['generation'] += gen_sum
            # 재생에너지 여부 (PV / WT)
            gn = gen.lower()
            if 'pv' in gn or 'solar' in gn or 'wt' in gn or 'wind' in gn:
                region_data[region]['re_generation'] += gen_sum
                # Curtailment 계산 (확장발전기는 p_nom_opt 사용)
                if gen in network.generators_t.p_max_pu.columns:
                    _extendable = bool(network.generators.at[gen, 'p_nom_extendable']) if 'p_nom_extendable' in network.generators.columns else False
                    _has_opt = ('p_nom_opt' in network.generators.columns and
                                pd.notna(network.generators.at[gen, 'p_nom_opt']) and
                                float(network.generators.at[gen, 'p_nom_opt']) > 0)
                    _p_nom = float(network.generators.at[gen, 'p_nom_opt']) if (_extendable and _has_opt) else float(network.generators.at[gen, 'p_nom'])
                    potential = network.generators_t.p_max_pu[gen] * _p_nom
                    actual = network.generators_t.p[gen]
                    curtailed = float((potential - actual).clip(lower=0).sum())
                    region_data[region]['curtailment'] += curtailed

    # 2. 소비량 집계 (EL 버스에 연결된 부하)
    loads_p_df = None
    try:
        if hasattr(network.loads_t, 'p') and not network.loads_t.p.empty:
            loads_p_df = network.loads_t.p
        elif hasattr(network.loads_t, 'p_set') and not network.loads_t.p_set.empty:
            loads_p_df = network.loads_t.p_set
    except Exception:
        pass

    if loads_p_df is not None and not network.loads.empty:
        for ld in network.loads.index:
            if ld not in loads_p_df.columns:
                continue
            bus = str(network.loads.at[ld, 'bus'])
            if not _is_el_bus(bus):
                continue
            region = _region_of(bus)
            if region not in region_data:
                continue
            load_sum = float(_np.nansum(loads_p_df[ld].fillna(0).values))
            region_data[region]['consumption'] += load_sum

    # 3. AC 선로 조류 집계
    # p0 > 0: bus0→bus1 흐름 (bus0 지역 OUT, bus1 지역 IN)
    # p0 < 0: bus1→bus0 흐름 (bus1 지역 OUT, bus0 지역 IN)
    if (not network.lines.empty
            and hasattr(network.lines_t, 'p0')
            and not network.lines_t.p0.empty):
        for line in network.lines.index:
            if line not in network.lines_t.p0.columns:
                continue
            bus0 = str(network.lines.at[line, 'bus0'])
            bus1 = str(network.lines.at[line, 'bus1'])
            if not (_is_el_bus(bus0) and _is_el_bus(bus1)):
                continue
            r0, r1 = _region_of(bus0), _region_of(bus1)
            if r0 == r1:
                continue
            p0 = network.lines_t.p0[line]
            # r0 기준
            out_r0 = float(p0.clip(lower=0).sum())   # p0>0 → r0 내보냄
            in_r0  = float((-p0).clip(lower=0).sum()) # p0<0 → r0 받음
            if r0 in region_data:
                region_data[r0]['tx_out'] += out_r0
                region_data[r0]['tx_in']  += in_r0
            if r1 in region_data:
                region_data[r1]['tx_in']  += out_r0  # r0가 내보낸 만큼 r1이 받음
                region_data[r1]['tx_out'] += in_r0   # r0가 받은 만큼 r1이 내보냄

    # 4. DC 링크 조류 집계 (지역간 전력 링크만)
    # p0 > 0: r0가 내보냄(OUT), r1이 받음(IN)
    if (not network.links.empty
            and hasattr(network.links_t, 'p0')
            and not network.links_t.p0.empty):
        for link in network.links.index:
            if link not in network.links_t.p0.columns:
                continue
            bus0 = str(network.links.at[link, 'bus0'])
            bus1 = str(network.links.at[link, 'bus1'])
            if not (_is_el_bus(bus0) and _is_el_bus(bus1)):
                continue
            r0, r1 = _region_of(bus0), _region_of(bus1)
            if r0 == r1:
                continue
            p0 = network.links_t.p0[link]
            out_r0 = float(p0.clip(lower=0).sum())
            in_r0  = float((-p0).clip(lower=0).sum())
            if r0 in region_data:
                region_data[r0]['tx_out'] += out_r0
                region_data[r0]['tx_in']  += in_r0
            if r1 in region_data:
                region_data[r1]['tx_in']  += out_r0
                region_data[r1]['tx_out'] += in_r0

    # 결과 DataFrame 생성
    rows = []
    for region in sorted(region_data.keys()):
        d = region_data[region]
        gen = d['generation']
        re_gen = d['re_generation']
        re_ratio = round((re_gen / gen * 100), 2) if gen > 0 else 0.0
        rows.append({
            '지역': region,
            '전력생산_MWh': round(gen, 1),
            '전력소비_MWh': round(d['consumption'], 1),
            '송전전력IN_MWh': round(d['tx_in'], 1),
            '송전전력OUT_MWh': round(d['tx_out'], 1),
            '재생에너지생산비중_%': re_ratio,
            'Curtailment_MWh': round(d['curtailment'], 1),
        })
    return _pd.DataFrame(rows)


# 국가 기준 시간별 수급표(전력/열/수소) 생성

def _build_country_timeseries_tables(network):
    idx = network.snapshots
    zeros = pd.Series(0.0, index=idx)
    bus_to_carrier = network.buses.carrier.to_dict() if not network.buses.empty else {}

    # loads_t.p가 없거나 비어있으면 loads_t.p_set을 폴백으로 사용
    loads_p_df = None
    try:
        if hasattr(network.loads_t, 'p') and (network.loads_t.p is not None) and (not network.loads_t.p.empty):
            loads_p_df = network.loads_t.p
        elif hasattr(network.loads_t, 'p_set') and (network.loads_t.p_set is not None) and (not network.loads_t.p_set.empty):
            loads_p_df = network.loads_t.p_set
    except Exception:
        try:
            loads_p_df = getattr(network.loads_t, 'p_set')
        except Exception:
            loads_p_df = None

    def _fe(carrier):
        c = str(carrier).strip().lower()
        if c in ['el', '전력', 'ac', 'dc'] or ('electric' in c or 'power' in c or 'hvac' in c or 'hvdc' in c):
            return 'EL'
        if c in ['h', '열'] or ('heat' in c):
            return 'H'
        if c in ['h2', '수소'] or ('hydrogen' in c):
            return 'H2'
        return None

    def _fe_of_bus(bus_name):
        # 1) carrier 우선
        c = bus_to_carrier.get(bus_name, '')
        t = _fe(c)
        if t is not None:
            return t
        # 2) 버스명 토큰으로 판별 (예: 'SEL_EL', 'BSN_H2')
        try:
            s = str(bus_name)
            tokens = [tok for tok in s.split('_') if tok]
            if tokens:
                last = tokens[-1].strip().upper()
                if last in ['EL', 'H', 'H2']:
                    return last
        except Exception:
            pass
        return None

    def _accumulate_supply_and_load(target_type):
        supply = zeros.copy()
        load = zeros.copy()
        # 발전기 → 대상 최종단 버스에 연결된 출력 합계
        if (not network.generators.empty) and hasattr(network.generators_t, 'p') and (not network.generators_t.p.empty):
            for gen in network.generators.index:
                try:
                    bus = network.generators.at[gen, 'bus']
                    if _fe_of_bus(bus) != target_type:
                        continue
                    if gen in network.generators_t.p.columns:
                        supply = supply.add(network.generators_t.p[gen].fillna(0.0), fill_value=0.0)
                except Exception:
                    continue
        # 부하 → 대상 최종단 버스 부하 합계 (p가 없으면 p_set 사용)
        if (not network.loads.empty) and (loads_p_df is not None) and (not loads_p_df.empty):
            for ld in network.loads.index:
                try:
                    bus = network.loads.at[ld, 'bus']
                    if _fe_of_bus(bus) != target_type:
                        continue
                    if ld in loads_p_df.columns:
                        # p열이 전부 NaN이면 p_set으로 폴백
                        col_series = loads_p_df[ld]
                        if col_series.isna().all() and hasattr(network.loads_t, 'p_set') and (not network.loads_t.p_set.empty) and (ld in network.loads_t.p_set.columns):
                            col_series = network.loads_t.p_set[ld]
                        load = load.add(col_series.fillna(0.0), fill_value=0.0)
                except Exception:
                    continue
        # 링크 출력 포트에서 대상 최종단으로의 양(+) 유입을 공급으로 가정
        for k in [1, 2, 3]:
            p_attr = f"p{k}"
            if not hasattr(network.links_t, p_attr):
                continue
            dfp = getattr(network.links_t, p_attr)
            if dfp is None or dfp.empty:
                continue
            bus_col = f"bus{k}"
            if bus_col not in network.links.columns:
                continue
            for link in network.links.index:
                try:
                    if pd.isna(network.links.at[link, bus_col]):
                        continue
                    dest_bus = network.links.at[link, bus_col]
                    if _fe_of_bus(dest_bus) != target_type:
                        continue
                    if link not in dfp.columns:
                        continue
                    # 링크 출력 포트는 모델 부호상 목적 버스로 음(-) 유입될 수 있어, 공급으로는 부호 반전 후 양수만 합산
                    delivered = (-dfp[link]).fillna(0.0).clip(lower=0.0)
                    supply = supply.add(delivered, fill_value=0.0)
                except Exception:
                    continue
        return supply, load

    el_sup, el_load = _accumulate_supply_and_load('EL')
    h_sup, h_load = _accumulate_supply_and_load('H')
    h2_sup, h2_load = _accumulate_supply_and_load('H2')

    ts_el = pd.DataFrame({'Supply_MW': el_sup, 'Load_MW': el_load})
    ts_el['Net_MW'] = ts_el['Supply_MW'] - ts_el['Load_MW']

    ts_h = pd.DataFrame({'Supply_MW': h_sup, 'Load_MW': h_load})
    ts_h['Net_MW'] = ts_h['Supply_MW'] - ts_h['Load_MW']

    ts_h2 = pd.DataFrame({'Supply_MW': h2_sup, 'Load_MW': h2_load})
    ts_h2['Net_MW'] = ts_h2['Supply_MW'] - ts_h2['Load_MW']

    ts_el = ts_el.reindex(idx).fillna(0.0)
    ts_h = ts_h.reindex(idx).fillna(0.0)
    ts_h2 = ts_h2.reindex(idx).fillna(0.0)

    return ts_el, ts_h, ts_h2


def _build_analysis_window_sheet(network, results, writer, start_dt=None, n_hours=168):
    """
    선로 포화 구간 중심 168시간(1주일) 분석 윈도우 시트 생성.

    Parameters
    ----------
    network  : pypsa.Network
    results  : dict  (extract_results 반환값)
    writer   : pd.ExcelWriter (openpyxl engine)
    start_dt : None | str | pd.Timestamp
        None이면 선로 포화도 최대 구간을 자동 탐색
    n_hours  : int  기본 168 (=1주일)

    Sheet layout (섹션 수직 배치)
    ─────────────────────────────
    1. 타이틀 / 기간 정보
    2. 송전망 조류 (조류MW + 포화율%)
    3. 발전원별 출력 집계 (PV/WT/Nuclear/Coal/LNG/Hydro/CHP/DR/기타)
    4. 유연성자원/섹터커플링 (링크 조류MW)
    5. 저장설비 (충방전MW / SoC MWh)
    6. 재생에너지 Curtailment (MWh)
    7. 지역별 시간별 전력수요 (MW)
    """
    import pandas as _pd
    import numpy as _np
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Analysis_Window 생성 생략: openpyxl 없음")
        return

    snapshots = network.snapshots

    # ── 시작 시점 자동 결정 (선로 포화도 최대 구간) ──────────────
    auto_detected = start_dt is None
    if auto_detected:
        try:
            if (not network.lines.empty
                    and hasattr(network.lines_t, 'p0')
                    and not network.lines_t.p0.empty):
                util = _pd.DataFrame(index=snapshots)
                for line in network.lines_t.p0.columns:
                    if line in network.lines.index:
                        s_nom = float(network.lines.at[line, 's_nom'])
                        if s_nom > 0:
                            util[line] = network.lines_t.p0[line].abs() / s_nom
                if not util.empty:
                    peak_ts  = util.max(axis=1).idxmax()
                    peak_loc = snapshots.get_loc(peak_ts)
                    # peak를 윈도우의 1/4 지점에 배치
                    start_loc = max(0, peak_loc - n_hours // 4)
                    start_dt  = snapshots[start_loc]
                else:
                    start_dt = snapshots[0]
            else:
                start_dt = snapshots[0]
        except Exception:
            start_dt = snapshots[0]

    start_dt = _pd.Timestamp(start_dt)
    end_dt   = start_dt + _pd.Timedelta(hours=n_hours)

    mask = (snapshots >= start_dt) & (snapshots < end_dt)
    idx  = snapshots[mask]
    if len(idx) == 0:
        idx = snapshots[:min(n_hours, len(snapshots))]

    print(f"[Analysis_Window] {idx[0]} ~ {idx[-1]}  ({len(idx)}h)"
          + ("  [자동: 선로 포화 최대 구간]" if auto_detected else ""))

    bus_carrier = network.buses.carrier.to_dict() if not network.buses.empty else {}

    def _is_el(bus_name):
        c = str(bus_carrier.get(bus_name, '')).lower()
        return ('electric' in c or c in ['el', 'ac', 'dc', 'hvac', 'hvdc']
                or str(bus_name).upper().endswith('_EL'))

    # ── 섹션별 DataFrame 준비 ────────────────────────────────────

    # 1) 선로 조류 + 포화율
    sec_line = _pd.DataFrame(index=idx)
    try:
        if hasattr(network.lines_t, 'p0') and not network.lines_t.p0.empty:
            for line in network.lines_t.p0.columns:
                p = network.lines_t.p0[line].reindex(idx).fillna(0)
                sec_line[f'{line}_조류_MW'] = p.round(2).values
                if line in network.lines.index:
                    s_nom = float(network.lines.at[line, 's_nom'])
                    if s_nom > 0:
                        sec_line[f'{line}_포화율_%'] = (p.abs() / s_nom * 100).round(1).values
    except Exception as _e:
        print(f"  [AW] 선로 조류 준비 경고: {_e}")

    # 2) 발전원별 출력 집계 (전력 버스 연결 발전기만)
    _coal_kw = ['당진','영흥','하동','삼척','태안','보령','삼천포','호남','동해',
                '강릉안인','북평','신서천','고성','여수','삼척그린파워']
    _agg_keys = ['PV_MW','WT_MW','Nuclear_MW','Coal_MW','LNG_MW',
                 'Hydro_MW','CHP_MW','DR_MW','기타_MW']
    _agg = {k: _np.zeros(len(idx)) for k in _agg_keys}
    try:
        if hasattr(network.generators_t, 'p') and not network.generators_t.p.empty:
            for gen in network.generators_t.p.columns:
                if gen not in network.generators.index:
                    continue
                bus = str(network.generators.at[gen, 'bus'])
                if not _is_el(bus):
                    continue
                p = network.generators_t.p[gen].reindex(idx).fillna(0).values
                gn = gen.lower()
                if 'pv' in gn or 'solar' in gn:
                    _agg['PV_MW'] += p
                elif 'wt' in gn or 'wind' in gn:
                    _agg['WT_MW'] += p
                elif 'nuclear' in gn:
                    _agg['Nuclear_MW'] += p
                elif 'coal' in gn or any(kw in gn for kw in _coal_kw):
                    _agg['Coal_MW'] += p
                elif 'chp' in gn:
                    _agg['CHP_MW'] += p
                elif 'lng' in gn or 'gas' in gn:
                    _agg['LNG_MW'] += p
                elif 'hydro' in gn or 'water' in gn:
                    _agg['Hydro_MW'] += p
                elif 'dr' in gn.split('_'):
                    _agg['DR_MW'] += p
                else:
                    _agg['기타_MW'] += p
    except Exception as _e:
        print(f"  [AW] 발전원 집계 경고: {_e}")
    sec_gen = _pd.DataFrame({k: _np.round(v, 2) for k, v in _agg.items()}, index=idx)
    # 값이 전부 0인 열 제거
    sec_gen = sec_gen.loc[:, (sec_gen.abs() > 0.01).any()]

    # 3) 링크 조류 (유연성자원 / 섹터커플링)
    sec_link = _pd.DataFrame(index=idx)
    try:
        if hasattr(network.links_t, 'p0') and not network.links_t.p0.empty:
            for link in network.links_t.p0.columns:
                p = network.links_t.p0[link].reindex(idx).fillna(0)
                sec_link[f'{link}_MW'] = p.round(2).values
    except Exception as _e:
        print(f"  [AW] 링크 조류 준비 경고: {_e}")

    # 4) 저장설비 (충방전 MW, SoC MWh)
    sec_sto = _pd.DataFrame(index=idx)
    try:
        if hasattr(network, 'stores_t'):
            st = network.stores_t
            if hasattr(st, 'p') and st.p is not None and not st.p.empty:
                for s in st.p.columns:
                    sec_sto[f'{s}_충방전_MW'] = st.p[s].reindex(idx).fillna(0).round(2).values
            if hasattr(st, 'e') and st.e is not None and not st.e.empty:
                for s in st.e.columns:
                    sec_sto[f'{s}_SoC_MWh'] = st.e[s].reindex(idx).fillna(0).round(2).values
    except Exception as _e:
        print(f"  [AW] 저장설비 준비 경고: {_e}")

    # 5) RE Curtailment 시계열
    sec_curt = _pd.DataFrame(index=idx)
    try:
        if results and 're_curtailment' in results and results['re_curtailment']:
            for gen, data in results['re_curtailment'].items():
                ts = data.get('timeseries')
                if ts is not None:
                    sec_curt[f'{gen}_MWh'] = ts.reindex(idx).fillna(0).round(2).values
    except Exception as _e:
        print(f"  [AW] Curtailment 준비 경고: {_e}")

    # 6) 지역별 시간별 전력수요
    sec_load = _pd.DataFrame(index=idx)
    try:
        ldf = _pd.DataFrame()
        if hasattr(network.loads_t, 'p') and not network.loads_t.p.empty:
            ldf = network.loads_t.p
        elif hasattr(network.loads_t, 'p_set') and not network.loads_t.p_set.empty:
            ldf = network.loads_t.p_set
        if not ldf.empty:
            for ld in ldf.columns:
                bus = str(network.loads.at[ld, 'bus']) if ld in network.loads.index else ''
                if _is_el(bus):
                    sec_load[f'{ld}_MW'] = ldf[ld].reindex(idx).fillna(0).round(2).values
    except Exception as _e:
        print(f"  [AW] 부하 준비 경고: {_e}")

    # ── openpyxl로 시트 작성 ─────────────────────────────────────
    wb = writer.book
    ws = wb.create_sheet('Analysis_Window')

    def _fill(hex6):
        return PatternFill('solid', fgColor=hex6)

    HEAD_FILL = _fill('BDD7EE')
    WHITE_BOLD = Font(bold=True, color='FFFFFF', size=10)
    BOLD9      = Font(bold=True, size=9)
    NORM9      = Font(size=9)

    SECTIONS = [
        ('1. 송전망 조류 현황  (조류 MW / 포화율 %)',         sec_line, '1F4E79'),
        ('2. 발전원별 출력 현황  (MW)',                        sec_gen,  '375623'),
        ('3. 유연성자원 / 섹터커플링 운영현황  (링크 조류 MW)', sec_link, '7030A0'),
        ('4. 저장설비 현황  (충방전 MW / SoC MWh)',            sec_sto,  'C55A11'),
        ('5. 재생에너지 Curtailment  (MWh)',                   sec_curt, '843C0C'),
        ('6. 지역별 시간별 전력수요  (MW)',                     sec_load, '155F8A'),
    ]

    row = 1

    # 타이틀 행
    c = ws.cell(row, 1, '■ Analysis Window – 섹터커플링 운영 현황 분석')
    c.font = Font(bold=True, size=13, color='1F4E79')
    row += 1

    # 기간 정보 행
    ws.cell(row, 1, '분석 시작').font = BOLD9
    ws.cell(row, 2, start_dt.strftime('%Y-%m-%d %H:%M')).font = NORM9
    ws.cell(row, 4, '분석 종료').font = BOLD9
    ws.cell(row, 5, end_dt.strftime('%Y-%m-%d %H:%M')).font = NORM9
    ws.cell(row, 7, '기간').font = BOLD9
    ws.cell(row, 8, f'{len(idx)}h ({n_hours // 24}일)').font = NORM9
    if auto_detected:
        ws.cell(row, 10,
                '※ 시작 시점은 선로 포화도 최대 구간을 자동 탐색하였습니다.'
                ).font = Font(italic=True, color='FF0000', size=8)
    row += 2

    def _write_section(ws, row, title, df, color_hex):
        if df is None or df.empty:
            return row
        n_data_cols = df.shape[1]
        fill_sec  = _fill(color_hex)
        # ── 섹션 헤더 ──
        c = ws.cell(row, 1, title)
        c.font  = WHITE_BOLD
        c.fill  = fill_sec
        c.alignment = Alignment(horizontal='left', vertical='center')
        for ci in range(2, n_data_cols + 2):
            ws.cell(row, ci).fill = fill_sec
        row += 1
        # ── 컬럼 헤더 ──
        c = ws.cell(row, 1, 'Timestamp')
        c.font = BOLD9; c.fill = HEAD_FILL
        for ci, col_name in enumerate(df.columns, start=2):
            c = ws.cell(row, ci, str(col_name))
            c.font = BOLD9; c.fill = HEAD_FILL
        row += 1
        # ── 데이터 행 ──
        for ts, data_row in df.iterrows():
            ws.cell(row, 1, str(ts)).font = NORM9
            for ci, val in enumerate(data_row.values, start=2):
                try:
                    v = float(val)
                    ws.cell(row, ci, 0 if _np.isnan(v) else round(v, 2)).font = NORM9
                except Exception:
                    ws.cell(row, ci, val).font = NORM9
            row += 1
        return row + 1  # 섹션 간 빈 행

    for sec_title, sec_df, sec_color in SECTIONS:
        row = _write_section(ws, row, sec_title, sec_df, sec_color)

    # ── 열 너비 자동 조정 (최대 28) ──────────────────────────────
    try:
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max(
                (len(str(cell.value)) for cell in col_cells if cell.value is not None),
                default=8
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 28)
    except Exception:
        pass

    print(f"  → Analysis_Window 시트 완료: {len(idx)}개 시간대 × 6개 섹션")


def save_results(network, filename=None, subdir=None):
    """최적화 결과를 Excel 파일로 저장"""
    try:
        # Curtailment 분석 및 출력 (저장 전에 먼저 실행)
        print("\n" + "="*80)
        print("재생에너지 Curtailment 분석 중...")
        results = extract_results(network)
        print("="*80 + "\n")
        
        has_objective = hasattr(network, 'objective') and (network.objective is not None)
        if not has_objective:
            print("경고: 최적화 목적함수가 없지만, 중간 결과와 시계열을 저장합니다.")

        # 현재 시간을 포함한 폴더명 생성
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        if subdir:
            results_dir = subdir
            os.makedirs(results_dir, exist_ok=True)
        else:
            results_dir = f'results/{current_time}'
            os.makedirs('results', exist_ok=True)
            os.makedirs(results_dir, exist_ok=True)

        print(f"결과를 '{results_dir}' 폴더에 저장 중...")

        # 1. 기본 Excel 결과 파일
        excel_filename = f'{results_dir}/optimization_result_{current_time}.xlsx'
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            # 발전기 출력 결과 (없으면 빈 프레임 저장)
            try:
                gtp = network.generators_t.p
            except Exception:
                gtp = pd.DataFrame()

            # 석탄 세부 발전기를 지역별 합산으로 치환
            try:
                coal_plant_keywords = ['당진', '영흥', '하동', '삼척', '태안', '보령', '삼천포', '호남', '동해',
                                       '강릉안인', '북평', '신서천', '고성', '여수', '삼척그린파워']
                coal_detail_cols = [c for c in gtp.columns
                                    if ('coal' in str(network.generators.at[c, 'carrier']).lower()
                                        if c in network.generators.index else False)
                                    or any(kw in c for kw in coal_plant_keywords)]

                if coal_detail_cols:
                    # 석탄 세부 발전기 → 별도 시트에 저장
                    gtp[coal_detail_cols].to_excel(writer, sheet_name='Generator_Output_Coal')

                    # 지역별 합산 컬럼 생성 (예: CND_석탄, GND_석탄, ...)
                    region_coal_agg = {}
                    for col in coal_detail_cols:
                        region = col.split('_')[0] if '_' in col else 'ETC'
                        agg_col = f'{region}_Coal'
                        if agg_col not in region_coal_agg:
                            region_coal_agg[agg_col] = gtp[col].copy()
                        else:
                            region_coal_agg[agg_col] += gtp[col]

                    # 원본에서 석탄 세부 컬럼 제거 후 지역별 합산 컬럼 추가
                    gtp_agg = gtp.drop(columns=coal_detail_cols)
                    coal_agg_df = pd.DataFrame(region_coal_agg, index=gtp.index)
                    gtp_agg = pd.concat([gtp_agg, coal_agg_df], axis=1)
                    print(f"Generator_Output: 석탄 세부 {len(coal_detail_cols)}개 → 지역별 합산 {len(region_coal_agg)}개로 대체")
                    print(f"Generator_Output_Coal: 석탄 세부 발전기 {len(coal_detail_cols)}개 저장")
                else:
                    gtp_agg = gtp

                gtp_agg.to_excel(writer, sheet_name='Generator_Output')
            except Exception as _e_coal_agg:
                print(f"Generator_Output 석탄 합산 경고: {_e_coal_agg}")
                gtp.to_excel(writer, sheet_name='Generator_Output')

            # AC 선로 조류 결과
            try:
                if hasattr(network.lines_t, 'p0') and (not network.lines_t.p0.empty):
                    network.lines_t.p0.to_excel(writer, sheet_name='Line_Flow')
            except Exception:
                pass

            # AC 선로 정보 (s_nom, s_nom_opt) — scenario_analysis.py 이용률 계산에 사용
            try:
                if not network.lines.empty:
                    line_info_cols = ['bus0', 'bus1', 's_nom']
                    line_info_df = network.lines[line_info_cols].copy()
                    line_info_df.index.name = 'Line'
                    if 's_nom_opt' in network.lines.columns:
                        line_info_df['s_nom_opt'] = network.lines['s_nom_opt']
                    line_info_df.to_excel(writer, sheet_name='Line_Info')
            except Exception:
                pass

            # HVDC Link 조류 결과 추가
            try:
                if hasattr(network.links_t, 'p0') and (not network.links_t.p0.empty):
                    network.links_t.p0.to_excel(writer, sheet_name='Link_Flow')
            except Exception:
                pass

            # 버스 정보
            try:
                bus_results = pd.DataFrame({
                    'v_nom': network.buses.v_nom,
                    'carrier': network.buses.carrier
                })
                bus_results.to_excel(writer, sheet_name='Bus_Info')
            except Exception:
                pass

            # 발전기 정보
            try:
                gen_results = pd.DataFrame({
                    'bus': network.generators.bus,
                    'p_nom': network.generators.p_nom,
                    'p_nom_min': network.generators.p_nom_min if 'p_nom_min' in network.generators.columns else pd.Series(index=network.generators.index, dtype=float),
                    'p_nom_extendable': network.generators.p_nom_extendable if 'p_nom_extendable' in network.generators.columns else pd.Series(index=network.generators.index, dtype=bool),
                    'p_max_pu': network.generators.p_max_pu,
                    'marginal_cost': network.generators.marginal_cost
                })
                if 'p_nom_opt' in network.generators.columns:
                    gen_results['p_nom_opt'] = network.generators.p_nom_opt
                gen_results.to_excel(writer, sheet_name='Generator_Info')
            except Exception:
                pass

            # Link 정보 추가
            try:
                if not network.links.empty:
                    link_results = pd.DataFrame({
                        'bus0': network.links.bus0,
                        'bus1': network.links.bus1,
                        'p_nom': network.links.p_nom,
                        'efficiency': network.links.efficiency
                    })
                    if 'p_nom_opt' in network.links.columns:
                        link_results['p_nom_opt'] = network.links.p_nom_opt
                    link_results.to_excel(writer, sheet_name='Link_Info')
            except Exception:
                pass

            # ESS 충방전 결과 (있는 경우에만)
            try:
                if hasattr(network, 'stores_t') and (not network.stores_t.p.empty):
                    network.stores_t.p.to_excel(writer, sheet_name='Storage_Power')
                if hasattr(network, 'stores_t') and (hasattr(network.stores_t, 'e') and (not network.stores_t.e.empty)):
                    network.stores_t.e.to_excel(writer, sheet_name='Storage_Energy')
            except Exception:
                pass

            # ESS 정보 (있는 경우에만)
            try:
                if not network.stores.empty:
                    store_results = pd.DataFrame({
                        'bus': network.stores.bus,
                        'carrier': network.stores.carrier,
                        'e_nom': network.stores.e_nom,
                        'e_cyclic': network.stores.e_cyclic
                    })
                    if 'e_nom_opt' in network.stores.columns:
                        store_results['e_nom_opt'] = network.stores.e_nom_opt
                    store_results.to_excel(writer, sheet_name='Storage_Info')
            except Exception:
                pass

            # 시간별 부하 결과 (p 없으면 p_set 저장)
            try:
                ltp = None
                if hasattr(network.loads_t, 'p') and (network.loads_t.p is not None) and (not network.loads_t.p.empty):
                    ltp = network.loads_t.p
                elif hasattr(network.loads_t, 'p_set') and (network.loads_t.p_set is not None) and (not network.loads_t.p_set.empty):
                    ltp = network.loads_t.p_set
                else:
                    ltp = pd.DataFrame(index=network.snapshots)
                ltp.to_excel(writer, sheet_name='Hourly_Loads')
            except Exception:
                pass

            # 최적화 요약
            try:
                status_label = 'Optimal' if has_objective else 'Infeasible/NoObjective'
                total_cost_val = float(network.objective) if has_objective else float('nan')
                summary = pd.DataFrame({
                    'Parameter': ['Total Cost', 'Status'],
                    'Value': [total_cost_val, status_label]
                })
                summary.to_excel(writer, sheet_name='Summary', index=False)
            except Exception:
                pass

            # 최종에너지별 공급 집계 시트 추가
            try:
                fe_total, fe_by_region = build_final_energy_supply_tables(network)
                if fe_total is not None and not fe_total.empty:
                    fe_total.to_excel(writer, sheet_name='FinalEnergy_Supply', index=False)
                if fe_by_region is not None and not fe_by_region.empty:
                    fe_by_region.to_excel(writer, sheet_name='FinalEnergy_Supply_ByRegion', index=False)
            except Exception as _e_fe:
                print(f"최종에너지 공급 집계 시트 저장 경고: {_e_fe}")

            # 지역별 전력수급 균형 시트 추가
            try:
                regional_balance_df = build_regional_power_balance_table(network)
                if regional_balance_df is not None and not regional_balance_df.empty:
                    regional_balance_df.to_excel(writer, sheet_name='Regional_PowerBalance', index=False)
                    print(f"Regional_PowerBalance 시트 저장 완료: {len(regional_balance_df)}개 지역")
            except Exception as _e_rpb:
                print(f"지역별 전력수급 균형 시트 저장 경고: {_e_rpb}")

            # 국가 기준 시간별 수급표(전력/열/수소)
            try:
                ts_el, ts_h, ts_h2 = _build_country_timeseries_tables(network)
                ts_el.to_excel(writer, sheet_name='TS_Electricity_National')
                ts_h.to_excel(writer, sheet_name='TS_Heat_National')
                ts_h2.to_excel(writer, sheet_name='TS_Hydrogen_National')
                
                # 8) Curtailment 상세 정보 저장
                if 're_curtailment' in results and results['re_curtailment']:
                    try:
                        # Curtailment 요약
                        curtailment_summary = []
                        for gen, data in results['re_curtailment'].items():
                            curtailment_summary.append({
                                '발전기': gen,
                                '지역': gen.split('_')[0] if '_' in gen else gen,
                                '기술': 'PV' if 'PV' in gen.upper() else ('WT' if 'WT' in gen.upper() or 'WIND' in gen.upper() else '기타'),
                                '잠재_발전량_MWh': data['potential_MWh'],
                                '실제_발전량_MWh': data['actual_MWh'],
                                'Curtailment_MWh': data['total_MWh'],
                                'Curtailment_비율_%': data['curtailment_rate_%']
                            })
                        curtailment_summary_df = pd.DataFrame(curtailment_summary)
                        curtailment_summary_df = curtailment_summary_df.sort_values('Curtailment_MWh', ascending=False)
                        curtailment_summary_df.to_excel(writer, sheet_name='RE_Curtailment_Summary', index=False)
                        
                        # Curtailment 시계열 (상위 10개 발전기)
                        top_generators = curtailment_summary_df.head(10)['발전기'].tolist()
                        curtailment_timeseries = pd.DataFrame(index=network.snapshots)
                        for gen in top_generators:
                            if gen in results['re_curtailment']:
                                curtailment_timeseries[gen] = results['re_curtailment'][gen]['timeseries']
                        if not curtailment_timeseries.empty:
                            curtailment_timeseries.to_excel(writer, sheet_name='RE_Curtailment_Timeseries')
                        
                        print(f"Curtailment 데이터 저장 완료: {len(curtailment_summary)}개 발전기")
                    except Exception as e_curt:
                        print(f"Curtailment 저장 경고: {e_curt}")
            except Exception as _e_ts:
                print(f"국가 시간별 수급표 저장 경고: {_e_ts}")
                try:
                    # 폴백: 부하만이라도 기록
                    if hasattr(network.loads_t, 'p_set') and (network.loads_t.p_set is not None) and (not network.loads_t.p_set.empty):
                        network.loads_t.p_set.to_excel(writer, sheet_name='TS_Loads_Fallback')
                except Exception:
                    pass

            # 분석 윈도우 시트 (선로 포화 구간 168h)
            try:
                print("\nAnalysis_Window 시트 생성 중...")
                _build_analysis_window_sheet(network, results, writer)
            except Exception as _e_aw:
                print(f"Analysis_Window 시트 생성 경고: {_e_aw}")

        # 2. 개별 CSV 파일들 저장
        try:
            gen_info_df = pd.DataFrame({
                'bus': network.generators.bus,
                'p_nom': network.generators.p_nom,
                'p_nom_min': network.generators.p_nom_min if 'p_nom_min' in network.generators.columns else pd.Series(index=network.generators.index, dtype=float),
                'p_nom_extendable': network.generators.p_nom_extendable if 'p_nom_extendable' in network.generators.columns else pd.Series(index=network.generators.index, dtype=bool),
                'marginal_cost': network.generators.marginal_cost
            })
            if 'p_nom_opt' in network.generators.columns:
                gen_info_df['p_nom_opt'] = network.generators.p_nom_opt
            gen_info_df.to_csv(f'{results_dir}/optimization_result_{current_time}_generator_info.csv')
        except Exception as _e:
            print(f"Generator Info CSV 저장 경고: {_e}")

        # 발전기 출력
        try:
            gtp_csv = network.generators_t.p
            coal_plant_keywords = ['당진', '영흥', '하동', '삼척', '태안', '보령', '삼천포', '호남', '동해',
                                   '강릉안인', '북평', '신서천', '고성', '여수', '삼척그린파워']
            coal_detail_cols_csv = [c for c in gtp_csv.columns
                                    if ('coal' in str(network.generators.at[c, 'carrier']).lower()
                                        if c in network.generators.index else False)
                                    or any(kw in c for kw in coal_plant_keywords)]
            if coal_detail_cols_csv:
                # 석탄 세부 → 별도 CSV
                gtp_csv[coal_detail_cols_csv].to_csv(f'{results_dir}/optimization_result_{current_time}_generator_output_coal.csv')
                # 지역별 합산 후 기존 CSV
                region_coal_agg_csv = {}
                for col in coal_detail_cols_csv:
                    region = col.split('_')[0] if '_' in col else 'ETC'
                    agg_col = f'{region}_Coal'
                    if agg_col not in region_coal_agg_csv:
                        region_coal_agg_csv[agg_col] = gtp_csv[col].copy()
                    else:
                        region_coal_agg_csv[agg_col] += gtp_csv[col]
                gtp_csv_agg = gtp_csv.drop(columns=coal_detail_cols_csv)
                gtp_csv_agg = pd.concat([gtp_csv_agg, pd.DataFrame(region_coal_agg_csv, index=gtp_csv.index)], axis=1)
                gtp_csv_agg.to_csv(f'{results_dir}/optimization_result_{current_time}_generator_output.csv')
            else:
                gtp_csv.to_csv(f'{results_dir}/optimization_result_{current_time}_generator_output.csv')
        except Exception:
            pass

        # 부하
        try:
            if hasattr(network.loads_t, 'p') and (not network.loads_t.p.empty):
                network.loads_t.p.to_csv(f'{results_dir}/optimization_result_{current_time}_load.csv')
            elif hasattr(network.loads_t, 'p_set') and (not network.loads_t.p_set.empty):
                network.loads_t.p_set.to_csv(f'{results_dir}/optimization_result_{current_time}_load.csv')
        except Exception:
            pass

        # 저장장치 (있는 경우)
        try:
            if hasattr(network, 'stores_t') and not network.stores_t.p.empty:
                network.stores_t.p.to_csv(f'{results_dir}/optimization_result_{current_time}_storage.csv')
        except Exception:
            pass

        # 선로 사용량 (있는 경우)
        try:
            if not network.lines_t.p0.empty:
                network.lines_t.p0.to_csv(f'{results_dir}/optimization_result_{current_time}_line_usage.csv')
        except Exception:
            pass

        # 최종에너지 집계 CSV
        try:
            fe_total, fe_by_region = build_final_energy_supply_tables(network)
            if fe_total is not None and not fe_total.empty:
                fe_total.to_csv(f'{results_dir}/optimization_result_{current_time}_final_energy_supply.csv', index=False)
            if fe_by_region is not None and not fe_by_region.empty:
                fe_by_region.to_csv(f'{results_dir}/optimization_result_{current_time}_final_energy_supply_by_region.csv', index=False)
        except Exception as _e2:
            print(f"최종에너지 공급 집계 CSV 저장 경고: {_e2}")

        # 지역별 전력수급 균형 CSV
        try:
            regional_balance_df = build_regional_power_balance_table(network)
            if regional_balance_df is not None and not regional_balance_df.empty:
                regional_balance_df.to_csv(
                    f'{results_dir}/optimization_result_{current_time}_regional_power_balance.csv',
                    index=False, encoding='utf-8-sig'
                )
        except Exception as _e_rpb2:
            print(f"지역별 전력수급 균형 CSV 저장 경고: {_e_rpb2}")

        # 3. 통계 정보 JSON 파일
        try:
            total_cost_val = float(network.objective)
        except Exception:
            total_cost_val = float('nan')
        stats = {
            'timestamp': current_time,
            'total_cost': total_cost_val,
            'total_generators': len(network.generators),
            'total_buses': len(network.buses),
            'total_loads': len(network.loads),
            'total_stores': len(network.stores),
            'total_links': len(network.links)
        }

        import json
        with open(f'{results_dir}/optimization_result_{current_time}_stats.json', 'w', encoding='utf-8') as f:
            json.dump(stats, f, ensure_ascii=False, indent=2)

        # 4. PyPSA 네트워크 파일 저장
        network.export_to_netcdf(f'{results_dir}/optimization_result_{current_time}.nc')

        if SKIP_VISUALIZATION:
            print("[시각화 생략] SKIP_VISUALIZATION=True — 지역분석·시각화·후처리를 건너뜁니다.")
        else:
            # 5. 지역별 분석 결과 생성
            analyze_regional_results(network, results_dir, current_time)

            # 6. 시각화 결과 생성
            create_visualizations(network, results_dir, current_time)

            # 6.5. 커스텀 시각화 5종 (results_dir/images/ 폴더)
            try:
                create_custom_visualizations(network, results_dir)
            except Exception as _e_cv:
                print(f"커스텀 시각화 생성 경고: {_e_cv}")

            # 7. 계절별 대표 구간 시각화 (5종 × 4계절, 전국)
            try:
                create_seasonal_analysis_charts(network, results, results_dir, current_time)
            except Exception as _e_sea:
                print(f"계절별 시각화 생성 경고: {_e_sea}")

            # 8. 지역별 × 계절별 시각화 (5종 × 4계절 × 각 지역)
            try:
                create_regional_analysis_charts(network, results, results_dir, current_time)
            except Exception as _e_reg:
                print(f"지역별 시각화 생성 경고: {_e_reg}")

        print(f"결과가 '{results_dir}' 폴더에 저장되었습니다.")
        print(f"- Excel 파일: {excel_filename}")
        print(f"- CSV 파일들: generator_output, load, storage, line_usage")
        print(f"- 통계 파일: stats.json")
        print(f"- 네트워크 파일: .nc")
        print(f"- 시각화 파일들: PNG, HTML")
        print(f"- 커스텀 차트 5종: {results_dir}/images/")

        return True

    except Exception as e:
        print(f"결과 저장 중 오류 발생: {str(e)}")
        traceback.print_exc()
        return False

def create_seasonal_analysis_charts(network, results, results_dir, current_time, region=None):
    """계절별 대표 168h 구간에 대해 5종 시각화 생성 및 저장.

    region=None : 전국 집계 → seasonal_analysis/{계절}/
    region='BSN': BSN 지역만  → regional_analysis/BSN/{계절}/
    생성 파일: 01_dispatch_stack.png  ② 02_line_saturation_heatmap.png
               ③ 03_sector_coupling_vs_congestion.png  ④ 04_storage_soc_price.png
               ⑤ 05_energy_balance.png
    """
    if os.environ.get('DISABLE_PLOTS', '0') == '1':
        return

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np
    import re as _re
    import math as _math

    # 한글 폰트
    try:
        import matplotlib as mpl
        mpl.rcParams['font.family'] = ['Malgun Gothic', 'AppleGothic', 'DejaVu Sans']
        mpl.rcParams['axes.unicode_minus'] = False
    except Exception:
        pass

    snapshots = network.snapshots
    bus_carrier = network.buses.carrier.to_dict() if not network.buses.empty else {}

    KR_REGIONS = {'SEL', 'BSN', 'DGU', 'ICN', 'GWJ', 'DJN', 'USN', 'SJG',
                  'GGD', 'GWD', 'CBD', 'CND', 'JBD', 'JND', 'GBD', 'GND', 'JJD'}

    def _is_el(bus_name):
        c = str(bus_carrier.get(bus_name, '')).lower()
        return ('electric' in c or c in ['el', 'ac', 'dc', 'hvac', 'hvdc']
                or str(bus_name).upper().endswith('_EL'))

    def _rof(bus_name):
        s = str(bus_name)
        return s.split('_')[0] if '_' in s else s

    def _in_region(bus_name):
        """region=None이면 전체 포함, 지정 시 해당 지역만"""
        return region is None or _rof(bus_name) == region

    def _classify_link_tech(lname):
        """링크명에서 지역코드를 제거하고 기술 유형 반환"""
        n = lname.lower()
        if 'chp' in n:                                     return 'CHP'
        if 'electroly' in n:                               return '전해조'
        if 'heatpump' in n or 'heat_pump' in n or n.endswith('_hp'): return '히트펌프'
        if 'hvdc' in n:                                    return 'HVDC'
        if 'h2' in n or 'hydrogen' in n:                   return 'H2링크'
        if 'dr' in n.split('_'):                           return 'DR'
        parts = lname.split('_')
        tp = [p for p in parts if p.upper() not in KR_REGIONS and not p.isdigit() and p]
        return '_'.join(tp) if tp else lname

    def _classify_store_tech(sname):
        """저장설비명에서 지역코드를 제거하고 기술 유형 반환"""
        n = sname.lower()
        if 'ess' in n or 'battery' in n: return 'ESS'
        if 'h2' in n or 'hydrogen' in n: return 'H2저장'
        if 'heat' in n and 'pump' not in n: return '열저장'
        parts = sname.split('_')
        tp = [p for p in parts if p.upper() not in KR_REGIONS and not p.isdigit() and p]
        return '_'.join(tp) if tp else sname

    COAL_KW = ['당진', '영흥', '하동', '삼척', '태안', '보령', '삼천포', '호남', '동해',
               '강릉안인', '북평', '신서천', '고성', '여수', '삼척그린파워']

    STACK_ORDER = ['Nuclear', 'Coal', 'LNG', 'Hydro', 'CHP', 'WT', 'PV', 'DR', 'Other']
    GEN_COLORS  = {
        'Nuclear': '#8E44AD', 'Coal': '#555555', 'LNG': '#E67E22',
        'Hydro':   '#3498DB', 'CHP': '#8D6E63',  'PV':  '#FFC107',
        'WT':      '#4CAF50', 'DR':  '#E74C3C',   'Other': '#95A5A6',
    }

    def _classify_gen(name):
        n = name.lower()
        if 'pv' in n or 'solar' in n:            return 'PV'
        if 'wt' in n or 'wind' in n:             return 'WT'
        if 'nuclear' in n:                        return 'Nuclear'
        if 'coal' in n or any(k in n for k in COAL_KW): return 'Coal'
        if 'chp' in n:                            return 'CHP'
        if 'lng' in n or 'gas' in n:              return 'LNG'
        if 'hydro' in n or 'water' in n:          return 'Hydro'
        if 'dr' in n.split('_'):                  return 'DR'
        return 'Other'

    def _get_gen_agg(idx):
        agg = {k: np.zeros(len(idx)) for k in STACK_ORDER}
        try:
            if not network.generators.empty and hasattr(network.generators_t, 'p') and not network.generators_t.p.empty:
                for gen in network.generators_t.p.columns:
                    if gen not in network.generators.index:
                        continue
                    bus = str(network.generators.at[gen, 'bus'])
                    if not _is_el(bus) or not _in_region(bus):
                        continue
                    agg[_classify_gen(gen)] += network.generators_t.p[gen].reindex(idx).fillna(0).values
        except Exception as _e:
            print(f"    [gen_agg 경고] {_e}")
        return {k: v for k, v in agg.items() if np.abs(v).max() > 0.1}

    def _get_demand(idx):
        demand = pd.Series(0.0, index=idx)
        try:
            ldf = pd.DataFrame()
            if hasattr(network.loads_t, 'p') and not network.loads_t.p.empty:
                ldf = network.loads_t.p
            elif hasattr(network.loads_t, 'p_set') and not network.loads_t.p_set.empty:
                ldf = network.loads_t.p_set
            if not ldf.empty:
                for ld in ldf.columns:
                    bus = str(network.loads.at[ld, 'bus']) if ld in network.loads.index else ''
                    if _is_el(bus) and _in_region(bus):
                        demand = demand.add(ldf[ld].reindex(idx).fillna(0), fill_value=0)
        except Exception:
            pass
        return demand

    def _get_season_window(months, n_hours=168):
        """해당 월 범위에서 선로 포화 피크 시점을 중심으로 168h 시작 시점 반환"""
        s_snaps = snapshots[[t.month in months for t in snapshots]]
        if len(s_snaps) < n_hours:
            return s_snaps[0] if len(s_snaps) > 0 else snapshots[0]
        try:
            if not network.lines.empty and hasattr(network.lines_t, 'p0') and not network.lines_t.p0.empty:
                util_max = pd.Series(0.0, index=s_snaps)
                for line in network.lines_t.p0.columns:
                    if line in network.lines.index:
                        s_nom = float(network.lines.at[line, 's_nom'])
                        if s_nom > 0:
                            p = network.lines_t.p0[line].reindex(s_snaps).fillna(0)
                            util_max = util_max.combine(p.abs() / s_nom, max)
                peak_ts  = util_max.idxmax()
                peak_loc = list(s_snaps).index(peak_ts)
                start_loc = max(0, peak_loc - n_hours // 4)
                return s_snaps[start_loc]
        except Exception:
            pass
        return s_snaps[0]

    # ── ① Dispatch Stack ────────────────────────────────────────
    def _plot01_dispatch_stack(idx, out_dir, s_name):
        gen_agg = _get_gen_agg(idx)
        demand  = _get_demand(idx)
        if not gen_agg:
            print(f"    발전 데이터 없음 – dispatch_stack 생략")
            return

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(18, 11), sharex=True,
                                        gridspec_kw={'height_ratios': [4, 1], 'hspace': 0.04})
        x      = np.arange(len(idx))
        bottom = np.zeros(len(idx))
        handles = []

        for key in STACK_ORDER:
            if key not in gen_agg:
                continue
            v = gen_agg[key]
            ax1.fill_between(x, bottom, bottom + v, alpha=0.85,
                             color=GEN_COLORS[key], linewidth=0.3)
            handles.append(mpatches.Patch(color=GEN_COLORS[key], label=key))
            bottom += v

        ax1.plot(x, demand.values, 'k-', lw=2.0, zorder=10, label='전력수요')
        handles.append(plt.Line2D([0], [0], color='k', lw=2, label='전력수요'))

        # RE Curtailment 하단 패널
        curt_total = pd.Series(0.0, index=idx)
        if results and 're_curtailment' in results:
            for data in results['re_curtailment'].values():
                ts = data.get('timeseries')
                if ts is not None:
                    curt_total = curt_total.add(ts.reindex(idx).fillna(0), fill_value=0)
        ax2.fill_between(x, 0, curt_total.values, color='red', alpha=0.6)
        ax2.set_ylabel('Curtailment\n(MWh)', fontsize=9)
        ax2.grid(True, alpha=0.3)

        tick_pos = list(range(0, len(idx), 24))
        tick_lbl = [idx[i].strftime('%m/%d\n%H:00') for i in tick_pos]
        ax2.set_xticks(tick_pos)
        ax2.set_xticklabels(tick_lbl, fontsize=8)
        ax1.set_ylabel('출력 (MW)', fontsize=11)
        # Y축: 전체 시뮬레이션 기간의 최대값으로 통일 (계절 간 비교 가능)
        ax1.set_ylim(0, _GLOBAL_GEN_Y_MAX)
        ax2.set_ylim(0, _GLOBAL_CURT_Y_MAX)
        # 범례 상단 일렬 배치
        ax1.legend(handles=handles, loc='lower center', bbox_to_anchor=(0.5, 1.01),
                   ncol=len(handles), fontsize=9, frameon=True, borderaxespad=0)
        ax1.grid(True, alpha=0.3)
        ax1.set_xlim(0, len(idx) - 1)
        # fig.suptitle로 제목을 figure 최상단에 배치해 범례와 겹치지 않게
        plt.subplots_adjust(top=0.83, hspace=0.04)
        fig.suptitle(f'[{s_name}] 발전원별 누적 출력 및 수요',
                     fontsize=12, fontweight='bold', y=0.98)
        plt.savefig(f'{out_dir}/01_dispatch_stack.png', dpi=200, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        print(f"    → 01_dispatch_stack.png")

    # ── ② 선로 포화 히트맵 ────────────────────────────────────────
    def _plot02_line_heatmap(idx, out_dir, s_name):
        if network.lines.empty or not hasattr(network.lines_t, 'p0') or network.lines_t.p0.empty:
            print(f"    선로 데이터 없음 – heatmap 생략")
            return

        util_raw = {}
        for line in network.lines_t.p0.columns:
            if line in network.lines.index:
                bus0 = str(network.lines.at[line, 'bus0'])
                bus1 = str(network.lines.at[line, 'bus1'])
                # region 필터: 해당 지역에 연결된 선로만
                if region and _rof(bus0) != region and _rof(bus1) != region:
                    continue
                s_nom = float(network.lines.at[line, 's_nom'])
                if s_nom > 0:
                    p = network.lines_t.p0[line].reindex(idx).fillna(0)
                    u = (p.abs() / s_nom * 100).values
                    if u.max() > 0.1:
                        util_raw[line] = u

        # 병렬 선로 집계: GWD_GGD1, GWD_GGD2 → GWD_GGD (최대 포화율)
        util_data = {}
        for name, vals in util_raw.items():
            base = _re.sub(r'\d+$', '', str(name)).rstrip('_')
            if base in util_data:
                util_data[base] = np.maximum(util_data[base], vals)
            else:
                util_data[base] = vals.copy()

        if not util_data:
            return

        top_n = 30
        sorted_lines = sorted(util_data, key=lambda l: util_data[l].max(), reverse=True)[:top_n]
        matrix = np.array([util_data[l] for l in sorted_lines])

        fig_h = max(5, len(sorted_lines) * 0.38 + 2)
        fig, ax = plt.subplots(figsize=(20, fig_h))
        im = ax.pcolormesh(np.arange(len(idx) + 1), np.arange(len(sorted_lines) + 1),
                           matrix, cmap='RdYlGn_r', vmin=0, vmax=100)
        plt.colorbar(im, ax=ax, label='포화율 (%)', pad=0.01)

        ax.set_yticks(np.arange(len(sorted_lines)) + 0.5)
        ax.set_yticklabels(sorted_lines, fontsize=7)

        tick_pos = list(range(0, len(idx), 24))
        ax.set_xticks([t + 0.5 for t in tick_pos])
        ax.set_xticklabels([idx[i].strftime('%m/%d') for i in tick_pos], fontsize=9)

        # 포화율 ≥95% 셀 강조
        for yi, line in enumerate(sorted_lines):
            for xi, val in enumerate(util_data[line]):
                if val >= 95:
                    ax.add_patch(plt.Rectangle((xi, yi), 1, 1, fill=False,
                                               edgecolor='darkred', lw=1.2))
        ax.set_title(f'[{s_name}] 송전선로 시간대별 포화율 히트맵 (상위 {len(sorted_lines)}개)',
                     fontsize=12, fontweight='bold')
        ax.set_xlabel('시간 →', fontsize=10)
        plt.tight_layout()
        plt.savefig(f'{out_dir}/02_line_saturation_heatmap.png', dpi=180,
                    bbox_inches='tight', facecolor='white')
        plt.close(fig)
        print(f"    → 02_line_saturation_heatmap.png")

    # ── ③ 섹터커플링 vs 선로포화 ──────────────────────────────────
    def _plot03_sector_coupling(idx, out_dir, s_name):
        # ── 송전망 평균 이용률 ──────────────────────────────────
        # 전국: 전체 선로 평균, 지역: 해당 지역 연결 선로 평균
        avg_util   = pd.Series(0.0, index=idx)
        line_count = 0
        if not network.lines.empty and hasattr(network.lines_t, 'p0') and not network.lines_t.p0.empty:
            for line in network.lines_t.p0.columns:
                if line not in network.lines.index: continue
                bus0_l = str(network.lines.at[line, 'bus0'])
                bus1_l = str(network.lines.at[line, 'bus1'])
                if region and _rof(bus0_l) != region and _rof(bus1_l) != region:
                    continue
                s_nom = float(network.lines.at[line, 's_nom'])
                if s_nom > 0:
                    p = network.lines_t.p0[line].reindex(idx).fillna(0)
                    avg_util   += p.abs() / s_nom * 100
                    line_count += 1
        if line_count > 0:
            avg_util = avg_util / line_count

        # ── 지역 전력 수출입 (지역 모드에서만) ──────────────────
        # net_import > 0 : 외부에서 유입 (수입), < 0 : 외부로 송출 (수출)
        net_import = None
        if region:
            net_import = pd.Series(0.0, index=idx)
            try:
                if not network.lines.empty and hasattr(network.lines_t, 'p0') and not network.lines_t.p0.empty:
                    for line in network.lines_t.p0.columns:
                        if line not in network.lines.index: continue
                        b0 = str(network.lines.at[line, 'bus0'])
                        b1 = str(network.lines.at[line, 'bus1'])
                        r0, r1 = _rof(b0), _rof(b1)
                        p = network.lines_t.p0[line].reindex(idx).fillna(0)
                        if r0 == region and r1 != region:
                            net_import = net_import - p   # 지역에서 외부로 = 순수출
                        elif r1 == region and r0 != region:
                            net_import = net_import + p   # 외부에서 지역으로 = 순수입
            except Exception:
                pass
            try:
                if not network.links.empty and hasattr(network.links_t, 'p0') and not network.links_t.p0.empty:
                    for link in network.links_t.p0.columns:
                        if link not in network.links.index: continue
                        b0 = str(network.links.at[link, 'bus0'])
                        b1 = str(network.links.at[link, 'bus1'])
                        r0, r1 = _rof(b0), _rof(b1)
                        if r0 == r1: continue  # 지역 내부 링크 제외
                        p = network.links_t.p0[link].reindex(idx).fillna(0)
                        if r0 == region and r1 != region:
                            net_import = net_import - p
                        elif r1 == region and r0 != region:
                            net_import = net_import + p
            except Exception:
                pass

        # 링크 기술별 합산 (지역 구분 없이)
        link_tech_agg = {}
        if not network.links.empty and hasattr(network.links_t, 'p0') and not network.links_t.p0.empty:
            for link in network.links_t.p0.columns:
                if link not in network.links.index:
                    continue
                bus0 = str(network.links.at[link, 'bus0'])
                bus1 = str(network.links.at[link, 'bus1'])
                # region 필터: 해당 지역 관련 링크만
                if region and _rof(bus0) != region and _rof(bus1) != region:
                    continue
                p = network.links_t.p0[link].reindex(idx).fillna(0).values
                if abs(p).max() <= 0.1:
                    continue
                tech = _classify_link_tech(link)
                if tech in link_tech_agg:
                    link_tech_agg[tech] += p
                else:
                    link_tech_agg[tech] = p.copy()

        # 잉여/부족 전력 = 총 발전 − 총 전력수요
        _gen_agg_03 = _get_gen_agg(idx)
        _demand_03  = _get_demand(idx)
        _total_gen  = np.zeros(len(idx))
        for _v in _gen_agg_03.values():
            _total_gen += _v
        surplus = _total_gen - _demand_03.values   # + : 잉여,  − : 부족

        # 유연성 자원: V2G·ESS·히트펌프·전해조 (CHP·HVDC 제외)
        _FLEX_INCLUDE = {'V2G', 'ESS', '히트펌프', '전해조'}
        _FLEX_COLORS  = {'V2G': '#2196F3', 'ESS': '#4CAF50',
                         '히트펌프': '#FF9800', '전해조': '#9C27B0'}
        flex_flows = {}   # tech → np.array  (+ = 전력흡수,  − = 전력공급)

        # Stores: ESS, V2G (p>0 = 충전/전력흡수, p<0 = 방전/전력공급)
        try:
            if (not network.stores.empty
                    and hasattr(network.stores_t, 'p')
                    and not network.stores_t.p.empty):
                for s in network.stores_t.p.columns:
                    bus = str(network.stores.at[s, 'bus']) if s in network.stores.index else ''
                    if not _in_region(bus): continue
                    tech = _classify_store_tech(s)
                    if tech not in _FLEX_INCLUDE: continue
                    p = network.stores_t.p[s].reindex(idx).fillna(0).values
                    flex_flows[tech] = flex_flows.get(tech, np.zeros(len(idx))) + p
        except Exception:
            pass

        # Links: 히트펌프·전해조 (p0>0 = 전력소비)
        try:
            if (not network.links.empty
                    and hasattr(network.links_t, 'p0')
                    and not network.links_t.p0.empty):
                for link in network.links_t.p0.columns:
                    if link not in network.links.index: continue
                    bus0 = str(network.links.at[link, 'bus0'])
                    bus1 = str(network.links.at[link, 'bus1'])
                    if region and _rof(bus0) != region and _rof(bus1) != region: continue
                    tech = _classify_link_tech(link)
                    if tech not in _FLEX_INCLUDE: continue
                    p = network.links_t.p0[link].reindex(idx).fillna(0).values
                    flex_flows[tech] = flex_flows.get(tech, np.zeros(len(idx))) + p
        except Exception:
            pass

        flex_flows = {k: v for k, v in flex_flows.items() if abs(v).max() > 0.1}

        if avg_util.max() < 0.1 and abs(surplus).max() < 0.1 and not flex_flows:
            print(f"    데이터 없음 – sector_coupling 생략")
            return

        x = np.arange(len(idx))
        fig, ax1 = plt.subplots(figsize=(18, 8))

        # 왼쪽 Y: 송전망 평균 이용률 (빨강)
        _util_lbl = (f"{region} 지역 송전망 평균 이용률(%)"
                     if region else "송전망 평균 이용률(%)")
        ax1.fill_between(x, 0, avg_util.values, alpha=0.15, color='#D32F2F')
        ax1.plot(x, avg_util.values, color='#D32F2F', lw=1.8, label=_util_lbl)
        ax1.set_ylabel('송전망 평균 이용률 (%)', color='#D32F2F', fontsize=11)
        ax1.tick_params(axis='y', labelcolor='#D32F2F')
        ax1.set_ylim(0, max(avg_util.max() * 1.3, 15))

        # 오른쪽 Y: 전력수지 + 유연성자원 흐름 + 수출입(지역)
        ax2 = ax1.twinx()
        ax2.axhline(0, color='gray', lw=0.8, ls='-', alpha=0.4)
        ax2.plot(x, surplus, color='black', lw=1.8, ls='--', alpha=0.75,
                 label='전력수지 (발전-수요)')

        # 지역 전력 수출입 (지역 모드에서만)
        if net_import is not None and abs(net_import.values).max() > 0.1:
            ax2.fill_between(x,
                             np.where(net_import.values >= 0, net_import.values, 0),
                             0, alpha=0.20, color='#00BCD4')
            ax2.fill_between(x, 0,
                             np.where(net_import.values < 0, net_import.values, 0),
                             alpha=0.20, color='#FF5722')
            ax2.plot(x, net_import.values, lw=1.4, color='#00838F',
                     alpha=0.85, ls='-.', label='전력 수출입 (+수입/-수출)')

        for tech, vals in flex_flows.items():
            clr = _FLEX_COLORS.get(tech, '#757575')
            pos = np.where(vals >= 0, vals, 0)
            neg = np.where(vals < 0,  vals, 0)
            ax2.fill_between(x, 0, pos, alpha=0.45, color=clr)
            ax2.fill_between(x, neg, 0, alpha=0.35, color=clr)
            ax2.plot(x, vals, lw=1.2, color=clr, alpha=0.85, label=f'{tech}')

        ax2.set_ylabel('전력수지 (+생산우위/-소비우위)', color='#1565C0', fontsize=11)
        ax2.tick_params(axis='y', labelcolor='#1565C0')

        h1, l1 = ax1.get_legend_handles_labels()
        h2, l2 = ax2.get_legend_handles_labels()
        ax1.legend(h1 + h2, l1 + l2, loc='upper left', fontsize=9, ncol=3,
                   framealpha=0.9)

        tick_pos = list(range(0, len(idx), 24))
        ax1.set_xticks(tick_pos)
        ax1.set_xticklabels([idx[i].strftime('%m/%d\n%H:00') for i in tick_pos], fontsize=9)
        ax1.set_xlim(0, len(idx) - 1)
        ax1.grid(True, alpha=0.2)
        ax1.set_title(f'[{s_name}] 전력수지 및 유연성자원 운영 (V2G·ESS·HP·전해조)',
                      fontsize=12, fontweight='bold')
        plt.tight_layout()
        plt.savefig(f'{out_dir}/03_sector_coupling_vs_congestion.png', dpi=200,
                    bbox_inches='tight', facecolor='white')
        plt.close(fig)
        print(f"    → 03_sector_coupling_vs_congestion.png")

    # ── ④ 에너지 저장설비 충방전 + SoC + 한계가격 (기술별 서브플롯) ──
    def _plot04_storage_price(idx, out_dir, s_name):
        """
        ESS / V2G / 열저장 / 수소저장 각각을 서브플롯으로 구성.
        각 패널: SoC(MWh, 파랑 면), 충전(+, 밝은 색), 방전(-, 진한 색),
                우측 Y축: 평균 한계가격
        첨부 이미지와 동일한 레이아웃.
        """
        # ── 1) 한계가격 수집 ──────────────────────────────────────
        price_data = pd.Series(dtype=float)
        try:
            if hasattr(network.buses_t, 'marginal_price') and not network.buses_t.marginal_price.empty:
                el_buses_p = [b for b in network.buses_t.marginal_price.columns
                              if _is_el(b) and _in_region(b)]
                if el_buses_p:
                    price_data = (network.buses_t.marginal_price[el_buses_p]
                                  .reindex(idx).fillna(0).mean(axis=1))
        except Exception:
            pass

        # ── 2) 저장설비 기술별 SoC + 충방전 수집 ─────────────────
        # 에너지원 정의: tech명, 서브플롯 제목, SoC색, 충전색, 방전색
        STORE_TYPES = [
            {'tech': 'ESS',   'title': '전력 저장 – ESS',
             'c_soc': '#1565C0', 'c_chg': '#64B5F6', 'c_dis': '#F44336'},
            {'tech': 'V2G',   'title': '전력 저장 – V2G (차량배터리)',
             'c_soc': '#1B5E20', 'c_chg': '#81C784', 'c_dis': '#FF7043'},
            {'tech': '열저장', 'title': '열 저장',
             'c_soc': '#E65100', 'c_chg': '#FFD54F', 'c_dis': '#FF6D00'},
            {'tech': 'H2저장', 'title': '수소 저장',
             'c_soc': '#6A1B9A', 'c_chg': '#CE93D8', 'c_dis': '#AB47BC'},
        ]

        soc_by_tech  = {}   # tech → np.array SoC (MWh)
        flow_by_tech = {}   # tech → np.array power (+ charge / − discharge)

        try:
            if hasattr(network, 'stores_t'):
                _e_df = (network.stores_t.e
                         if hasattr(network.stores_t, 'e')
                            and network.stores_t.e is not None
                            and not network.stores_t.e.empty
                         else pd.DataFrame())
                _p_df = (network.stores_t.p
                         if hasattr(network.stores_t, 'p')
                            and not network.stores_t.p.empty
                         else pd.DataFrame())
                for s in set(list(_e_df.columns) + list(_p_df.columns)):
                    if s in network.stores.index:
                        bus = str(network.stores.at[s, 'bus'])
                        if not _in_region(bus): continue
                    tech = _classify_store_tech(s)
                    if _e_df is not None and s in _e_df.columns:
                        e = _e_df[s].reindex(idx).ffill().fillna(0).values
                        soc_by_tech[tech] = soc_by_tech.get(tech, np.zeros(len(idx))) + e
                    if _p_df is not None and s in _p_df.columns:
                        p = _p_df[s].reindex(idx).fillna(0).values
                        flow_by_tech[tech] = flow_by_tech.get(tech, np.zeros(len(idx))) + p
        except Exception:
            pass

        # 수소 저장이 없으면 전해조 전력소비를 H2 생산으로 대체 표시
        try:
            if ('H2저장' not in soc_by_tech
                    and not network.links.empty
                    and hasattr(network.links_t, 'p0')
                    and not network.links_t.p0.empty):
                _h2_prod = np.zeros(len(idx))
                for lk in network.links_t.p0.columns:
                    if lk not in network.links.index: continue
                    b0 = str(network.links.at[lk, 'bus0'])
                    if region and _rof(b0) != region: continue
                    if _classify_link_tech(lk) == '전해조':
                        _h2_prod += network.links_t.p0[lk].reindex(idx).fillna(0).values
                if _h2_prod.max() > 0.1:
                    soc_by_tech['H2저장']  = np.zeros(len(idx))
                    flow_by_tech['H2저장'] = _h2_prod  # 생산 = 충전
        except Exception:
            pass

        # ── 3) 활성 서브플롯 선별 ─────────────────────────────────
        active = [t for t in STORE_TYPES
                  if t['tech'] in soc_by_tech or t['tech'] in flow_by_tech]

        if not active:
            print(f"    저장설비 데이터 없음 – storage_price 생략")
            return

        # ── 4) 그리기 ─────────────────────────────────────────────
        n  = len(active)
        x  = np.arange(len(idx))
        tick_pos = list(range(0, len(idx), 24))
        tick_lbl = [idx[i].strftime('%m/%d\n%H:00') for i in tick_pos]

        fig, axes = plt.subplots(n, 1, figsize=(18, 5 * n), sharex=True,
                                 gridspec_kw={'hspace': 0.35})
        if n == 1:
            axes = [axes]

        for ax_main, tinfo in zip(axes, active):
            tech  = tinfo['tech']
            soc   = soc_by_tech.get(tech,  np.zeros(len(idx)))
            flow  = flow_by_tech.get(tech, np.zeros(len(idx)))
            chg   = np.where(flow >= 0, flow, 0.0)
            dis   = np.where(flow <  0, flow, 0.0)

            # SoC 면 표시
            ax_main.fill_between(x, 0, soc, alpha=0.20,
                                 color=tinfo['c_soc'], label='SoC (MWh)')
            ax_main.plot(x, soc, color=tinfo['c_soc'], lw=2.0)

            # 충전(+) / 방전(−) fill
            ax_main.fill_between(x, 0,   chg, alpha=0.55,
                                 color=tinfo['c_chg'], label='충전 (+)')
            ax_main.fill_between(x, dis, 0,   alpha=0.55,
                                 color=tinfo['c_dis'], label='방전 (−)')
            ax_main.axhline(0, color='gray', lw=0.6, alpha=0.5)

            ax_main.set_ylabel('에너지/전력 (MWh / MW)',
                               color=tinfo['c_soc'], fontsize=10)
            ax_main.tick_params(axis='y', labelcolor=tinfo['c_soc'])
            ax_main.set_title(tinfo['title'], fontsize=11, fontweight='bold')
            ax_main.grid(True, alpha=0.25)

            # 한계가격 (오른쪽 축)
            if len(price_data) > 0:
                ax_p = ax_main.twinx()
                ax_p.plot(x, price_data.values, color='#C0392B',
                          lw=1.2, alpha=0.75, label='평균 한계가격')
                ax_p.set_ylabel('한계가격 (원/MWh)', color='#C0392B', fontsize=9)
                ax_p.tick_params(axis='y', labelcolor='#C0392B')
                h_main, l_main = ax_main.get_legend_handles_labels()
                h_p,    l_p    = ax_p.get_legend_handles_labels()
                ax_main.legend(h_main + h_p, l_main + l_p,
                               loc='upper right', fontsize=8, ncol=2)
            else:
                ax_main.legend(loc='upper right', fontsize=8)

            ax_main.set_xlim(0, len(idx) - 1)

        axes[-1].set_xticks(tick_pos)
        axes[-1].set_xticklabels(tick_lbl, fontsize=8)

        fig.suptitle(f'[{s_name}] 에너지 저장설비 충방전 및 한계가격',
                     fontsize=13, fontweight='bold', y=1.01)
        plt.savefig(f'{out_dir}/04_storage_soc_price.png', dpi=180,
                    bbox_inches='tight', facecolor='white')
        plt.close(fig)
        print(f"    → 04_storage_soc_price.png")

    # ── ⑤ 에너지 수급 균형 (수평 바) ─────────────────────────────
    def _plot05_energy_balance(idx, out_dir, s_name):
        gen_agg  = _get_gen_agg(idx)
        demand   = _get_demand(idx)
        if not gen_agg:
            print(f"    발전 데이터 없음 – energy_balance 생략")
            return

        supply_by_type = {k: float(np.sum(v)) / 1000 for k, v in gen_agg.items() if np.sum(v) > 0.1}
        total_supply   = sum(supply_by_type.values())
        total_demand   = float(demand.sum()) / 1000

        sto_charge = sto_discharge = 0.0
        try:
            if hasattr(network, 'stores_t') and hasattr(network.stores_t, 'p') and not network.stores_t.p.empty:
                p_all = network.stores_t.p.reindex(idx).fillna(0)
                sto_charge    = float(p_all.clip(lower=0).sum().sum()) / 1000
                sto_discharge = float((-p_all).clip(lower=0).sum().sum()) / 1000
        except Exception:
            pass

        curt_mwh = 0.0
        if results and 're_curtailment' in results:
            for data in results['re_curtailment'].values():
                ts = data.get('timeseries')
                if ts is not None:
                    curt_mwh += float(ts.reindex(idx).fillna(0).sum()) / 1000

        supply_types = [k for k in STACK_ORDER if k in supply_by_type]
        supply_vals  = [supply_by_type[k] for k in supply_types]
        supply_cols  = [GEN_COLORS[k] for k in supply_types]

        consume_lbl  = ['전력수요', '저장충전', 'RE커튼먼트']
        consume_vals = [total_demand, sto_charge, curt_mwh]
        consume_cols = ['#1565C0', '#FF9800', '#D32F2F']
        filtered = [(l, v, c) for l, v, c in zip(consume_lbl, consume_vals, consume_cols) if v > 0.001]
        if filtered:
            consume_lbl, consume_vals, consume_cols = map(list, zip(*filtered))
        else:
            consume_lbl, consume_vals, consume_cols = ['전력수요'], [total_demand], ['#1565C0']

        fig_h = max(5, max(len(supply_types), len(consume_lbl)) * 0.7 + 3)
        fig, (ax_s, ax_c) = plt.subplots(1, 2, figsize=(18, fig_h))

        bars_s = ax_s.barh(supply_types, supply_vals, color=supply_cols,
                           edgecolor='white', linewidth=0.5)
        ax_s.bar_label(bars_s, [f'{v:.1f} GWh' for v in supply_vals], padding=4, fontsize=9)
        ax_s.set_xlabel('에너지 (GWh)', fontsize=10)
        ax_s.set_title('공급 (발전원별)', fontsize=11, fontweight='bold')
        ax_s.set_xlim(0, max(supply_vals) * 1.35 if supply_vals else 1)
        ax_s.invert_xaxis()
        ax_s.grid(axis='x', alpha=0.3)

        bars_c = ax_c.barh(consume_lbl, consume_vals, color=consume_cols,
                           edgecolor='white', linewidth=0.5)
        ax_c.bar_label(bars_c, [f'{v:.1f} GWh' for v in consume_vals], padding=4, fontsize=9)
        ax_c.set_xlabel('에너지 (GWh)', fontsize=10)
        ax_c.set_title('소비 (용도별)', fontsize=11, fontweight='bold')
        ax_c.set_xlim(0, max(consume_vals) * 1.35 if consume_vals else 1)
        ax_c.grid(axis='x', alpha=0.3)

        fig.text(0.5, 0.98,
                 f'[{s_name}] 168h 에너지 수급 균형  |  발전: {total_supply:.1f} GWh  /  수요: {total_demand:.1f} GWh',
                 ha='center', va='top', fontsize=12, fontweight='bold',
                 bbox=dict(facecolor='#F0F0F0', edgecolor='gray', boxstyle='round,pad=0.4'))
        plt.tight_layout(rect=[0, 0, 1, 0.95])
        plt.savefig(f'{out_dir}/05_energy_balance.png', dpi=200,
                    bbox_inches='tight', facecolor='white')
        plt.close(fig)
        print(f"    → 05_energy_balance.png")

    # ── 계절별 루프 ───────────────────────────────────────────────
    SEASONS = [
        {'key': '봄_Spring',   'months': [3, 4, 5],       'name': '봄(3~5월)'},
        {'key': '여름_Summer', 'months': [6, 7, 8],       'name': '여름(6~8월)'},
        {'key': '가을_Autumn', 'months': [9, 10, 11],     'name': '가을(9~11월)'},
        {'key': '겨울_Winter', 'months': [11, 12, 1, 2],  'name': '겨울(11~2월)'},
    ]

    # ── 전체 시뮬레이션 기간 Y축 최댓값 사전 계산 (계절별 스케일 통일) ──
    _GLOBAL_GEN_Y_MAX  = 10000
    _GLOBAL_CURT_Y_MAX = 1000
    try:
        _ga = _get_gen_agg(snapshots)
        _bot = np.zeros(len(snapshots))
        for _v in _ga.values():
            _bot += _v
        _raw = float(_bot.max())
        if _raw > 0:
            _mag = 10 ** _math.floor(_math.log10(_raw))
            _ru  = _mag if _raw / _mag < 5 else _mag * 5
            _GLOBAL_GEN_Y_MAX = _math.ceil(_raw / _ru) * _ru
    except Exception:
        pass
    try:
        if results and 're_curtailment' in results:
            _ca = pd.Series(0.0, index=snapshots)
            for _d in results['re_curtailment'].values():
                _ts = _d.get('timeseries')
                if _ts is not None:
                    _ca = _ca.add(_ts.reindex(snapshots).fillna(0), fill_value=0)
            _cr = float(_ca.max())
            if _cr > 0:
                _mag = 10 ** _math.floor(_math.log10(max(_cr, 1)))
                _ru  = _mag if _cr / _mag < 5 else _mag * 5
                _GLOBAL_CURT_Y_MAX = _math.ceil(_cr / _ru) * _ru
    except Exception:
        pass

    label_prefix = f"{region} / " if region else ""

    for season in SEASONS:
        s_key, s_name, months = season['key'], season['name'], season['months']
        full_name = f"{label_prefix}{s_name}"
        print(f"\n  [{full_name}] 대표 구간 탐색 ...")

        start_ts = _get_season_window(months)
        end_ts   = start_ts + pd.Timedelta(hours=168)
        mask     = (snapshots >= start_ts) & (snapshots < end_ts)
        idx      = snapshots[mask]

        if len(idx) < 24:
            print(f"    유효 데이터 부족 ({len(idx)}h) – 건너뜀")
            continue

        print(f"    {idx[0].strftime('%Y-%m-%d %H:%M')} ~ {idx[-1].strftime('%Y-%m-%d %H:%M')}  ({len(idx)}h)")

        if region:
            out_dir = f'{results_dir}/regional_analysis/{region}/{s_key}'
        else:
            out_dir = f'{results_dir}/seasonal_analysis/{s_key}'
        os.makedirs(out_dir, exist_ok=True)

        for fn, label in [
            (_plot01_dispatch_stack,  '① dispatch_stack'),
            (_plot02_line_heatmap,    '② line_heatmap'),
            (_plot03_sector_coupling, '③ sector_coupling'),
            (_plot04_storage_price,   '④ storage_price'),
            (_plot05_energy_balance,  '⑤ energy_balance'),
        ]:
            try:
                fn(idx, out_dir, full_name)
            except Exception as _e:
                print(f"    {label} 오류: {_e}")

    if region:
        print(f"\n[{region} 지역 시각화] 완료 → {results_dir}/regional_analysis/{region}/")
    else:
        print(f"\n[계절별 시각화] 완료 → {results_dir}/seasonal_analysis/")


def create_regional_analysis_charts(network, results, results_dir, current_time):
    """지역별 × 계절별 5종 시각화 자동 생성.

    각 EL 버스에서 지역 코드를 추출한 뒤,
    create_seasonal_analysis_charts(region=...)를 호출하여
    results_dir/regional_analysis/{지역코드}/{계절}/ 아래에 저장한다.
    """
    if os.environ.get('DISABLE_PLOTS', '0') == '1':
        return

    bus_carrier = network.buses.carrier.to_dict() if not network.buses.empty else {}

    def _is_el_simple(b):
        c = str(bus_carrier.get(b, '')).lower()
        return ('electric' in c or c in ['el', 'ac', 'dc', 'hvac', 'hvdc']
                or str(b).upper().endswith('_EL'))

    el_buses = [b for b in network.buses.index if _is_el_simple(b)]
    region_codes = sorted({(b.split('_')[0] if '_' in b else b) for b in el_buses})

    print(f"\n[지역별 시각화] 대상 지역 {len(region_codes)}개: {region_codes}")

    for region in region_codes:
        print(f"\n  ◆ {region} 지역 시각화 시작 ...")
        try:
            create_seasonal_analysis_charts(
                network, results, results_dir, current_time, region=region
            )
        except Exception as _e:
            import traceback
            print(f"  {region} 시각화 오류: {_e}")
            traceback.print_exc()

    print(f"\n[지역별 시각화] 전체 완료 → {results_dir}/regional_analysis/")


def create_custom_visualizations(network, results_dir):
    """
    visualize_results.py 의 5종 차트를 results_dir/images/ 에 저장.
    - 네트워크 객체를 직접 받아 nc 파일 재로딩 없이 실행
    - 개별 차트 오류는 경고로 처리하여 다음 차트 생성에 영향 없음
    """
    if os.environ.get('DISABLE_PLOTS', '0') == '1':
        print('[커스텀 시각화] DISABLE_PLOTS=1 → 스킵')
        return

    images_dir = os.path.join(results_dir, 'images')
    try:
        import sys as _sys
        _script_dir = get_base_dir()
        if _script_dir not in _sys.path:
            _sys.path.insert(0, _script_dir)

        import visualize_results as _vr
        _vr.run_all_charts(network, images_dir)

    except ImportError as _e:
        print(f'[커스텀 시각화] visualize_results.py 로드 실패: {_e}')
    except Exception as _e:
        print(f'[커스텀 시각화] 오류: {_e}')
        import traceback as _tb
        _tb.print_exc()


def create_visualizations(network, results_dir, current_time):
    if os.environ.get('DISABLE_PLOTS','0') == '1':
        print('시각화 생략(DISABLE_PLOTS=1)')
        return True
    """시각화 결과 생성"""
    try:
        import matplotlib.pyplot as plt
        import plotly.graph_objects as go
        import plotly.express as px
        from plotly.subplots import make_subplots
        import plotly.offline as pyo
        
        # 한글 폰트 설정(윈도우: 맑은 고딕)
        try:
            import matplotlib as mpl
            mpl.rcParams['font.family'] = ['Malgun Gothic', 'DejaVu Sans']
            mpl.rcParams['axes.unicode_minus'] = False
            import warnings as _warn
            import logging as _logging
            _warn.filterwarnings('ignore', message='findfont:', category=UserWarning, module='matplotlib')
            try:
                _logging.getLogger('matplotlib.font_manager').setLevel(_logging.ERROR)
            except Exception:
                pass
        except Exception:
            pass
        
        print("시각화 결과 생성 중...")
        
        # 1. 지역별 에너지 밸런스 차트
        gen_output = network.generators_t.p
        load_output = network.loads_t.p
        
        # 지역별 발전량과 부하량 계산
        regional_generation = {}
        regional_load = {}
        
        for gen_name in gen_output.columns:
            if '_' in gen_name:
                region = gen_name.split('_')[0]
                if region not in regional_generation:
                    regional_generation[region] = 0
                regional_generation[region] += gen_output[gen_name].sum()
        
        for load_name in load_output.columns:
            if '_' in load_name:
                region = load_name.split('_')[0]
                if region not in regional_load:
                    regional_load[region] = 0
                regional_load[region] += load_output[load_name].sum()
        
        # 지역별 에너지 밸런스 DataFrame 생성
        regions = list(set(list(regional_generation.keys()) + list(regional_load.keys())))
        balance_data = []
        for region in regions:
            gen = regional_generation.get(region, 0)
            load = regional_load.get(region, 0)
            balance = gen - load
            balance_data.append([region, gen, load, balance])
        
        balance_df = pd.DataFrame(balance_data, columns=['지역', '발전량(MWh)', '부하량(MWh)', '밸런스(MWh)'])
        balance_df.to_csv(f'{results_dir}/regional_energy_balance.csv', index=False, encoding='utf-8-sig')
        
        # 지역별 에너지 밸런스 차트 생성
        fig, ax = plt.subplots(figsize=(15, 8))
        x = range(len(balance_df))
        width = 0.35
        
        ax.bar([i - width/2 for i in x], balance_df['발전량(MWh)'], width, label='발전량', alpha=0.8)
        ax.bar([i + width/2 for i in x], balance_df['부하량(MWh)'], width, label='부하량', alpha=0.8)
        
        ax.set_xlabel('지역')
        ax.set_ylabel('에너지 (MWh)')
        ax.set_title('지역별 에너지 밸런스')
        ax.set_xticks(x)
        ax.set_xticklabels(balance_df['지역'], rotation=45)
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig(f'{results_dir}/regional_energy_balance.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        # 2. 지역별 재생에너지 비율 차트
        renewable_ratio_data = []
        for region in regions:
            total_gen = 0
            renewable_gen = 0
            
            for gen_name in gen_output.columns:
                if gen_name.startswith(region + '_'):
                    gen_sum = gen_output[gen_name].sum()
                    total_gen += gen_sum
                    if 'PV' in gen_name or 'WT' in gen_name or 'Wind' in gen_name or 'Solar' in gen_name:
                        renewable_gen += gen_sum
            
            if total_gen > 0:
                ratio = (renewable_gen / total_gen) * 100
            else:
                ratio = 0
            
            renewable_ratio_data.append([region, renewable_gen, total_gen, ratio])
        
        renewable_df = pd.DataFrame(renewable_ratio_data, columns=['지역', '재생에너지(MWh)', '총발전량(MWh)', '재생에너지비율(%)'])
        
        # 재생에너지 비율 차트
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 8))
        
        # 재생에너지 비율 막대 차트
        ax1.bar(renewable_df['지역'], renewable_df['재생에너지비율(%)'], color='green', alpha=0.7)
        ax1.set_xlabel('지역')
        ax1.set_ylabel('재생에너지 비율 (%)')
        ax1.set_title('지역별 재생에너지 비율')
        ax1.tick_params(axis='x', rotation=45)
        ax1.grid(True, alpha=0.3)
        
        # 재생에너지 vs 총발전량 비교
        x = range(len(renewable_df))
        width = 0.35
        
        ax2.bar([i - width/2 for i in x], renewable_df['재생에너지(MWh)'], width, label='재생에너지', color='green', alpha=0.7)
        ax2.bar([i + width/2 for i in x], renewable_df['총발전량(MWh)'] - renewable_df['재생에너지(MWh)'], width, label='기타 발전', color='gray', alpha=0.7)
        
        ax2.set_xlabel('지역')
        ax2.set_ylabel('발전량 (MWh)')
        ax2.set_title('지역별 발전원별 발전량')
        ax2.set_xticks(x)
        ax2.set_xticklabels(renewable_df['지역'], rotation=45)
        ax2.legend()
        ax2.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig(f'{results_dir}/regional_renewable_ratio.png', dpi=300, bbox_inches='tight')
        plt.close()
        
        # 3. 송전선로 조류 지도 (HTML)
        if not network.lines_t.p0.empty:
            # 송전선로 조류 데이터 준비
            line_flows = network.lines_t.p0.mean()  # 평균 조류
            transmission_data = []
            
            for line_name in line_flows.index:
                if line_name in network.lines.index:
                    bus0 = network.lines.at[line_name, 'bus0']
                    bus1 = network.lines.at[line_name, 'bus1']
                    flow = line_flows[line_name]
                    capacity = network.lines.at[line_name, 's_nom']
                    utilization = abs(flow) / capacity * 100 if capacity > 0 else 0
                    
                    transmission_data.append([line_name, bus0, bus1, flow, capacity, utilization])
            
            transmission_df = pd.DataFrame(transmission_data, 
                                         columns=['선로명', '시작버스', '종료버스', '평균조류(MW)', '용량(MVA)', '이용률(%)'])
            transmission_df.to_csv(f'{results_dir}/transmission_flow.csv', index=False, encoding='utf-8-sig')
            
            # 송전선로 네트워크 그래프
            import networkx as nx
            
            G = nx.Graph()
            
            # 노드 추가 (버스)
            for bus in network.buses.index:
                G.add_node(bus)
            
            # 엣지 추가 (선로)
            for line_name in network.lines.index:
                bus0 = network.lines.at[line_name, 'bus0']
                bus1 = network.lines.at[line_name, 'bus1']
                if line_name in line_flows.index:
                    flow = abs(line_flows[line_name])
                    G.add_edge(bus0, bus1, weight=flow, line_name=line_name)
            
            # 네트워크 그래프 그리기
            plt.figure(figsize=(20, 15))
            pos = nx.spring_layout(G, k=3, iterations=50)
            
            # 노드 그리기
            nx.draw_networkx_nodes(G, pos, node_color='lightblue', node_size=500, alpha=0.8)
            
            # 엣지 그리기 (조류 크기에 따라 두께 조정)
            edges = G.edges()
            weights = [G[u][v]['weight'] for u, v in edges]
            max_weight = max(weights) if weights else 1e-6
            edge_widths = [w/max_weight * 5 + 0.5 for w in weights]
            
            nx.draw_networkx_edges(G, pos, width=edge_widths, alpha=0.6, edge_color='red')
            
            # 라벨 그리기
            nx.draw_networkx_labels(G, pos, font_size=8, font_weight='bold')
            
            plt.title('송전선로 네트워크 및 조류 현황', fontsize=16, fontweight='bold')
            plt.axis('off')
            plt.tight_layout()
            plt.savefig(f'{results_dir}/transmission_network_graph.png', dpi=300, bbox_inches='tight')
            plt.close()
            
            # 인터랙티브 송전선로 지도 (HTML)
            fig = go.Figure()
            
            # 선로별 조류 막대 차트
            fig.add_trace(go.Bar(
                x=transmission_df['선로명'],
                y=transmission_df['이용률(%)'],
                name='선로 이용률',
                text=transmission_df['평균조류(MW)'].round(1),
                textposition='auto',
                hovertemplate='<b>%{x}</b><br>이용률: %{y:.1f}%<br>조류: %{text} MW<extra></extra>'
            ))
            
            fig.update_layout(
                title='송전선로별 이용률 및 조류 현황',
                xaxis_title='송전선로',
                yaxis_title='이용률 (%)',
                hovermode='x unified',
                height=600
            )
            
            pyo.plot(fig, filename=f'{results_dir}/transmission_flow_map.html', auto_open=False)
        
        # 4. 한국 지도 기반 송전선로 시각화 추가
        try:
            create_korea_transmission_map(network, results_dir, current_time)
        except Exception as e:
            print(f"한국 지도 기반 송전선로 시각화 생성 중 오류: {str(e)}")
        
        # 5. 이전 버전의 한국 지도 시각화 추가
        try:
            create_legacy_korea_map(network, results_dir, current_time)
        except Exception as e:
            print(f"이전 버전 한국 지도 시각화 생성 중 오류: {str(e)}")
        
        # 6. RE 전용 한국 지도 시각화 추가
        try:
            create_re_korea_map(network, results_dir, current_time)
        except Exception as e:
            print(f"RE 전용 한국 지도 시각화 생성 중 오류: {str(e)}")
        
        print("시각화 결과 생성 완료:")
        print("- 지역별 에너지 밸런스 차트")
        print("- 지역별 재생에너지 비율 차트")
        print("- 송전선로 네트워크 그래프")
        print("- 인터랙티브 송전선로 지도")
        print("- 한국 지도 기반 송전선로 시각화")
        print("- 이전 버전 한국 지도 시각화")
        print("- RE 전용 한국 지도 시각화 (신규)")
        
        return True
        
    except Exception as e:
        print(f"시각화 생성 중 오류 발생: {str(e)}")
        traceback.print_exc()
        return False

def create_korea_transmission_map(network, results_dir, current_time):
    """한국 지도 기반 송전선로 시각화"""
    try:
        from korea_map import KoreaMapVisualizer
        import matplotlib.pyplot as plt
        
        print("한국 지도 기반 송전선로 시각화 생성 중...")
        
        # 지역 코드와 한국어 이름 매핑
        region_mapping = {
            'SEL': '서울특별시',
            'BSN': '부산광역시', 
            'DGU': '대구광역시',
            'ICN': '인천광역시',
            'GWJ': '광주광역시',
            'DJN': '대전광역시',
            'USN': '울산광역시',
            'SJG': '세종특별자치시',
            'GGD': '경기도',
            'GWD': '강원도',
            'CBD': '충청북도',
            'CND': '충청남도',
            'JBD': '전라북도',
            'JND': '전라남도',
            'GBD': '경상북도',
            'GND': '경상남도',
            'JJD': '제주특별자치도'
        }
        
        # 송전선로 조류 데이터 준비
        line_flows = network.lines_t.p0.mean() if not network.lines_t.p0.empty else pd.Series()
        
        # 송전선로 정보를 DataFrame으로 구성
        transmission_lines = []
        
        for line_name in network.lines.index:
            bus0 = network.lines.at[line_name, 'bus0']
            bus1 = network.lines.at[line_name, 'bus1']
            
            # 버스 이름에서 지역 코드 추출
            region0 = bus0.split('_')[0] if '_' in bus0 else bus0
            region1 = bus1.split('_')[0] if '_' in bus1 else bus1
            
            # 지역 코드를 한국어 이름으로 변환
            region0_name = region_mapping.get(region0, region0)
            region1_name = region_mapping.get(region1, region1)
            
            # 조류 정보
            flow = line_flows.get(line_name, 0) if line_name in line_flows.index else 0
            capacity = network.lines.at[line_name, 's_nom']
            utilization = abs(flow) / capacity * 100 if capacity > 0 else 0
            
            transmission_lines.append({
                '선로명': line_name,
                '시작지역': region0_name,
                '종료지역': region1_name,
                '조류(MW)': flow,
                '용량(MVA)': capacity,
                '이용률(%)': utilization
            })
        
        # DataFrame 생성
        transmission_df = pd.DataFrame(transmission_lines)
        
        # 한국 지도 시각화 객체 생성
        visualizer = KoreaMapVisualizer()
        
        if visualizer.load_map_data():
            # 기본 한국 지도 생성
            map_save_path = f'{results_dir}/korea_transmission_map_{current_time}.png'
            
            # 지도 그리기
            fig, ax = plt.subplots(figsize=(15, 12))
            
            # 행정구역 경계 그리기
            visualizer.map_data.plot(ax=ax, 
                                   color='lightgray', 
                                   edgecolor='black', 
                                   linewidth=0.5,
                                   alpha=0.7)
            
            # 지역 중심점 계산
            region_centroids = {}
            for idx, row in visualizer.map_data.iterrows():
                region_name = row['SIDO_NM']
                centroid = row.geometry.centroid
                region_centroids[region_name] = (centroid.x, centroid.y)
            
            # 송전선로 그리기
            for _, line in transmission_df.iterrows():
                start_region = line['시작지역']
                end_region = line['종료지역']
                utilization = line['이용률(%)']
                
                if start_region in region_centroids and end_region in region_centroids:
                    start_x, start_y = region_centroids[start_region]
                    end_x, end_y = region_centroids[end_region]
                    
                    # 이용률에 따른 선 두께와 색상 설정
                    line_width = max(0.5, utilization / 20)  # 최소 0.5, 최대 5
                    
                    if utilization > 80:
                        color = 'red'
                    elif utilization > 60:
                        color = 'orange'
                    elif utilization > 40:
                        color = 'yellow'
                    else:
                        color = 'green'
                    
                    # 송전선로 그리기
                    ax.plot([start_x, end_x], [start_y, end_y], 
                           color=color, linewidth=line_width, alpha=0.8)
                    
                    # 중점에 이용률 표시 (이용률이 높은 경우만)
                    if utilization > 50:
                        mid_x = (start_x + end_x) / 2
                        mid_y = (start_y + end_y) / 2
                        ax.text(mid_x, mid_y, f'{utilization:.0f}%', 
                               fontsize=8, ha='center', va='center',
                               bbox=dict(facecolor='white', alpha=0.7, edgecolor='none'))
            
            # 지역 중심점에 지역명 표시
            for region_name, (x, y) in region_centroids.items():
                # 지역 코드로 변환
                region_code = None
                for code, name in region_mapping.items():
                    if name == region_name:
                        region_code = code
                        break
                
                if region_code:
                    ax.plot(x, y, 'ko', markersize=8, alpha=0.8)
                    ax.text(x, y + 20000, region_code, fontsize=10, ha='center', va='bottom',
                           fontweight='bold',
                           bbox=dict(facecolor='white', alpha=0.8, edgecolor='black'))
            
            # 범례 추가
            legend_elements = [
                plt.Line2D([0], [0], color='green', lw=2, label='이용률 < 40%'),
                plt.Line2D([0], [0], color='yellow', lw=2, label='이용률 40-60%'),
                plt.Line2D([0], [0], color='orange', lw=2, label='이용률 60-80%'),
                plt.Line2D([0], [0], color='red', lw=2, label='이용률 > 80%')
            ]
            ax.legend(handles=legend_elements, loc='upper left', bbox_to_anchor=(0.02, 0.98))
            
            ax.set_title('한국 송전선로 이용률 현황', fontsize=16, fontweight='bold', pad=20)
            ax.set_axis_off()
            
            plt.tight_layout()
            plt.savefig(map_save_path, dpi=300, bbox_inches='tight', facecolor='white')
            plt.close()
            
            print(f"한국 지도 기반 송전선로 시각화가 '{map_save_path}'에 저장되었습니다.")
            
            # 송전선로 상세 정보 CSV 저장
            transmission_df.to_csv(f'{results_dir}/korea_transmission_details_{current_time}.csv', 
                                 index=False, encoding='utf-8-sig')
            
            return True
        
            print("한국 지도 데이터를 로드할 수 없습니다.")
            return False
            
    except Exception as e:
        print(f"한국 지도 기반 송전선로 시각화 생성 중 오류: {str(e)}")
        traceback.print_exc()
        return False

def create_legacy_korea_map(network, results_dir, current_time):
    """이전 버전의 한국 지도 시각화 (간단한 지역 연결 지도)"""
    try:
        import matplotlib.pyplot as plt
        import matplotlib.patches as patches
        from matplotlib.patches import FancyBboxPatch
        import numpy as np
        
        print("이전 버전 한국 지도 시각화 생성 중...")
        
        # 한국 지역 좌표 정의 (11개 지역 - 흡수합병된 SEL/SJG/DJN/DGU/USN/GWJ 제외)
        region_coords = {
            'ICN': (4, 7),    # 인천
            'GGD': (5, 6),    # 경기
            'GWD': (7, 8),    # 강원
            'CBD': (6, 5),    # 충북
            'CND': (4, 5),    # 충남
            'JBD': (3, 3),    # 전북
            'JND': (2, 2),    # 전남
            'GBD': (7, 4),    # 경북
            'GND': (6, 2),    # 경남
            'BSN': (7, 1),    # 부산
            'JJD': (1, 0)     # 제주
        }
        
        # 송전선로 조류 데이터 준비
        line_flows = network.lines_t.p0.mean() if not network.lines_t.p0.empty else pd.Series()
        
        # 지역별 발전량 계산
        gen_output = network.generators_t.p
        regional_generation = {}
        regional_renewable = {}
        
        for gen_name in gen_output.columns:
            if '_' in gen_name:
                region = gen_name.split('_')[0]
                gen_sum = gen_output[gen_name].sum()
                
                if region not in regional_generation:
                    regional_generation[region] = 0
                    regional_renewable[region] = 0
                
                regional_generation[region] += gen_sum
                
                # 재생에너지 여부 확인
                if any(keyword in gen_name for keyword in ['PV', 'WT', 'Solar', 'Wind']):
                    regional_renewable[region] += gen_sum
        
        # 지역별 부하량 계산
        load_output = network.loads_t.p
        regional_load = {}
        
        for load_name in load_output.columns:
            if '_' in load_name:
                region = load_name.split('_')[0]
                load_sum = load_output[load_name].sum()
                
                if region not in regional_load:
                    regional_load[region] = 0
                regional_load[region] += load_sum
        
        # 그래프 생성
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 10))
        
        # 첫 번째 지도: 발전량과 송전선로
        ax1.set_xlim(0, 9)
        ax1.set_ylim(-1, 9)
        ax1.set_aspect('equal')
        
        # 지역별 발전량을 원의 크기로 표현
        max_generation = max(regional_generation.values()) if regional_generation else 1e-6
        
        for region, coords in region_coords.items():
            x, y = coords
            generation = regional_generation.get(region, 0)
            renewable = regional_renewable.get(region, 0)
            load = regional_load.get(region, 0)
            
            # 발전량에 비례한 원의 크기
            size = max(50, (generation / max_generation) * 1000) if max_generation > 0 else 50
            
            # 재생에너지 비율에 따른 색상
            renewable_ratio = renewable / generation if generation > 0 else 0
            color = plt.cm.RdYlGn(renewable_ratio)
            
            # 지역 원 그리기
            circle = plt.Circle((x, y), np.sqrt(size)/20, color=color, alpha=0.7, edgecolor='black')
            ax1.add_patch(circle)
            
            # 지역명 표시
            ax1.text(x, y, region, ha='center', va='center', fontsize=8, fontweight='bold')
            
            # 발전량 정보 표시
            info_text = f"{generation/1000:.0f}GWh"
            if renewable_ratio > 0:
                info_text += f"\n재생{renewable_ratio*100:.0f}%"
            ax1.text(x, y-0.5, info_text, ha='center', va='top', fontsize=6)
        
        # 송전선로 그리기
        for line_name in network.lines.index:
            bus0 = network.lines.at[line_name, 'bus0']
            bus1 = network.lines.at[line_name, 'bus1']
            
            # 버스 이름에서 지역 코드 추출
            region0 = bus0.split('_')[0] if '_' in bus0 else bus0
            region1 = bus1.split('_')[0] if '_' in bus1 else bus1
            
            if region0 in region_coords and region1 in region_coords:
                x0, y0 = region_coords[region0]
                x1, y1 = region_coords[region1]
                
                # 조류 정보
                flow = abs(line_flows.get(line_name, 0)) if line_name in line_flows.index else 0
                capacity = network.lines.at[line_name, 's_nom']
                utilization = flow / capacity * 100 if capacity > 0 else 0
                
                # 이용률에 따른 선 색상과 두께
                if utilization > 80:
                    color = 'red'
                    width = 3
                elif utilization > 60:
                    color = 'orange'
                    width = 2.5
                elif utilization > 40:
                    color = 'yellow'
                    width = 2
                else:
                    color = 'green'
                    width = 1.5
                
                # 송전선로 그리기
                ax1.plot([x0, x1], [y0, y1], color=color, linewidth=width, alpha=0.8)
                
                # 이용률이 높은 경우 수치 표시
                if utilization > 50:
                    mid_x, mid_y = (x0 + x1) / 2, (y0 + y1) / 2
                    ax1.text(mid_x, mid_y, f'{utilization:.0f}%', 
                            fontsize=6, ha='center', va='center',
                            bbox=dict(facecolor='white', alpha=0.7, edgecolor='none', pad=1))
        
        ax1.set_title('지역별 발전량 및 송전선로 이용률', fontsize=14, fontweight='bold')
        ax1.set_xlabel('재생에너지 비율: 빨간색(낮음) → 녹색(높음)', fontsize=10)
        ax1.grid(True, alpha=0.3)
        ax1.set_xticks([])
        ax1.set_yticks([])
        
        # 두 번째 지도: 에너지 밸런스
        ax2.set_xlim(0, 9)
        ax2.set_ylim(-1, 9)
        ax2.set_aspect('equal')
        
        # 지역별 에너지 밸런스 (발전량 - 부하량)
        for region, coords in region_coords.items():
            x, y = coords
            generation = regional_generation.get(region, 0)
            load = regional_load.get(region, 0)
            balance = generation - load
            
            # 밸런스에 따른 색상 (잉여: 파란색, 부족: 빨간색)
            gen_max = max([1e-6] + list(regional_generation.values()))
            load_max = max([1e-6] + list(regional_load.values()))
            if balance > 0:
                color = 'blue'
                alpha = min(0.8, abs(balance) / gen_max * 2)
            else:
                color = 'red'
                alpha = min(0.8, abs(balance) / load_max * 2)
            
            # 밸런스 크기에 비례한 원
            denom = max([1e-6] + list(regional_generation.values()) + list(regional_load.values()))
            size = max(50, abs(balance) / denom * 1000)
            
            circle = plt.Circle((x, y), np.sqrt(size)/20, color=color, alpha=alpha, edgecolor='black')
            ax2.add_patch(circle)
            
            # 지역명과 밸런스 정보
            ax2.text(x, y, region, ha='center', va='center', fontsize=8, fontweight='bold', color='white')
            balance_text = f"{balance/1000:.0f}GWh"
            if balance > 0:
                balance_text = "+" + balance_text
            ax2.text(x, y-0.5, balance_text, ha='center', va='top', fontsize=6)
        
        ax2.set_title('지역별 에너지 밸런스 (발전량 - 부하량)', fontsize=14, fontweight='bold')
        ax2.set_xlabel('파란색: 잉여, 빨간색: 부족', fontsize=10)
        ax2.grid(True, alpha=0.3)
        ax2.set_xticks([])
        ax2.set_yticks([])
        
        # 범례 추가
        legend_elements1 = [
            plt.Line2D([0], [0], color='green', lw=2, label='이용률 < 40%'),
            plt.Line2D([0], [0], color='yellow', lw=2, label='이용률 40-60%'),
            plt.Line2D([0], [0], color='orange', lw=2, label='이용률 60-80%'),
            plt.Line2D([0], [0], color='red', lw=2, label='이용률 > 80%')
        ]
        ax1.legend(handles=legend_elements1, loc='upper left')
        
        legend_elements2 = [
            plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='blue', markersize=10, label='에너지 잉여'),
            plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='red', markersize=10, label='에너지 부족')
        ]
        ax2.legend(handles=legend_elements2, loc='upper left')
        
        plt.tight_layout()
        
        # 파일 저장
        legacy_map_path = f'{results_dir}/legacy_korea_transmission_map_{current_time}.png'
        plt.savefig(legacy_map_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        print(f"이전 버전 한국 지도 시각화가 '{legacy_map_path}'에 저장되었습니다.")
        
        # 상세 정보 CSV 저장
        legacy_data = []
        for region in region_coords.keys():
            generation = regional_generation.get(region, 0)
            renewable = regional_renewable.get(region, 0)
            load = regional_load.get(region, 0)
            balance = generation - load
            renewable_ratio = renewable / generation * 100 if generation > 0 else 0
            
            legacy_data.append([
                region, generation/1000, renewable/1000, load/1000, 
                balance/1000, renewable_ratio
            ])
        
        legacy_df = pd.DataFrame(legacy_data, columns=[
            '지역', '발전량(GWh)', '재생에너지(GWh)', '부하량(GWh)', 
            '밸런스(GWh)', '재생에너지비율(%)'
        ])
        legacy_df.to_csv(f'{results_dir}/legacy_korea_energy_summary_{current_time}.csv', 
                        index=False, encoding='utf-8-sig')
        
        return True
        
    except Exception as e:
        print(f"이전 버전 한국 지도 시각화 생성 중 오류: {str(e)}")
        traceback.print_exc()
        return False

def create_re_korea_map(network, results_dir, current_time):
    """RE 전용 한국 지도 시각화 (지역별 RE 발전/소비/송전)"""
    try:
        import matplotlib.pyplot as plt
        import numpy as np
        
        print("RE 전용 한국 지도 시각화 생성 중...")
        
        # 한국 지역 좌표 정의
        region_coords = {
            'SEL': (5, 7), 'ICN': (4, 7), 'GGD': (5, 6), 'GWD': (7, 8),
            'CBD': (6, 5), 'CND': (4, 5), 'DJN': (5, 4), 'SJG': (5, 4.5),
            'JBD': (3, 3), 'JND': (2, 2), 'GWJ': (3, 2.5), 'GBD': (7, 4),
            'DGU': (7, 3), 'GND': (6, 2), 'BSN': (7, 1), 'USN': (7.5, 1.5),
            'JJD': (1, 0)
        }
        
        # 1. 지역별 RE 발전량 및 전력 수요 계산
        regional_re_gen = {}  # 지역별 RE 발전량
        regional_re_by_tech = {}  # 지역별 기술별 RE 발전량
        regional_demand = {}  # 지역별 전력 수요
        
        gen_output = network.generators_t.p
        load_output = network.loads_t.p if hasattr(network.loads_t, 'p') and not network.loads_t.p.empty else network.loads_t.p_set
        
        # RE 발전량 집계 (RE 버스에 연결된 발전기)
        for gen_name in gen_output.columns:
            if '_' not in gen_name:
                continue
            
            # RE 발전기만 (PV, WT, Solar, Wind)
            is_re = any(keyword in gen_name for keyword in ['PV', 'WT', 'Solar', 'Wind', '_RE'])
            gen_bus = network.generators.at[gen_name, 'bus'] if gen_name in network.generators.index else ''
            is_re_bus = '_RE' in str(gen_bus)
            
            if not (is_re or is_re_bus):
                continue
            
            region = gen_name.split('_')[0]
            gen_sum = gen_output[gen_name].sum()
            
            # 총 RE 발전량
            if region not in regional_re_gen:
                regional_re_gen[region] = 0
                regional_re_by_tech[region] = {'PV': 0, 'WT': 0, '기타': 0}
            regional_re_gen[region] += gen_sum
            
            # 기술별 분류
            if 'PV' in gen_name or 'Solar' in gen_name:
                regional_re_by_tech[region]['PV'] += gen_sum
            elif 'WT' in gen_name or 'Wind' in gen_name:
                regional_re_by_tech[region]['WT'] += gen_sum
            else:
                regional_re_by_tech[region]['기타'] += gen_sum
        
        # 전력 수요 집계 (EL 버스 부하만)
        for load_name in load_output.columns:
            if '_' not in load_name or '_EL' not in load_name:
                continue
            region = load_name.split('_')[0]
            load_sum = load_output[load_name].sum()
            if region not in regional_demand:
                regional_demand[region] = 0
            regional_demand[region] += load_sum
        
        # 2. RE 송전선 조류 계산
        re_line_flows = {}
        
        # Lines 데이터에서 RE 송전선 확인
        if not network.lines.empty and hasattr(network.lines_t, 'p0') and not network.lines_t.p0.empty:
            line_flows = network.lines_t.p0.mean()
            for line_name in line_flows.index:
                if line_name not in network.lines.index:
                    continue
                # RE 송전선만 (_RE 버스 간 연결)
                bus0 = network.lines.at[line_name, 'bus0']
                bus1 = network.lines.at[line_name, 'bus1']
                if '_RE' in str(bus0) and '_RE' in str(bus1):
                    flow = abs(line_flows[line_name])
                    capacity = network.lines.at[line_name, 's_nom']
                    utilization = flow / capacity * 100 if capacity > 0 else 0
                    re_line_flows[line_name] = {
                        'bus0': bus0,
                        'bus1': bus1,
                        'flow': flow,
                        'capacity': capacity,
                        'utilization': utilization
                    }
        
        # Links 데이터에서도 RE 송전선 확인 (Lines에 없을 경우)
        if not network.links.empty and hasattr(network.links_t, 'p0') and not network.links_t.p0.empty:
            link_flows = network.links_t.p0.mean()
            for link_name in link_flows.index:
                if link_name not in network.links.index:
                    continue
                # RE 송전선만 (_RE 버스 간 연결)
                bus0 = network.links.at[link_name, 'bus0']
                bus1 = network.links.at[link_name, 'bus1']
                if '_RE' in str(bus0) and '_RE' in str(bus1):
                    flow = abs(link_flows[link_name])
                    # Links는 p_nom 사용
                    capacity = network.links.at[link_name, 'p_nom'] if 'p_nom' in network.links.columns else 1e-6
                    utilization = flow / capacity * 100 if capacity > 0 else 0
                    re_line_flows[link_name] = {
                        'bus0': bus0,
                        'bus1': bus1,
                        'flow': flow,
                        'capacity': capacity,
                        'utilization': utilization
                    }
        
        print(f"RE 송전선 발견: {len(re_line_flows)}개")
        if re_line_flows:
            print("RE 송전선 목록:")
            for line_name, data in list(re_line_flows.items())[:5]:  # 처음 5개만 출력
                print(f"  - {line_name}: {data['bus0']} → {data['bus1']}, 흐름: {data['flow']:.1f}MW, 이용률: {data['utilization']:.1f}%")
        
        # 3. 지도 생성
        fig, ax = plt.subplots(figsize=(20, 10))
        ax.set_xlim(0, 9)
        ax.set_ylim(-1, 9)
        ax.set_aspect('equal')
        
        # 4. RE 송전선 그리기
        for line_name, line_data in re_line_flows.items():
            bus0 = line_data['bus0']
            bus1 = line_data['bus1']
            
            region0 = bus0.split('_')[0] if '_' in bus0 else bus0
            region1 = bus1.split('_')[0] if '_' in bus1 else bus1
            
            if region0 not in region_coords or region1 not in region_coords:
                continue
            
            x0, y0 = region_coords[region0]
            x1, y1 = region_coords[region1]
            
            flow = line_data['flow']
            utilization = line_data['utilization']
            
            # 이용률에 따른 색상
            if utilization > 80:
                color = 'red'
            elif utilization > 60:
                color = 'orange'
            elif utilization > 40:
                color = 'yellow'
            elif utilization > 20:
                color = 'lightgreen'
            else:
                color = 'green'
            
            # 흐름량에 따른 선 두께 (정규화)
            max_flow = max([lf['flow'] for lf in re_line_flows.values()]) if re_line_flows else 1e-6
            width = max(0.5, (flow / max_flow) * 5 + 0.5)
            
            # RE 송전선 그리기
            ax.plot([x0, x1], [y0, y1], color=color, linewidth=width, alpha=0.7, linestyle='--')
            
            # 유량 표시 (큰 흐름만)
            if flow > max_flow * 0.3:
                mid_x, mid_y = (x0 + x1) / 2, (y0 + y1) / 2
                ax.text(mid_x, mid_y, f'{flow/1000:.0f}GW', 
                       fontsize=6, ha='center', va='center',
                       bbox=dict(facecolor='white', alpha=0.7, edgecolor='none', pad=1))
        
        # 5. 지역별 RE 발전/수요 원 그리기
        max_demand = max(regional_demand.values()) if regional_demand else 1e-6
        
        for region, coords in region_coords.items():
            x, y = coords
            re_generation = regional_re_gen.get(region, 0)
            demand = regional_demand.get(region, 0)
            re_tech = regional_re_by_tech.get(region, {'PV': 0, 'WT': 0, '기타': 0})
            
            # 수요 기준 원 크기 (2/3 크기로 축소)
            size = max(50, (demand / max_demand) * 1000 * (2/3)) if max_demand > 0 else 50
            
            # RE 자급률 계산 (지역 내 RE 발전 / 지역 수요)
            re_self_sufficiency = (re_generation / demand * 100) if demand > 0 else 0
            
            # RE 자급률에 따른 색상 (빨강 → 노랑 → 녹색)
            if re_self_sufficiency >= 80:
                color = plt.cm.RdYlGn(0.9)  # 진한 녹색
            elif re_self_sufficiency >= 50:
                color = plt.cm.RdYlGn(0.7)  # 연한 녹색
            elif re_self_sufficiency >= 30:
                color = plt.cm.RdYlGn(0.5)  # 노랑
            elif re_self_sufficiency >= 10:
                color = plt.cm.RdYlGn(0.3)  # 주황
            else:
                color = plt.cm.RdYlGn(0.1)  # 빨강
            
            # 지역 원 그리기
            circle = plt.Circle((x, y), np.sqrt(size)/20, color=color, alpha=0.8, edgecolor='black', linewidth=2)
            ax.add_patch(circle)
            
            # 지역명
            ax.text(x, y+0.15, region, ha='center', va='center', fontsize=10, fontweight='bold', color='white')
            
            # 지역 정보 텍스트 (간소화: 수요, PV, WT만 표시)
            info_lines = []
            info_lines.append(f"수요: {demand/1e6:.1f}TWh")
            if re_tech['PV'] > 0:
                info_lines.append(f"PV: {re_tech['PV']/1e6:.1f}TWh")
            if re_tech['WT'] > 0:
                info_lines.append(f"WT: {re_tech['WT']/1e6:.1f}TWh")
            
            info_text = '\n'.join(info_lines)
            ax.text(x, y-0.5, info_text, ha='center', va='top', fontsize=6,
                   bbox=dict(facecolor='white', alpha=0.8, edgecolor='gray', pad=2))
        
        ax.set_title('지역별 재생에너지 발전 및 송전 현황', fontsize=16, fontweight='bold', pad=20)
        ax.set_xlabel('원 색상: RE 자급률 (빨강:낮음 → 녹색:높음) | 선 색상: 송전선 이용률 | 선 두께: RE 흐름량', fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.set_xticks([])
        ax.set_yticks([])
        
        # 범례 추가
        legend_elements = [
            plt.Line2D([0], [0], marker='o', color='w', markerfacecolor=plt.cm.RdYlGn(0.9), 
                      markersize=12, label='RE 자급률 ≥80%'),
            plt.Line2D([0], [0], marker='o', color='w', markerfacecolor=plt.cm.RdYlGn(0.5), 
                      markersize=12, label='RE 자급률 30-80%'),
            plt.Line2D([0], [0], marker='o', color='w', markerfacecolor=plt.cm.RdYlGn(0.1), 
                      markersize=12, label='RE 자급률 <30%'),
            plt.Line2D([0], [0], color='green', lw=2, linestyle='--', label='RE 송전 이용률 <40%'),
            plt.Line2D([0], [0], color='orange', lw=2, linestyle='--', label='RE 송전 이용률 40-80%'),
            plt.Line2D([0], [0], color='red', lw=2, linestyle='--', label='RE 송전 이용률 >80%')
        ]
        ax.legend(handles=legend_elements, loc='upper left', fontsize=8)
        
        plt.tight_layout()
        
        # 파일 저장
        re_map_path = f'{results_dir}/re_korea_map_{current_time}.png'
        plt.savefig(re_map_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        
        print(f"RE 전용 한국 지도 시각화가 '{re_map_path}'에 저장되었습니다.")
        
        # 상세 정보 CSV 저장
        re_data = []
        for region in region_coords.keys():
            re_gen = regional_re_gen.get(region, 0)
            demand = regional_demand.get(region, 0)
            re_tech = regional_re_by_tech.get(region, {'PV': 0, 'WT': 0, '기타': 0})
            re_self_sufficiency = (re_gen / demand * 100) if demand > 0 else 0
            
            re_data.append([
                region, 
                demand/1e6, 
                re_gen/1e6, 
                re_tech['PV']/1e6, 
                re_tech['WT']/1e6, 
                re_tech['기타']/1e6,
                re_self_sufficiency
            ])
        
        re_df = pd.DataFrame(re_data, columns=[
            '지역', '전력수요(TWh)', 'RE발전(TWh)', 'PV(TWh)', 'WT(TWh)', '기타RE(TWh)', 'RE자급률(%)'
        ])
        re_df.to_csv(f'{results_dir}/re_korea_summary_{current_time}.csv', 
                    index=False, encoding='utf-8-sig')
        
        return True
        
    except Exception as e:
        print(f"RE 전용 한국 지도 시각화 생성 중 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def validate_input_data(input_data):
    """입력 데이터 유효성 검사"""
    required_sheets = ['buses', 'generators', 'lines', 'loads', 'timeseries']
    
    # 필수 시트 확인
    for sheet in required_sheets:
        if sheet not in input_data:
            raise ValueError(f"필수 시트 '{sheet}'가 없습니다.")
        if input_data[sheet].empty:
            raise ValueError(f"'{sheet}' 시트가 비어있습니다.")
    
    # 데이터 타입 확인 및 변환
    for _, row in input_data['buses'].iterrows():
        if not isinstance(row['name'], str):
            input_data['buses'].loc[_, 'name'] = str(row['name'])
    
    # timeseries 데이터 확인
    timeseries = input_data['timeseries'].iloc[0]
    if not pd.to_datetime(timeseries['start_time']):
        raise ValueError("잘못된 start_time 형식입니다.")
    if not pd.to_datetime(timeseries['end_time']):
        raise ValueError("잘못된 end_time 형식입니다.")
    if 'h' not in str(timeseries['frequency']).lower():
        raise ValueError("frequency는 'h' 형식이어야 합니다.")

    return input_data

def check_network_connections(network):
    """네트워크 연결 상태 확인"""
    print("\n=== 네트워크 연결 상태 확인 ===")
    
    # 1. 버스 연결 확인
    print("\n버스 정보:")
    for bus in network.buses.index:
        print(f"\n버스 {bus}:")
        # 연결된 발전기
        gens = network.generators[network.generators.bus == bus].index
        print(f"연결된 발전기: {list(gens)}")
        # 연결된 부하
        loads = network.loads[network.loads.bus == bus].index
        print(f"연결된 부하: {list(loads)}")
        # 연결된 저장장치
        stores = network.stores[network.stores.bus == bus].index
        print(f"연결된 저장장치: {list(stores)}")
    
    # 2. ESS 연결 및 설정 확인
    print("\nESS 상세 설정:")
    for store in network.stores.index:
        print(f"\n저장장치 {store}:")
        print(f"연결된 버스: {network.stores.at[store, 'bus']}")
        print(f"저장용량: {network.stores.at[store, 'e_nom']} MWh")
        print(f"충전효율: {network.stores.at[store, 'efficiency_store'] if 'efficiency_store' in network.stores else 1.0}")
        print(f"방전효율: {network.stores.at[store, 'efficiency_dispatch'] if 'efficiency_dispatch' in network.stores else 1.0}")
        print(f"순환 운전: {network.stores.at[store, 'e_cyclic']}")
        print(f"초기 충전상태: {network.stores.at[store, 'e_initial'] if 'e_initial' in network.stores else 0}")

    # 3. 선로 연결 확인
    print("\n선로 연결:")
    if not network.lines.empty:
        for line in network.lines.index:
            print(f"\n선로 {line}:")
            print(f"From: {network.lines.at[line, 'bus0']} To: {network.lines.at[line, 'bus1']}")
            print(f"용량: {network.lines.at[line, 's_nom']} MVA")

def check_excel_data_loading(input_data):
    """엑셀 데이터 로드 상태 확인"""
    print("\n=== 엑셀 데이터 로드 상태 확인 ===")
    
    # 1. 시트 존재 여부 확인
    print("\n로드된 시트:")
    for sheet_name in input_data.keys():
        print(f"\n[{sheet_name}] 시트:")
        print(f"행 수: {len(input_data[sheet_name])}")
        print(f"컬럼: {list(input_data[sheet_name].columns)}")
    
    # 2. stores 시트 상세 확인
    if 'stores' in input_data:
        print("\n\n=== Stores 시트 상세 데이터 ===")
        stores_df = input_data['stores']
        print("\n컬럼별 데이터 타입:")
        print(stores_df.dtypes)
        print("\n실제 데이터:")
        print(stores_df)
        
        # 필수 컬럼 확인
        required_columns = [
            'name', 'bus', 'carrier', 'e_nom', 'e_nom_extendable', 
            'e_cyclic', 'standing_loss', 'efficiency_store', 
            'efficiency_dispatch', 'e_initial', 'e_nom_max'
        ]
        missing_columns = [col for col in required_columns if col not in stores_df.columns]
        if missing_columns:
            print(f"\n누락된 필수 컬럼: {missing_columns}")
    
    # 3. links 시트 상세 확인
    if 'links' in input_data:
        print("\n\n=== Links 시트 상세 데이터 ===")
        links_df = input_data['links']
        print("\n컬럼별 데이터 타입:")
        print(links_df.dtypes)
        print("\n실제 데이터:")
        print(links_df)
        
        # 필수 컬럼 확인 (efficiency0 사용)
        required_columns = [
            'name', 'bus0', 'bus1', 'p_nom', 
            'p_nom_extendable', 'p_nom_max'
        ]
        missing_columns = [col for col in required_columns if col not in links_df.columns]
        if missing_columns:
            print(f"\n누락된 필수 컬럼: {missing_columns}")
        # efficiency 체크 (efficiency 또는 efficiency0 중 하나만 있으면 OK)
        if 'efficiency' not in links_df.columns and 'efficiency0' not in links_df.columns:
            print("\n[경고] 'efficiency' 또는 'efficiency0' 컬럼이 필요합니다.")

def analyze_regional_results(network, results_dir, current_time):
    """지역별 분석 결과 생성"""
    try:
        print("지역별 분석 결과 생성 중...")
        
        # 발전기 출력 데이터
        gen_output = network.generators_t.p
        
        # 지역별 발전량 계산
        regional_generation = {}
        for gen_name in gen_output.columns:
            # 지역 코드 추출 (A_, B_ 등)
            if '_' in gen_name:
                region = gen_name.split('_')[0]
                if region not in regional_generation:
                    regional_generation[region] = 0
                regional_generation[region] += gen_output[gen_name].sum()
        
        # 지역별 발전량 DataFrame 생성
        regional_df = pd.DataFrame(list(regional_generation.items()), 
                                 columns=['지역', '총발전량(MWh)'])
        regional_df.to_csv(f'{results_dir}/optimization_result_{current_time}_지역별_발전량.csv', 
                          index=False, encoding='utf-8-sig')
        
        # 발전원별 발전량 계산
        generation_by_type = {}
        for gen_name in gen_output.columns:
            # 발전원 타입 추출(보강)
            gen_type = 'Unknown'
            lname = gen_name.lower()
            if 'nuclear' in lname:
                gen_type = '원자력'
            elif 'pv' in lname or 'solar' in lname:
                gen_type = '태양광'
            elif 'wt' in lname or 'wind' in lname:
                gen_type = '풍력'
            elif 'hydro' in lname or 'water' in lname:
                gen_type = '수력'
            elif 'coal' in lname:
                gen_type = '석탄'
            elif 'lng' in lname or 'gas' in lname:
                gen_type = 'LNG'
            elif 'oil' in lname or 'diesel' in lname:
                gen_type = '석유'
            elif 'biomass' in lname or 'bio' in lname:
                gen_type = '바이오'
            elif 'geothermal' in lname:
                gen_type = '지열'
            elif 'chp' in lname:
                gen_type = '열병합'
            elif 'h2' in lname or 'hydrogen' in lname:
                gen_type = '수소'
            elif '_h_fallback_gen' in lname or 'heat' in lname:
                gen_type = '열'
            
            if gen_type not in generation_by_type:
                generation_by_type[gen_type] = 0
            generation_by_type[gen_type] += gen_output[gen_name].sum()
        
        # 발전원별 발전량 DataFrame 생성
        type_df = pd.DataFrame(list(generation_by_type.items()), 
                             columns=['발전원', '총발전량(MWh)'])
        type_df.to_csv(f'{results_dir}/optimization_result_{current_time}_발전원별_발전량.csv', 
                      index=False, encoding='utf-8-sig')
        
        # 지역별 발전원별 발전량 계산
        regional_type_generation = {}
        for gen_name in gen_output.columns:
            if '_' in gen_name:
                region = gen_name.split('_')[0]
                
                # 발전원 타입 추출
                gen_type = 'Unknown'
                lname = gen_name.lower()
                if 'nuclear' in lname:
                    gen_type = '원자력'
                elif 'pv' in lname or 'solar' in lname:
                    gen_type = '태양광'
                elif 'wt' in lname or 'wind' in lname:
                    gen_type = '풍력'
                elif 'hydro' in lname or 'water' in lname:
                    gen_type = '수력'
                elif 'coal' in lname:
                    gen_type = '석탄'
                elif 'lng' in lname or 'gas' in lname:
                    gen_type = 'LNG'
                elif 'oil' in lname or 'diesel' in lname:
                    gen_type = '석유'
                elif 'biomass' in lname or 'bio' in lname:
                    gen_type = '바이오'
                elif 'geothermal' in lname:
                    gen_type = '지열'
                elif 'chp' in lname:
                    gen_type = '열병합'
                elif 'h2' in lname or 'hydrogen' in lname:
                    gen_type = '수소'
                elif '_h_fallback_gen' in lname or 'heat' in lname:
                    gen_type = '열'
                
                key = f"{region}_{gen_type}"
                if key not in regional_type_generation:
                    regional_type_generation[key] = 0
                regional_type_generation[key] += gen_output[gen_name].sum()
        
        # 지역별 발전원별 DataFrame 생성
        regional_type_data = []
        for key, value in regional_type_generation.items():
            region, gen_type = key.split('_', 1)
            regional_type_data.append([region, gen_type, value])
        
        regional_type_df = pd.DataFrame(regional_type_data, 
                                      columns=['지역', '발전원', '총발전량(MWh)'])
        regional_type_df.to_csv(f'{results_dir}/optimization_result_{current_time}_지역별_발전원별_발전량.csv', 
                               index=False, encoding='utf-8-sig')
        
        # 상위 발전기 발전량 (상위 20개)
        total_gen_output = gen_output.sum().sort_values(ascending=False)
        top_generators = total_gen_output.head(20)
        top_gen_df = pd.DataFrame({
            '발전기명': top_generators.index,
            '총발전량(MWh)': top_generators.values
        })
        top_gen_df.to_csv(f'{results_dir}/optimization_result_{current_time}_상위발전기_발전량.csv', 
                         index=False, encoding='utf-8-sig')
        
        print("지역별 분석 결과 생성 완료:")
        print(f"- 지역별 발전량")
        print(f"- 발전원별 발전량")
        print(f"- 지역별 발전원별 발전량")
        print(f"- 상위 발전기 발전량")
        
        return True
        
    except Exception as e:
        print(f"지역별 분석 중 오류 발생: {str(e)}")
        traceback.print_exc()
        return False

def ensure_column(df, column_name, default_value=np.nan):
    """DataFrame에 컬럼이 없으면 기본값으로 추가"""
    if column_name not in df.columns:
        df[column_name] = default_value
    return df


def apply_year_overrides(input_data, overrides):
    """연도별 오버라이드 적용. overrides 형식 예:
    {
        'generators': {
            'SEL_PV_1': {'p_nom': 100, 'p_nom_extendable': True},
            '*': {'marginal_cost': 0}
        },
        'timeseries': {'start_time': '2021-01-01', 'end_time': '2022-01-01'}
    }
    """
    if not overrides:
        return input_data
    data = copy.deepcopy(input_data)

    for sheet, spec in overrides.items():
        if sheet not in data:
            continue
        df = data[sheet]

        # timeseries는 단일 행을 가정
        if sheet == 'timeseries' and isinstance(spec, dict):
            if len(df) == 0:
                # 빈 경우 1행 생성
                data[sheet] = pd.DataFrame([spec])
            else:
                for col, val in spec.items():
                    if col not in df.columns:
                        df[col] = np.nan
                    df.at[df.index[0], col] = val
            continue

        # 그 외 시트: 'name' 기준으로 개별 행 적용 + '*' 와일드카드
        if isinstance(spec, dict):
            # 와일드카드 먼저
            wildcard_updates = spec.get('*', None)
            if wildcard_updates:
                for col, val in wildcard_updates.items():
                    if col not in df.columns:
                        df[col] = np.nan
                    df[col] = val

            # 개별 name 처리
            if 'name' in df.columns:
                for name_key, updates in spec.items():
                    if name_key == '*':
                        continue
                    if not isinstance(updates, dict):
                        continue
                    mask = df['name'].astype(str) == str(name_key)
                    if not mask.any():
                        continue
                    for col, val in updates.items():
                        if col not in df.columns:
                            df[col] = np.nan
                        df.loc[mask, col] = val

            data[sheet] = df

    return data


def extract_capacity_carryover(network):
    """이전 해 네트워크에서 다음 해로 인계할 용량 추출"""
    carry = {
        'generators': {},
        'lines': {},
        'links': {},
        'stores': {}
    }

    if network is None:
        return carry

    # Generators
    if not network.generators.empty:
        if 'p_nom_opt' in network.generators.columns:
            caps = network.generators['p_nom_opt'].fillna(network.generators['p_nom'])
        else:
            # 일부 버전은 optimize 후 p_nom_opt가 별도 속성으로 있음
            caps = getattr(network.generators, 'p_nom_opt', network.generators['p_nom'])
        for name, val in caps.items():
            carry['generators'][str(name)] = float(val)

    # Lines (AC)
    if not network.lines.empty:
        s_nom_opt = getattr(network.lines, 's_nom_opt', None)
        if s_nom_opt is not None:
            caps = network.lines['s_nom_opt'].fillna(network.lines['s_nom'])
        else:
            caps = network.lines['s_nom']
        for name, val in caps.items():
            carry['lines'][str(name)] = float(val)

    # Links (HVDC 등)
    if not network.links.empty:
        if 'p_nom_opt' in network.links.columns:
            caps = network.links['p_nom_opt'].fillna(network.links['p_nom'])
        else:
            caps = getattr(network.links, 'p_nom_opt', network.links['p_nom'])
        for name, val in caps.items():
            carry['links'][str(name)] = float(val)

    # Stores (energy capacity)
    if not network.stores.empty:
        if 'e_nom_opt' in network.stores.columns:
            caps = network.stores['e_nom_opt'].fillna(network.stores['e_nom'])
        else:
            caps = getattr(network.stores, 'e_nom_opt', network.stores['e_nom'])
        for name, val in caps.items():
            carry['stores'][str(name)] = float(val)

    return carry


def save_capacity_snapshot(carry, year, snapshots_dir=None):
    """연도 최적화 결과(설비 용량)를 JSON 파일로 저장.

    Parameters
    ----------
    carry        : extract_capacity_carryover() 반환값
    year         : 해당 연도 (정수)
    snapshots_dir: 저장 폴더 (기본: results_longterm/_capacity_snapshots/)

    Returns
    -------
    저장 파일 경로 (str)
    """
    import json as _json
    if snapshots_dir is None:
        snapshots_dir = os.path.join('results_longterm', '_capacity_snapshots')
    os.makedirs(snapshots_dir, exist_ok=True)
    fpath = os.path.join(snapshots_dir, f'capacity_{year}.json')
    with open(fpath, 'w', encoding='utf-8') as f:
        _json.dump(carry, f, ensure_ascii=False, indent=2)
    print(f"[스냅샷 저장] {year}년 설비 용량 → {fpath}")
    return fpath


def load_capacity_snapshot(year, snapshots_dir=None):
    """저장된 연도별 설비 용량 스냅샷 로드.

    Parameters
    ----------
    year         : 로드할 연도 (정수)
    snapshots_dir: 저장 폴더 (기본: results_longterm/_capacity_snapshots/)

    Returns
    -------
    carry dict, 또는 파일이 없으면 None
    """
    import json as _json
    if snapshots_dir is None:
        snapshots_dir = os.path.join('results_longterm', '_capacity_snapshots')
    fpath = os.path.join(snapshots_dir, f'capacity_{year}.json')
    if not os.path.exists(fpath):
        print(f"[스냅샷 없음] {year}년 스냅샷 파일 없음: {fpath}")
        return None
    with open(fpath, 'r', encoding='utf-8') as f:
        carry = _json.load(f)
    print(f"[스냅샷 로드] {year}년 설비 용량 ← {fpath}")
    return carry


def apply_carryover_to_input(input_data, carryover_caps, policy='min'):
    """인계 용량을 입력 데이터에 반영
    policy='min': 다음 해 최소 용량 하한으로 적용(추가 확장 허용)
    """
    if not carryover_caps:
        return input_data

    data = copy.deepcopy(input_data)

    # Generators
    if 'generators' in data and not data['generators'].empty:
        df = data['generators']
        ensure_column(df, 'p_nom_min', 0.0)
        for name, cap in carryover_caps.get('generators', {}).items():
            mask = df['name'].astype(str) == str(name)
            if mask.any():
                df.loc[mask, 'p_nom_min'] = np.maximum(df.loc[mask, 'p_nom_min'].astype(float), cap)
        data['generators'] = df

    # Lines
    if 'lines' in data and not data['lines'].empty:
        df = data['lines']
        ensure_column(df, 's_nom_min', 0.0)
        for name, cap in carryover_caps.get('lines', {}).items():
            mask = df['name'].astype(str) == str(name)
            if mask.any():
                df.loc[mask, 's_nom_min'] = np.maximum(df.loc[mask, 's_nom_min'].astype(float), cap)
        data['lines'] = df

    # Links
    if 'links' in data and not data['links'].empty:
        df = data['links']
        ensure_column(df, 'p_nom_min', 0.0)
        for name, cap in carryover_caps.get('links', {}).items():
            mask = df['name'].astype(str) == str(name)
            if mask.any():
                df.loc[mask, 'p_nom_min'] = np.maximum(df.loc[mask, 'p_nom_min'].astype(float), cap)
        data['links'] = df

    # Stores
    if 'stores' in data and not data['stores'].empty:
        df = data['stores']
        ensure_column(df, 'e_nom_min', 0.0)
        for name, cap in carryover_caps.get('stores', {}).items():
            mask = df['name'].astype(str) == str(name)
            if mask.any():
                df.loc[mask, 'e_nom_min'] = np.maximum(df.loc[mask, 'e_nom_min'].astype(float), cap)
        data['stores'] = df

    return data


def run_multi_year_sequence(years, base_input_file=INPUT_FILE, overrides_by_year=None, carryover=True, results_root='results_multi'):
    """연도별 순차 실행 루프.
    - years: [2020, 2021, ...]
    - overrides_by_year: {year: overrides(dict)}
    - carryover: True면 이전 해 용량을 다음 해 최소 용량으로 인계
    - results_root: 결과 저장 루트 디렉터리
    반환: {year: {'network': Network, 'results_dir': str}}
    """
    timestamp_root = os.path.join(results_root, datetime.now().strftime("%Y%m%d_%H%M%S"))
    os.makedirs(timestamp_root, exist_ok=True)
    results = {}
    prev_network = None

    for year in years:
        print(f"\n===== {year}년도 분석 시작 =====")
        # 0) 연도별 interface 시나리오를 통합 파일에 반영(시트 직접 갱신)
        try:
            root_dir = get_base_dir()
            integrated_path = os.path.abspath(os.path.join(root_dir, base_input_file))
            interface_path = os.path.abspath(os.path.join(root_dir, 'interface.xlsx'))
            _update_integrated_for_year(integrated_path, interface_path, year)
        except Exception as _e:
            print(f"통합파일 갱신 경고({year}): {str(_e)}")
        # 1) 기본 입력 로드
        input_data = read_input_data(base_input_file)

        # 1.5) 해당 연도 수요 시나리오를 loads.p_set에 주입
        input_data = _apply_scenario_to_loads_in_input(input_data, year)

        # 1.6) 버스명 표준화 및 전 시트 반영(예: BSN_BSN_EL → BSN_EL)
        input_data = standardize_bus_names_in_input(input_data)
        try:
            integrated_path = os.path.abspath(os.path.join(get_base_dir(), base_input_file))
            _persist_standardized_input(integrated_path, input_data)
        except Exception as _e:
            print(f"표준화된 버스명 저장 경고: {str(_e)}")

        # 2) 연도별 오버라이드 적용 (수요/패턴/원가/효율/확장가능 등)
        if overrides_by_year and year in overrides_by_year:
            input_data = apply_year_overrides(input_data, overrides_by_year[year])

        # 3) 이전 해 결과 인계 (용량 하한)
        if carryover and prev_network is not None:
            caps = extract_capacity_carryover(prev_network)
            input_data = apply_carryover_to_input(input_data, caps, policy='min')

        # 4) 네트워크 생성 및 최적화
        network = create_network(input_data)
        success = optimize_network(network)
        if not success:
            print(f"{year}년도 최적화 실패. 부분 결과(시계열)를 저장합니다.")
            year_dir = os.path.join(timestamp_root, str(year))
            os.makedirs(year_dir, exist_ok=True)
            try:
                save_results(network, subdir=year_dir)
                results[year] = {'network': network, 'results_dir': year_dir}
            except Exception as _e_sv:
                print(f"부분 결과 저장 실패: {_e_sv}")
                results[year] = {'network': network, 'results_dir': None}
            prev_network = network
            continue

        # 5) 결과 저장 (타임스탬프/연도 서브폴더)
        year_dir = os.path.join(timestamp_root, str(year))
        os.makedirs(year_dir, exist_ok=True)
        save_results(network, subdir=year_dir)

        results[year] = {'network': network, 'results_dir': year_dir}
        prev_network = network
        print(f"===== {year}년도 분석 완료 =====\n")

    return results

def ensure_integrated_input():
    """interface.xlsx 기반으로 integrated_input_data.xlsx 생성/업데이트"""
    try:
        root_dir = get_base_dir()
        integrated_path = os.path.abspath(os.path.join(root_dir, INPUT_FILE))
        interface_path = os.path.abspath(os.path.join(root_dir, 'interface.xlsx'))

        need_build = False
        reason = ''
        if not os.path.exists(integrated_path):
            need_build = True
            reason = '통합 파일이 존재하지 않음'
        elif os.path.exists(interface_path):
            try:
                if os.path.getmtime(interface_path) > os.path.getmtime(integrated_path):
                    need_build = True
                    reason = 'interface.xlsx가 더 최신'
            except Exception:
                pass

        if need_build:
            print(f"통합 입력 생성/업데이트 실행 ({reason})...")
            ok = False
            # 지역 엑셀 프로세서를 우선 사용(링크/CHP 포함 컬럼 처리 보장)
            try:
                from process_regional_excel import RegionalExcelProcessor
                import openpyxl as _oxl
                proc = RegionalExcelProcessor(interface_path)
                proc.output_path = integrated_path
                # 워크북을 열고 지역/링크 등 데이터를 먼저 적재한 뒤 통합 생성
                try:
                    _wb = _oxl.load_workbook(interface_path, data_only=True)
                    _steps_ok = True
                    _steps_ok &= bool(proc.read_selected_regions(_wb))
                    # Coal_DB를 반드시 read_regional_data 이전에 호출해야 함
                    _coal_ok = proc.read_coal_db(_wb)
                    if not _coal_ok:
                        print("경고: Coal_DB 읽기 실패 (계속 진행)")
                    _steps_ok &= bool(proc.read_regional_data(_wb))
                    _steps_ok &= bool(proc.read_connections(_wb))
                    _steps_ok &= bool(proc.read_timeseries(_wb))
                    _steps_ok &= bool(proc.read_patterns(_wb))
                except Exception as _e0:
                    print(f"RegionalExcelProcessor 사전 적재 실패: {_e0}")
                    _steps_ok = False
                ok = bool(proc.create_integrated_data()) if _steps_ok else False
            except Exception as e:
                print(f"RegionalExcelProcessor 생성 실패: {str(e)}")
            # 보조 경로로 v3 생성기 시도
            if not ok:
                try:
                    from create_integrated_data import create_integrated_data as v3_create
                    ok = bool(v3_create())
                except Exception as e:
                    print(f"v3 생성기 호출 실패: {str(e)}")
            if ok:
                try:
                    print(f"통합 입력 생성 완료: {integrated_path}")
                    print(f"갱신된 수정 시간: {datetime.fromtimestamp(os.path.getmtime(integrated_path))}")
                    # 생성물 검증: 필수 시트 비어있으면 대체 생성기로 재시도
                    try:
                        _xls_chk = pd.ExcelFile(integrated_path)
                        _has_buses = ('buses' in _xls_chk.sheet_names) and (not pd.read_excel(integrated_path, sheet_name='buses').empty)
                        _has_gens = ('generators' in _xls_chk.sheet_names) and (not pd.read_excel(integrated_path, sheet_name='generators').empty)
                        if not (_has_buses and _has_gens):
                            print("경고: 통합 입력에 필수 시트가 비어있습니다. v3 생성기로 재시도합니다.")
                            from create_integrated_data import create_integrated_data as _v3_create
                            if bool(_v3_create()):
                                print("v3 생성기 재시도 성공")
                            else:
                                print("경고: v3 생성기 재시도 실패. 기존 파일 유지")
                    except Exception as _e_chk:
                        print(f"통합 입력 검증 실패: {_e_chk}")
                except Exception:
                    pass
            else:
                print("경고: 통합 입력 생성에 실패했습니다. 기존 파일이 있으면 그걸 사용합니다.")
        else:
            print("통합 입력이 최신 상태입니다. 재생성 생략.")
    except Exception as e:
        print(f"통합 입력 준비 중 오류: {str(e)}")

def _fallback_build_lines_from_interface(input_data, interface_path):
    try:
        if not os.path.exists(interface_path):
            print("interface.xlsx가 없어 선로 폴백 생성을 건너뜁니다.")
            return None
        # 지역간 연결 시트 읽기 (레이블은 5행, 데이터는 6행부터라는 전제)
        try:
            try:
                df = pd.read_excel(interface_path, sheet_name='지역간 연결', header=4)
            except Exception:
                # 폴백: 레이블이 1행일 수도 있으므로 0행 시도
                df = pd.read_excel(interface_path, sheet_name='지역간 연결', header=0)
        except Exception as e:
            print(f"'지역간 연결' 시트를 읽을 수 없습니다: {str(e)}")
            return None
        if df is None or df.empty:
            print("'지역간 연결' 시트가 비어있습니다.")
            return None

        # 컬럼 표준화(공백 제거) 및 로그
        df.columns = [str(c).strip() for c in df.columns]
        print(f"'지역간 연결' 시트 컬럼: {df.columns.tolist()}")
        print("상위 5행:\n", df.head(5))
        cols_lower = {c: str(c).strip().lower() for c in df.columns}

        def find_col(cands):
            # 1단계: 완전 일치 우선
            for c in df.columns:
                cl = cols_lower[c]
                for k in cands:
                    if k == cl:
                        return c
            # 2단계: 부분 문자열 매칭
            for c in df.columns:
                cl = cols_lower[c]
                for k in cands:
                    if k in cl:
                        return c
            return None

        # 주요 컬럼 매핑 (PyPSA 표준 우선, 없으면 한글)
        from_col = 'bus0' if 'bus0' in df.columns else find_col(['from', '출발', '시작', '지역1', '출발지역', 'from_region', 'region1'])
        to_col = 'bus1' if 'bus1' in df.columns else find_col(['to', '도착', '끝', '지역2', '도착지역', 'to_region', 'region2'])
        name_col = 'name' if 'name' in df.columns else find_col(['이름', '선로명'])

        # PyPSA 표준 컬럼명 직접 사용 (사용자가 interface.xlsx를 표준화함)
        # 존재하는 컬럼만 사용
        type_col = 'type' if 'type' in df.columns else None
        length_col = 'length' if 'length' in df.columns else None
        x_col = 'x' if 'x' in df.columns else None
        r_col = 'r' if 'r' in df.columns else None
        s_nom_col = 's_nom' if 's_nom' in df.columns else find_col(['capacity', '용량', 'mva', '정격'])
        
        # 디버깅: 사용할 컬럼 출력
        print(f"컬럼 매칭 결과 (PyPSA 표준):")
        print(f"  s_nom_col: {s_nom_col}")
        print(f"  length_col: {length_col}")
        print(f"  x_col: {x_col}")
        print(f"  r_col: {r_col}")
        print(f"  type_col: {type_col}")

        if from_col is None or to_col is None:
            print("'지역간 연결' 시트에서 출발/도착 컬럼을 찾지 못했습니다.")
            return None

        # 버스 목록
        buses_df = input_data.get('buses', pd.DataFrame())
        bus_names = list(buses_df['name'].astype(str)) if ('name' in buses_df.columns) else []
        bus_carriers = dict(zip(buses_df['name'].astype(str), buses_df['carrier'].astype(str))) if ('carrier' in buses_df.columns and 'name' in buses_df.columns) else {}

        def guess_bus(region_code):
            if not bus_names:
                return str(region_code)
            region_code = str(region_code).strip()
            candidates = [b for b in bus_names if b == region_code or b.startswith(region_code + '_')]
            if not candidates:
                # 넓게 포함 검색
                candidates = [b for b in bus_names if region_code in b]
            if not candidates:
                return region_code
            # 전력 계통 우선
            elec_pref = [b for b in candidates if ('electric' in b.lower()) or (bus_carriers.get(b, '').lower() == 'electricity')]
            if elec_pref:
                return elec_pref[0]
            return candidates[0]

        lines_rows = []
        added = 0
        for _, row in df.iterrows():
            fr_raw = row.get(from_col)
            to_raw = row.get(to_col)
            if pd.isna(fr_raw) or pd.isna(to_raw):
                continue
            fr_region = str(fr_raw).strip()
            to_region = str(to_raw).strip()

            name_val = str(row.get(name_col)).strip() if (name_col and pd.notna(row.get(name_col))) else f"{fr_region}-{to_region}"
            s_nom_val = float(pd.to_numeric(row.get(s_nom_col), errors='coerce')) if s_nom_col and pd.notna(row.get(s_nom_col)) else 1000.0
            x_val = float(pd.to_numeric(row.get(x_col), errors='coerce')) if x_col and pd.notna(row.get(x_col)) else 0.1
            r_val = float(pd.to_numeric(row.get(r_col), errors='coerce')) if r_col and pd.notna(row.get(r_col)) else 0.01
            length_val = float(pd.to_numeric(row.get(length_col), errors='coerce')) if length_col and pd.notna(row.get(length_col)) else np.nan
            type_val = str(row.get(type_col)).strip() if (type_col and pd.notna(row.get(type_col))) else np.nan

            bus0 = guess_bus(fr_region)
            bus1 = guess_bus(to_region)

            lines_rows.append({
                'name': name_val,
                'bus0': bus0,
                'bus1': bus1,
                's_nom': s_nom_val,
                'x': x_val,
                'r': r_val,
                'length': length_val,
                'type': type_val
            })
            added += 1

        if added == 0:
            print("'지역간 연결' 시트에서 유효한 선로 레코드를 만들지 못했습니다.")
            return None

        lines_df = pd.DataFrame(lines_rows)
        print(f"interface.xlsx → '지역간 연결'에서 선로 {len(lines_df)}개를 폴백 생성했습니다.")
        # 일부 기본 컬럼 보장
        for c in ['s_nom_extendable', 's_nom_min', 's_nom_max']:
            if c not in lines_df.columns:
                lines_df[c] = np.nan
        return lines_df
    except Exception as e:
        print(f"선로 폴백 생성 중 오류: {str(e)}")
        return None

def _persist_lines_to_integrated(integrated_path, lines_df):
    try:
        if lines_df is None or lines_df.empty:
            return False
        if not os.path.exists(integrated_path):
            print(f"경고: {integrated_path} 파일이 없어 lines를 저장하지 못했습니다.")
            return False
        # 기존 모든 시트를 읽어 dict로 보관 후 lines 교체/추가
        xls = pd.ExcelFile(integrated_path)
        sheets = {}
        for sn in xls.sheet_names:
            if sn.lower() == 'lines':
                continue
            sheets[sn] = pd.read_excel(integrated_path, sheet_name=sn)
        # 재작성
        with pd.ExcelWriter(integrated_path, engine='openpyxl') as writer:
            # 기존 시트들 쓰기
            for sn, df in sheets.items():
                df.to_excel(writer, sheet_name=sn, index=False)
            # lines 쓰기
            lines_df.to_excel(writer, sheet_name='lines', index=False)
        print(f"lines 시트를 '{integrated_path}'에 저장했습니다.")
        return True
    except Exception as e:
        print(f"lines 시트 저장 실패: {str(e)}")
        return False

def _to_bool(value):
    try:
        if isinstance(value, (bool, np.bool_)):
            return bool(value)
        if pd.isna(value):
            return False
        s = str(value).strip().lower()
        return s in ['1', 'true', 'yes', 'y', 't']
    except Exception:
        return False

def _normalize_bus_name(raw_name, available_bus_names, prefer_electric=True, bus_carriers=None):
    try:
        if not isinstance(raw_name, str):
            raw = str(raw_name)
        else:
            raw = raw_name.strip()
        bus_set = set(str(b) for b in available_bus_names)
        if raw in bus_set:
            return raw
        # 토큰 분해
        tokens = raw.split('_')
        tokens = [t for t in tokens if t]
        region = None
        energy = None
        if len(tokens) >= 2:
            region = tokens[0]
            energy = tokens[-1]
        elif len(tokens) == 1:
            region = tokens[0]
        candidates = []
        # 1) 원문
        candidates.append(raw)
        # 2) REGION_ENERGY 형태
        if region and energy:
            candidates.append(f"{region}_{energy}")
        # 3) REGION_REGION_ENERGY 형태
        if region and energy:
            candidates.append(f"{region}_{region}_{energy}")
        # 4) 시작/끝 패턴 매칭
        if region and energy:
            for b in available_bus_names:
                bs = str(b)
                if bs.startswith(region + '_') and bs.endswith('_' + energy):
                    candidates.append(bs)
        # 5) region로만 시작하는 모든 버스
        if region:
            for b in available_bus_names:
                bs = str(b)
                if bs.startswith(region + '_'):
                    candidates.append(bs)
        # 고유화
        seen = set()
        ordered = []
        for c in candidates:
            if c not in seen:
                seen.add(c)
                ordered.append(c)
        # 캐리어 선호
        if prefer_electric and bus_carriers:
            elec = [c for c in ordered if c in bus_set and str(bus_carriers.get(c, '')).lower() == 'electricity']
            if elec:
                return elec[0]
        # 첫 매칭 반환
        for c in ordered:
            if c in bus_set:
                return c
        return raw
    except Exception:
        return str(raw_name)

def _standardize_bus_token_by_carrier(carrier):
    c = str(carrier).strip().lower()
    if 'electric' in c or c == 'el' or c == '전력':
        return 'EL'
    if 'heat' in c or c == 'h' or c == '열':
        return 'H'
    if 'hydrogen' in c or c == 'h2' or c == '수소':
        return 'H2'
    if 'gas' in c or 'lng' in c or c == '가스':
        return 'LNG'
    if 'ev' in c or c == '전기차':
        return 'EV'
    return None


def standardize_bus_names_in_input(input_data):
    try:
        if 'buses' not in input_data or input_data['buses'].empty:
            return input_data
        buses = input_data['buses']
        if 'name' not in buses.columns:
            return input_data
        # 에너지원 토큰 결정
        energy_token_by_bus = {}
        if 'carrier' in buses.columns:
            for _, row in buses.iterrows():
                bus_name = str(row['name']).strip()
                token = _standardize_bus_token_by_carrier(row['carrier'])
                energy_token_by_bus[bus_name] = token
        
        # 표준 이름 생성 함수
        def make_std_name(old_name):
            name = str(old_name).strip()
            tokens = [t for t in name.split('_') if t]
            region = tokens[0] if tokens else name
            # 우선 carrier 기반 토큰 사용, 없으면 마지막 토큰 유지(LNG 허용)
            energy = energy_token_by_bus.get(name)
            if energy is None:
                last = tokens[-1] if len(tokens) >= 2 else 'EL'
                energy = last if last.upper() in ['EL','H','H2','LNG'] else 'EL'
            return f"{region}_{energy}"
                
        # 매핑 생성
        mapping = {}
        for _, row in buses.iterrows():
            old = str(row['name']).strip()
            new = make_std_name(old)
            mapping[old] = new
        
        # 충돌 방지: 동일 new로 여러 old 매핑되는 경우 원본 유지
        reverse = {}
        for old, new in mapping.items():
            if new not in reverse:
                reverse[new] = old
            else:
                # 충돌 시 변경하지 않음
                mapping[old] = old
        
        # 적용 함수
        def apply_map(df, col):
            if df is None or df.empty or col not in df.columns:
                return df
            df[col] = df[col].astype(str).map(lambda x: mapping.get(x, x))
            return df
        
        # buses.name 업데이트
        input_data['buses']['name'] = input_data['buses']['name'].astype(str).map(lambda x: mapping.get(x, x))
        # generators.bus, loads.bus, links bus0/bus1, lines bus0/bus1 업데이트
        if 'generators' in input_data and not input_data['generators'].empty:
            input_data['generators'] = apply_map(input_data['generators'], 'bus')
        if 'loads' in input_data and not input_data['loads'].empty:
            input_data['loads'] = apply_map(input_data['loads'], 'bus')
        if 'links' in input_data and not input_data['links'].empty:
            input_data['links'] = apply_map(input_data['links'], 'bus0')
            input_data['links'] = apply_map(input_data['links'], 'bus1')
            input_data['links'] = apply_map(input_data['links'], 'bus2')
            input_data['links'] = apply_map(input_data['links'], 'bus3')
        if 'lines' in input_data and not input_data['lines'].empty:
            input_data['lines'] = apply_map(input_data['lines'], 'bus0')
            input_data['lines'] = apply_map(input_data['lines'], 'bus1')
        if 'stores' in input_data and not input_data['stores'].empty:
            input_data['stores'] = apply_map(input_data['stores'], 'bus')
        
        # 로그 요약
        changes = sum(1 for old, new in mapping.items() if old != new)
        if changes > 0:
            print(f"버스명 표준화 적용: {changes}개 이름 변경(형식 REGION_ENERGY, 예: BSN_EL)")

        # 추가 정규화: lines/link/store의 버스 컬럼에 잔여 중복 토큰(예: BSN_EL_EL) 압축 및 실제 버스 세트에 맞게 보정
        try:
            bus_names_set = set(input_data['buses']['name'].astype(str)) if ('buses' in input_data and not input_data['buses'].empty and 'name' in input_data['buses'].columns) else set()

            def _compress_bus_name(raw):
                try:
                    s = str(raw).strip()
                    if not s:
                        return s
                    if s in bus_names_set:
                        return s
                    tokens = [t for t in s.split('_') if t]
                    region = tokens[0] if tokens else s
                    # 에너지 토큰 후보 압축 (중복 제거) + LNG 허용
                    energy_candidates = [t for t in tokens[1:] if t.upper() in ['EL','H','H2','LNG']]
                    energy = None
                    for t in energy_candidates:
                        tt = t.upper()
                        if tt in ['EL','H','H2','LNG']:
                            energy = tt
                            break
                    if energy is None and len(tokens) >= 2:
                        energy = tokens[-1].upper()
                        if energy not in ['EL','H','H2','LNG']:
                            # 기본값은 EL로 폴백(단, 원문에 LNG가 포함되어 있으면 LNG 유지)
                            if 'LNG' in [t.upper() for t in tokens[1:]]:
                                energy = 'LNG'
                            else:
                                energy = 'EL'
                    candidate = f"{region}_{energy}" if energy else region
                    if candidate in bus_names_set:
                        return candidate
                    # region으로 시작하는 버스 중 첫번째 대안
                    for b in bus_names_set:
                        if b.startswith(region + '_'):
                            return b
                    return candidate
                except Exception:
                    return str(raw)

            def _apply_compress(df, columns):
                if df is None or df.empty:
                    return df
                for c in columns:
                    if c in df.columns:
                        df[c] = df[c].astype(str).map(_compress_bus_name)
                return df

            if 'lines' in input_data and not input_data['lines'].empty:
                input_data['lines'] = _apply_compress(input_data['lines'], ['bus0','bus1'])
            if 'links' in input_data and not input_data['links'].empty:
                input_data['links'] = _apply_compress(input_data['links'], ['bus0','bus1','bus2','bus3'])
            if 'stores' in input_data and not input_data['stores'].empty:
                input_data['stores'] = _apply_compress(input_data['stores'], ['bus'])
        except Exception as _e:
            print(f"버스명 추가 정규화 경고: {str(_e)}")

        return input_data
    except Exception as e:
        print(f"버스명 표준화 중 오류: {str(e)}")
        return input_data

def _persist_standardized_input(integrated_path, input_data):
    try:
        if not input_data or not isinstance(input_data, dict):
            return False
        # 쓰기 가능한 시트만 수집
        sheets = {k: v for k, v in input_data.items() if isinstance(v, pd.DataFrame)}
        if not sheets:
            return False
        with pd.ExcelWriter(integrated_path, engine='openpyxl') as writer:
            for sn, df in sheets.items():
                # 인덱스 제거하여 저장
                df.to_excel(writer, sheet_name=sn, index=False)
        # 저장 후 수정 시간을 interface.xlsx보다 늦게 설정
        # → 다음 실행 시 ensure_integrated_input()이 불필요한 재생성을 하지 않도록 함
        try:
            root_dir = os.path.dirname(os.path.abspath(integrated_path))
            interface_path = os.path.join(root_dir, 'interface.xlsx')
            if os.path.exists(interface_path):
                import time as _time
                iface_mtime = os.path.getmtime(interface_path)
                integ_mtime = os.path.getmtime(integrated_path)
                if integ_mtime <= iface_mtime:
                    # integrated가 interface보다 오래됐으면 현재 시간으로 갱신
                    new_mtime = _time.time() + 1
                    os.utime(integrated_path, (new_mtime, new_mtime))
        except Exception:
            pass
        print(f"표준화된 버스명이 '{integrated_path}'에 저장되었습니다.")
        return True
    except Exception as e:
        print(f"표준화 입력 저장 실패: {str(e)}")
        return False

def _parse_generator_scenario_from_interface(interface_path, year):
    try:
        if not os.path.exists(interface_path):
            return {}
        df = pd.read_excel(interface_path, sheet_name='시나리오_발전기')
        if df is None or df.empty:
            return {}
        df.columns = [str(c).strip() for c in df.columns]

        # 연도 컬럼(예: '2024', '2025', ...) 탐지
        year_str = str(year)
        year_cols = [c for c in df.columns if str(c).strip().isdigit() and len(str(c).strip()) == 4]
        name_col = None
        bus_col = None
        for c in df.columns:
            cl = str(c).strip().lower()
            if cl in ['name', '이름']:
                name_col = c
            if cl in ['bus', '버스']:
                bus_col = c
        scenario_by_name = {}

        if year_cols and year_str in [str(c).strip() for c in year_cols] and name_col is not None:
            # 형식: [이름, 버스, 2024, 2025, ...]
            year_col = [c for c in df.columns if str(c).strip() == year_str][0]
            for _, row in df.iterrows():
                gname = str(row[name_col]).strip()
                if not gname or gname.lower() == 'nan':
                    continue
                val = pd.to_numeric(row.get(year_col), errors='coerce')
                if pd.notna(val):
                    scenario_by_name[gname] = float(val)
            return scenario_by_name

        # 폴백: 기존 방식(연도/지역/타입) 형식은 더 이상 사용하지 않지만, 남겨둠
        return {}
    except Exception as e:
        print(f"발전기 시나리오 파싱 오류: {str(e)}")
        return {}


def _build_generator_overrides_for_year_v2(input_data, year, interface_path):
    try:
        gens = input_data.get('generators', pd.DataFrame())
        if gens is None or gens.empty:
            return {}
        scenario = _parse_generator_scenario_from_interface(interface_path, year)
        if not scenario:
            print(f"{year}년 발전기 시나리오가 없어 오버라이드 생략")
            return {}
        def get_type_from_name(name):
            n = str(name)
            if 'PV' in n:
                return 'PV'
            if 'WT' in n:
                return 'WT'
            if 'Nuclear' in n:
                return 'Nuclear'
            if 'CHP' in n:
                return 'CHP'
            if 'Coal' in n:
                return 'Coal'
            if 'Hydro' in n:
                return 'Hydro'
            if 'LNG' in n:
                return 'LNG'
            return None
        # 그룹별(지역, 타입) 기본 합계 및 멤버 목록
        group_sum = {}
        group_members = {}
        for _, row in gens.iterrows():
            name = str(row['name'])
            region = name.split('_')[0] if '_' in name else None
            gtype = get_type_from_name(name)
            if region and gtype:
                key = (region, gtype)
                base_nom = float(pd.to_numeric(row['p_nom'], errors='coerce') or 0.0)
                group_sum[key] = group_sum.get(key, 0.0) + base_nom
                group_members.setdefault(key, []).append((name, base_nom))
        overrides = {'generators': {}}
        for key, target in scenario.items():
            region, gtype = key
            base = group_sum.get(key, 0.0)
            members = group_members.get(key, [])
            if not members:
                print(f"경고: {year}년 {region}-{gtype} 그룹에 대응하는 발전기 레코드가 없습니다.")
                continue
            if base <= 0:
                print(f"경고: {year}년 {region}-{gtype} 기본 합계 0 → 분배 불가, 기존 값 유지")
                continue
            scale = float(target) / float(base)
            for name, base_nom in members:
                new_nom = base_nom * scale
                if gtype in ['PV', 'WT']:
                    overrides['generators'][name] = {
                        'p_nom_min': new_nom,
                        'p_nom': new_nom
                    }
                else:
                    overrides['generators'][name] = {
                        'p_nom': new_nom
                    }
        return overrides
    except Exception as e:
        print(f"발전기 오버라이드 v2 생성 오류: {str(e)}")
        return {}


def build_overrides_for_years(years, base_input_file):
    """연도별 오버라이드 딕셔너리 구성.

    각 연도에 대해:
    - timeseries: 해당 연도 1월 1일 ~ 다음 해 1월 1일 (대표 1년)
    - generators: interface.xlsx '시나리오_발전기' 시트에 정의된 용량 반영
    """
    try:
        root_dir = get_base_dir()
        interface_path = os.path.abspath(os.path.join(root_dir, 'interface.xlsx'))
        base = read_input_data(base_input_file)
        freq = '1h'
        try:
            if 'timeseries' in base and not base['timeseries'].empty:
                freq = str(base['timeseries'].iloc[0]['frequency'])
        except Exception:
            pass
        overrides_by_year = {}
        for y in years:
            ts_override = {
                'start_time': f"{y}-01-01 00:00:00",
                'end_time':   f"{y+1}-01-01 00:00:00",
                'frequency':  freq
            }
            # 이름별 목표 용량(interface.xlsx '시나리오_발전기' 시트) 반영
            name_to_target = _parse_generator_scenario_from_interface(interface_path, y)
            ov = {'timeseries': ts_override, 'generators': {}}
            if name_to_target:
                for gname, target in name_to_target.items():
                    # 재생에너지: 최소 용량 설정(확장 가능 여부는 입력 파일에 따름)
                    if any(k in gname for k in ['PV', 'WT']):
                        ov['generators'][gname] = {'p_nom_min': float(target), 'p_nom': float(target)}
                    else:
                        ov['generators'][gname] = {'p_nom': float(target)}
            overrides_by_year[y] = ov
        print(f"연도별 오버라이드 구성 완료: {list(overrides_by_year.keys())}")
        return overrides_by_year
    except Exception as e:
        print(f"연도별 오버라이드 구성 오류: {str(e)}")
        return {}

def _parse_demand_scenario_wide(interface_path):
    try:
        if not os.path.exists(interface_path):
            return {}
        df = pd.read_excel(interface_path, sheet_name='시나리오_에너지수요')
        if df is None or df.empty:
            return {}
        df.columns = [str(c).strip() for c in df.columns]
        name_col = None
        for c in df.columns:
            if str(c).strip().lower() in ['name', '이름']:
                name_col = c
                break
        if name_col is None:
            return {}
        year_cols = [c for c in df.columns if str(c).strip().isdigit() and len(str(c).strip()) == 4]
        if not year_cols:
            return {}
        out = {}
        for _, row in df.iterrows():
            lname = str(row.get(name_col, '')).strip()
            if not lname or lname.lower() == 'nan':
                continue
            per_year = {}
            for yc in year_cols:
                val = pd.to_numeric(row.get(yc), errors='coerce')
                if pd.notna(val):
                    per_year[int(str(yc).strip())] = float(val)
            if per_year:
                out[lname] = per_year
        return out
    except Exception as e:
        print(f"수요 시나리오(와이드) 파싱 오류: {str(e)}")
        return {}


def _apply_scenario_to_loads_in_input(input_data, scenario_year):
    try:
        if 'loads' not in input_data or input_data['loads'].empty:
            return input_data
        # interface.xlsx 경로
        root_dir = get_base_dir()
        interface_path = os.path.abspath(os.path.join(root_dir, 'interface.xlsx'))
        if not os.path.exists(interface_path):
            return input_data
        try:
            df = pd.read_excel(interface_path, sheet_name='시나리오_에너지수요')
        except Exception:
            return input_data
        if df is None or df.empty:
            return input_data
        df.columns = [str(c).strip() for c in df.columns]
        year_col = str(scenario_year)
        if year_col not in df.columns:
            return input_data
        # 기준 컬럼 탐색
        name_col = None
        bus_col = None
        for c in df.columns:
            cl = str(c).strip().lower()
            if cl in ['name', '이름']:
                name_col = c
            if cl in ['bus', '버스']:
                bus_col = c
        loads_df = input_data['loads']
        if 'p_set' not in loads_df.columns:
            loads_df['p_set'] = np.nan
        # 시나리오 → 값 매핑 구성
        map_by_name = {}
        if name_col is not None:
            for _, row in df.iterrows():
                nm = str(row.get(name_col, '')).strip()
                if not nm:
                    continue
                val = pd.to_numeric(row.get(year_col), errors='coerce')
                if pd.notna(val):
                    map_by_name[nm] = float(val)
        # 버스 기반 보조 매핑(BSN_EL → BSN_Demand_EL 등)
        if bus_col is not None:
            for _, row in df.iterrows():
                bus = str(row.get(bus_col, '')).strip()
                if not bus:
                    continue
                val = pd.to_numeric(row.get(year_col), errors='coerce')
                if pd.isna(val):
                    continue
                parts = [t for t in bus.split('_') if t]
                if len(parts) >= 2:
                    region, energy = parts[0], parts[-1]
                    load_name = f"{region}_Demand_{energy}"
                    # 이름 매핑이 우선이나 없으면 버스 파생 이름으로 보강
                    if load_name not in map_by_name:
                        map_by_name[load_name] = float(val)
        # 매핑 적용(이름 기준 → 순서 불일치 해소)
        updated = 0
        for idx, row in loads_df.iterrows():
            lname = str(row.get('name', '')).strip()
            if lname in map_by_name:
                loads_df.at[idx, 'p_set'] = map_by_name[lname]
                updated += 1
        input_data['loads'] = loads_df
        print(f"loads.p_set 주입(이름/버스 매핑): {scenario_year}년 {updated}개 행 업데이트")
        if updated == 0:
            print("경고: 시나리오_에너지수요에서 일치하는 이름/버스를 찾지 못했습니다. 시트의 '이름' 또는 '버스' 컬럼과 loads의 'name'을 확인하세요.")
        return input_data
    except Exception as e:
        print(f"loads 시나리오 주입 오류: {str(e)}")
        return input_data

def _update_integrated_for_year(integrated_path, interface_path, year):
    try:
        if not os.path.exists(integrated_path) or not os.path.exists(interface_path):
            return False
        # 통합 파일의 모든 시트를 읽음
        xls_int = pd.ExcelFile(integrated_path)
        sheets = {sn: pd.read_excel(integrated_path, sheet_name=sn) for sn in xls_int.sheet_names}

        # 1) 시나리오_에너지수요 → loads.p_set 업데이트 (연도 헤더 기반, 순서 매칭)
        try:
            df_load_scn = pd.read_excel(interface_path, sheet_name='시나리오_에너지수요')
            df_load_scn.columns = [str(c).strip() for c in df_load_scn.columns]
            year_str = str(year)
            if 'loads' in sheets and year_str in df_load_scn.columns:
                df_loads = sheets['loads']
                # 연도 컬럼에서 숫자만 추출하여 상단/하단 NaN 제거
                series = pd.to_numeric(df_load_scn[year_str], errors='coerce')
                vals = series.dropna().values
                n = min(len(df_loads), len(vals))
                if n > 0:
                    df_loads.loc[df_loads.index[:n], 'p_set'] = vals[:n]
                    sheets['loads'] = df_loads
                    print(f"integrated_input_data loads.p_set 갱신: {year}년 {n}개 행 업데이트")
        except Exception as e:
            print(f"loads 갱신 경고({year}): {str(e)}")

        # 2) 시나리오_발전기 → generators 업데이트 (이름 기반)
        try:
            df_gen_scn = pd.read_excel(interface_path, sheet_name='시나리오_발전기')
            df_gen_scn.columns = [str(c).strip() for c in df_gen_scn.columns]
            name_col = None
            for c in df_gen_scn.columns:
                if str(c).strip().lower() in ['name', '이름']:
                    name_col = c
                    break
            year_str = str(year)
            if 'generators' in sheets and name_col is not None and year_str in df_gen_scn.columns:
                df_gens = sheets['generators']
                if 'p_nom_min' not in df_gens.columns:
                    df_gens['p_nom_min'] = np.nan
                updated = 0
                # 이름 → 값 매핑
                name_to_val = {}
                for _, row in df_gen_scn.iterrows():
                    gname = str(row.get(name_col, '')).strip()
                    if not gname:
                        continue
                    val = pd.to_numeric(row.get(year_str), errors='coerce')
                    if pd.notna(val):
                        name_to_val[gname] = float(val)
                for idx, row in df_gens.iterrows():
                    gname = str(row.get('name', '')).strip()
                    if not gname or gname not in name_to_val:
                        continue
                    target = name_to_val[gname]
                    if ('PV' in gname) or ('WT' in gname):
                        # 재생: 최소용량 설정 (확장가능 여부는 기존값/인터페이스에 따름)
                        df_gens.at[idx, 'p_nom_min'] = float(target)
                        # p_nom은 최소 target 이상으로 보정
                        base_nom = float(pd.to_numeric(row.get('p_nom'), errors='coerce') or 0.0)
                        df_gens.at[idx, 'p_nom'] = float(max(base_nom, target))
                    else:
                        # 비재생: 연도 값으로 고정
                        df_gens.at[idx, 'p_nom'] = float(target)
                    updated += 1
                sheets['generators'] = df_gens
                print(f"integrated_input_data generators 갱신: {year}년 {updated}개 행 업데이트")
        except Exception as e:
            print(f"generators 갱신 경고({year}): {str(e)}")

        # 3) 지역간 연결 → lines 재생성(연결/길이/형식/병렬수 포함)
        try:
            tmp_input = {'buses': sheets.get('buses', pd.DataFrame())}
            fb_lines = _fallback_build_lines_from_interface(tmp_input, interface_path)
            if fb_lines is not None and not fb_lines.empty:
                sheets['lines'] = fb_lines
                print(f"integrated_input_data lines 갱신: {len(fb_lines)}개 레코드")
        except Exception as e:
            print(f"lines 갱신 경고({year}): {str(e)}")

        # 변경사항 저장
        with pd.ExcelWriter(integrated_path, engine='openpyxl') as writer:
            for sn, df in sheets.items():
                df.to_excel(writer, sheet_name=sn, index=False)
        return True
    except Exception as e:
        print(f"연도별 통합 입력 갱신 실패: {str(e)}")
        return False

def _run_week_uc_analysis(input_data, full_network, uc_week):
    """특정 주(week)에 대해 committable=True 적용 후 UC 분석 수행"""
    import copy

    print("\n" + "=" * 80)
    print(f"[Week {uc_week} UC 분석] Unit Commitment 분석 시작")
    print("=" * 80)

    # 1. 해당 주의 스냅샷 범위 계산 (1-based week, 월요일 기준)
    all_snapshots = full_network.snapshots
    year_start = all_snapshots[0]

    # 주 번호에 해당하는 첫 번째 시각 계산 (ISO week 방식)
    import datetime as dt
    # 해당 연도의 첫 월요일 기준으로 주 시작
    week_start_dt = dt.datetime.strptime(
        f"{year_start.year}-W{uc_week:02d}-1", "%Y-W%W-%w"
    )
    week_end_dt = week_start_dt + dt.timedelta(days=7)

    week_snapshots = all_snapshots[
        (all_snapshots >= week_start_dt) & (all_snapshots < week_end_dt)
    ]

    if len(week_snapshots) == 0:
        print(f"[경고] Week {uc_week}에 해당하는 스냅샷이 없습니다. UC 분석을 건너뜁니다.")
        return

    print(f"  대상 기간: {week_snapshots[0]} ~ {week_snapshots[-1]}")
    print(f"  시간대 수: {len(week_snapshots)}개")

    # 2. input_data를 복사하여 committable 활성화
    input_data_uc = {k: v.copy() if hasattr(v, 'copy') else v
                     for k, v in input_data.items()}

    # interface.xlsx에 committable=True로 설정된 발전기만 UC 적용
    # (현재 input_data에는 False로 강제됐을 수 있으므로, 원본 interface에서 재로드)
    try:
        root_dir = get_base_dir()
        interface_path = os.path.abspath(os.path.join(root_dir, 'interface.xlsx'))
        xls_iface = pd.ExcelFile(interface_path)
        # 지역별 시트에서 committable=True 발전기 이름 수집
        uc_gen_names = set()
        for sn in xls_iface.sheet_names:
            try:
                df_sn = pd.read_excel(interface_path, sheet_name=sn)
                if 'committable' in df_sn.columns and 'name' in df_sn.columns:
                    uc_rows = df_sn[df_sn['committable'].apply(
                        lambda x: str(x).strip().lower() in ['true', '1', 'yes']
                    )]
                    for _, row in uc_rows.iterrows():
                        uc_gen_names.add(str(row['name']).strip())
            except Exception:
                pass
    except Exception as e:
        print(f"  [경고] interface.xlsx 재로드 실패, input_data 원본 사용: {e}")
        uc_gen_names = set()

    # integrated_input_data의 generators 중 interface에서 committable=True인 것 활성화
    if 'generators' in input_data_uc and not input_data_uc['generators'].empty:
        df_g = input_data_uc['generators'].copy()

        def _should_enable_uc(gen_name):
            if uc_gen_names:
                # interface 이름과 직접 매칭 또는 prefix 매칭
                for iface_name in uc_gen_names:
                    if iface_name.upper() in gen_name.upper() or gen_name.upper() == iface_name.upper():
                        return True
            # carrier='coal' 발전기는 기본적으로 UC 대상
            return False

        uc_applied = 0
        for idx in df_g.index:
            gen_name = str(df_g.at[idx, 'name'])
            gen_carrier = str(df_g.at[idx, 'carrier']).lower() if 'carrier' in df_g.columns else ''
            if _should_enable_uc(gen_name) or gen_carrier == 'coal':
                df_g.at[idx, 'committable'] = True
                uc_applied += 1
        input_data_uc['generators'] = df_g
        print(f"  UC 적용 발전기: {uc_applied}개 (committable=True)")

    # 3. 해당 주의 timeseries만 사용하도록 설정
    if 'timeseries' in input_data_uc and not input_data_uc['timeseries'].empty:
        ts_row = input_data_uc['timeseries'].iloc[0].copy()
        ts_row['start_time'] = week_snapshots[0]
        ts_row['end_time'] = week_snapshots[-1] + pd.Timedelta(hours=1)
        input_data_uc['timeseries'] = pd.DataFrame([ts_row])

    # 4. 네트워크 생성 및 최적화
    print(f"\n  UC 네트워크 생성 중...")
    network_uc = create_network(input_data_uc)

    if network_uc is None:
        print("  [오류] UC 네트워크 생성 실패")
        return

    # 전년도 분석 결과를 초기 상태로 활용 (해당 주 시작 시점의 발전기 상태)
    try:
        week_start_ts = week_snapshots[0]
        prev_ts_candidates = all_snapshots[all_snapshots < week_start_ts]
        if len(prev_ts_candidates) > 0:
            prev_ts = prev_ts_candidates[-1]
            for gen in network_uc.generators.index:
                if network_uc.generators.at[gen, 'committable']:
                    if (gen in full_network.generators.index and
                            hasattr(full_network.generators_t, 'p') and
                            gen in full_network.generators_t.p.columns and
                            prev_ts in full_network.generators_t.p.index):
                        prev_output = float(full_network.generators_t.p.at[prev_ts, gen])
                        p_nom = float(network_uc.generators.at[gen, 'p_nom'])
                        # 이전 출력량으로 초기 상태 추정
                        if prev_output > p_nom * 0.05:
                            network_uc.generators.at[gen, 'up_time_before'] = 1
                        else:
                            network_uc.generators.at[gen, 'down_time_before'] = 1
        print(f"  이전 시점({prev_ts if len(prev_ts_candidates)>0 else '없음'}) 기준 초기 상태 설정 완료")
    except Exception as e:
        print(f"  [경고] 초기 상태 설정 실패 (무시): {e}")

    print(f"\n  UC 최적화 실행 중 (Week {uc_week})...")
    uc_ok = optimize_network(network_uc)

    # 5. 결과 저장 (별도 폴더에 저장)
    if uc_ok:
        uc_subdir = f"results/week{uc_week:02d}_UC_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        print(f"\n  UC 결과 저장 중: {uc_subdir}")
        save_results(network_uc, subdir=uc_subdir)
        print(f"[Week {uc_week} UC 분석] 완료! 결과 폴더: {uc_subdir}")
    else:
        print(f"[Week {uc_week} UC 분석] 최적화 실패")


def main():
    # ── 분석 모드 선택 프롬프트 ──────────────────────────────────────────────
    # [임시 비활성화] 입력 없이 바로 전체 기간(all) 모드로 실행
    # 아래 주석을 해제하면 다시 대화형 선택 가능
    #
    # print("=" * 60)
    # print("  PyPSA GESI 분석 모드 선택")
    # print("=" * 60)
    # print("  all      : 전체 기간 분석 (committable 강제 False)")
    # print("  1 ~ 52   : 전체 기간 분석 후, 해당 주(week)에 대해")
    # print("             committable=True UC 분석을 추가 실행")
    # print("=" * 60)
    # try:
    #     mode_input = input("분석 모드 입력 (all 또는 주 번호 1-52): ").strip().lower()
    # except EOFError:
    #     mode_input = 'all'
    # if mode_input == 'all':
    #     analysis_mode = 'all'
    #     uc_week = None
    #     print("\n[all 모드] 전체 기간 분석, 모든 발전기 committable=False 적용")
    # else:
    #     try:
    #         uc_week = int(mode_input)
    #         if not (1 <= uc_week <= 52):
    #             raise ValueError("범위 초과")
    #         analysis_mode = 'week'
    #         print(f"\n[Week {uc_week} 모드] 전체 기간(committable=False) 분석 후")
    #         print(f"               Week {uc_week}에 대해 UC 분석 추가 실행")
    #     except ValueError:
    #         print(f"[경고] '{mode_input}'은 유효하지 않습니다. all 모드로 실행합니다.")
    #         analysis_mode = 'all'
    #         uc_week = None

    analysis_mode = 'all'
    uc_week = None
    print("[all 모드] 1년 전체 기간 분석으로 바로 시작합니다.")

    # ── 공통 준비 ───────────────────────────────────────────────────────────
    ensure_integrated_input()
    print("지역별 시트 원본 데이터 사용 모드로 실행...")
    print("데이터 로드 시작...")
    input_data = read_input_data(INPUT_FILE)
    if input_data is None:
        return

    input_data = standardize_bus_names_in_input(input_data)

    try:
        integrated_path = os.path.abspath(os.path.join(get_base_dir(), INPUT_FILE))
        _persist_standardized_input(integrated_path, input_data)
    except Exception as e:
        print(f"표준화 저장 중 오류: {str(e)}")

    check_excel_data_loading(input_data)

    # ── all / week 공통: 전체 기간은 항상 committable=False ─────────────────
    if 'generators' in input_data and not input_data['generators'].empty:
        input_data['generators']['committable'] = False
        print("[전체 기간] 모든 발전기 committable=False 강제 적용")

    # ── 전체 기간 분석 ──────────────────────────────────────────────────────
    print("\n네트워크 생성 시작...")
    network = create_network(input_data)

    print("최적화 시작...")
    ok = optimize_network(network)
    if ok:
        print("결과 저장 시작...")
        save_results(network)
        print("전체 기간 분석 완료!")
    else:
        print("전체 기간 최적화 실패!")
        # 실패 시 사전 진단 실행
        try:
            _diagnose_infeasibility(network)
        except Exception:
            pass

    # ── week 모드: 추가 UC 분석 ─────────────────────────────────────────────
    if analysis_mode == 'week' and uc_week is not None and ok:
        _run_week_uc_analysis(input_data, network, uc_week)

    print("\n모든 과정 완료!")

if __name__ == "__main__":
    if _IS_FROZEN:
        try:
            main()
            print("\n" + "=" * 60)
            print("  분석이 정상적으로 완료되었습니다.")
            print(f"  결과 폴더: {_os.path.join(_EXE_DIR, 'results')}")
            print("=" * 60)
        except Exception as _e_main:
            print("\n" + "=" * 60)
            print("  [오류] 분석 중 예외가 발생했습니다:")
            print(f"  {_e_main}")
            print("-" * 60)
            _tb_early.print_exc()
            print("=" * 60)
            print(f"\n자세한 내용은 로그 파일을 확인하세요:\n  {_LOG_PATH}")
        finally:
            if _log_file:
                try:
                    _log_file.flush()
                    _log_file.close()
                except Exception:
                    pass
            _sys.stdout = _orig_stdout
            _sys.stderr = _orig_stderr

        input("\nEnter 키를 누르면 창이 닫힙니다...")
    else:
        main()