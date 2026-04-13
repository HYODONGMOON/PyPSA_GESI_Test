"""
interface.xlsx의 모든 지역별 시트에서 발전기 carrier를 자동으로 수정
"""
import openpyxl
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime

# 백업 생성
backup_name = f'interface_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
shutil.copy2('interface.xlsx', backup_name)
print(f'백업 생성됨: {backup_name}')

# Excel 파일 열기
wb = openpyxl.load_workbook('interface.xlsx')

# 발전기 이름 → carrier 매핑
def infer_carrier_from_name(gen_name):
    """발전기 이름에서 carrier 추론"""
    name_lower = gen_name.lower()
    
    if 'nuclear' in name_lower or '원자력' in name_lower:
        return 'nuclear'
    elif 'coal' in name_lower or '석탄' in name_lower:
        return 'coal'
    elif 'lng' in name_lower or 'gas' in name_lower:
        return 'gas'
    elif 'hydro' in name_lower or '수력' in name_lower:
        return 'hydro'
    elif 'pv' in name_lower or 'solar' in name_lower or '태양광' in name_lower:
        return 'solar'
    elif 'wt' in name_lower or 'wind' in name_lower or '풍력' in name_lower:
        return 'wind'
    elif 'slack' in name_lower or '슬랙' in name_lower:
        return 'electricity'  # 슬랙은 electricity 유지
    else:
        return None  # 변경하지 않음

# 지역 코드 목록
region_codes = ['BSN', 'CBD', 'CND', 'DGU', 'DJN', 'GBD', 'GGD', 'GND', 
                'GWD', 'GWJ', 'ICN', 'JBD', 'JJD', 'JND', 'SEL', 'SJN', 'USN']

total_updated = 0
regions_updated = 0

print('\n'+'='*70)
print('interface.xlsx carrier 자동 수정')
print('='*70)

for region in region_codes:
    sheet_name = f'지역_{region}'
    
    if sheet_name not in wb.sheetnames:
        print(f'[건너뜀] {sheet_name} 시트 없음')
        continue
    
    ws = wb[sheet_name]
    
    # 발전기 섹션 찾기
    found_generator_section = False
    for row in range(1, min(200, ws.max_row + 1)):
        cell = ws.cell(row=row, column=1)
        if cell.value and '발전기' in str(cell.value):
            header_row = row + 2
            
            # 헤더 읽기
            headers = []
            for i in range(1, 30):
                h = ws.cell(row=header_row, column=i).value
                if h:
                    headers.append(h)
                else:
                    break
            
            # carrier 컬럼 찾기
            carrier_col = None
            name_col = 1
            
            for i, h in enumerate(headers):
                if '캐리어' in str(h) or 'carrier' in str(h).lower():
                    carrier_col = i + 1
                    break
            
            if not carrier_col:
                print(f'[건너뜀] {sheet_name}: carrier 컬럼 없음')
                break
            
            # 데이터 행 처리
            region_updated = 0
            for data_row in range(header_row+1, min(header_row+50, ws.max_row+1)):
                name = ws.cell(row=data_row, column=name_col).value
                if not name:
                    break
                
                current_carrier = ws.cell(row=data_row, column=carrier_col).value
                inferred_carrier = infer_carrier_from_name(str(name))
                
                # carrier가 'electricity'인데 추론된 carrier가 있으면 수정
                # 또는 carrier가 비어있는 경우에도 추론된 carrier 적용
                current_lower = str(current_carrier).lower() if current_carrier else ''
                
                if inferred_carrier:
                    # electricity이거나 비어있으면 수정 (단, 슬랙은 electricity 유지)
                    if (current_lower in ['electricity', '', 'none']) and inferred_carrier != 'electricity':
                        ws.cell(row=data_row, column=carrier_col).value = inferred_carrier
                        print(f'  {region}_{name}: {current_carrier} → {inferred_carrier}')
                        region_updated += 1
                        total_updated += 1
            
            if region_updated > 0:
                regions_updated += 1
                print(f'  [{region}] {region_updated}개 발전기 carrier 수정됨')
            
            found_generator_section = True
            break
    
    if not found_generator_section:
        print(f'[건너뜀] {sheet_name}: 발전기 섹션 없음')

print('\n' + '='*70)
print(f'수정 완료: {regions_updated}개 지역, 총 {total_updated}개 발전기')
print('='*70)

# 저장
wb.save('interface.xlsx')
print('\ninterface.xlsx 저장 완료!')
print(f'백업 파일: {backup_name}')
