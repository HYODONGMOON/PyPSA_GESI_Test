# -*- coding: utf-8 -*-
"""
PyPSA-GESI 시나리오 후처리 분석 스크립트
=========================================
폴더 구조:
    scenario_result/
        REF/
            2030/REF2030.xlsx
            2035/REF2035.xlsx
            2040/REF2040.xlsx
            2045/REF2045.xlsx
            2050/REF2050.xlsx
        CZ/
            2035/CZ2035.xlsx   (2030은 REF/2030 공유)
            2040/CZ2040.xlsx
            2045/CZ2045.xlsx
            2050/CZ2050.xlsx

분석 내용:
    1. 연도별 에너지 믹스 (설비용량 + 저장용량) 묶음막대그래프
    2. Sankey 에너지 흐름 다이어그램 (연도별)
    3. Curtailment · 전력망 포화 · 섹터커플링 효과 종합표
    4. 송전망 이용률 분포 (연도별 묶음 박스플롯)

사용법:
    python scenario_analysis.py
    또는
    python scenario_analysis.py --scenario REF   (특정 시나리오만)
    python scenario_analysis.py --year 2035      (특정 연도만)
"""

import os
import sys
import warnings

# Windows cp949 콘솔에서 한글/유니코드 출력 오류 방지
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(errors="replace")
        sys.stderr.reconfigure(errors="replace")
    except Exception:
        pass
import argparse
import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
from matplotlib.gridspec import GridSpec
import plotly.graph_objects as go
import plotly.io as pio

warnings.filterwarnings("ignore", category=UserWarning)
matplotlib.rcParams["font.family"] = "Malgun Gothic"
matplotlib.rcParams["axes.unicode_minus"] = False
plt.rcParams["figure.dpi"] = 150

# ─────────────────────────────────────────────
# 0. 설정
# ─────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCENARIO_ROOT = os.path.join(BASE_DIR, "scenario_result")
OUTPUT_DIR    = os.path.join(SCENARIO_ROOT, "analysis_output")
LINE_CAP_FILE = os.path.join(BASE_DIR, "integrated_input_data.xlsx")

# 시나리오 정의
SCENARIOS = {
    "REF": [2030, 2035, 2040, 2045, 2050],
    "CZ":  [2030, 2035, 2040, 2045, 2050],  # 2030은 REF 공유
}
# CZ 2030은 REF 2030 파일 사용
CZ_2030_FROM_REF = True

# 발전기 분류 키워드
CAT_RE         = ["_PV", "_WT", "_Hydro", "_Pumped"]
CAT_NUCLEAR    = ["_Nuclear"]
CAT_LNG        = ["_LNG"]
CAT_COAL       = ["_석탄", "_무연탄", "_유연탄", "_태안", "_보령", "_하동", "_당진",
                  "_삼천포", "_영흥", "_한울", "_호남", "_여수", "_울산", "_광양",
                  "_Coal", "_coal"]
CAT_DR         = ["_DR"]
CAT_SKIP       = ["_slack", "_Fallback", "_Fuel_Supply", "_Slack_Failsafe"]

# Link 분류
LINK_ELECTROLYSER = ["Electrolyser"]
LINK_HP           = ["_HP"]
LINK_FUELCELL     = ["Fuelcell"]
LINK_EV_CHARGER   = ["EV_Charger"]
LINK_EV_V2G       = ["EV_V2G"]
LINK_ESS_CH       = ["ESS_Ch"]
LINK_ESS_DISCH    = ["ESS_Disch"]
LINK_HEAT_CH      = ["Heat_Ch"]
LINK_HEAT_DISCH   = ["Heat_Disch"]
LINK_CHP          = ["_CHP"]
LINK_HVDC         = ["_HVDC"]

# 색상 팔레트
COLORS = {
    "PV":          "#FFD700",
    "WT":          "#87CEEB",
    "Hydro":       "#4169E1",
    "Nuclear":     "#FF6347",
    "LNG":         "#FFA500",
    "Coal":        "#808080",
    "DR":          "#9370DB",
    "Electrolyser":"#20B2AA",
    "HP":          "#FF69B4",
    "Fuelcell":    "#32CD32",
    "ESS":         "#DAA520",
    "H2_Storage":  "#00CED1",
    "H_Storage":   "#FF7F50",
    "EV":          "#6495ED",
    "CHP":         "#BC8F8F",
}

REGION_NAMES = {
    "BSN": "부산·경남", "CBD": "서울·경기", "CND": "충청남도",
    "GBD": "경상북도",  "GGD": "경기도",    "GND": "강원도",
    "GWD": "강원서부",  "ICN": "인천",      "JBD": "전라북도",
    "JJD": "제주",     "JND": "전라남도",
}

YEAR_COLORS = {
    2030: "#2196F3", 2035: "#4CAF50", 2040: "#FF9800",
    2045: "#9C27B0", 2050: "#F44336",
}


# ─────────────────────────────────────────────
# 1. 데이터 로딩
# ─────────────────────────────────────────────

def get_file_path(scenario: str, year: int) -> str | None:
    """시나리오+연도에 해당하는 Excel 결과 파일 경로 반환"""
    if scenario == "CZ" and year == 2030 and CZ_2030_FROM_REF:
        folder = os.path.join(SCENARIO_ROOT, "REF", str(year))
        fname  = f"REF{year}.xlsx"
    else:
        folder = os.path.join(SCENARIO_ROOT, scenario, str(year))
        fname  = f"{scenario}{year}.xlsx"
    path = os.path.join(folder, fname)
    return path if os.path.exists(path) else None


def load_scenario_data(scenario: str, year: int) -> dict | None:
    """결과 Excel에서 필요한 시트를 모두 로드하여 dict 반환"""
    path = get_file_path(scenario, year)
    if path is None:
        return None
    print(f"  로딩: {os.path.relpath(path)}")
    xl = pd.ExcelFile(path)
    sheets_avail = xl.sheet_names

    def read(name, **kw):
        if name in sheets_avail:
            return pd.read_excel(xl, sheet_name=name, **kw)
        return pd.DataFrame()

    return {
        "scenario": scenario,
        "year": year,
        "path": path,
        "generator_info":       read("Generator_Info"),
        "link_info":            read("Link_Info"),
        "storage_info":         read("Storage_Info"),
        "generator_output":     read("Generator_Output"),
        "link_flow":            read("Link_Flow"),
        "line_flow":            read("Line_Flow"),
        "final_energy":         read("FinalEnergy_Supply"),
        "regional_balance":     read("Regional_PowerBalance"),
        "curtailment_summary":  read("RE_Curtailment_Summary"),
        "curtailment_ts":       read("RE_Curtailment_Timeseries"),
        "hourly_loads":         read("Hourly_Loads"),
        "line_info":            read("Line_Info"),        # s_nom (실제 최적화 사용값)
    }


def load_all_data(filter_scenario=None, filter_year=None) -> list[dict]:
    """사용 가능한 모든 시나리오 데이터 로드"""
    results = []
    for sc, years in SCENARIOS.items():
        if filter_scenario and sc != filter_scenario:
            continue
        for yr in years:
            if filter_year and yr != filter_year:
                continue
            # CZ의 경우 2030은 REF에서 가져오되 CZ로 표기
            data = load_scenario_data(sc, yr)
            if data:
                results.append(data)
            else:
                print(f"  [건너뜀] {sc} {yr}: 파일 없음")
    return results


def load_line_capacities() -> pd.DataFrame:
    """integrated_input_data.xlsx에서 선로 용량 로드"""
    if not os.path.exists(LINE_CAP_FILE):
        return pd.DataFrame(columns=["name", "bus0", "bus1", "s_nom"])
    df = pd.read_excel(LINE_CAP_FILE, sheet_name="lines")[["name", "bus0", "bus1", "s_nom"]]
    df["region0"] = df["bus0"].str.replace("_EL", "", regex=False)
    df["region1"] = df["bus1"].str.replace("_EL", "", regex=False)
    df["pair"] = df.apply(
        lambda r: "_".join(sorted([r["region0"], r["region1"]])), axis=1
    )
    return df


# ─────────────────────────────────────────────
# 유틸리티
# ─────────────────────────────────────────────

def classify_generator(name: str) -> str | None:
    """발전기 이름 → 카테고리 문자열"""
    if any(k in name for k in CAT_SKIP):
        return None
    if any(k in name for k in CAT_RE):
        if "_PV" in name:   return "PV"
        if "_WT" in name:   return "WT"
        return "Hydro"
    if any(k in name for k in CAT_NUCLEAR): return "Nuclear"
    if any(k in name for k in CAT_LNG):     return "LNG"
    if any(k in name for k in CAT_COAL):    return "Coal"
    if any(k in name for k in CAT_DR):      return "DR"
    return "Other"


def classify_link(name: str) -> str | None:
    if any(k in name for k in LINK_ELECTROLYSER): return "Electrolyser"
    if any(k in name for k in LINK_HP):            return "HP"
    if any(k in name for k in LINK_FUELCELL):      return "Fuelcell"
    if any(k in name for k in LINK_EV_CHARGER):    return "EV_Charger"
    if any(k in name for k in LINK_EV_V2G):        return "EV_V2G"
    if any(k in name for k in LINK_ESS_CH):        return "ESS_Ch"
    if any(k in name for k in LINK_ESS_DISCH):     return "ESS_Disch"
    if any(k in name for k in LINK_HEAT_CH):       return "Heat_Ch"
    if any(k in name for k in LINK_HEAT_DISCH):    return "Heat_Disch"
    if any(k in name for k in LINK_CHP):           return "CHP"
    if any(k in name for k in LINK_HVDC):          return "HVDC"
    return None


def gw(mw): return mw / 1_000
def twh(mwh): return mwh / 1_000_000


# ─────────────────────────────────────────────
# 2. 분석 1: 에너지 믹스 – 설비용량 + 저장용량
# ─────────────────────────────────────────────

def extract_capacity(data: dict) -> dict:
    """발전/유연성/저장 설비용량(MW, MWh) 추출"""
    gi = data["generator_info"]
    li = data["link_info"]
    si = data["storage_info"]

    # 발전설비(MW): PV, WT, Hydro, Nuclear, LNG, Coal, CHP, Fuelcell
    gen_cap   = {k: 0.0 for k in ["PV", "WT", "Hydro", "Nuclear", "LNG", "Coal", "CHP", "Fuelcell"]}
    # 유연성설비(MW): Electrolyser, HP, DR, EV_Charger
    flex_cap  = {k: 0.0 for k in ["Electrolyser", "HP", "DR", "EV_Charger"]}
    # 저장설비(MWh): ESS, Pumped, H2_Storage, H_Storage
    store_cap = {k: 0.0 for k in ["ESS", "Pumped", "H2_Storage", "H_Storage"]}

    # ── Generator: PV/WT/Hydro/Nuclear/LNG/Coal → gen_cap, DR → flex_cap ──
    if not gi.empty:
        col = "p_nom_opt" if "p_nom_opt" in gi.columns else "p_nom"
        for _, row in gi.iterrows():
            name = row["Generator"]
            cap  = float(row.get(col, 0) or 0)
            if any(k in name for k in CAT_SKIP):
                continue
            if "_PV"      in name: gen_cap["PV"]      += cap
            elif "_WT"    in name: gen_cap["WT"]       += cap
            elif "_Hydro" in name: gen_cap["Hydro"]    += cap
            elif any(k in name for k in CAT_NUCLEAR): gen_cap["Nuclear"] += cap
            elif "_LNG"   in name and "Fuel_Supply" not in name: gen_cap["LNG"] += cap
            elif any(k in name for k in CAT_COAL):    gen_cap["Coal"]    += cap
            elif "_DR"    in name: flex_cap["DR"]      += cap   # DR → 유연성

    # ── Link: CHP/Fuelcell → gen_cap, 나머지 → flex_cap ──
    if not li.empty:
        col = "p_nom_opt" if "p_nom_opt" in li.columns else "p_nom"
        for _, row in li.iterrows():
            link_name = row["Link"]
            cap       = float(row.get(col, 0) or 0)
            cat       = classify_link(link_name)
            if cat == "CHP":           gen_cap["CHP"]            += cap
            elif cat == "Fuelcell":    gen_cap["Fuelcell"]        += cap
            elif cat == "Electrolyser":flex_cap["Electrolyser"]   += cap
            elif cat == "HP":          flex_cap["HP"]             += cap
            elif cat == "EV_Charger":  flex_cap["EV_Charger"]     += cap

    # ── Store: ESS/H2/Heat/Pumped → store_cap ──
    if not si.empty:
        col = "e_nom_opt" if "e_nom_opt" in si.columns else "e_nom"
        for _, row in si.iterrows():
            name    = row["Store"]
            carrier = str(row.get("carrier", "")).lower()
            cap     = float(row.get(col, 0) or 0)
            if "ESS" in name:                                   store_cap["ESS"]        += cap
            elif "H2_Storage" in name:                          store_cap["H2_Storage"] += cap
            elif "H_Storage"  in name or "Heat" in name:       store_cap["H_Storage"]  += cap
            elif "Pumped"     in name or carrier == "electricity": store_cap["Pumped"]  += cap

    return {"gen": gen_cap, "flex": flex_cap, "store": store_cap}


def plot_energy_mix(all_data: list[dict], scenario: str, out_dir: str):
    """
    연도별 에너지 믹스 - 단일 통합 차트
    - 발전설비(PV/WT/Hydro/Nuclear/LNG/Coal/CHP/Fuelcell): 좌측 Y축(GW), 왼쪽 바
    - 유연성설비(Electrolyser/HP/DR/EV_Charger):            좌측 Y축(GW), 가운데 바
    - 저장설비(ESS/Pumped/H2_Storage/H_Storage):             우측 Y축(GWh), 오른쪽 바(빗금)
    """
    sc_data = sorted([d for d in all_data if d["scenario"] == scenario], key=lambda d: d["year"])
    if not sc_data:
        print(f"  [{scenario}] 데이터 없음, 에너지 믹스 건너뜀")
        return

    years = [d["year"] for d in sc_data]
    caps  = [extract_capacity(d) for d in sc_data]
    n     = len(years)

    GEN_CATS   = ["PV", "WT", "Hydro", "Nuclear", "LNG", "Coal", "CHP", "Fuelcell"]
    FLEX_CATS  = ["Electrolyser", "HP", "DR", "EV_Charger"]
    STORE_CATS = ["ESS", "Pumped", "H2_Storage", "H_Storage"]

    GEN_COLORS = {
        "PV": "#FFD700", "WT": "#87CEEB", "Hydro": "#4169E1",
        "Nuclear": "#FF6347", "LNG": "#FFA500", "Coal": "#808080",
        "CHP": "#BC8F8F", "Fuelcell": "#32CD32",
    }
    FLEX_COLORS = {
        "Electrolyser": "#20B2AA", "HP": "#FF69B4",
        "DR": "#9370DB", "EV_Charger": "#6495ED",
    }
    STORE_COLORS = {
        "ESS": "#DAA520", "Pumped": "#5F9EA0",
        "H2_Storage": "#00CED1", "H_Storage": "#FF7F50",
    }

    fig, ax1 = plt.subplots(figsize=(14, 8))
    ax2 = ax1.twinx()

    w     = 0.22   # 개별 바 너비
    pitch = 1.1    # 연도 간 간격
    x_ctr = np.arange(n) * pitch
    pos_gen   = x_ctr - w
    pos_flex  = x_ctr
    pos_store = x_ctr + w

    # ── 발전설비 바 ──
    bot_g = np.zeros(n)
    hdl_g = []
    for cat in GEN_CATS:
        vals  = np.array([gw(c["gen"].get(cat, 0)) for c in caps])
        color = GEN_COLORS.get(cat, "#AAAAAA")
        ax1.bar(pos_gen, vals, w * 0.92, bottom=bot_g, color=color,
                edgecolor="white", linewidth=0.4)
        bot_g += vals
        if any(v > 0 for v in vals):
            hdl_g.append(mpatches.Patch(color=color, label=cat))

    # ── 유연성설비 바 ──
    bot_f = np.zeros(n)
    hdl_f = []
    for cat in FLEX_CATS:
        vals  = np.array([gw(c["flex"].get(cat, 0)) for c in caps])
        color = FLEX_COLORS.get(cat, "#BBBBBB")
        ax1.bar(pos_flex, vals, w * 0.92, bottom=bot_f, color=color,
                edgecolor="white", linewidth=0.4, alpha=0.88)
        bot_f += vals
        if any(v > 0 for v in vals):
            hdl_f.append(mpatches.Patch(color=color, label=cat, alpha=0.88))

    # ── 저장설비 바 (우측 Y축, GWh) ──
    bot_s = np.zeros(n)
    hdl_s = []
    for cat in STORE_CATS:
        vals  = np.array([c["store"].get(cat, 0) / 1_000 for c in caps])  # GWh
        color = STORE_COLORS.get(cat, "#CCCCCC")
        ax2.bar(pos_store, vals, w * 0.92, bottom=bot_s, color=color,
                edgecolor="white", linewidth=0.4, hatch="//", alpha=0.75)
        bot_s += vals
        if any(v > 0 for v in vals):
            hdl_s.append(mpatches.Patch(color=color, label=cat, hatch="//", alpha=0.75))

    # ── 바 그룹 레이블 (첫 연도에만) ──
    y_lbl = -ax1.get_ylim()[1] * 0.06 if ax1.get_ylim()[1] > 0 else -5
    for xi in range(n):
        ax1.text(pos_gen[xi],   y_lbl, "발전",  ha="center", fontsize=7.5, color="#555555", fontweight="bold")
        ax1.text(pos_flex[xi],  y_lbl, "유연성", ha="center", fontsize=7.5, color="#555555", fontweight="bold")
        ax1.text(pos_store[xi], y_lbl, "저장",  ha="center", fontsize=7.5, color="#555555", fontweight="bold")

    # ── 축 설정 ──
    ax1.set_xticks(x_ctr)
    ax1.set_xticklabels([f"'{str(y)[-2:]}" for y in years], fontsize=11)
    ax1.set_ylabel("설비용량 (GW)", fontsize=11)
    ax2.set_ylabel("저장용량 (GWh)", fontsize=11)
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}"))
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax1.grid(axis="y", linestyle="--", alpha=0.35)
    ax1.set_xlim(-pitch * 0.5, x_ctr[-1] + pitch * 0.6)

    # ── 범례 (3개 구역) ──
    leg1 = ax1.legend(handles=hdl_g, loc="upper left",
                      fontsize=8, title="발전설비", title_fontsize=9, ncol=2, framealpha=0.9)
    ax1.add_artist(leg1)
    leg2 = ax1.legend(handles=hdl_f, loc="upper center",
                      fontsize=8, title="유연성설비", title_fontsize=9, ncol=2,
                      bbox_to_anchor=(0.48, 1.0), framealpha=0.9)
    ax1.add_artist(leg2)
    ax2.legend(handles=hdl_s, loc="upper right",
               fontsize=8, title="저장설비(빗금)", title_fontsize=9, ncol=1, framealpha=0.9)

    fig.suptitle(f"[{scenario}] 연도별 에너지 믹스 - 설비용량",
                 fontsize=14, fontweight="bold")
    plt.tight_layout()
    out_path = os.path.join(out_dir, f"01_energy_mix_{scenario}.png")
    fig.savefig(out_path, bbox_inches="tight", dpi=150)
    plt.close(fig)
    print(f"  저장: {os.path.relpath(out_path)}")


# ─────────────────────────────────────────────
# 3. 분석 2: Sankey 에너지 흐름 다이어그램
# ─────────────────────────────────────────────

def build_sankey_data(data: dict) -> dict | None:
    """결과 데이터에서 Sankey용 국가 합산 에너지 흐름(TWh) 추출"""
    go_ts = data.get("generator_output", pd.DataFrame())
    lf_ts = data.get("link_flow",        pd.DataFrame())
    fe    = data.get("final_energy",     pd.DataFrame())

    if go_ts.empty or lf_ts.empty:
        return None

    # 발전기 출력 합산 (TWh)
    gen_total = {}
    for col in go_ts.columns:
        if col == "snapshot":
            continue
        cat = classify_generator(col)
        if cat is None:
            continue
        val = twh(float(go_ts[col].clip(lower=0).sum()))
        if val > 0:
            gen_total[cat] = gen_total.get(cat, 0) + val

    # 링크 유량 합산 (TWh)
    link_total = {}
    for col in lf_ts.columns:
        if col == "snapshot":
            continue
        cat = classify_link(col)
        if cat is None:
            continue
        val = twh(float(lf_ts[col].clip(lower=0).sum()))
        if val > 0:
            link_total[cat] = link_total.get(cat, 0) + val

    # 최종 에너지 수요 (FinalEnergy_Supply)
    demand_elec = demand_heat = demand_h2 = demand_ev = 0.0
    if not fe.empty:
        fe_clean = fe.dropna(subset=[fe.columns[0]])
        for _, row in fe_clean.iterrows():
            try:
                carrier = str(row.iloc[0]).strip()
                tech    = str(row.iloc[1]).strip()
                mwh_val = float(row.iloc[2]) if pd.notna(row.iloc[2]) else 0.0
                val_twh = twh(mwh_val)
                if carrier in ("전력", "electricity"):
                    demand_elec += val_twh
                elif carrier in ("열", "heat", "온열"):
                    demand_heat += val_twh
                elif carrier in ("수소", "hydrogen"):
                    demand_h2 += val_twh
                elif carrier in ("전기차", "EV"):
                    demand_ev += val_twh
            except Exception:
                continue

    # Curtailment (출력제한)
    curt_total = 0.0
    cs = data.get("curtailment_summary", pd.DataFrame())
    if not cs.empty:
        curt_col = [c for c in cs.columns if "Curtailment_MWh" in c or "curtailment" in c.lower()]
        if curt_col:
            curt_total = twh(float(cs[curt_col[0]].sum()))

    return {
        "gen":         gen_total,
        "link":        link_total,
        "demand_elec": demand_elec,
        "demand_heat": demand_heat,
        "demand_h2":   demand_h2,
        "demand_ev":   demand_ev,
        "curtailment": curt_total,
    }


def _build_sankey_trace(sd: dict, domain_x: list) -> go.Sankey | None:
    """Sankey 트레이스 1개 생성 (REF 또는 CZ 한 연도)"""
    if sd is None:
        return None

    NODES = [
        "태양광(PV)", "풍력(WT)", "수력", "원자력", "LNG", "석탄", "CHP",
        "전력망",
        "전력수요", "전기차 충전",
        "전해조(H2생산)", "히트펌프(열)", "연료전지",
        "수소 수요", "열 수요", "출력제한",
    ]
    C = {n: i for i, n in enumerate(NODES)}
    NODE_COLORS = [
        "#FFD700", "#87CEEB", "#4169E1", "#FF6347", "#FFA500", "#808080", "#BC8F8F",
        "#3399CC",
        "#5577BB", "#6495ED",
        "#20B2AA", "#FF69B4", "#32CD32",
        "#00CED1", "#FF7F50", "#CC0000",
    ]

    def rgba(r, g, b, a=0.4):
        return f"rgba({r},{g},{b},{a})"

    src, tgt, val, clr = [], [], [], []

    def add(s, t, v, c="rgba(180,180,180,0.4)"):
        if v > 0.01:
            src.append(C[s]); tgt.append(C[t])
            val.append(round(v, 2)); clr.append(c)

    # 발전 -> 전력망
    add("태양광(PV)",  "전력망", sd["gen"].get("PV",      0), rgba(255,215,  0))
    add("풍력(WT)",    "전력망", sd["gen"].get("WT",      0), rgba(135,206,235))
    add("수력",        "전력망", sd["gen"].get("Hydro",   0), rgba( 65,105,225))
    add("원자력",      "전력망", sd["gen"].get("Nuclear", 0), rgba(255, 99, 71))
    add("LNG",         "전력망", sd["gen"].get("LNG",     0), rgba(255,165,  0))
    add("석탄",        "전력망", sd["gen"].get("Coal",    0), rgba(128,128,128))
    add("CHP",         "전력망", sd["gen"].get("CHP",     0), rgba(188,143,143))
    add("연료전지",    "전력망", sd["link"].get("Fuelcell", 0), rgba(50,205, 50))

    # 전력망 -> 수요/P2X
    direct = max(0, sd["demand_elec"] - sd["link"].get("EV_Charger", 0))
    add("전력망", "전력수요",   direct,                             rgba( 68, 68,170))
    add("전력망", "전기차 충전", sd["link"].get("EV_Charger",   0), rgba(100,149,237))
    add("전력망", "전해조(H2생산)", sd["link"].get("Electrolyser",0), rgba(32,178,170))
    add("전력망", "히트펌프(열)",   sd["link"].get("HP",          0), rgba(255,105,180))
    add("전력망", "출력제한",        sd["curtailment"],               rgba(204,  0,  0))

    # P2X -> 최종수요
    add("전해조(H2생산)", "수소 수요", sd["demand_h2"],               rgba( 32,178,170,0.3))
    add("전해조(H2생산)", "연료전지",  sd["link"].get("Fuelcell", 0), rgba(  0,206,209))
    add("히트펌프(열)",   "열 수요",   sd["demand_heat"],              rgba(255,105,180,0.3))

    if not val:
        return None

    return go.Sankey(
        arrangement="snap",
        domain={"x": domain_x, "y": [0, 1]},
        node=dict(
            pad=15, thickness=16,
            line=dict(color="black", width=0.4),
            label=NODES,
            color=NODE_COLORS,
        ),
        link=dict(source=src, target=tgt, value=val, color=clr),
    )


def plot_sankey(all_data: list[dict], _scenario_unused: str, out_dir: str):
    """
    REF / CZ 좌우 비교 인터랙티브 Sankey HTML
    - 연도 선택 드롭다운으로 두 시나리오를 동시에 전환
    - 파일: 02_sankey_comparison.html
    """
    years = sorted(set(d["year"] for d in all_data))
    if not years:
        return

    fig = go.Figure()
    n_traces_per_year = 2  # REF + CZ

    trace_idx = 0
    trace_map  = {}  # (scenario, year) -> trace index

    for yr in years:
        for sc, dom_x in [("REF", [0.0, 0.46]), ("CZ", [0.54, 1.0])]:
            matches = [d for d in all_data if d["scenario"] == sc and d["year"] == yr]
            sd = build_sankey_data(matches[0]) if matches else None
            trace = _build_sankey_trace(sd, dom_x)

            if trace is None:
                trace = go.Sankey(
                    domain={"x": dom_x, "y": [0, 1]},
                    node=dict(label=["(데이터 없음)"], pad=10, thickness=10),
                    link=dict(source=[], target=[], value=[]),
                )

            trace.visible = (yr == years[0])
            fig.add_trace(trace)
            trace_map[(sc, yr)] = trace_idx
            trace_idx += 1

    # ── 드롭다운 버튼 ──
    buttons = []
    total_traces = len(years) * n_traces_per_year
    for yr in years:
        vis = [False] * total_traces
        for sc in ("REF", "CZ"):
            idx = trace_map.get((sc, yr))
            if idx is not None:
                vis[idx] = True
        buttons.append(dict(
            label=str(yr),
            method="update",
            args=[
                {"visible": vis},
                {"title": {"text":
                    f"에너지 흐름 비교 - {yr}년  |  좌: REF  /  우: CZ  (단위: TWh)",
                 "font": {"size": 15}}},
            ],
        ))

    fig.update_layout(
        title=dict(
            text=f"에너지 흐름 비교 - {years[0]}년  |  좌: REF  /  우: CZ  (단위: TWh)",
            font=dict(size=15),
            x=0.5, xanchor="center",
        ),
        updatemenus=[dict(
            active=0,
            buttons=buttons,
            direction="down",
            x=0.5, y=1.08,
            xanchor="center", yanchor="top",
            showactive=True,
            bgcolor="#EEEEEE",
            bordercolor="#555555",
            font=dict(size=12),
            pad=dict(r=10, t=5),
        )],
        annotations=[
            dict(text="<b>REF 시나리오</b>", x=0.23, y=1.04,
                 xref="paper", yref="paper", showarrow=False,
                 font=dict(size=14, color="#1144AA")),
            dict(text="<b>CZ 시나리오</b>", x=0.77, y=1.04,
                 xref="paper", yref="paper", showarrow=False,
                 font=dict(size=14, color="#AA4411")),
            dict(text="연도 선택:", x=0.35, y=1.085,
                 xref="paper", yref="paper", showarrow=False,
                 font=dict(size=12, color="#333333")),
        ],
        height=680,
        font_size=11,
        margin=dict(t=120, b=20, l=20, r=20),
    )

    html_path = os.path.join(out_dir, "02_sankey_comparison.html")
    fig.write_html(html_path, include_plotlyjs="cdn")
    print(f"  저장: {os.path.relpath(html_path)}")


# ─────────────────────────────────────────────
# 4. 분석 3: Curtailment · 전력망 포화 · 섹터커플링 종합표
# ─────────────────────────────────────────────

def build_curtailment_grid_table(data: dict, line_caps: pd.DataFrame) -> pd.DataFrame:
    """
    지역별 연도별:
      - 총 Curtailment (GWh)
      - Curtailment 발생 시간수 (시간)
      - 전력망 포화(>95%) 동시 발생 시간수
      - 섹터커플링(전해조+HP) 활성 동시 시간수
      - 전력망 포화 & 섹터커플링 동시 활성 시간수
    """
    curt_ts  = data.get("curtailment_ts",  pd.DataFrame())
    lf_ts    = data.get("line_flow",       pd.DataFrame())
    link_ts  = data.get("link_flow",       pd.DataFrame())
    rb       = data.get("regional_balance", pd.DataFrame())

    rows = []

    # ── 지역 목록 (Regional_PowerBalance 기준) ──
    if not rb.empty:
        region_col = rb.columns[0]
        regions = list(rb[region_col].dropna().astype(str).str.strip())
    else:
        regions = list(REGION_NAMES.keys())

    # ── 시간별 전력망 포화 여부 (어떤 권역 간 선로가 95% 초과) ──
    grid_sat_hours: pd.Series | None = None
    if not lf_ts.empty and not line_caps.empty:
        lf = lf_ts.set_index("snapshot") if "snapshot" in lf_ts.columns else lf_ts
        sat_any = pd.Series(False, index=lf.index)
        for _, cap_row in line_caps.iterrows():
            line_name = cap_row["name"]
            if line_name not in lf.columns:
                continue
            s_nom = float(cap_row["s_nom"])
            if s_nom <= 0:
                continue
            util = lf[line_name].abs() / s_nom
            sat_any = sat_any | (util >= 0.95)
        grid_sat_hours = sat_any

    # ── 시간별 섹터커플링 활성 여부 ──
    sc_active: pd.Series | None = None
    if not link_ts.empty:
        lnk = link_ts.set_index("snapshot") if "snapshot" in link_ts.columns else link_ts
        sc_cols = [c for c in lnk.columns
                   if any(k in c for k in ["Electrolyser", "_HP"])]
        if sc_cols:
            sc_sum    = lnk[sc_cols].clip(lower=0).sum(axis=1)
            sc_active = sc_sum > 1.0  # 1 MW 이상 가동

    # ── Curtailment 시계열 ──
    if not curt_ts.empty:
        ct = curt_ts.set_index("snapshot") if "snapshot" in curt_ts.columns else curt_ts
    else:
        ct = pd.DataFrame()

    for region in regions:
        # 해당 지역 curtailment 컬럼
        region_curt_cols = [c for c in ct.columns if c.startswith(region + "_")]
        if region_curt_cols:
            region_curt = ct[region_curt_cols].sum(axis=1)
        else:
            region_curt = pd.Series(0.0, index=ct.index if not ct.empty else [])

        total_curt_gwh   = twh(float(region_curt.clip(lower=0).sum())) * 1000  # GWh
        curt_hours       = int((region_curt > 1.0).sum())

        # 전력망 포화 & curtailment 동시 발생
        grid_and_curt = 0
        sc_and_curt   = 0
        grid_sc_curt  = 0

        if grid_sat_hours is not None and not region_curt.empty:
            common_idx = region_curt.index.intersection(grid_sat_hours.index)
            if len(common_idx) > 0:
                grid_and_curt = int(
                    (grid_sat_hours.loc[common_idx] & (region_curt.loc[common_idx] > 1.0)).sum()
                )

        if sc_active is not None and not region_curt.empty:
            common_idx2 = region_curt.index.intersection(sc_active.index)
            if len(common_idx2) > 0:
                sc_and_curt = int(
                    (sc_active.loc[common_idx2] & (region_curt.loc[common_idx2] > 1.0)).sum()
                )

        if grid_sat_hours is not None and sc_active is not None and not region_curt.empty:
            common_idx3 = region_curt.index.intersection(
                grid_sat_hours.index.intersection(sc_active.index)
            )
            if len(common_idx3) > 0:
                grid_sc_curt = int(
                    (grid_sat_hours.loc[common_idx3]
                     & sc_active.loc[common_idx3]
                     & (region_curt.loc[common_idx3] > 1.0)).sum()
                )

        rows.append({
            "지역":                  REGION_NAMES.get(region, region),
            "총Curtailment(GWh)":   round(total_curt_gwh, 1),
            "Curtailment발생(시간)": curt_hours,
            "전력망포화+Curt(시간)": grid_and_curt,
            "섹터커플링+Curt(시간)": sc_and_curt,
            "포화+섹커+Curt(시간)":  grid_sc_curt,
        })

    return pd.DataFrame(rows)


def export_curtailment_table(all_data: list[dict], line_caps: pd.DataFrame,
                              out_dir: str):
    """시나리오별 Curtailment · 전력망포화 · 섹터커플링 Excel 종합표 저장"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    default_sheet = wb.active

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    sub_fill    = PatternFill("solid", fgColor="BDD7EE")
    sub_font    = Font(bold=True, size=10)
    thin        = Side(border_style="thin", color="AAAAAA")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheets_created = 0
    for sc, years in SCENARIOS.items():
        for yr in sorted(years):
            matches = [d for d in all_data if d["scenario"] == sc and d["year"] == yr]
            if not matches:
                continue
            d   = matches[0]
            df  = build_curtailment_grid_table(d, line_caps)
            if df.empty:
                continue

            ws_name = f"{sc}{yr}"
            ws = wb.create_sheet(ws_name)
            sheets_created += 1

            # 제목
            ws.merge_cells("A1:F1")
            ws["A1"] = f"[{sc} {yr}] Curtailment · 전력망포화 · 섹터커플링 분석"
            ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
            ws["A1"].fill      = header_fill
            ws["A1"].alignment = center_aln

            # 헤더
            for ci, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=2, column=ci, value=col_name)
                cell.font      = sub_font
                cell.fill      = sub_fill
                cell.alignment = center_aln
                cell.border    = border

            # 데이터
            for ri, (_, row) in enumerate(df.iterrows(), start=3):
                for ci, val in enumerate(row, start=1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                               vertical="center")
                    cell.border = border

            # 컬럼 너비
            col_widths = [14, 18, 18, 20, 20, 20]
            for ci, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(ci)].width = w

    # 주석 시트
    ws_note = wb.create_sheet("분석방법론")
    notes = [
        ["항목", "설명"],
        ["총Curtailment(GWh)",   "해당 지역에서 1년간 발생한 재생에너지 출력제한 합계"],
        ["Curtailment발생(시간)", "시간당 Curtailment > 1MW인 시간수"],
        ["전력망포화+Curt(시간)", "동시에 어떤 선로든 이용률 ≥95% & 해당지역 Curtailment 발생 시간수"],
        ["섹터커플링+Curt(시간)", "동시에 전해조 또는 HP가 1MW 이상 가동 & 해당지역 Curtailment 발생 시간수"],
        ["포화+섹커+Curt(시간)",  "위 3개 조건 모두 동시 충족 시간수"],
        ["", "※ 전력망 포화 기준: 개별 선로 |유량| / s_nom ≥ 95%"],
    ]
    for ri, row_data in enumerate(notes, start=1):
        for ci, v in enumerate(row_data, start=1):
            ws_note.cell(row=ri, column=ci, value=v).font = (
                Font(bold=True) if ri == 1 else Font()
            )

    if default_sheet.title not in [ws.title for ws in wb.worksheets if ws != default_sheet]:
        wb.remove(default_sheet)

    xlsx_path = os.path.join(out_dir, "03_curtailment_grid_sectorcoupling.xlsx")
    wb.save(xlsx_path)
    print(f"  저장: {os.path.relpath(xlsx_path)}")


# ─────────────────────────────────────────────
# 5. 분석 4: 송전망 이용률 분포 (연도별 묶음 박스플롯)
# ─────────────────────────────────────────────

REGION_PAIR_LABELS = {
    "GGD_GWD": "경기↔강원서부",
    "GBD_GWD": "경북↔강원서부",
    "CBD_GWD": "서울↔강원서부",
    "GGD_ICN": "경기↔인천",
    "CND_GGD": "충남↔경기",
    "CBD_GGD": "서울↔경기",
    "GBD_GND": "경북↔강원",
    "BSN_GND": "부산↔강원",
    "GND_JND": "강원↔전남",
    "BSN_GBD": "부산↔경북",
    "CBD_GBD": "서울↔경북",
    "JBD_JND": "전북↔전남",
    "JJD_JND": "제주↔전남",
    "CND_JBD": "충남↔전북",
    "CBD_CND": "서울↔충남",
}

# 주요 송전 회랑 정의
# pair key: load_line_capacities()에서 sorted([region0, region1]) 형식
CORRIDORS = [
    {
        "name": "서남 → 수도권  (JND - JBD - CND - GGD)",
        "nodes": ["JND", "JBD", "CND", "GGD"],   # x축 하단 노드명
        "segments": [
            ("JBD_JND", "JND↔JBD"),
            ("CND_JBD", "JBD↔CND"),
            ("CND_GGD", "CND↔GGD"),
        ],
    },
    {
        "name": "동남 → 수도권  (BSN - GBD - CBD - GGD)",
        "nodes": ["BSN", "GBD", "CBD", "GGD"],
        "segments": [
            ("BSN_GBD", "BSN↔GBD"),
            ("CBD_GBD", "GBD↔CBD"),
            ("CBD_GGD", "CBD↔GGD"),
        ],
    },
]


def compute_pair_utilization(lf_ts: pd.DataFrame, line_caps: pd.DataFrame,
                              line_info: pd.DataFrame | None = None) -> pd.DataFrame:
    """
    권역 쌍별 시간별 이용률(%) 계산.

    우선순위:
      1) line_info (결과 Excel의 Line_Info 시트 — 실제 최적화 사용 s_nom)
      2) line_caps (integrated_input_data.xlsx 의 lines 시트)

    line_caps는 쌍(pair) 키 생성과 선로명 매핑에만 사용.
    """
    if lf_ts.empty or line_caps.empty:
        return pd.DataFrame()

    lf = lf_ts.copy()
    if "snapshot" in lf.columns:
        lf = lf.set_index("snapshot")

    # line_info 에서 s_nom dict 구성 (line_name → s_nom)
    li_snom: dict[str, float] = {}
    if line_info is not None and not line_info.empty:
        li_copy = line_info.copy()
        if "Line" in li_copy.columns:
            li_copy = li_copy.set_index("Line")
        cap_col = "s_nom_opt" if "s_nom_opt" in li_copy.columns else "s_nom"
        if cap_col in li_copy.columns:
            li_snom = li_copy[cap_col].dropna().to_dict()

    pairs = line_caps.groupby("pair")
    result = {}
    for pair_key, grp in pairs:
        lines_in_pair = [l for l in grp["name"] if l in lf.columns]
        if not lines_in_pair:
            continue

        # 용량: line_info 우선, 없으면 line_caps 사용
        total_cap = 0.0
        for ln in lines_in_pair:
            if ln in li_snom:
                total_cap += float(li_snom[ln])
            else:
                row_s = grp.loc[grp["name"] == ln, "s_nom"]
                total_cap += float(row_s.iloc[0]) if not row_s.empty else 0.0

        if total_cap <= 0:
            continue

        total_flow = lf[lines_in_pair].abs().sum(axis=1)
        result[pair_key] = (total_flow / total_cap * 100).clip(lower=0, upper=100)  # 0~100%

    return pd.DataFrame(result)


def _util_color(mean_util: float) -> str:
    """평균 이용률에 따른 박스 색상 (기존 이미지 기준)"""
    if mean_util >= 70:  return "#CC2222"   # 진빨 >=70%
    if mean_util >= 44:  return "#E88000"   # 주황 44~70%
    if mean_util >= 30:  return "#F0C040"   # 노랑 30~44%
    if mean_util >= 15:  return "#50A050"   # 초록 15~30%
    return "#4488CC"                         # 파랑 <15%


def plot_grid_utilization(all_data: list[dict], line_caps: pd.DataFrame,
                           scenario: str, out_dir: str):
    """
    주요 송전 회랑 이용률 분포
    - 회랑 1: 서남->수도권 (JND-JBD-CND-GGD)
    - 회랑 2: 동남->수도권 (BSN-GBD-CBD-GGD)
    - 각 구간마다 연도별 묶음 박스플롯 (기존 이미지 스타일 유지)
    - 박스 색상 = 이용률 수준별 색상, 연도 구분은 테두리 및 해칭
    """
    sc_data = sorted(
        [d for d in all_data if d["scenario"] == scenario],
        key=lambda d: d["year"]
    )
    if not sc_data:
        return
    if line_caps.empty:
        print(f"  [{scenario}] 선로 용량 데이터 없음, 이용률 건너뜀")
        return

    years   = [d["year"] for d in sc_data]
    n_years = len(years)

    # 연도별 박스 색상 (YEAR_COLORS 사용, 없으면 기본 팔레트)
    _YR_PALETTE = ["#2196F3", "#4CAF50", "#FF9800", "#9C27B0", "#F44336",
                   "#00BCD4", "#8BC34A", "#FF5722", "#673AB7", "#E91E63"]
    yr_color_map = {yr: (YEAR_COLORS.get(yr) or _YR_PALETTE[i % len(_YR_PALETTE)])
                    for i, yr in enumerate(years)}

    fig, axes = plt.subplots(
        nrows=len(CORRIDORS), ncols=1,
        figsize=(14, 6 * len(CORRIDORS)),
        squeeze=False,
    )

    for ci, corridor in enumerate(CORRIDORS):
        ax = axes[ci][0]
        segments = corridor["segments"]
        n_segs   = len(segments)

        # 구간별 박스 배치 위치 계산
        # 각 구간 그룹의 중심 간격 = n_years * box_w + gap
        box_w    = 0.14
        inner_gap = 0.04   # 연도간 간격
        group_gap = 0.55   # 구간간 추가 여백
        step = n_years * (box_w + inner_gap) + group_gap

        seg_centers = [si * step for si in range(n_segs)]

        # 연도 오프셋 (그룹 내 중앙 정렬)
        yr_offsets = [(yi - (n_years - 1) / 2) * (box_w + inner_gap)
                      for yi in range(n_years)]

        for si, (pair_key, seg_label) in enumerate(segments):
            cx = seg_centers[si]

            for yi, d in enumerate(sc_data):
                year = years[yi]
                util_df = compute_pair_utilization(
                    d["line_flow"], line_caps,
                    line_info=d.get("line_info")   # 결과 파일의 실제 s_nom 사용
                )

                if pair_key in util_df.columns:
                    series = util_df[pair_key].dropna().values
                else:
                    series = np.array([0.0])

                pos      = cx + yr_offsets[yi]
                yr_color = yr_color_map[year]

                bps = ax.boxplot(
                    [series],
                    positions=[pos],
                    widths=box_w,
                    patch_artist=True,
                    showfliers=False,
                    showmeans=True,
                    medianprops=dict(color="black",   linewidth=1.8),
                    whiskerprops=dict(color=yr_color, linewidth=1.2, linestyle="-"),
                    capprops=dict(color=yr_color,     linewidth=1.5),
                    meanprops=dict(
                        marker="D", markersize=6,
                        markerfacecolor="white",
                        markeredgecolor=yr_color,
                        markeredgewidth=1.8,
                    ),
                )
                box_patch = bps["boxes"][0]
                box_patch.set_facecolor(yr_color)   # 박스 채우기 = 연도 색상
                box_patch.set_alpha(0.65)
                box_patch.set_edgecolor(yr_color)
                box_patch.set_linewidth(2.0)

                # ── 포화/미사용 시간 주석 (모든 연도 동일 높이에 배치) ──
                sat_h  = int((series >= 99).sum())  # 99% 이상 = 포화
                zero_h = int((series <= 1).sum())   # 1% 이하 = 미사용

                # 포화: 100% 기준선 바로 위 고정 위치 (연도별 가로 오프셋만)
                if sat_h > 0:
                    ax.text(pos, 101.5, f"{sat_h}h",
                            ha="center", va="bottom",
                            fontsize=6.5, color="#CC0000", fontweight="bold")

                # 미사용: 0% 기준선 바로 아래 고정 위치
                if zero_h > 0:
                    ax.text(pos, -2.0, f"{zero_h}h",
                            ha="center", va="top",
                            fontsize=6.0, color="#336699")

        # ── 100% 기준선 ──
        ax.axhline(y=100, color="#CC0000", linestyle="--",
                   linewidth=1.3, alpha=0.9, zorder=0)
        ax.text(seg_centers[-1] + yr_offsets[-1] + box_w * 0.7,
                100, " 100%", color="#CC0000", fontsize=8, va="center")

        # ── Y축 설정 ──
        # 데이터 최솟값/최댓값 기반 자동 스케일 (최소 0~100 범위 확보)
        all_vals: list[float] = []
        for pk, _ in segments:
            for d in sc_data:
                u_df = compute_pair_utilization(
                    d["line_flow"], line_caps, line_info=d.get("line_info"))
                if pk in u_df.columns:
                    all_vals.extend(u_df[pk].dropna().tolist())
        y_data_max = max(max(all_vals), 100.0) if all_vals else 100.0
        y_data_min = min(min(all_vals), 0.0)   if all_vals else 0.0
        top_margin    = max(16, (y_data_max - y_data_min) * 0.18)  # 포화 주석 공간
        bottom_margin = max(10, (y_data_max - y_data_min) * 0.12)  # 미사용 주석 공간
        ax.set_ylim(bottom=y_data_min - bottom_margin,
                    top=y_data_max + top_margin)

        ax.set_ylabel("이용률 (%)", fontsize=10)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}%"))
        ax.grid(axis="y", linestyle="--", alpha=0.35)

        # ── 구간 경계 세로선 & 그 위치에 역 노드명 배치 ──
        # 회랑 노드 목록 (CORRIDORS의 "nodes" key 사용)
        corridor_nodes = corridor.get("nodes", [])
        divider_xs: list[float] = []
        for si in range(1, n_segs):
            mid = (seg_centers[si - 1] + seg_centers[si]) / 2
            ax.axvline(x=mid, color="#AAAAAA", linewidth=0.8, linestyle=":")
            divider_xs.append(mid)

        # X축: 기존 tick 제거 후 숨김
        ax.set_xticks([])
        ax.set_xticklabels([])
        ax.spines["bottom"].set_visible(True)

        # 회랑 노드명을 x축 아래에 그림 (좌끝/분할선/우끝)
        x_left  = seg_centers[0]  + yr_offsets[0]  - box_w * 0.5
        x_right = seg_centers[-1] + yr_offsets[-1] + box_w * 0.5

        node_xs = [x_left] + divider_xs + [x_right]
        for ni, (nx_, node_name) in enumerate(zip(node_xs, corridor_nodes)):
            ha = "left" if ni == 0 else ("right" if ni == len(corridor_nodes) - 1 else "center")
            ax.text(nx_, ax.get_ylim()[0] - (y_data_max - y_data_min) * 0.05,
                    node_name, ha=ha, va="top",
                    fontsize=11, fontweight="bold", color="#222222",
                    transform=ax.transData, clip_on=False)

        # 노드 간 화살표 (→) 그리기
        y_arrow = ax.get_ylim()[0] - (y_data_max - y_data_min) * 0.04
        for ni in range(len(node_xs) - 1):
            x_a = node_xs[ni]   + 0.04
            x_b = node_xs[ni+1] - 0.04
            ax.annotate(
                "", xy=(x_b, y_arrow), xytext=(x_a, y_arrow),
                arrowprops=dict(arrowstyle="->", color="#555555", lw=1.2),
                annotation_clip=False,
            )

        ax.set_title(f"{corridor['name']}", fontsize=12, fontweight="bold", pad=8)

        # 연도별 색상 범례
        year_legend = [
            mpatches.Patch(facecolor=yr_color_map[yr], edgecolor=yr_color_map[yr],
                           alpha=0.8, label=str(yr))
            for yr in years
        ]
        ax.legend(handles=year_legend, loc="upper right",
                  fontsize=8, title="연도", title_fontsize=8,
                  framealpha=0.9)

    fig.suptitle(
        f"[{scenario}] 주요 송전 회랑 이용률 분포\n"
        f"(IQR 범위, ◆ 평균, — 중앙값 | 위숫자: 포화시간, 아래숫자: 미사용시간)",
        fontsize=13, fontweight="bold",
    )
    plt.tight_layout(rect=[0, 0.02, 1, 0.97])

    out_path = os.path.join(out_dir, f"04_grid_utilization_{scenario}.png")
    fig.savefig(out_path, bbox_inches="tight", dpi=150)
    plt.close(fig)
    print(f"  저장: {os.path.relpath(out_path)}")


# ─────────────────────────────────────────────
# 6. 분석 1b: 에너지 수지 (생산/소비 상하 막대)
# ─────────────────────────────────────────────

PROD_CATS = ["PV", "WT", "Hydro", "Nuclear", "LNG", "Coal", "CHP", "Fuelcell"]
PROD_COLORS_BAL = {
    "PV": "#FFD700", "WT": "#87CEEB", "Hydro": "#4169E1",
    "Nuclear": "#FF6347", "LNG": "#FFA500", "Coal": "#808080",
    "CHP": "#BC8F8F", "Fuelcell": "#32CD32",
    "Curtailment": "#C8A000",   # 출력제한 (금색 빗금)
}
CONS_CATS = ["전력수요", "DC수요", "FAB수요", "EV수요", "열수요", "수소수요"]
CONS_COLORS_BAL = {
    "전력수요": "#CC3333", "DC수요": "#EE7777",
    "FAB수요": "#FF9966", "EV수요": "#6495ED",
    "열수요":  "#FF69B4", "수소수요": "#20B2AA",
}


def _extract_energy_balance(data: dict) -> tuple[dict, dict]:
    """연도별 에너지 수지: (생산 TWh dict, 소비 TWh dict)"""
    go_ts = data.get("generator_output", pd.DataFrame())
    lf_ts = data.get("link_flow",        pd.DataFrame())
    hl    = data.get("hourly_loads",     pd.DataFrame())
    cs    = data.get("curtailment_summary", pd.DataFrame())

    prod = {cat: 0.0 for cat in PROD_CATS + ["Curtailment"]}
    cons = {cat: 0.0 for cat in CONS_CATS}

    # 발전기 출력 (Generator_Output)
    if not go_ts.empty:
        for col in go_ts.columns:
            if col == "snapshot":
                continue
            cat = classify_generator(col)
            if cat in prod:
                prod[cat] += twh(float(go_ts[col].clip(lower=0).sum()))

    # 링크 출력 (Link_Flow: CHP, Fuelcell)
    if not lf_ts.empty:
        for col in lf_ts.columns:
            if col == "snapshot":
                continue
            cat = classify_link(col)
            if cat in ("CHP", "Fuelcell"):
                prod[cat] = prod.get(cat, 0) + twh(float(lf_ts[col].clip(lower=0).sum()))

    # 출력제한 (Curtailment)
    if not cs.empty:
        cc = [c for c in cs.columns if "Curtailment_MWh" in c or "curtailment" in c.lower()]
        if cc:
            prod["Curtailment"] = twh(float(cs[cc[0]].sum()))

    # 소비: Hourly_Loads에서 추출
    if not hl.empty:
        # 컬럼명 공백 제거
        hl.columns = [c.strip() for c in hl.columns]
        load_cols = [c for c in hl.columns if c != "snapshot"]

        # 중복 컬럼 제거: 같은 이름 중 값이 있는 쪽 유지
        hl = hl.loc[:, ~hl.columns.duplicated(keep="last")]
        load_cols = [c for c in hl.columns if c != "snapshot"]

        el_cols  = [c for c in load_cols if "_Demand_EL"  in c]
        dc_cols  = [c for c in load_cols if "_Demand_DC"  in c and c.endswith("_DC")]
        fab_cols = [c for c in load_cols if "_Demand_Fab" in c or "_Demand_FAB" in c]
        ev_cols  = [c for c in load_cols if "_Demand_EV"  in c]
        h_cols   = [c for c in load_cols if "_Demand_H"   in c and "_Demand_H2" not in c
                    and "_Demand_EL" not in c and "_Demand_HP" not in c]
        h2_cols  = [c for c in load_cols if "_Demand_H2"  in c]

        if el_cols:  cons["전력수요"] = twh(float(hl[el_cols].sum().sum()))
        if dc_cols:  cons["DC수요"]   = twh(float(hl[dc_cols].sum().sum()))
        if fab_cols: cons["FAB수요"]  = twh(float(hl[fab_cols].sum().sum()))
        if ev_cols:  cons["EV수요"]   = twh(float(hl[ev_cols].sum().sum()))
        if h_cols:   cons["열수요"]   = twh(float(hl[h_cols].sum().sum()))
        if h2_cols:  cons["수소수요"] = twh(float(hl[h2_cols].sum().sum()))

    return prod, cons


def plot_energy_balance(all_data: list[dict], scenario: str, out_dir: str):
    """
    연도별 에너지 수지 차트 (두 번째 이미지 참고)
    - 위: 생산 (PV/WT/Hydro/Nuclear/LNG/Coal/CHP/Fuelcell + 출력제한)
    - 아래: 소비 (전력수요/DC수요/FAB수요/EV수요/열수요/수소수요)
    - 순수지 수치 표시
    """
    sc_data = sorted([d for d in all_data if d["scenario"] == scenario], key=lambda d: d["year"])
    if not sc_data:
        return

    years = [d["year"] for d in sc_data]
    n     = len(years)
    x     = np.arange(n)
    w     = 0.65

    # 연도별 생산/소비 추출
    all_prod = []
    all_cons = []
    for d in sc_data:
        p, c = _extract_energy_balance(d)
        all_prod.append(p)
        all_cons.append(c)

    fig, ax = plt.subplots(figsize=(12, 8))

    # ── 생산 (위, +) ──
    bot_p = np.zeros(n)
    hdl_p = []
    for cat in PROD_CATS:
        vals  = np.array([p.get(cat, 0.0) for p in all_prod])
        color = PROD_COLORS_BAL.get(cat, "#AAAAAA")
        ax.bar(x, vals, w, bottom=bot_p, color=color, edgecolor="white", linewidth=0.3)
        bot_p += vals
        if any(v > 0 for v in vals):
            hdl_p.append(mpatches.Patch(color=color, label=cat))

    # 출력제한 (빗금)
    curt_vals = np.array([p.get("Curtailment", 0.0) for p in all_prod])
    if curt_vals.sum() > 0:
        ax.bar(x, curt_vals, w, bottom=bot_p,
               color=PROD_COLORS_BAL["Curtailment"], edgecolor="white",
               linewidth=0.3, hatch="//", label="출력제한")
        hdl_p.append(mpatches.Patch(color=PROD_COLORS_BAL["Curtailment"],
                                    hatch="//", label="출력제한"))
        bot_p += curt_vals

    # ── 소비 (아래, -) ──
    bot_c = np.zeros(n)
    hdl_c = []
    for cat in CONS_CATS:
        vals  = np.array([-c.get(cat, 0.0) for c in all_cons])  # 음수
        color = CONS_COLORS_BAL.get(cat, "#BBBBBB")
        mask  = vals < 0
        if mask.any():
            ax.bar(x, vals, w, bottom=bot_c, color=color, edgecolor="white", linewidth=0.3)
            bot_c += vals
            hdl_c.append(mpatches.Patch(color=color, label=cat))

    # ── 순수지 주석 ──
    for xi in range(n):
        net = bot_p[xi] + bot_c[xi]   # 생산 합계 + 소비 합계(음수)
        y_annot = (bot_p[xi] + bot_c[xi]) / 2
        ax.text(x[xi], y_annot, f"{net:+.0f}",
                ha="center", va="center", fontsize=8,
                fontweight="bold", color="white",
                bbox=dict(boxstyle="round,pad=0.2", fc="#333333", alpha=0.6, ec="none"))

    # ── 제로 기준선 ──
    ax.axhline(0, color="black", linewidth=1.2)

    # ── 축 설정 ──
    ax.set_xticks(x)
    ax.set_xticklabels([f"'{str(y)[-2:]}" for y in years], fontsize=11)
    ax.set_ylabel("에너지 (TWh)", fontsize=11)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax.grid(axis="y", linestyle="--", alpha=0.35)

    # ── 범례 ──
    leg1 = ax.legend(handles=hdl_p, loc="upper left",
                     fontsize=8, title="생산(+)", title_fontsize=9, ncol=2, framealpha=0.9)
    ax.add_artist(leg1)
    ax.legend(handles=hdl_c, loc="lower left",
              fontsize=8, title="소비(-)", title_fontsize=9, ncol=2, framealpha=0.9)

    ax.set_title(f"[{scenario}] 연도별 에너지 수지\n(생산: +, 소비: -, 가운데 숫자: 순수지 TWh)",
                 fontsize=12, fontweight="bold")
    plt.tight_layout()

    out_path = os.path.join(out_dir, f"01b_energy_balance_{scenario}.png")
    fig.savefig(out_path, bbox_inches="tight", dpi=150)
    plt.close(fig)
    print(f"  저장: {os.path.relpath(out_path)}")


# ─────────────────────────────────────────────
# 7. 분석 3b: Curtailment 요약 연도별 추이 그래프
# ─────────────────────────────────────────────

def plot_curtailment_trend(all_data: list[dict], scenario: str, out_dir: str):
    """연도별 Curtailment 총량 추이 + 지역별 분해"""
    sc_data = sorted(
        [d for d in all_data if d["scenario"] == scenario],
        key=lambda d: d["year"]
    )
    if not sc_data:
        return

    years   = [d["year"] for d in sc_data]
    regions = list(REGION_NAMES.keys())
    x       = np.arange(len(years))
    width   = 0.6

    # 지역 × 연도 Curtailment (TWh)
    region_curt: dict[str, list[float]] = {r: [] for r in regions}

    for d in sc_data:
        cs = d.get("curtailment_summary", pd.DataFrame())
        yr_reg: dict[str, float] = {r: 0.0 for r in regions}

        if not cs.empty:
            reg_col  = [c for c in cs.columns if "지역" in c or "Region" in c or c == "region"]
            curt_col = [c for c in cs.columns if "Curtailment_MWh" in c]
            if reg_col and curt_col:
                for _, row in cs.iterrows():
                    region = str(row[reg_col[0]]).strip()
                    if region in yr_reg:
                        yr_reg[region] += twh(float(row[curt_col[0]] or 0)) * 1000  # → TWh

        for r in regions:
            region_curt[r].append(yr_reg.get(r, 0.0))

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

    # 스택 막대 – 지역별 Curtailment
    region_colors = plt.cm.tab20(np.linspace(0, 1, len(regions)))
    bottom = np.zeros(len(years))
    for ri, region in enumerate(regions):
        vals = np.array(region_curt[region])
        ax1.bar(x, vals, width, bottom=bottom,
                color=region_colors[ri], label=REGION_NAMES.get(region, region),
                edgecolor="white", linewidth=0.4)
        bottom += vals

    ax1.set_xticks(x)
    ax1.set_xticklabels([str(y) for y in years])
    ax1.set_ylabel("출력제한량 (TWh)")
    ax1.set_title(f"[{scenario}] 지역별 Curtailment 추이")
    ax1.legend(loc="upper left", fontsize=7, ncol=2)
    ax1.grid(axis="y", linestyle="--", alpha=0.4)

    # 선 그래프 – 비율(발전량 대비 curtailment)
    curt_rates = []
    for d in sc_data:
        cs  = d.get("curtailment_summary", pd.DataFrame())
        fe  = d.get("final_energy",        pd.DataFrame())

        total_curt = 0.0
        if not cs.empty:
            curt_col = [c for c in cs.columns if "Curtailment_MWh" in c]
            if curt_col:
                total_curt = twh(float(cs[curt_col[0]].sum()))

        total_re = 0.0
        if not fe.empty:
            fe_clean = fe.dropna(subset=[fe.columns[1]])
            for _, row in fe_clean.iterrows():
                tech = str(row.iloc[1]).strip()
                if any(k in tech for k in ["PV", "WT", "태양광", "풍력", "수력", "Hydro"]):
                    try:
                        total_re += twh(float(row.iloc[2] or 0))
                    except Exception:
                        pass

        rate = total_curt / (total_re + total_curt) * 100 if (total_re + total_curt) > 0 else 0
        curt_rates.append(rate)

    ax2.plot(years, curt_rates, "o-", color="#CC0000", linewidth=2, markersize=7)
    for y, r in zip(years, curt_rates):
        ax2.annotate(f"{r:.1f}%", (y, r), textcoords="offset points",
                     xytext=(0, 10), ha="center", fontsize=9)
    ax2.set_ylabel("Curtailment 비율 (RE 잠재발전량 대비 %)")
    ax2.set_title(f"[{scenario}] Curtailment 비율 추이")
    ax2.set_xticks(years)
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.1f}%"))
    ax2.grid(linestyle="--", alpha=0.4)

    plt.tight_layout()
    out_path = os.path.join(out_dir, f"03b_curtailment_trend_{scenario}.png")
    fig.savefig(out_path, bbox_inches="tight")
    plt.close(fig)
    print(f"  저장: {os.path.relpath(out_path)}")


# ─────────────────────────────────────────────
# 8. 메인
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="PyPSA-GESI 시나리오 분석")
    parser.add_argument("--scenario", type=str, default=None,
                        help="분석 시나리오 (REF 또는 CZ). 미지정 시 전체")
    parser.add_argument("--year",     type=int, default=None,
                        help="분석 연도 (예: 2035). 미지정 시 전체")
    parser.add_argument("--skip",     nargs="*", default=[],
                        help="건너뛸 분석 번호 목록 (예: --skip 2 4)")
    args = parser.parse_args()

    skip = set(str(s) for s in args.skip)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"\n{'='*60}")
    print(f"PyPSA-GESI 시나리오 분석")
    print(f"출력 폴더: {OUTPUT_DIR}")
    print(f"{'='*60}\n")

    # 데이터 로드
    print("[1/5] 데이터 로딩 중...")
    all_data = load_all_data(
        filter_scenario=args.scenario,
        filter_year=args.year,
    )
    if not all_data:
        print("  [경고] 분석 가능한 데이터가 없습니다.")
        print(f"  폴더 구조를 확인하세요: {SCENARIO_ROOT}")
        print("  예) scenario_result/REF/2035/REF2035.xlsx")
        return
    print(f"  => {len(all_data)}개 시나리오-연도 로드 완료\n")

    # 선로 용량 로드
    line_caps = load_line_capacities()
    if line_caps.empty:
        print("  [경고] 선로 용량 데이터 없음. 이용률 분석 제한됨.")

    # 시나리오 목록 결정
    scenarios_to_run = [args.scenario] if args.scenario else list(SCENARIOS.keys())

    for sc in scenarios_to_run:
        sc_count = sum(1 for d in all_data if d["scenario"] == sc)
        if sc_count == 0:
            continue
        print(f"{'─'*50}")
        print(f"시나리오: {sc}  ({sc_count}개 연도)")
        print(f"{'─'*50}")

        if "1" not in skip:
            print("[분석 1] 에너지 믹스 - 설비용량 (통합 차트)")
            plot_energy_mix(all_data, sc, OUTPUT_DIR)
            print("[분석 1b] 에너지 수지 (생산/소비)")
            plot_energy_balance(all_data, sc, OUTPUT_DIR)

        if "3" not in skip:
            print("[분석 3] Curtailment 추이 그래프")
            plot_curtailment_trend(all_data, sc, OUTPUT_DIR)

        print()

    if "2" not in skip:
        print("[분석 2] Sankey 에너지 흐름 (REF/CZ 비교 HTML)")
        # Sankey는 전체 데이터를 한 번에 처리 (시나리오 무관)
        plot_sankey(all_data, "", OUTPUT_DIR)

    if "3" not in skip:
        print("[분석 3b] Curtailment 종합 Excel 저장")
        export_curtailment_table(all_data, line_caps, OUTPUT_DIR)

    if "4" not in skip:
        for sc in scenarios_to_run:
            sc_count = sum(1 for d in all_data if d["scenario"] == sc)
            if sc_count == 0:
                continue
            print(f"[분석 4] 송전망 이용률 분포 [{sc}]")
            plot_grid_utilization(all_data, line_caps, sc, OUTPUT_DIR)

    print(f"\n{'='*60}")
    print(f"분석 완료. 결과 저장 위치: {OUTPUT_DIR}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
