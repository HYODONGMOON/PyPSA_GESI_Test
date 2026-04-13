import pandas as pd
import openpyxl

print("=" * 80)
print("1. integrated_input_data.xlsx의 Coal 발전기 carrier 확인")
print("=" * 80)

df_gens = pd.read_excel('integrated_input_data.xlsx', sheet_name='generators')

# Coal 발전기 필터링
coal_gens = df_gens[df_gens['name'].str.contains('Coal|coal|석탄|당진|영흥|하동|삼척|태안', case=False, na=False)]

print(f"\nCoal 관련 발전기: {len(coal_gens)}개")
print("\nCarrier 값 분포:")
print(coal_gens['carrier'].value_counts())

print("\n샘플:")
print(coal_gens[['name', 'carrier', 'p_nom']].head(10))

print("\n" + "=" * 80)
print("2. interface.xlsx의 지역별 시트에서 Coal 발전기 carrier 확인")
print("=" * 80)

wb = openpyxl.load_workbook('interface.xlsx', data_only=True)

# CND 지역 확인
if '지역_CND' in wb.sheetnames:
    ws = wb['지역_CND']
    
    # 발전기 섹션 찾기
    print("\n지역_CND의 발전기 섹션에서 Coal 찾기:")
    for row in range(1, min(ws.max_row + 1, 500)):
        cell_a = ws.cell(row, 1).value
        if cell_a and 'Coal' in str(cell_a):
            # 헤더 찾기 - carrier 열 위치 확인
            # 일반적으로 name, bus, carrier, p_nom 순서
            print(f"\n  Row {row}: {cell_a}")
            # 주변 10개 행 확인
            for offset in range(-2, 8):
                r = row + offset
                if r > 0:
                    vals = [ws.cell(r, c).value for c in range(1, 15)]
                    print(f"    Row {r}: {vals[:10]}")
            break

wb.close()
