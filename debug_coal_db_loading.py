import openpyxl
import pandas as pd

print("=" * 80)
print("Coal_DB 데이터 상세 확인")
print("=" * 80)

wb = openpyxl.load_workbook('interface.xlsx', data_only=True)
ws = wb['Coal_DB']

# 선택연도
selected_year = ws['D2'].value
print(f"선택연도: {selected_year}")

# Row 5 헤더 확인
print(f"\nRow 5 헤더:")
for col in range(1, 10):
    val = ws.cell(5, col).value
    print(f"  {chr(64+col)}: {val}")

# Row 6부터 데이터 샘플 확인
print(f"\nRow 6-15 데이터 샘플:")
print(f"{'Row':<5} {'지역(A)':<10} {'발전소(D)':<15} {'호기(E)':<10} {'용량(F)':<10} {'효율(AM)':<10}")
print("-" * 70)

for row in range(6, 16):
    region = ws.cell(row, 1).value
    plant = ws.cell(row, 4).value
    unit = ws.cell(row, 5).value
    capacity = ws.cell(row, 6).value  # F열 (2024년)
    efficiency = ws.cell(row, 39).value  # AM열
    
    region_str = str(region) if region else "None"
    plant_str = str(plant) if plant else "None"
    unit_str = str(unit) if unit else "None"
    cap_str = str(capacity) if capacity else "None"
    eff_str = str(efficiency) if efficiency else "None"
    
    print(f"{row:<5} {region_str:<10} {plant_str:<15} {unit_str:<10} {cap_str:<10} {eff_str:<10}")

wb.close()

print("\n" + "=" * 80)
print("통합 데이터 generators 시트 확인")
print("=" * 80)

df_gens = pd.read_excel('integrated_input_data.xlsx', sheet_name='generators')
print(f"\n총 발전기 수: {len(df_gens)}")

# Carrier별 분포
print(f"\nCarrier별 발전기 수:")
print(df_gens['carrier'].value_counts())

# 모든 발전기 이름 확인 (처음 30개)
print(f"\n발전기 이름 샘플 (처음 30개):")
for idx, name in enumerate(df_gens['name'].head(30)):
    print(f"  {idx+1}. {name}")

# Coal이나 석탄발전소 이름이 있는지 확인
print(f"\n\n'Coal' 포함 발전기:")
coal_named = df_gens[df_gens['name'].str.contains('Coal', case=False, na=False)]
print(f"개수: {len(coal_named)}")
if not coal_named.empty:
    print(coal_named[['name', 'carrier', 'p_nom']])

# 지역별 발전기 수
print(f"\n\n지역별 발전기 수:")
region_counts = {}
for name in df_gens['name']:
    region = str(name).split('_')[0] if '_' in str(name) else 'Unknown'
    region_counts[region] = region_counts.get(region, 0) + 1

for region, count in sorted(region_counts.items()):
    print(f"  {region}: {count}개")
