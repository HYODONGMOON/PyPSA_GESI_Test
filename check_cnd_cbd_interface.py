import openpyxl

wb = openpyxl.load_workbook('interface.xlsx')

print('='*70)
print('CND, CBD 시트의 Coal/Hydro 확인')
print('='*70)

for sheet_name in ['지역_CND', '지역_CBD']:
    ws = wb[sheet_name]
    
    print(f'\n[{sheet_name}]')
    
    # 발전기 섹션 찾기
    for row in range(1, min(200, ws.max_row + 1)):
        cell = ws.cell(row=row, column=1)
        if cell.value and '발전기' in str(cell.value):
            header_row = row + 2
            
            # 헤더 읽기
            headers = []
            for i in range(1, 30):
                h = ws.cell(row=header_row, column=i).value
                if h:
                    headers.append(h)
                else:
                    break
            
            print(f'  헤더: {headers[:8]}')
            
            # carrier 컬럼 위치
            carrier_col = None
            for i, h in enumerate(headers):
                if '캐리어' in str(h):
                    carrier_col = i + 1
                    break
            
            # 모든 발전기 출력
            for data_row in range(header_row+1, min(header_row+30, ws.max_row+1)):
                name = ws.cell(row=data_row, column=1).value
                if not name:
                    break
                
                carrier = ws.cell(row=data_row, column=carrier_col).value if carrier_col else None
                p_nom = ws.cell(row=data_row, column=4).value  # 4번째 컬럼은 p_nom
                print(f'  {name}: carrier={carrier}, p_nom={p_nom}')
            
            break

print('\n' + '='*70)
