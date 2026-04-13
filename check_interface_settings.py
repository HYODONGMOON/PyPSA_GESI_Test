import openpyxl

wb = openpyxl.load_workbook('interface.xlsx', data_only=True)

print("=" * 80)
print("interface.xlsx의 지역별 시트에서 발전기 설정 확인")
print("=" * 80)

# 샘플로 CND 지역 확인
if '지역_CND' in wb.sheetnames:
    ws = wb['지역_CND']
    
    print("\n지역_CND의 발전기 섹션:")
    
    # 발전기 섹션 찾기 (Coal, LNG 등)
    for row in range(1, min(ws.max_row + 1, 100)):
        cell_a = ws.cell(row, 1).value
        
        if cell_a and str(cell_a).strip() in ['Coal', 'LNG']:
            print(f"\n{cell_a} 발전기 (Row {row}):")
            
            # 헤더 찾기 - 보통 2행 위에 헤더가 있음
            header_row = row - 2
            headers = []
            for col in range(1, 15):
                h = ws.cell(header_row, col).value
                if h:
                    headers.append((col, h))
            
            print(f"  헤더: {[h[1] for h in headers]}")
            
            # 데이터 읽기
            data = {}
            for col, header in headers:
                value = ws.cell(row, col).value
                data[header] = value
            
            print(f"  데이터:")
            for k, v in data.items():
                if k in ['이름', 'name', '버스', 'bus', 'carrier', 'committable', 'p_min_pu', 'marginal_cost', 'p_nom', '명목용량(MW)']:
                    print(f"    {k}: {v}")

print("\n" + "=" * 80)
print("Coal_DB 시트 확인")
print("=" * 80)

if 'Coal_DB' in wb.sheetnames:
    ws = wb['Coal_DB']
    selected_year = ws['D2'].value
    print(f"\n선택연도 (D2): {selected_year}")
    
    # 샘플 데이터 확인
    print("\n샘플 석탄발전기 (상위 5개):")
    for row in range(6, 11):
        region = ws.cell(row, 1).value
        plant = ws.cell(row, 4).value
        unit = ws.cell(row, 5).value
        capacity = ws.cell(row, 6).value  # F열 (2024년)
        efficiency = ws.cell(row, 39).value  # AM열
        
        if region and plant:
            print(f"  {region}_{plant}_{unit}: {capacity} MW, 효율 {efficiency}")

wb.close()
