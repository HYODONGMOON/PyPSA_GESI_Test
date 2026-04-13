import openpyxl

wb = openpyxl.load_workbook('interface.xlsx', data_only=True)

if '지역_CND' in wb.sheetnames:
    ws = wb['지역_CND']
    
    print("=" * 80)
    print("지역_CND 발전기 섹션 헤더 확인 (Row 19)")
    print("=" * 80)
    
    # 헤더 행 (Row 19)
    headers = []
    for col in range(1, 20):
        val = ws.cell(19, col).value
        if val:
            headers.append((col, val))
            print(f"  Col {col:2d}: {val}")
    
    print("\n" + "=" * 80)
    print("Coal 발전기 데이터 (Row 25)")
    print("=" * 80)
    
    for col, header in headers:
        val = ws.cell(25, col).value
        print(f"  {header}: {val}")
    
    print("\n" + "=" * 80)
    print("LNG 발전기 데이터 (Row 22)")
    print("=" * 80)
    
    for col, header in headers:
        val = ws.cell(22, col).value
        print(f"  {header}: {val}")

wb.close()
