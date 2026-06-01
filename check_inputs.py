import openpyxl

# 1) interface.xlsx 인터페이스_1 - E/F 열 (이용률, 한도)
wb1 = openpyxl.load_workbook('interface.xlsx', data_only=True)
ws1 = wb1['인터페이스_1']
print('=== interface.xlsx 인터페이스_1 : 연결별 설정 ===')
print(f'  A1 (목표연도) = {ws1["A1"].value}')
print()
print('  [연결 입력표 B53:F75]  (B=출발, C=도착, D=용량, E=이용률%, F=한도h)')
for row in ws1.iter_rows(min_row=54, max_row=75, min_col=2, max_col=6):
    b, c, d, e, f = [cell.value for cell in row]
    if b and c:
        print(f'    {str(b):<6} -> {str(c):<6}  D={d}  E(이용률%)={e}  F(한도h)={f}')

# 2) 인터페이스_1 의 GGD_CND 선로 J열(2030) s_nom
print()
print('  [개별 선로 s_nom (J열=2030)] - GGD_CND / GND_JND / GBD_CBD / JND_JJD / JBD_CND')
target = ['GGD_CND', 'GND_JND', 'GBD_CBD', 'JND_JJD', 'JBD_CND']
for row in ws1.iter_rows(min_row=6, max_row=51):
    name = row[0].value
    if name and any(k in str(name) for k in target):
        j_val = row[9].value  # J열 = index 9 (0-based)
        f_val = row[5].value  # F열 = index 5
        print(f'    {str(name):<15}  F(s_nom)={f_val}   J(2030)={j_val}')

# 3) integrated_input_data.xlsx - 선로 s_nom
wb2 = openpyxl.load_workbook('integrated_input_data.xlsx', data_only=True)
print()
print(f'=== integrated_input_data.xlsx 시트 목록 ===')
print(' ', wb2.sheetnames)

lines_sheet = None
for sn in wb2.sheetnames:
    sl = sn.lower()
    if 'line' in sl or '선로' in sl or 'transmission' in sl:
        lines_sheet = sn
        break

if lines_sheet:
    ws2 = wb2[lines_sheet]
    headers = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
    print(f'\n  Lines 시트: "{lines_sheet}"')
    print(f'  헤더: {headers[:8]}')
    print(f'\n  [GGD_CND / GND_JND / GBD_CBD / JND_JJD / JBD_CND 관련 행]')
    for row in ws2.iter_rows(min_row=2, values_only=True):
        name = str(row[0]) if row[0] else ''
        if any(k in name for k in target):
            print(f'    {row[:8]}')
else:
    print('  Lines 시트를 찾지 못했습니다.')
    print('  시트명을 직접 확인하세요:', wb2.sheetnames)
