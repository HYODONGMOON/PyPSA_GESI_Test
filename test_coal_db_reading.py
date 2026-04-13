import pandas as pd
import openpyxl

def read_coal_db(interface_file='interface.xlsx'):
    """Coal_DB 시트에서 개별 석탄발전기 정보를 읽어옴"""
    
    # D2 셀에서 선택연도 읽기
    wb = openpyxl.load_workbook(interface_file, data_only=True)  # data_only=True로 수식 결과값 읽기
    ws = wb['Coal_DB']
    selected_year = ws['D2'].value
    
    # 선택연도가 문자열일 수도 있음
    if isinstance(selected_year, str):
        try:
            selected_year = int(selected_year)
        except ValueError:
            pass
    
    print(f"선택연도: {selected_year}")
    
    # 5번 행이 헤더
    # A5=지역, D5=발전소, E5=호기, F5~AF5=연도(2024~2050), AM5=효율
    
    # 연도와 열 번호 매핑 (5번 행에서 읽기, F열부터 AF열까지)
    year_to_col = {}
    for col_idx in range(6, 33):  # F(6)부터 AF(32)까지
        year_val = ws.cell(5, col_idx).value
        if year_val:
            try:
                # 문자열로 저장된 연도를 정수로 변환
                year_int = int(str(year_val).strip())
                year_to_col[year_int] = col_idx
            except (ValueError, AttributeError):
                continue
    
    print(f"연도-열 매핑: {year_to_col}")
    
    # 선택연도에 해당하는 열 찾기
    if selected_year not in year_to_col:
        print(f"경고: 선택연도 {selected_year}가 테이블에 없습니다.")
        print(f"사용 가능한 연도: {sorted(year_to_col.keys())}")
        wb.close()
        return None
    
    capacity_col = year_to_col[selected_year]
    col_letter = openpyxl.utils.get_column_letter(capacity_col)
    print(f"{selected_year}년 용량은 {col_letter}열 (col {capacity_col})")
    
    # 데이터 읽기 (6번 행부터)
    coal_generators = []
    for row_idx in range(6, ws.max_row + 1):
        region = ws.cell(row_idx, 1).value  # A열: 지역
        plant_name = ws.cell(row_idx, 4).value  # D열: 발전소
        unit = ws.cell(row_idx, 5).value  # E열: 호기
        capacity = ws.cell(row_idx, capacity_col).value  # 선택연도 용량
        efficiency = ws.cell(row_idx, 39).value  # AM열: 효율
        
        # 유효한 데이터만 추가
        if region and plant_name:
            try:
                capacity_val = float(capacity) if capacity else 0.0
                efficiency_val = float(efficiency) if efficiency else 0.45  # 기본값
                
                if capacity_val > 0:  # 용량이 0보다 큰 경우만
                    coal_generators.append({
                        'region': str(region).strip(),
                        'plant_name': str(plant_name).strip(),
                        'unit': str(unit).strip() if unit else '1',
                        'capacity_MW': capacity_val,
                        'efficiency': efficiency_val
                    })
            except (ValueError, TypeError) as e:
                print(f"Row {row_idx} 처리 오류: {e}")
                continue
    
    wb.close()
    
    return pd.DataFrame(coal_generators)

# 테스트
df_coal = read_coal_db()
if df_coal is not None:
    print(f"\n총 석탄발전기 수: {len(df_coal)}")
    print(f"\n지역별 발전기 수:")
    print(df_coal['region'].value_counts())
    print(f"\n상위 10개:")
    print(df_coal.head(10))
    print(f"\n총 용량: {df_coal['capacity_MW'].sum():.1f} MW")
