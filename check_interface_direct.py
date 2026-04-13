import openpyxl

wb = openpyxl.load_workbook('interface.xlsx')

print('='*70)
print('interface.xlsx 직접 확인 - 문제 발전기들')
print('='*70)

problem_regions = [
    ('지역_BSN', ['Coal', 'Hydro']),
    ('지역_CBD', ['Hydro']),
    ('지역_CND', ['Coal']),
    ('지역_DGU', ['Coal', 'Hydro'])
]

for sheet_name, gen_names in problem_regions:
    ws = wb[sheet_name]
    
    print(f'\n[{sheet_name}]')
    
    # 발전기 섹션 찾기
    for row in range(1, min(200, ws.max_row + 1)):
        cell = ws.cell(row=row, column=1)
        if cell.value and '발전기' in str(cell.value):
            header_row = row + 2
            
            # 헤더 읽기
            headers = []
            for i in range(1, 30):
                h = ws.cell(row=header_row, column=i).value
                if h:
                    headers.append(h)
                else:
                    break
            
            # carrier 컬럼 위치
            carrier_col = None
            for i, h in enumerate(headers):
                if '캐리어' in str(h):
                    carrier_col = i + 1
                    break
            
            # 데이터 행 읽기 - 모든 발전기 출력
            print(f'  헤더: {headers[:8]}')
            for data_row in range(header_row+1, min(header_row+20, ws.max_row+1)):
                name = ws.cell(row=data_row, column=1).value
                if not name:
                    break
                
                carrier = ws.cell(row=data_row, column=carrier_col).value if carrier_col else None
                p_nom = ws.cell(row=data_row, column=4).value  # 4번째 컬럼은 p_nom
                print(f'  {name}: carrier={carrier}, p_nom={p_nom}')
            
            break

print('\n' + '='*70)
