import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('interface.xlsx', data_only=True)
print("=== Unit Cost Projection 시트 ===")
ws = wb.worksheets[0]
print('시트명:', repr(ws.title))
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 30:
        break
    vals = [str(v)[:30] if v is not None else '' for v in row[:10]]
    if any(v.strip() for v in vals):
        print(f'  행{i+1}: {vals}')

print()
print("=== 발전_BSN 시트 헤더 행 찾기 ===")
for ws in wb.worksheets:
    if 'BSN' in ws.title and ws.title.startswith('\ubc1c\uc804'):
        print('시트:', ws.title)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 80:
                break
            vals = [str(v)[:25] if v is not None else '' for v in row[:16]]
            if any(v.strip() for v in vals):
                # 헤더 행이나 데이터 행 출력
                first = vals[0].strip()
                if any(k in ' '.join(vals) for k in ['\uc774\ub984', '\ube44\uc6a9', 'cost', '\ud55c\uacc4', '\uc790\ubcf8', 'MW', '\uc6d0', '\uc8fc\uc218']):
                    print(f'  행{i+1}: {vals}')
        break
