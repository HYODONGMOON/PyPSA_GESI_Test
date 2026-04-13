import openpyxl
import pandas as pd

print("=" * 80)
print("1. interface.xlsx의 Coal과 LNG committable 설정 확인")
print("=" * 80)

wb = openpyxl.load_workbook('interface.xlsx', data_only=True)

if '지역_CND' in wb.sheetnames:
    ws = wb['지역_CND']
    
    print("\n지역_CND 발전기 섹션 (Row 15~30):")
    
    for row in range(15, 31):
        cell_a = ws.cell(row, 1).value
        cell_b = ws.cell(row, 2).value
        cell_c = ws.cell(row, 3).value
        cell_d = ws.cell(row, 4).value
        cell_e = ws.cell(row, 5).value
        cell_f = ws.cell(row, 6).value
        
        # 중요한 행만 출력
        if cell_a or (row >= 20 and row <= 28):
            print(f"  Row {row:2d}: A={cell_a} | B={cell_b} | C={cell_c} | D={cell_d} | E={cell_e} | F={cell_f}")

wb.close()

print("\n" + "=" * 80)
print("2. 권장 사항")
print("=" * 80)

print("""
문제 분석:
- Committable=True인 발전기는 최소 가동 제약(p_min_pu)과 기동정지 비용이 적용됨
- 석탄과 LNG가 모두 Committable=True이면 최적화가 복잡해져 infeasible 발생 가능
- LNG만 Committable=False로 하면 석탄이 가동되지 않을 수 있음 (경제성 때문)

해결 방안:
1. 석탄: Committable=False로 변경 (기본 운영)
2. LNG: Committable=False 유지 (유연한 운영)
   
또는

1. 석탄: Committable=True, p_min_pu=0.4 정도로 설정
2. LNG: Committable=False (더 유연하게)
3. 석탄 marginal_cost < LNG marginal_cost 확인 (석탄이 더 싸야 함)
""")
