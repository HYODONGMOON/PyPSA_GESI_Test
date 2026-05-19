#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PyPSA GESI  —  시나리오 입력 인터페이스
=====================================================
선택 흐름:
  ① 구분         → 지역설정 / 지역간연결 / 제약조건
  ② (지역설정)   → 지역 선택
  ③              → 카테고리 선택
  ④              → 기술 선택
  ⑤              → 변수 입력 (현재: 용량)

  지역간연결 선택 시 → 스크롤 테이블 (연결별 이용률/한도 입력)
"""

import tkinter as tk
from tkinter import ttk, messagebox
import os, sys, math, warnings

warnings.filterwarnings("ignore", category=UserWarning)

try:
    import openpyxl
except ImportError:
    _r = tk.Tk(); _r.withdraw()
    messagebox.showerror("모듈 없음", "openpyxl 필요:\n  pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────
# 경로 / 상수
# ─────────────────────────────────────────────────────────────────────
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
INTERFACE_PATH  = os.path.join(BASE_DIR, "interface.xlsx")
INTERFACE1_SHEET = "인터페이스_1"
SHEET_NAME      = "인터페이스"
PYPSA_GUI_PATH  = os.path.join(BASE_DIR, "PyPSA_GUI.py")

# 상위 구분
SECTION_KEYS   = ["region_settings", "connections", "constraints"]
SECTION_LABELS = {
    "region_settings": "지역 설정",
    "connections"    : "지역간 연결",
    "constraints"    : "제약 조건",
}

CATEGORY_KEYS   = ["Generators", "Storages", "Links", "Demand"]
CATEGORY_LABELS = {
    "Generators": "발전설비  (Generators)",
    "Storages"  : "저장설비  (Storages)",
    "Links"     : "링크설비  (Links)",
    "Demand"    : "수요      (Demand)",
}
UNIT_MAP = {
    "Generators": "MW",
    "Storages"  : "MWh",
    "Links"     : "MW",
    "Demand"    : "GWh/yr",
}
VARIABLE_DEFS = [
    ("capacity",     "설치 용량",  "",       True),
    ("marginal_cost","한계 비용",  "원/MWh", False),
    ("capital_cost", "자본 비용",  "원/MW",  False),
    ("efficiency",   "효율",       "%",      False),
]

# 색상 팔레트
C_BLUE   = "#1565C0"
C_GREEN  = "#1B5E20"
C_ORANGE = "#E65100"
C_RED    = "#B71C1C"
C_PURPLE = "#4A148C"
C_TEAL   = "#006064"
C_CHG    = "#FFF9C4"
C_CALC   = "#E8F5E9"   # 계산값 배경
FONT     = "Malgun Gothic"

# 인터페이스_1 지역간연결 범위
# 컬럼 레이아웃: B=출발, C=도착, D=용량(무시), E=이용률(%), F=한도(시간)
CONN_HEADER_ROW = 53
CONN_DATA_START = 54
CONN_DATA_END   = 75
CONN_COL_FROM   = 2   # B
CONN_COL_TO     = 3   # C
CONN_COL_CAP    = 4   # D  용량 (읽기 전용)
CONN_COL_UTIL   = 5   # E  이용률(%)
CONN_COL_LIMIT  = 6   # F  한도(시간)

# 선로 데이터 범위 (행5~51)
LINE_DATA_START  = 6
LINE_DATA_END    = 51
LINE_COL_NAME    = 1   # A
LINE_COL_FROM    = 2   # B
LINE_COL_TO      = 4   # D  (to_region, not bus)
LINE_COL_SNOM    = 6   # F


# ─────────────────────────────────────────────────────────────────────
# 데이터 모델 — 지역 설정 (기존)
# ─────────────────────────────────────────────────────────────────────
class SheetData:
    """'인터페이스' 시트 읽기/쓰기"""

    # 화면 표시명 → 지역 코드 매핑
    # 인터페이스 시트 헤더(통합 지역명 포함) → integrated_input_data 기준 지역 코드
    DISP_TO_CODE: dict = {
        # 인터페이스 시트 통합 지역명 (D~N열 헤더)
        '부산울산':     'BSN',
        '충북':         'CBD',
        '대전세종충남': 'CND',
        '대구경북':     'GBD',
        '서울경기':     'GGD',
        '경남':         'GND',
        '강원':         'GWD',
        '인천':         'ICN',
        '전북':         'JBD',
        '제주':         'JJD',
        '광주전남':     'JND',
        # 개별 지역명 (이전 호환)
        '부산': 'BSN', '釜山': 'BSN',
        '경기': 'GGD', '경기도': 'GGD', '경기·서울': 'GGD',
        '仁川': 'ICN',
        '강원도': 'GWD',
        '충청북도': 'CBD', '충청북도권': 'CBD',
        '충남': 'CND', '충청남도': 'CND', '충청남도권': 'CND',
        '전라북도': 'JBD', '전라북도권': 'JBD',
        '전남': 'JND', '전라남도': 'JND', '전라남도권': 'JND',
        '경북': 'GBD', '경상북도': 'GBD', '경상북도권': 'GBD',
        '경상남도': 'GND', '경상남도권': 'GND',
        '제주도': 'JJD', '제주특별자치도': 'JJD',
        # 코드 직접 사용
        'BSN': 'BSN', 'GGD': 'GGD', 'ICN': 'ICN', 'GWD': 'GWD',
        'CBD': 'CBD', 'CND': 'CND', 'JBD': 'JBD', 'JND': 'JND',
        'GBD': 'GBD', 'GND': 'GND', 'JJD': 'JJD',
    }

    # 카테고리 → integrated_input_data 시트명 / 용량 컬럼명 매핑
    CAT_SHEET = {
        'Generators': ('generators', 'p_nom'),
        'Storages':   ('stores',     'e_nom'),
        'Links':      ('links',      'p_nom'),
        'Demand':     ('loads',      'p_set'),
    }

    def __init__(self):
        self.regions:  list = []
        self.items:    dict = {k: [] for k in CATEGORY_KEYS}
        self.data:     dict = {}
        self.cells:    dict = {}

    def load(self, path: str = INTERFACE_PATH):
        """
        인터페이스 시트 로드.

        인터페이스 시트 구조:
        - 1행:    헤더 [None, 2030, None, '부산울산', '충북', ...]
        - 2~29행: 요약 섹션 (INDEX/MATCH 배열수식 → openpyxl data_only=True 시 None 반환)
        - 30행:   빈 행
        - 31행:   연도 헤더 [None, None, None, 2030, 2035, ...]
        - 32행~:  데이터 섹션 [None, 지역명, 기술, 값_2030, 값_2035, ...]

        읽기:  데이터 섹션(32행~)에서 실제 값 취득
        쓰기:  데이터 섹션(32행~)에 기록 → INDEX/MATCH 수식이 자동으로 참조
        """
        wb = openpyxl.load_workbook(path, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            raise KeyError(f"'{SHEET_NAME}' 시트를 찾을 수 없습니다.")
        ws = wb[SHEET_NAME]
        all_rows = list(ws.iter_rows(values_only=True))

        # ── 1행: 지역 표시명 → 열 인덱스(1-based) ──────────────────────
        hdr = all_rows[0] if all_rows else []
        for c0, val in enumerate(hdr):
            if c0 >= 3 and val is not None and str(val).strip():
                self.regions.append((str(val).strip(), c0 + 1))

        if len(all_rows) < 31:
            return

        # ── 2~29행: 요약 섹션 → 기술명·카테고리 목록 추출 ─────────────
        cur_cat = None
        for r_idx, row in enumerate(all_rows[1:29], start=2):  # 2~29행(1-based)
            col_a = row[0]   # '용량' / '부하' / None
            col_b = row[1]   # 카테고리 또는 None
            col_c = row[2]   # 기술명
            if col_b is not None and str(col_b).strip() in CATEGORY_KEYS:
                cur_cat = str(col_b).strip()
            if col_a is not None and cur_cat is not None and col_c is not None:
                tech = str(col_c).strip()
                if tech and tech not in [n for n, _ in self.items[cur_cat]]:
                    self.items[cur_cat].append((tech, r_idx))

        # ── 31행: 연도 헤더 → 목표 연도 컬럼(0-based) 결정 ────────────
        year_hdr = all_rows[30]   # 0-based index 30 = Excel row 31

        target_year = 2030
        b1 = all_rows[0][1] if len(all_rows[0]) > 1 else None
        if isinstance(b1, (int, float)):
            target_year = int(b1)

        year_col_map = {}   # year(int) → 0-based 컬럼 인덱스
        for c0, val in enumerate(year_hdr):
            if isinstance(val, (int, float)):
                year_col_map[int(val)] = c0

        if target_year not in year_col_map and year_col_map:
            target_year = min(year_col_map.keys(),
                              key=lambda y: abs(y - target_year))
        target_col_0 = year_col_map.get(target_year, 3)   # 기본 D열(0-based=3)
        target_col_1 = target_col_0 + 1                   # 1-based (openpyxl 쓰기용)

        # 유효 지역 집합
        valid_regions = {disp for disp, _ in self.regions}

        # ── 32행~: 데이터 섹션 → 실제 값 취득 ─────────────────────────
        for r_idx, row in enumerate(all_rows[31:], start=32):  # Excel row 32~
            col_b = row[1] if len(row) > 1 else None
            col_c = row[2] if len(row) > 2 else None
            if col_b is None or col_c is None:
                continue
            region_disp = str(col_b).strip()
            tech        = str(col_c).strip()
            if not region_disp or not tech:
                continue
            if region_disp not in valid_regions:
                continue

            raw = row[target_col_0] if target_col_0 < len(row) else None
            try:
                num = float(raw) if raw is not None else 0.0
            except (TypeError, ValueError):
                num = 0.0

            key = (tech, region_disp)
            self.data[key]  = num
            # 쓰기 위치 = 데이터 섹션 행(r_idx), 목표 연도 열(target_col_1)
            self.cells[key] = (r_idx, target_col_1)

    def save_values(self, updates: dict, path: str = INTERFACE_PATH):
        wb = openpyxl.load_workbook(path)
        ws = wb[SHEET_NAME]
        for key, new_v in updates.items():
            pos = self.cells.get(key)
            if pos:
                r, c = pos
                ws.cell(row=r, column=c).value = new_v
                self.data[key] = new_v
        wb.save(path)


# ─────────────────────────────────────────────────────────────────────
# 데이터 모델 — 지역간 연결 (신규)
# ─────────────────────────────────────────────────────────────────────
class ConnectionData:
    """
    '인터페이스_1' 시트의 선로 데이터(행6~51)와
    지역간연결 입력표(행53~75)를 파싱합니다.

    connections: [
        {
          'from':       'GWD',
          'to':         'GGD',
          'row':        54,              ← Excel row
          'total_cap':  7299.9,          ← 해당 연결 s_nom 합계 (MW)
          'util_rate':  0.0,             ← 이용률 (%, 엑셀 D열)
          'time_limit': 0.0,             ← 한도 (시간/yr, 엑셀 E열)
          'lines':      ['GWD_GGD1', ...] ← 해당 연결의 개별 선로 목록
        }, ...
    ]
    """
    def __init__(self):
        self.connections: list = []
        self._line_s_nom: dict = {}   # line_name → s_nom

    def load(self, path: str = INTERFACE_PATH):
        self.connections.clear()
        self._line_s_nom.clear()

        if not os.path.exists(path):
            raise FileNotFoundError(f"파일 없음: {path}")
        wb = openpyxl.load_workbook(path, data_only=True)
        if INTERFACE1_SHEET not in wb.sheetnames:
            raise KeyError(f"'{INTERFACE1_SHEET}' 시트를 찾을 수 없습니다.")
        ws = wb[INTERFACE1_SHEET]

        # ── A1에서 목표 연도 읽기
        try:
            target_year = int(ws['A1'].value)
        except (TypeError, ValueError):
            target_year = None

        # ── 행5 헤더에서 목표 연도 컬럼 인덱스 찾기 (I=2024, J=2030, ...)
        snom_col_idx = None   # 0-based index within full row
        if target_year is not None:
            hdr_row = list(ws[5])
            for cell in hdr_row:
                try:
                    if int(cell.value) == target_year:
                        snom_col_idx = cell.column - 1   # 0-based
                        break
                except (TypeError, ValueError):
                    pass

        # ── 개별 선로 s_nom 집계 (행6~51)
        # 컬럼: A=name, B=from_region, D=to_region, 연도열=s_nom
        line_by_pair: dict = {}   # (from, to) → [(line_name, s_nom)]
        for row in ws.iter_rows(min_row=LINE_DATA_START, max_row=LINE_DATA_END):
            def _v(c_idx): return row[c_idx - 1].value
            name  = _v(LINE_COL_NAME)
            frm   = _v(LINE_COL_FROM)
            to    = _v(LINE_COL_TO)
            # 연도별 컬럼 우선 사용, 없으면 F열(LINE_COL_SNOM) 폴백
            if snom_col_idx is not None and snom_col_idx < len(row):
                s_nom = row[snom_col_idx].value
            else:
                s_nom = _v(LINE_COL_SNOM)
            if not name or not frm or not to:
                continue
            try:
                s_nom_f = float(s_nom) if s_nom is not None else 0.0
            except (TypeError, ValueError):
                s_nom_f = 0.0
            name_s = str(name).strip()
            frm_s  = str(frm).strip()
            to_s   = str(to).strip()
            self._line_s_nom[name_s] = s_nom_f
            key = (frm_s, to_s)
            line_by_pair.setdefault(key, []).append((name_s, s_nom_f))

        # ── 지역간연결 입력표 (행53~75)
        # B=출발, C=도착, D=용량(Excel 수식으로 자동 표시), E=이용률, F=한도
        for row in ws.iter_rows(min_row=CONN_DATA_START, max_row=CONN_DATA_END,
                                min_col=CONN_COL_FROM, max_col=CONN_COL_LIMIT):
            frm_cell, to_cell, cap_cell, util_cell, limit_cell = row
            frm_v = frm_cell.value
            to_v  = to_cell.value
            if not frm_v or not to_v:
                continue
            frm_s = str(frm_v).strip()
            to_s  = str(to_v).strip()
            try:
                util  = float(util_cell.value)  if util_cell.value  is not None else 0.0
            except (TypeError, ValueError):
                util = 0.0
            try:
                limit = float(limit_cell.value) if limit_cell.value is not None else 0.0
            except (TypeError, ValueError):
                limit = 0.0

            # total_cap: D열(Excel 수식 집계값) 우선, 없으면 개별 선로 합산
            try:
                cap_d = float(cap_cell.value) if cap_cell.value is not None else None
            except (TypeError, ValueError):
                cap_d = None

            if cap_d is not None and cap_d > 0:
                total_cap = cap_d
            else:
                key_fw   = (frm_s, to_s)
                key_rv   = (to_s, frm_s)
                lines_fw = line_by_pair.get(key_fw, [])
                lines_rv = line_by_pair.get(key_rv, [])
                total_cap = sum(s for _, s in lines_fw + lines_rv)

            key      = (frm_s, to_s)
            rev_key  = (to_s, frm_s)
            lines_fw = line_by_pair.get(key, [])
            lines_rv = line_by_pair.get(rev_key, [])
            all_lines = lines_fw + lines_rv

            self.connections.append({
                'from':       frm_s,
                'to':         to_s,
                'row':        frm_cell.row,
                'total_cap':  total_cap,
                'util_rate':  util,
                'time_limit': limit,
                'lines':      [n for n, _ in all_lines],
            })

    def save_values(self, updates: dict, path: str = INTERFACE_PATH):
        """
        updates = { row_num: {'util_rate': x, 'time_limit': y} }
        """
        wb = openpyxl.load_workbook(path)
        ws = wb[INTERFACE1_SHEET]
        for row_num, vals in updates.items():
            if 'util_rate' in vals:
                ws.cell(row=row_num, column=CONN_COL_UTIL ).value = vals['util_rate']
            if 'time_limit' in vals:
                ws.cell(row=row_num, column=CONN_COL_LIMIT).value = vals['time_limit']
        wb.save(path)

    @staticmethod
    def calc_applied(total_cap: float, util_rate: float):
        """
        이용률(%) 적용 결과 계산.
        Returns: (applied_cap, max_pu_normal, overload_extra_cap)
        """
        r = util_rate / 100.0
        applied   = total_cap * (1 + r)
        max_pu    = 1.0 / (1 + r) if r > 0 else 1.0
        extra_cap = total_cap * r
        return applied, max_pu, extra_cap


# ─────────────────────────────────────────────────────────────────────
# GUI 메인 클래스
# ─────────────────────────────────────────────────────────────────────
class ScenarioGUI:

    def __init__(self, master: tk.Tk):
        self.master = master
        master.title("PyPSA GESI  —  시나리오 입력 인터페이스")
        master.geometry("1400x780")
        master.resizable(True, True)
        master.configure(bg="white")
        master.bind("<Control-s>", lambda e: self._save())

        self.db   = SheetData()
        self.cdb  = ConnectionData()
        self.changed: set = set()
        self._entry_widgets: dict = {}
        self._conn_entries:  dict = {}   # row → {'util': Entry, 'limit': Entry}

        try:
            self.db.load()
            self.cdb.load()
        except FileNotFoundError as e:
            messagebox.showerror("파일 없음", str(e))
            master.after(200, master.destroy); return
        except KeyError as e:
            messagebox.showerror("시트 없음", str(e))
            master.after(200, master.destroy); return
        except Exception as e:
            messagebox.showerror("로딩 오류", str(e))
            master.after(200, master.destroy); return

        self._build_ui()

    # ── UI 구축 ──────────────────────────────────────────────────────
    def _build_ui(self):
        # 상단 타이틀
        top = tk.Frame(self.master, bg=C_BLUE, pady=9)
        top.pack(fill="x")
        tk.Label(top, text="PyPSA GESI  |  시나리오 입력 인터페이스",
                 bg=C_BLUE, fg="white", font=(FONT, 14, "bold")).pack()
        tk.Label(top,
                 text=f"파일: {os.path.basename(INTERFACE_PATH)}   ·   Ctrl+S = 저장",
                 bg=C_BLUE, fg="#BBDEFB", font=(FONT, 9)).pack()

        # 본문
        body = tk.Frame(self.master, bg="white")
        body.pack(fill="both", expand=True, padx=8, pady=6)

        # ① 구분 listbox (항상 표시)
        self.lb_section = self._make_listbox(
            body, "① 구분", [(SECTION_LABELS[k], k) for k in SECTION_KEYS],
            C_PURPLE, width=14, on_select=self._on_section_select)

        # 콘텐츠 영역 (동적 전환)
        self.frm_content = tk.Frame(body, bg="white")
        self.frm_content.pack(side="left", fill="both", expand=True)

        # ── 지역설정 패널 ────────────────────────────────────────────
        self.panel_region = tk.Frame(self.frm_content, bg="white")
        self.lb_region = self._make_listbox(
            self.panel_region, "② 지역 선택", self.db.regions,
            C_BLUE, width=16, on_select=self._on_region_select)
        cat_items = [(CATEGORY_LABELS[k], k) for k in CATEGORY_KEYS]
        self.lb_cat = self._make_listbox(
            self.panel_region, "③ 카테고리", cat_items,
            C_GREEN, width=22, on_select=self._on_cat_select)
        self.lb_tech = self._make_listbox(
            self.panel_region, "④ 기술 선택", [],
            C_ORANGE, width=18, on_select=self._on_tech_select)
        self._build_input_panel(self.panel_region)

        # ── 지역간연결 패널 ──────────────────────────────────────────
        self.panel_conn = tk.Frame(self.frm_content, bg="white")
        self._build_connection_panel(self.panel_conn)

        # ── 제약조건 패널 (placeholder) ──────────────────────────────
        self.panel_const = tk.Frame(self.frm_content, bg="white")
        tk.Label(self.panel_const,
                 text="제약 조건 입력 (향후 구현 예정)",
                 fg="#AAA", font=(FONT, 13, "italic"), bg="white").pack(expand=True)

        # 하단 버튼 바
        bot = tk.Frame(self.master, bg="#F0F4F8", relief="ridge", bd=1, pady=7)
        bot.pack(fill="x", side="bottom")
        def btn(parent, text, cmd, bg):
            return tk.Button(parent, text=text, command=cmd, bg=bg, fg="white",
                             font=(FONT, 10, "bold"), relief="flat", cursor="hand2",
                             activebackground=bg, padx=14, pady=6)
        btn(bot, "💾  저장  (Ctrl+S)", self._save,   C_BLUE  ).pack(side="left",  padx=8)
        btn(bot, "▶  분석 실행",        self._run,    C_GREEN ).pack(side="left",  padx=4)
        btn(bot, "↺  새로고침",          self._reload, C_ORANGE).pack(side="left",  padx=4)
        btn(bot, "✕  닫기",              self.master.destroy, C_RED).pack(side="right", padx=8)
        self.lbl_status = tk.Label(bot, text="준비", bg="#F0F4F8", fg="#444",
                                   font=(FONT, 9))
        self.lbl_status.pack(side="right", padx=14)

        # 초기 선택: 지역설정
        self.lb_section.selection_set(0)
        self._on_section_select()
        self.lb_cat.selection_set(0)
        self._on_cat_select()

    # ── 재사용 리스트박스 ────────────────────────────────────────────
    def _make_listbox(self, parent, title, items, sel_color,
                      width=18, on_select=None):
        frm = tk.LabelFrame(parent, text=f"  {title}  ",
                             font=(FONT, 10, "bold"), bg="white", padx=4, pady=4)
        frm.pack(side="left", fill="y", padx=(0, 5))
        lb = tk.Listbox(frm, width=width, height=24,
                        font=(FONT, 10), selectmode="single", activestyle="dotbox",
                        selectbackground=sel_color, selectforeground="white",
                        relief="flat", bd=1, exportselection=False)
        sb = ttk.Scrollbar(frm, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        lb.pack(side="left", fill="both", expand=True)
        for entry in items:
            lb.insert(tk.END, entry[0] if isinstance(entry, tuple) else entry)
        if on_select:
            lb.bind("<<ListboxSelect>>", lambda e: on_select())
        return lb

    # ── 입력 변수 패널 (지역설정용) ──────────────────────────────────
    def _build_input_panel(self, parent):
        self.frm_input = tk.LabelFrame(parent, text="  ⑤ 입력 변수  ",
                                        font=(FONT, 10, "bold"),
                                        bg="white", padx=10, pady=8)
        self.frm_input.pack(side="left", fill="both", expand=True)
        self.lbl_path = tk.Label(self.frm_input,
                                  text="← 지역 / 카테고리 / 기술을 순서대로 선택하세요",
                                  fg="#888", font=(FONT, 10, "italic"),
                                  bg="white", anchor="w")
        self.lbl_path.pack(fill="x", pady=(0, 10))
        ttk.Separator(self.frm_input, orient="horizontal").pack(fill="x", pady=4)
        self.frm_vars = tk.Frame(self.frm_input, bg="white")
        self.frm_vars.pack(fill="both", expand=True, pady=6)
        self._show_placeholder()

    def _show_placeholder(self):
        for w in self.frm_vars.winfo_children():
            w.destroy()
        self._entry_widgets.clear()
        tk.Label(self.frm_vars,
                 text="기술을 선택하면 입력 변수가 표시됩니다.",
                 fg="#999", font=(FONT, 10, "italic"), bg="white").pack(pady=20)

    # ── 지역간연결 패널 ──────────────────────────────────────────────
    def _build_connection_panel(self, parent):
        # 안내 레이블
        info_frm = tk.Frame(parent, bg="white")
        info_frm.pack(fill="x", padx=10, pady=(8, 2))
        tk.Label(info_frm,
                 text="  지역간 송전망  —  이용률(%) 및 한도(시간/yr) 입력",
                 bg=C_TEAL, fg="white", font=(FONT, 11, "bold"),
                 anchor="w", pady=6, padx=10).pack(fill="x")
        tk.Label(info_frm,
                 text=(
                     "• 이용률(%): 기본용량 대비 증설비율.  10 입력 → 설비용량 ×1.10,  정상 최대이용률 = 1/1.10 ≈ 90.9%\n"
                     "• 한도(시간): 증설분(이용률 초과분)을 사용할 수 있는 연간 시간 제한."
                 ),
                 fg="#444", bg="#E0F2F1", font=(FONT, 9),
                 anchor="w", justify="left", pady=4, padx=10).pack(fill="x")

        # 스크롤 캔버스
        canvas = tk.Canvas(parent, bg="white", highlightthickness=0)
        vsb    = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y", padx=(0, 8))
        canvas.pack(fill="both", expand=True, padx=10, pady=(0, 4))

        inner = tk.Frame(canvas, bg="white")
        canvas_window = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_resize(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind("<Configure>", _on_resize)
        inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(
            int(-1 * (e.delta / 120)), "units"))

        # ── grid 기반 테이블 (헤더 + 데이터 행 동일 컨테이너) ────────────
        # 컬럼 정의: (헤더 텍스트, 픽셀 최소 너비, anchor)
        COLS = [
            ("연결  (출발 → 도착)",   220, "w"),
            ("현재 총 용량\n(MW)",     120, "e"),
            ("이용률\n(%)",             80, "center"),
            ("한도\n(시간/yr)",         95, "center"),
            ("적용 용량\n(MW) ▶",      120, "e"),
            ("정상 최대\n이용률",       100, "center"),
        ]
        NUM_COLS = len(COLS)

        tbl = tk.Frame(inner, bg="white")
        tbl.pack(fill="x", pady=(2, 0))
        for ci, (_, px, _) in enumerate(COLS):
            tbl.columnconfigure(ci, minsize=px, weight=0)

        # 헤더 행 (row 0)
        for ci, (txt, _, _) in enumerate(COLS):
            tk.Label(tbl, text=txt,
                     bg="#546E7A", fg="white",
                     font=(FONT, 9, "bold"),
                     anchor="center", justify="center",
                     pady=5, padx=4).grid(
                row=0, column=ci, sticky="nsew",
                padx=(0, 1), pady=(0, 2))

        # 구분선
        ttk.Separator(tbl, orient="horizontal").grid(
            row=1, column=0, columnspan=NUM_COLS, sticky="ew", pady=0)

        # 데이터 행
        self._conn_entries.clear()
        for idx, conn in enumerate(self.cdb.connections):
            row = idx + 2          # 0=헤더, 1=구분선, 2~=데이터
            bg_row = "#FAFAFA" if idx % 2 == 0 else "#F0F0F0"

            # 열 0: 연결명
            tk.Label(tbl, text=f"{conn['from']}  →  {conn['to']}",
                     bg=bg_row, font=(FONT, 10, "bold"),
                     anchor="w", padx=10, pady=4).grid(
                row=row, column=0, sticky="nsew")

            # 열 1: 현재 총 용량
            tk.Label(tbl, text=f"{conn['total_cap']:,.1f}",
                     bg=bg_row, font=(FONT, 10),
                     anchor="e", padx=10).grid(
                row=row, column=1, sticky="nsew")

            # 열 2: 이용률 입력
            ent_util = tk.Entry(tbl, font=(FONT, 10),
                                justify="right", relief="solid", bd=1)
            util_init = f"{conn['util_rate']:.1f}" if conn['util_rate'] else ""
            ent_util.insert(0, util_init)
            ent_util.grid(row=row, column=2, sticky="ew", padx=6, pady=3)

            # 열 3: 한도 입력
            ent_lim = tk.Entry(tbl, font=(FONT, 10),
                               justify="right", relief="solid", bd=1)
            lim_init = f"{conn['time_limit']:.0f}" if conn['time_limit'] else ""
            ent_lim.insert(0, lim_init)
            ent_lim.grid(row=row, column=3, sticky="ew", padx=6, pady=3)

            # 열 4: 적용 용량 (계산값, 이용률 미입력 시 빈 칸)
            lbl_cap = tk.Label(tbl, text="",
                               bg=C_CALC, font=(FONT, 10),
                               anchor="e", padx=10)
            lbl_cap.grid(row=row, column=4, sticky="nsew")

            # 열 5: 정상 최대 이용률 (계산값, 이용률 미입력 시 빈 칸)
            lbl_mpu = tk.Label(tbl, text="",
                               bg=C_CALC, font=(FONT, 10),
                               anchor="center", padx=4)
            lbl_mpu.grid(row=row, column=5, sticky="nsew")

            row_key = conn['row']
            self._conn_entries[row_key] = {
                'util':      ent_util,
                'limit':     ent_lim,
                'lbl_cap':   lbl_cap,
                'lbl_mpu':   lbl_mpu,
                'total_cap': conn['total_cap'],
            }
            # 이용률이 있을 때만 초기 계산값 표시
            self._update_conn_calc(row_key)
            ent_util.bind("<KeyRelease>",
                          lambda e, rk=row_key: self._update_conn_calc(rk))
            ent_util.bind("<FocusOut>",
                          lambda e, rk=row_key: self._update_conn_calc(rk))

        # 저장 버튼
        save_frm = tk.Frame(parent, bg="white", pady=6)
        save_frm.pack(fill="x", padx=10)
        tk.Button(save_frm, text="💾  지역간 연결 설정 저장",
                  command=self._save_connections,
                  bg=C_TEAL, fg="white", font=(FONT, 10, "bold"),
                  relief="flat", cursor="hand2", padx=14, pady=6).pack(side="left")

    def _update_conn_calc(self, row_key: int):
        """이용률 입력값 기반으로 계산결과 레이블 업데이트.
        이용률이 0 또는 미입력이면 빈 칸으로 표시."""
        entry_info = self._conn_entries.get(row_key)
        if not entry_info:
            return
        try:
            util = float(entry_info['util'].get().strip() or "0")
        except ValueError:
            util = 0.0
        if util <= 0:
            entry_info['lbl_cap'].config(text="")
            entry_info['lbl_mpu'].config(text="")
            return
        total_cap = entry_info['total_cap']
        applied, max_pu, _ = ConnectionData.calc_applied(total_cap, util)
        entry_info['lbl_cap'].config(text=f"{applied:,.1f}")
        entry_info['lbl_mpu'].config(text=f"{max_pu*100:.1f}%")

    # ── 패널 전환 ────────────────────────────────────────────────────
    def _on_section_select(self):
        idx = self.lb_section.curselection()
        if not idx:
            return
        key = SECTION_KEYS[idx[0]]

        self.panel_region.pack_forget()
        self.panel_conn.pack_forget()
        self.panel_const.pack_forget()

        if key == "region_settings":
            self.panel_region.pack(fill="both", expand=True)
        elif key == "connections":
            self.panel_conn.pack(fill="both", expand=True)
        else:
            self.panel_const.pack(fill="both", expand=True)

    # ── 지역설정 이벤트 ──────────────────────────────────────────────
    def _on_region_select(self):
        self._refresh_tech_list()
        self._show_placeholder()
        self._update_path_label()

    def _on_cat_select(self):
        self.lb_tech.selection_clear(0, tk.END)
        self._refresh_tech_list()
        self._show_placeholder()
        self._update_path_label()

    def _on_tech_select(self):
        if not self.lb_tech.curselection():
            return
        self._refresh_input_panel()
        self._update_path_label()

    def _refresh_tech_list(self):
        cat = self._sel_cat()
        self.lb_tech.delete(0, tk.END)
        if cat:
            for item_name, _ in self.db.items.get(cat, []):
                self.lb_tech.insert(tk.END, f"  {item_name}")

    def _refresh_input_panel(self):
        for w in self.frm_vars.winfo_children():
            w.destroy()
        self._entry_widgets.clear()

        disp, _ = self._sel_region()
        cat      = self._sel_cat()
        item     = self._sel_item()
        if not disp or not cat or not item:
            self._show_placeholder(); return

        unit       = UNIT_MAP.get(cat, "")
        is_changed = (item, disp) in self.changed

        row_idx = 0
        for var_key, var_label, var_unit, active in VARIABLE_DEFS:
            effective_unit = unit if var_key == "capacity" else var_unit
            fg_lbl = "#222" if active else "#AAAAAA"
            fg_ent = "#222" if active else "#BBBBBB"
            bg_ent = (C_CHG if is_changed and active else
                      "#FAFAFA" if active else "#F5F5F5")

            tk.Label(self.frm_vars,
                     text=f"{var_label}  ({effective_unit})" if effective_unit else var_label,
                     fg=fg_lbl, bg="white",
                     font=(FONT, 11, "bold" if active else "normal"),
                     anchor="w").grid(row=row_idx, column=0, columnspan=2,
                                      sticky="w", pady=(10 if row_idx == 0 else 16, 2))
            row_idx += 1

            if active:
                cur_val = self.db.data.get((item, disp), 0.0)
                tk.Label(self.frm_vars,
                         text=f"현재값:   {cur_val:>14,.2f}  {effective_unit}",
                         fg="#555", bg="white",
                         font=(FONT, 10)).grid(row=row_idx, column=0, columnspan=2,
                                               sticky="w", padx=4, pady=(0, 4))
                row_idx += 1
                tk.Label(self.frm_vars, text="수정값:", bg="white",
                         font=(FONT, 10), fg=fg_lbl).grid(
                    row=row_idx, column=0, sticky="w", padx=4)
                ent = tk.Entry(self.frm_vars, font=(FONT, 12), width=20,
                               justify="right", relief="solid", bd=1,
                               bg=bg_ent, fg=fg_ent)
                ent.insert(0, f"{cur_val:.2f}")
                ent.grid(row=row_idx, column=1, sticky="ew", padx=(4, 0), pady=2)
                ent.focus_set()
                self._entry_widgets[var_key] = ent
                row_idx += 1
                tk.Label(self.frm_vars, text=effective_unit,
                         fg="#777", bg="white", font=(FONT, 9)).grid(
                    row=row_idx, column=1, sticky="e", padx=4)
                row_idx += 1
            else:
                tk.Label(self.frm_vars,
                         text="(향후 추가 예정)",
                         fg="#AAAAAA", bg="white",
                         font=(FONT, 9, "italic")).grid(
                    row=row_idx, column=0, columnspan=2,
                    sticky="w", padx=4, pady=(0, 2))
                row_idx += 1

            ttk.Separator(self.frm_vars, orient="horizontal").grid(
                row=row_idx, column=0, columnspan=2,
                sticky="ew", pady=4)
            row_idx += 1

        self.frm_vars.columnconfigure(1, weight=1)

    def _update_path_label(self):
        disp, _ = self._sel_region()
        cat      = self._sel_cat()
        item     = self._sel_item()
        parts = []
        if disp:  parts.append(disp)
        if cat:   parts.append(CATEGORY_LABELS[cat].split("(")[0].strip())
        if item:  parts.append(item)
        if len(parts) == 3:
            txt = f"  📍  {parts[0]}   ›   {parts[1]}   ›   {parts[2]}"
            fg  = C_BLUE; fnt = (FONT, 10, "bold")
        else:
            txt = "← 지역 / 카테고리 / 기술을 순서대로 선택하세요"
            fg  = "#888";  fnt = (FONT, 10, "italic")
        self.lbl_path.config(text=txt, fg=fg, font=fnt)

    # ── 선택값 추출 ──────────────────────────────────────────────────
    def _sel_region(self):
        idx = self.lb_region.curselection()
        return self.db.regions[idx[0]] if idx else (None, None)

    def _sel_cat(self):
        idx = self.lb_cat.curselection()
        return CATEGORY_KEYS[idx[0]] if idx else None

    def _sel_item(self):
        idx = self.lb_tech.curselection()
        if not idx: return None
        return self.lb_tech.get(idx[0]).strip() or None

    # ── 저장 (지역설정) ──────────────────────────────────────────────
    def _save(self, _event=None):
        idx = self.lb_section.curselection()
        section = SECTION_KEYS[idx[0]] if idx else "region_settings"

        if section == "connections":
            self._save_connections(); return

        disp, _ = self._sel_region()
        cat      = self._sel_cat()
        item     = self._sel_item()
        if not disp or not cat or not item:
            messagebox.showwarning("선택 없음",
                                   "지역, 카테고리, 기술을 모두 선택한 후 저장하세요."); return
        ent = self._entry_widgets.get("capacity")
        if ent is None:
            messagebox.showwarning("입력창 없음", "입력 변수를 찾을 수 없습니다."); return
        raw = ent.get().strip().replace(",", "")
        try:
            new_val = float(raw)
        except ValueError:
            messagebox.showwarning("입력 오류",
                                   f"'{ent.get()}' 은(는) 유효한 숫자가 아닙니다."); return
        cur_val = self.db.data.get((item, disp), 0.0)
        if abs(new_val - cur_val) < 1e-9:
            messagebox.showinfo("변경 없음", "값이 변경되지 않았습니다."); return
        unit    = UNIT_MAP.get(cat, "")
        cat_lbl = CATEGORY_LABELS[cat].split("(")[0].strip()
        if not messagebox.askyesno("저장 확인",
                                    f"지역:  {disp}\n항목:  {cat_lbl}  ›  {item}\n\n"
                                    f"   {cur_val:>12,.2f}  {unit}   →   {new_val:>12,.2f}  {unit}\n\n"
                                    "저장하시겠습니까?"): return
        try:
            self.db.save_values({(item, disp): new_val})
        except PermissionError:
            messagebox.showerror("저장 실패",
                                 "파일이 Excel에서 열려 있습니다.\n닫은 후 다시 시도하세요."); return
        except Exception as e:
            messagebox.showerror("저장 실패", str(e)); return
        self.changed.add((item, disp))
        self.lbl_status.config(text=f"저장 완료  —  {item} ({disp}): {new_val:,.2f} {unit}")
        messagebox.showinfo("저장 완료",
                            f"{item}  ({disp})\n설치 용량  {new_val:,.2f} {unit}  저장 완료")
        self._refresh_input_panel()

    # ── 저장 (지역간연결) ────────────────────────────────────────────
    def _save_connections(self):
        updates = {}
        errors  = []
        changed_rows = []

        for conn in self.cdb.connections:
            row_key = conn['row']
            info    = self._conn_entries.get(row_key)
            if not info:
                continue
            try:
                new_util  = float(info['util'].get().strip()  or "0")
                new_limit = float(info['limit'].get().strip() or "0")
            except ValueError:
                errors.append(f"{conn['from']}→{conn['to']}: 숫자 형식 오류")
                continue
            if new_util < 0 or new_limit < 0:
                errors.append(f"{conn['from']}→{conn['to']}: 음수 불가"); continue

            old_util  = conn['util_rate']
            old_limit = conn['time_limit']
            if abs(new_util - old_util) > 1e-9 or abs(new_limit - old_limit) > 1e-9:
                updates[row_key] = {'util_rate': new_util, 'time_limit': new_limit}
                changed_rows.append(
                    f"  {conn['from']} → {conn['to']}: "
                    f"이용률 {old_util:.1f}→{new_util:.1f}%,  "
                    f"한도 {old_limit:.0f}→{new_limit:.0f}시간"
                )

        if errors:
            messagebox.showwarning("입력 오류", "\n".join(errors)); return
        if not updates:
            messagebox.showinfo("변경 없음", "변경된 값이 없습니다."); return

        msg = f"다음 {len(updates)}개 연결 설정을 저장합니다:\n\n"
        msg += "\n".join(changed_rows[:15])
        if len(changed_rows) > 15:
            msg += f"\n  ... 외 {len(changed_rows)-15}건"
        if not messagebox.askyesno("저장 확인", msg):
            return

        try:
            self.cdb.save_values(updates)
        except PermissionError:
            messagebox.showerror("저장 실패",
                                 "파일이 Excel에서 열려 있습니다.\n닫은 후 다시 시도하세요.")
            return
        except Exception as e:
            messagebox.showerror("저장 실패", str(e)); return

        # 내부 데이터 업데이트
        for conn in self.cdb.connections:
            if conn['row'] in updates:
                upd = updates[conn['row']]
                conn['util_rate']  = upd['util_rate']
                conn['time_limit'] = upd['time_limit']

        self.lbl_status.config(text=f"지역간연결 {len(updates)}건 저장 완료")
        messagebox.showinfo("저장 완료",
                            f"지역간 연결 설정 {len(updates)}건이 저장되었습니다.\n"
                            "분석 실행 시 자동으로 적용됩니다.")

    # ── 새로고침 ─────────────────────────────────────────────────────
    def _reload(self):
        try:
            self.db  = SheetData();  self.db.load()
            self.cdb = ConnectionData(); self.cdb.load()
        except Exception as e:
            messagebox.showerror("로딩 오류", str(e)); return
        self.changed.clear()
        self._refresh_tech_list()
        self._show_placeholder()
        # 지역간연결 패널 재구성
        for w in self.panel_conn.winfo_children():
            w.destroy()
        self._build_connection_panel(self.panel_conn)
        self.lbl_status.config(text="파일 새로고침 완료")

    # ── 분석 실행 ────────────────────────────────────────────────────
    def _run(self):
        import subprocess
        if not os.path.exists(PYPSA_GUI_PATH):
            messagebox.showerror("파일 없음",
                                 f"PyPSA_GUI.py 를 찾을 수 없습니다:\n{PYPSA_GUI_PATH}")
            return
        try:
            subprocess.Popen([sys.executable, PYPSA_GUI_PATH], cwd=BASE_DIR)
            self.lbl_status.config(text="PyPSA_GUI.py 실행 중 ...")
        except Exception as e:
            messagebox.showerror("실행 오류", str(e))


# ─────────────────────────────────────────────────────────────────────
# 진입점
# ─────────────────────────────────────────────────────────────────────
def main():
    root = tk.Tk()
    root.configure(bg="white")
    ScenarioGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
